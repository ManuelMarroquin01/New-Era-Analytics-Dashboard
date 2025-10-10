import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import os
import time
import warnings
import logging
from typing import Dict, Optional, List, Tuple, Any
from dataclasses import dataclass
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import openpyxl
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

# Configuración de logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class StockAnalysisConfig:
    """Configuración para el análisis de stock"""
    fecha_reporte: str
    colores_semaforo: Dict[str, str]
    umbrales: Dict[str, Tuple[float, float]]
    
    def __post_init__(self):
        self.fecha_reporte = datetime.now().strftime('%Y%m%d_%H%M')
        self.colores_semaforo = {
            "verde": "#28a745",
            "amarillo": "#ffc107",
            "rojo": "#dc3545",
            "gris": "#6c757d"
        }
        self.umbrales = {
            "verde": (1.0, 1.15),
            "amarillo": (1.15, float('inf')),
            "rojo": (0.0, 1.0)
        }

# Configuración inicial
warnings.filterwarnings("ignore", message="missing ScriptRunContext")
st.set_page_config(
    page_title="New Era Analytics Dashboard",
    page_icon="👑",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={
        'Get Help': 'https://www.newera.com/support',
        'Report a bug': "https://www.newera.com/support",
        'About': "# New Era Stock Analytics Dashboard\nSistema profesional de análisis de inventario"
    }
)

# Instancia de configuración
config = StockAnalysisConfig(fecha_reporte="", colores_semaforo={}, umbrales={})

class ProfessionalDesign:
    """Gestor de diseño profesional para la aplicación"""
    
    def __init__(self):
        self.primary_color = "#6b7280"  # Gris medio
        self.secondary_color = "#9ca3af"  # Gris claro
        self.accent_color = "#374151"  # Gris oscuro
        self.success_color = "#10b981"  # Verde
        self.warning_color = "#f59e0b"  # Amarillo
        self.danger_color = "#ef4444"  # Rojo
        self.neutral_color = "#e5e7eb"  # Gris muy claro
        self.background_color = "#f9fafb"  # Gris de fondo
        
    def inject_custom_css(self):
        """Inyecta CSS personalizado para el diseño profesional"""
        st.markdown("""
        <style>
        /* Import Google Fonts */
        @import url(\https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap\);
        
        /* Global Styles */
        .stApp {
            font-family: \Inter\, sans-serif;
            animation: fadeIn 0.8s ease-in-out;
        }
        
        /* Animaciones globales */
        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(20px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        @keyframes slideInFromLeft {
            from { transform: translateX(-100%); opacity: 0; }
            to { transform: translateX(0); opacity: 1; }
        }
        
        @keyframes slideInFromRight {
            from { transform: translateX(100%); opacity: 0; }
            to { transform: translateX(0); opacity: 1; }
        }
        
        @keyframes scaleIn {
            from { transform: scale(0.9); opacity: 0; }
            to { transform: scale(1); opacity: 1; }
        }
        
        @keyframes pulse {
            0%, 100% { transform: scale(1); }
            50% { transform: scale(1.05); }
        }
        
        /* Header ultra moderno y cool */
        .main-header {
            background: linear-gradient(to right, #000000 0%, #1a1a1a 50%, #0a0a0a 100%);
            padding: 5rem 0;
            border-radius: 0;
            margin: 0 -5rem 3rem -5rem;
            box-shadow: 
                0 25px 60px rgba(0, 0, 0, 0.3),
                0 10px 30px rgba(0, 0, 0, 0.2),
                inset 0 1px 0 rgba(255, 255, 255, 0.1);
            animation: headerSlideIn 1s ease-out;
            transition: all 0.5s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            overflow: hidden;
            border: none;
            backdrop-filter: blur(10px);
            min-height: 260px;
        }
        
        .main-header::before {
            content: \\;
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, 
                transparent 0%, 
                rgba(255, 255, 255, 0.08) 25%, 
                rgba(255, 255, 255, 0.15) 50%, 
                rgba(255, 255, 255, 0.08) 75%, 
                transparent 100%
            );
            animation: headerShine 8s ease-in-out infinite;
            z-index: 1;
        }
        
        @keyframes headerShine {
            0% { left: -100%; }
            15% { left: -100%; }
            50% { left: 100%; }
            85% { left: 100%; }
            100% { left: 100%; }
        }
        
        .main-header::after {
            content: \\;
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: radial-gradient(circle at 20% 80%, rgba(156, 163, 175, 0.1) 0%, transparent 50%),
                        radial-gradient(circle at 80% 20%, rgba(209, 213, 219, 0.1) 0%, transparent 50%);
            opacity: 0.6;
            animation: floatingParticles 8s ease-in-out infinite;
        }
        
        @keyframes headerSlideIn {
            from { 
                opacity: 0; 
                transform: translateY(-30px) scale(0.95); 
            }
            to { 
                opacity: 1; 
                transform: translateY(0) scale(1); 
            }
        }
        
        @keyframes floatingParticles {
            0%, 100% { 
                background: radial-gradient(circle at 20% 80%, rgba(156, 163, 175, 0.1) 0%, transparent 50%),
                            radial-gradient(circle at 80% 20%, rgba(209, 213, 219, 0.1) 0%, transparent 50%);
            }
            50% { 
                background: radial-gradient(circle at 30% 70%, rgba(156, 163, 175, 0.15) 0%, transparent 50%),
                            radial-gradient(circle at 70% 30%, rgba(209, 213, 219, 0.15) 0%, transparent 50%);
            }
        }
        
        /* Efecto hover removido para mantener encabezado fijo */
        
        /* Efecto de luz continuo - removido hover para que sea automático */
        
        .header-content {
            display: flex;
            align-items: center;
            justify-content: space-between;
            position: relative;
            z-index: 2;
            padding: 0 4rem;
        }
        
        .logo-section {
            display: flex;
            align-items: center;
            gap: 3rem;
            position: relative;
            flex: 1;
        }
        
        .logo-section::before {
            content: \\;
            position: absolute;
            left: -1rem;
            top: 50%;
            transform: translateY(-50%);
            width: 4px;
            height: 60px;
            background: linear-gradient(to bottom, transparent, #9ca3af, transparent);
            border-radius: 2px;
        }
        
        .logo-icon {
            font-size: 4rem;
            filter: drop-shadow(0 8px 16px rgba(0,0,0,0.3));
            position: relative;
            min-width: 140px;
        }
        
        .logo-section > div {
            margin-left: 1.5rem;
            flex: 1;
        }
        
        .header-title {
            color: #ffffff !important;
            font-size: 3rem;
            font-weight: 800;
            margin: 0;
            text-shadow: 0 4px 8px rgba(0,0,0,0.5);
        }
        
        
        /* Ocultar cualquier icono o flecha que pueda aparecer */
        .main-header .header-title + *:not(.header-subtitle) {
            display: none !important;
        }
        
        /* Ocultar elementos de navegación automáticos */
        .main-header [data-testid="stElementToolbar"] {
            display: none !important;
        }
        
        .header-subtitle {
            color: #ffffff !important;
            font-size: 1.2rem;
            font-weight: 500;
            margin: 0.5rem 0 0 0;
            text-shadow: 0 1px 2px rgba(0,0,0,0.3);
            letter-spacing: 0.5px;
        }
        
        .header-stats {
            display: flex;
            gap: 2rem;
            align-items: center;
        }
        
        .stat-item {
            text-align: center;
            color: #ffffff !important;
        }
        
        .stat-number {
            font-size: 1.8rem;
            font-weight: 700;
            display: block;
            color: #ffffff !important;
        }
        
        .stat-label {
            font-size: 0.9rem;
            font-weight: 300;
            color: #ffffff !important;
        }
        
        /* Pestañas profesionales */
        .stTabs [data-baseweb="tab-list"] {
            gap: 0;
            background: white;
            border-radius: 12px;
            padding: 8px;
            box-shadow: 0 4px 20px rgba(0,0,0,0.1);
            margin-bottom: 2rem;
        }
        
        .stTabs [data-baseweb="tab"] {
            height: 60px;
            padding: 0 2rem;
            background: transparent;
            border-radius: 8px;
            color: #6b7280;
            font-weight: 500;
            border: none;
            font-size: 1rem;
            transition: all 0.3s ease;
        }
        
        .stTabs [aria-selected="true"] {
            background: linear-gradient(135deg, #6b7280, #9ca3af) !important;
            color: white !important;
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(107, 114, 128, 0.3) !important;
        }
        
        /* Cards de bienvenida personalizadas por país con animaciones */
        .country-card {
            background: linear-gradient(135deg, #ffffff 0%, #f8fafc 100%);
            border: 1px solid #e2e8f0;
            border-radius: 24px;
            padding: 3rem 2rem;
            text-align: center;
            margin: 1.5rem 0;
            box-shadow: 0 4px 20px rgba(0,0,0,0.08), 0 1px 3px rgba(0,0,0,0.05);
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            overflow: hidden;
            cursor: pointer;
        }
        
        .country-card::before {
            content: \\;
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.4), transparent);
            transition: left 0.6s ease;
        }
        
        .country-card:hover::before {
            left: 100%;
        }
        
        /* Guatemala - Celeste */
        .country-card-gt {
            border-top: 4px solid #0ea5e9;
        }
        
        .country-card-gt:hover {
            transform: translateY(-8px) scale(1.02);
            box-shadow: 0 20px 40px rgba(14, 165, 233, 0.3), 0 8px 20px rgba(0, 0, 0, 0.08);
            border-color: rgba(14, 165, 233, 0.3);
            background: linear-gradient(135deg, #ffffff 0%, rgba(186, 230, 253, 0.15) 100%);
        }
        
        /* Panamá - Rojo */
        .country-card-pa {
            border-top: 4px solid #dc2626;
        }
        
        .country-card-pa:hover {
            transform: translateY(-8px) scale(1.02);
            box-shadow: 0 20px 40px rgba(220, 38, 38, 0.3), 0 8px 20px rgba(0, 0, 0, 0.08);
            border-color: rgba(220, 38, 38, 0.3);
            background: linear-gradient(135deg, #ffffff 0%, rgba(254, 202, 202, 0.15) 100%);
        }
        
        /* Honduras - Azul */
        .country-card-hn {
            border-top: 4px solid #1e40af;
        }
        
        .country-card-hn:hover {
            transform: translateY(-8px) scale(1.02);
            box-shadow: 0 20px 40px rgba(30, 64, 175, 0.3), 0 8px 20px rgba(0, 0, 0, 0.08);
            border-color: rgba(30, 64, 175, 0.3);
            background: linear-gradient(135deg, #ffffff 0%, rgba(191, 219, 254, 0.15) 100%);
        }
        
        /* El Salvador - Azul oscuro */
        .country-card-sv {
            border-top: 4px solid #1e3a8a;
        }
        
        .country-card-sv:hover {
            transform: translateY(-8px) scale(1.02);
            box-shadow: 0 20px 40px rgba(30, 58, 138, 0.3), 0 8px 20px rgba(0, 0, 0, 0.08);
            border-color: rgba(30, 58, 138, 0.3);
            background: linear-gradient(135deg, #ffffff 0%, rgba(191, 219, 254, 0.15) 100%);
        }
        
        /* Costa Rica - Verde */
        .country-card-cr {
            border-top: 4px solid #16a34a;
        }
        
        .country-card-cr:hover {
            transform: translateY(-8px) scale(1.02);
            box-shadow: 0 20px 40px rgba(22, 163, 74, 0.3), 0 8px 20px rgba(0, 0, 0, 0.08);
            border-color: rgba(22, 163, 74, 0.3);
            background: linear-gradient(135deg, #ffffff 0%, rgba(187, 247, 208, 0.15) 100%);
        }
        
        /* Efectos de hover para banderas */
        .country-flag {
            font-size: 4rem;
            margin-bottom: 1.5rem;
            display: block;
            transition: all 0.4s ease;
            filter: drop-shadow(0 4px 8px rgba(0,0,0,0.1));
        }
        
        .country-card:hover .country-flag {
            transform: scale(1.15) rotate(5deg);
            filter: drop-shadow(0 8px 16px rgba(0,0,0,0.15));
        }
        
        /* Efectos de hover para títulos */
        .country-title {
            transition: all 0.3s ease;
        }
        
        .country-card:hover .country-title {
            transform: translateY(-2px);
        }
        
        /* Efectos de hover para descripciones */
        .country-description {
            transition: all 0.3s ease;
        }
        
        .country-card:hover .country-description {
            transform: translateY(-1px);
        }
        
        /* Pestañas minimalistas y elegantes */
        .stTabs [data-baseweb="tab-list"] {
            background: rgba(255, 255, 255, 0.95);
            border-radius: 16px;
            padding: 6px;
            margin: 2rem 0;
            box-shadow: 
                0 1px 3px rgba(0, 0, 0, 0.05),
                0 1px 2px rgba(0, 0, 0, 0.1);
            border: 1px solid rgba(0, 0, 0, 0.08);
            position: relative;
            display: flex;
            justify-content: space-between;
            align-items: center;
            width: 100%;
            gap: 4px;
            backdrop-filter: blur(8px);
        }
        
        .stTabs [data-baseweb="tab"] {
            background: transparent;
            border: none;
            border-radius: 12px;
            padding: 12px 20px;
            margin: 0;
            font-weight: 400;
            font-size: 1.1rem;
            transition: all 0.4s ease;
            position: relative;
            overflow: hidden;
            z-index: 1;
            color: #64748b;
            letter-spacing: 0.25px;
            flex: 1;
            text-align: center;
            white-space: nowrap;
            min-width: 0;
            display: flex;
            align-items: center;
            justify-content: center;
            cursor: pointer;
            font-family: -apple-system, BlinkMacSystemFont, \Segoe UI\, Roboto, sans-serif;
            text-shadow: 0 2px 4px rgba(255,255,255,0.8);
            transform: scaleY(0.9);
            text-transform: uppercase;
        }
        
        .stTabs [data-baseweb="tab"]:hover {
            background: rgba(248, 250, 252, 0.8) !important;
            color: currentColor !important;
            font-weight: 400 !important;
            text-shadow: 0 2px 4px rgba(255,255,255,0.9) !important;
            transform: translateY(-12px) scale(1.05) !important;
            box-shadow: 
                0 25px 50px rgba(0, 0, 0, 0.15),
                0 12px 24px rgba(0, 0, 0, 0.12),
                0 4px 8px rgba(0, 0, 0, 0.08) !important;
        }
        
        .stTabs [data-baseweb="tab"][aria-selected="true"] {
            background: rgba(248, 250, 252, 0.2) !important;
            color: currentColor !important;
            font-weight: 400 !important;
            text-shadow: 0 2px 4px rgba(255,255,255,0.9) !important;
            position: relative;
        }
        
        .stTabs [data-baseweb="tab"][data-teststate="active"] {
            background: rgba(248, 250, 252, 0.2) !important;
            color: currentColor !important;
            font-weight: 400 !important;
            text-shadow: 0 2px 4px rgba(255,255,255,0.9) !important;
        }
        
        .stTabs [data-baseweb="tab"]:focus {
            background: rgba(248, 250, 252, 0.2) !important;
            color: currentColor !important;
            font-weight: 400 !important;
            text-shadow: 0 2px 4px rgba(255,255,255,0.9) !important;
        }
        
        .stTabs [data-baseweb="tab"][aria-selected="true"]:after {
            content: \\;
            position: absolute;
            bottom: 0;
            left: 50%;
            transform: translateX(-50%);
            width: 12px;
            height: 1px;
            background: rgba(156, 163, 175, 0.25);
            border-radius: 0.5px;
            opacity: 1;
        }
        
        .stTabs [data-baseweb="tab"][aria-selected="true"]:hover {
            background: rgba(249, 250, 251, 0.4) !important;
            transform: translateY(-8px) scale(1.04) !important;
            color: #52525b !important;
            box-shadow: 
                0 20px 40px rgba(0, 0, 0, 0.12),
                0 8px 16px rgba(0, 0, 0, 0.08),
                0 2px 4px rgba(0, 0, 0, 0.04) !important;
        }
        
        .stTabs [data-baseweb="tab"][aria-selected="true"]:hover:after {
            width: 16px;
            background: rgba(156, 163, 175, 0.3);
        }
        
        .stTabs [data-baseweb="tab-highlight"] {
            display: none;
        }
        
        .stTabs [data-baseweb="tab-panel"] {
            padding: 2.5rem 0;
            animation: fadeIn 0.5s ease-in-out;
        }
        
        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        /* Métricas KPI mejoradas */
        .metric-card {
            background: white;
            border-radius: 16px;
            padding: 1.5rem;
            box-shadow: 0 4px 20px rgba(0,0,0,0.08);
            border: 1px solid #f1f5f9;
            transition: all 0.4s cubic-bezier(0.175, 0.885, 0.32, 1.275);
            animation: scaleIn 0.6s ease-out;
            position: relative;
            overflow: hidden;
        }
        
        .metric-card::before {
            content: \\;
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.4), transparent);
            transition: left 0.6s ease;
        }
        
        .metric-card:hover::before {
            left: 100%;
        }
        
        .metric-card:hover {
            transform: translateY(-5px) scale(1.02);
            box-shadow: 0 12px 40px rgba(0,0,0,0.15);
        }
        
        /* Botones mejorados */
        .stDownloadButton > button {
            background: linear-gradient(135deg, #f59e0b, #d97706) !important;
            color: white !important;
            border: none !important;
            border-radius: 12px !important;
            font-weight: 600 !important;
            padding: 0.75rem 2rem !important;
            font-size: 1rem !important;
            transition: all 0.3s ease !important;
            box-shadow: 0 4px 15px rgba(245, 158, 11, 0.3) !important;
        }
        
        .stDownloadButton > button:hover {
            transform: translateY(-3px) scale(1.05) !important;
            box-shadow: 0 8px 25px rgba(245, 158, 11, 0.5) !important;
        }
        
        /* Animaciones para pestañas */
        .stTabs [data-baseweb="tab-list"] {
            animation: slideInFromRight 0.6s ease-out;
        }
        
        .stTabs [data-baseweb="tab"] {
            transition: all 0.3s ease !important;
            border-radius: 10px !important;
            margin: 0 5px !important;
        }
        
        .stTabs [data-baseweb="tab"]:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1) !important;
        }
        
        /* Animaciones para contenido de pestañas */
        .stTabs [data-baseweb="tab-panel"] {
            animation: fadeIn 0.8s ease-out;
        }
        
        /* Alertas mejoradas */
        .stAlert {
            border-radius: 12px !important;
            border: none !important;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1) !important;
            animation: scaleIn 0.5s ease-out;
            transition: all 0.3s ease !important;
        }
        
        .stAlert:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 6px 20px rgba(0,0,0,0.15) !important;
        }
        
        /* Animaciones para tablas */
        .stDataFrame {
            animation: slideInFromLeft 0.8s ease-out;
            transition: all 0.3s ease !important;
        }
        
        .stDataFrame:hover {
            box-shadow: 0 8px 25px rgba(0,0,0,0.1) !important;
        }
        
        /* Animaciones para gráficos */
        .js-plotly-plot {
            animation: fadeIn 1s ease-out;
            transition: all 0.3s ease !important;
        }
        
        .js-plotly-plot:hover {
            transform: translateY(-3px) !important;
            box-shadow: 0 10px 30px rgba(0,0,0,0.15) !important;
        }
        
        /* Animaciones para elementos de carga */
        .stSpinner {
            animation: pulse 2s infinite;
        }
        
        /* Animaciones para elementos de entrada */
        .stFileUploader {
            animation: slideInFromRight 0.6s ease-out;
        }
        
        /* Contenedor de archivo subido */
        .stFileUploader .uploadedFile {
            background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
            border: 1px solid #0284c7;
            border-radius: 12px;
            padding: 0.75rem 1rem;
            margin-top: 1rem;
            color: #0284c7;
            font-weight: 500;
        }
        
        /* Mensajes de éxito mejorados */
        .stSuccess {
            background: linear-gradient(135deg, #f0fdf4 0%, #dcfce7 100%);
            border: 1px solid #16a34a;
            border-radius: 12px;
            padding: 1rem;
            color: #16a34a;
            font-weight: 500;
        }
        
        /* Mensajes de error mejorados */
        .stError {
            background: linear-gradient(135deg, #fef2f2 0%, #fee2e2 100%);
            border: 1px solid #dc2626;
            border-radius: 12px;
            padding: 1rem;
            color: #dc2626;
            font-weight: 500;
        }
        
        /* Tablas mejoradas */
        .stDataFrame {
            border-radius: 16px !important;
            overflow: hidden !important;
            box-shadow: 0 8px 30px rgba(0,0,0,0.1) !important;
        }
        
        /* Sidebar mejorado */
        .stSidebar {
            background: linear-gradient(180deg, #f8fafc 0%, #e2e8f0 100%);
        }
        
        /* File uploader minimalista mejorado */
        .stFileUploader {
            border: 2px dashed #e5e7eb;
            border-radius: 20px;
            padding: 3rem 2rem;
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
            position: relative;
            overflow: hidden;
        }
        
        .stFileUploader::before {
            content: \\;
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255,255,255,0.6), transparent);
            transition: left 0.5s ease;
        }
        
        .stFileUploader:hover::before {
            left: 100%;
        }
        
        .stFileUploader:hover {
            border-color: #3b82f6;
            background: linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%);
            transform: translateY(-2px) scale(1.02);
            box-shadow: 0 20px 40px rgba(59, 130, 246, 0.1);
        }
        
        .stFileUploader > div {
            text-align: center;
            color: #6b7280;
            font-weight: 500;
        }
        
        .stFileUploader:hover > div {
            color: #3b82f6;
        }
        
        /* Spinner personalizado minimalista */
        .stSpinner {
            color: #3b82f6 !important;
        }
        
        .stSpinner > div {
            border-color: #3b82f6 !important;
        }
        
        /* Contenedor de loading mejorado */
        .stSpinner {
            background: rgba(255, 255, 255, 0.9);
            border-radius: 12px;
            padding: 2rem;
            backdrop-filter: blur(10px);
            box-shadow: 0 10px 30px rgba(0,0,0,0.1);
        }
        
        /* Ocultar elementos de Streamlit */
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        .stDeployButton {display: none;}
        
        </style>
        """, unsafe_allow_html=True)
    
    def create_main_header(self):
        """Crea el header principal profesional con hora en tiempo real y última actividad"""
        # Obtener conteos dinámicos
        total_countries = self._get_total_countries()
        total_stores = self._get_total_stores()
        
        # Obtener fecha de último trabajo con stock
        last_work_date = self._get_last_stock_work_date()
        
        # Crear contenedor para la hora que se puede actualizar
        header_container = st.container()
        
        with header_container:
            # Cargar logo de New Era
            try:
                logo_path = "LOGO NE NUEVO.png"
                with open(logo_path, "rb") as logo_file:
                    logo_data = logo_file.read()
                import base64
                logo_base64 = base64.b64encode(logo_data).decode()
                logo_html = f'<img src="data:image/png;base64,{logo_base64}" class="logo-icon" style="height: 100px; width: auto; filter: drop-shadow(0 6px 12px rgba(0,0,0,0.3));">'
            except FileNotFoundError:
                # Fallback a corona si no encuentra la imagen
                logo_html = '<span class="logo-icon" style="font-size: 5rem; display: inline-block;">👑</span>'
            
            st.markdown(f"""
            <div class="main-header">
                <div class="header-content">
                    <div class="logo-section">
                        {logo_html}
                        <div>
                            <h1 class="header-title">New Era Analytics</h1>
                            <p class="header-subtitle">Dashboard de Análisis de Inventario</p>
                        </div>
                    </div>
                    <div class="header-stats">
                        <div class="stat-item">
                            <span class="stat-number">{total_countries}</span>
                            <span class="stat-label">Países</span>
                        </div>
                        <div class="stat-item">
                            <span class="stat-number">{total_stores}</span>
                            <span class="stat-label">Tiendas</span>
                        </div>
                        <div class="stat-item">
                            <span class="stat-number" id="current-time">{datetime.now().strftime('%H:%M')}</span>
                            <span class="stat-label">Hora actual</span>
                        </div>
                        <div class="stat-item">
                            <span class="stat-number">{last_work_date}</span>
                            <span class="stat-label">Último trabajo</span>
                        </div>
                    </div>
                </div>
            </div>
            
            <script>
            function updateTime() {{
                const now = new Date();
                const timeString = now.toLocaleTimeString(\es-ES\, {{
                    hour: \2-digit\,
                    minute: \2-digit\,
                    hour12: false
                }});
                const timeElement = document.getElementById(\current-time\);
                if (timeElement) {{
                    timeElement.textContent = timeString;
                }}
            }}
            
            // Actualizar cada segundo
            setInterval(updateTime, 1000);
            
            // Actualizar inmediatamente
            updateTime();
            </script>
            """, unsafe_allow_html=True)
    
    def _get_last_stock_work_date(self):
        """Obtiene la fecha del último trabajo con stock"""
        # Verificar si hay registro de última actividad en session_state
        if 'last_stock_work_date' in st.session_state:
            return st.session_state.last_stock_work_date
        else:
            # Si no hay actividad previa, mostrar "Sin actividad"
            return "Sin actividad"
    
    def _update_last_stock_work_date(self):
        """Actualiza la fecha del último trabajo con stock"""
        current_date = datetime.now().strftime('%d/%m/%Y')
        st.session_state.last_stock_work_date = current_date
        return current_date
    
    def _get_league_logo(self, logo_filename: str, league_name: str, fallback: str) -> str:
        """Carga el logo de la liga o usa fallback si no encuentra la imagen"""
        try:
            with open(logo_filename, "rb") as logo_file:
                logo_data = logo_file.read()
            import base64
            logo_base64 = base64.b64encode(logo_data).decode()
            return f'<img src="data:image/png;base64,{logo_base64}" alt="{league_name}" style="height: 75px; width: auto; margin-bottom: 0.5rem;">'
        except FileNotFoundError:
            # Fallback al emoji si no encuentra la imagen
            return f'<span style="font-size: 1.5rem;">{fallback}</span>'
    
    def _get_total_countries(self) -> int:
        """Obtiene el número total de países dinámicamente desde CountryManager"""
        global country_manager
        if country_manager is None:
            country_manager = CountryManager()
        return len(country_manager.countries)
    
    def _get_total_stores(self) -> int:
        """Obtiene el número total de tiendas excluyendo bodegas centrales"""
        global country_manager
        if country_manager is None:
            country_manager = CountryManager()
        
        # Bodegas centrales a excluir
        central_warehouses = {
            "CENTRAL NEW ERA",  # Guatemala
            "New Era Central",  # El Salvador  
            "Bodega Central NEW ERA",  # Costa Rica
            "Bodega Central Albrook",  # Panama
            "Almacén general"  # Panama
        }
        
        total_stores = 0
        for country_name, country_data in country_manager.countries.items():
            # Contar todas las bodegas menos las centrales
            for bodega in country_data.bodegas:
                if bodega not in central_warehouses:
                    total_stores += 1
        
        return total_stores
    
    def create_leagues_section(self):
        """Crea la sección de ligas deportivas con logos y estilo de pestañas de países"""
        # Cargar todos los logos
        mlb_logo = self._get_league_logo("LOGO_MLB.png", "MLB", "⚾")
        nba_logo = self._get_league_logo("LOGO_NBA.png", "NBA", "🏀")
        nfl_logo = self._get_league_logo("LOGO_NFL.png", "NFL", "🏈")
        f1_logo = self._get_league_logo("LOGO_F1.png", "MOTORSPORT", "🏎️")
        ne_logo = self._get_league_logo("LOGO_NE 2.png", "NEW ERA", "👑")
        
        # Estilos flotantes sin bordes para las tarjetas de ligas
        st.markdown("""
        <style>
        .league-card {
            background: transparent;
            border-radius: 12px;
            padding: 1.5rem 1rem 1rem 1rem;
            text-align: center;
            cursor: pointer;
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            border: none;
            margin-bottom: 0.5rem;
            position: relative;
            overflow: visible;
        }
        
        /* Elementos ::before y ::after removidos para eliminar iluminación de toda la tarjeta */
        
        .league-card:hover {
            /* Solo iluminación, sin movimiento */
        }
        
        .league-card.selected {
            /* Solo mantener la iluminación cuando esté seleccionada */
        }
        
        .league-card img {
            height: 95px;
            width: auto;
            margin-bottom: 0.75rem;
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            filter: drop-shadow(0 4px 8px rgba(0,0,0,0.1));
            position: relative;
            z-index: 1;
        }
        
        .league-card:hover img {
            transform: scale(1.1);
            filter: drop-shadow(0 6px 12px rgba(0,0,0,0.15));
        }
        
        .league-card.selected img {
            transform: scale(1.05);
            filter: drop-shadow(0 4px 8px currentColor);
        }
        
        .league-card p {
            margin: 0.75rem 0 0 0;
            font-weight: 400;
            color: #64748b;
            font-size: 1.1rem;
            transition: all 0.4s ease;
            position: relative;
            z-index: 1;
            text-shadow: 0 2px 4px rgba(255,255,255,0.8);
            transform: scaleY(0.9);
        }
        
        .league-card:hover p,
        .league-card.selected p {
            color: currentColor;
            font-weight: 400;
            text-shadow: 0 2px 4px rgba(255,255,255,0.9);
        }
        
        /* Colores específicos para cada liga */
        .mlb-card {
            color: #1e3a8a;
        }
        
        .nba-card {
            color: #ea580c;
        }
        
        .nfl-card {
            color: #16a34a;
        }
        
        .f1-card {
            color: #dc2626;
        }
        
        .newera-card {
            color: #000000;
        }
        
        /* Estilo minimalista para el selectbox */
        [data-testid="league_selector"] {
            margin: 0.25rem 0 !important;
        }
        
        [data-testid="league_selector"] select {
            background: #fafafa !important;
            border: none !important;
            border-radius: 6px !important;
            padding: 0.4rem 0.6rem !important;
            font-size: 0.85rem !important;
            font-weight: 400 !important;
            color: #6b7280 !important;
            box-shadow: none !important;
            transition: all 0.15s ease !important;
        }
        
        [data-testid="league_selector"] label {
            display: none !important;
        }
        </style>
        """, unsafe_allow_html=True)
        
        # Descripción del sistema con imagen de fondo
        # Cargar imagen de fondo IMAGEN_NE
        try:
            with open("IMAGEN_NE.png", "rb") as bg_file:
                bg_data = bg_file.read()
            import base64
            bg_base64 = base64.b64encode(bg_data).decode()
            bg_image = f"data:image/png;base64,{bg_base64}"
        except FileNotFoundError:
            try:
                with open("IMAGEN_NE.jpg", "rb") as bg_file:
                    bg_data = bg_file.read()
                import base64
                bg_base64 = base64.b64encode(bg_data).decode()
                bg_image = f"data:image/jpeg;base64,{bg_base64}"
            except FileNotFoundError:
                # Si no encuentra la imagen, usar fondo blanco
                bg_image = None
        
        if bg_image:
            st.markdown(f"""
            <div style="
                position: relative;
                border-radius: 16px; 
                padding: 2rem; 
                margin-bottom: 2rem; 
                box-shadow: 0 4px 20px rgba(0,0,0,0.08);
                background-image: url({bg_image});
                background-size: cover;
                background-position: center;
                background-repeat: no-repeat;
                overflow: hidden;
            ">
                <div style="
                    position: absolute;
                    top: 0;
                    left: 0;
                    right: 0;
                    bottom: 0;
                    background: rgba(255, 255, 255, 0.85);
                    border-radius: 16px;
                    z-index: 1;
                "></div>
                <div style="position: relative; z-index: 2;">
                    <h3 style="color: #374151; font-weight: 600; margin-bottom: 1rem;">Análisis Integral de Inventario</h3>
                    <p style="color: #6b7280; font-size: 1.1rem; line-height: 1.6; margin-bottom: 1.5rem;">
                        Sistema profesional de análisis de stock por bodega, categorizado por ligas deportivas principales:
                    </p>
                </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            # Fallback si no encuentra la imagen
            st.markdown("""
            <div style="background: white; border-radius: 16px; padding: 2rem; margin-bottom: 2rem; box-shadow: 0 4px 20px rgba(0,0,0,0.08);">
                <h3 style="color: #374151; font-weight: 600; margin-bottom: 1rem;">Análisis Integral de Inventario</h3>
                <p style="color: #6b7280; font-size: 1.1rem; line-height: 1.6; margin-bottom: 1.5rem;">
                    Sistema profesional de análisis de stock por bodega, categorizado por ligas deportivas principales:
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        # Inicializar session state para la liga seleccionada
        if 'selected_league' not in st.session_state:
            st.session_state.selected_league = None
        
        # Selectbox con placeholder para manejar la selección
        league_selection = st.selectbox(
            "Liga",
            ["Selecciona la Liga", "Todas", "MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT"],
            index=0,
            key="league_selector",
            label_visibility="collapsed"
        )
        
        # Actualizar session state basado en la selección
        if league_selection == "Selecciona la Liga":
            st.session_state.selected_league = None
        elif league_selection == "Todas":
            st.session_state.selected_league = "Todas"
        else:
            st.session_state.selected_league = league_selection
        
        # Crear tarjetas clickeables con estilo de pestañas
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            selected_class = "selected" if (st.session_state.selected_league == "MLB" or st.session_state.selected_league == "Todas") else ""
            st.markdown(f"""
            <div class="league-card mlb-card {selected_class}" onclick="
                const selectbox = window.parent.document.querySelector(\[data-testid=\\\"league_selector\\\"] select\);
                if (selectbox) {{
                    selectbox.value = MLB;
                    selectbox.dispatchEvent(new Event(change, {{ bubbles: true }}));
                }}
            ">
                {mlb_logo}
                <p>MLB</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            selected_class = "selected" if (st.session_state.selected_league == "NBA" or st.session_state.selected_league == "Todas") else ""
            st.markdown(f"""
            <div class="league-card nba-card {selected_class}" onclick="
                const selectbox = window.parent.document.querySelector(\[data-testid=\\\"league_selector\\\"] select\);
                if (selectbox) {{
                    selectbox.value = NBA;
                    selectbox.dispatchEvent(new Event(change, {{ bubbles: true }}));
                }}
            ">
                {nba_logo}
                <p>NBA</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            selected_class = "selected" if (st.session_state.selected_league == "NFL" or st.session_state.selected_league == "Todas") else ""
            st.markdown(f"""
            <div class="league-card nfl-card {selected_class}" onclick="
                const selectbox = window.parent.document.querySelector(\[data-testid=\\\"league_selector\\\"] select\);
                if (selectbox) {{
                    selectbox.value = NFL;
                    selectbox.dispatchEvent(new Event(change, {{ bubbles: true }}));
                }}
            ">
                {nfl_logo}
                <p>NFL</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            selected_class = "selected" if (st.session_state.selected_league == "MOTORSPORT" or st.session_state.selected_league == "Todas") else ""
            st.markdown(f"""
            <div class="league-card f1-card {selected_class}" onclick="
                const selectbox = window.parent.document.querySelector(\[data-testid=\\\"league_selector\\\"] select\);
                if (selectbox) {{
                    selectbox.value = MOTORSPORT;
                    selectbox.dispatchEvent(new Event(change, {{ bubbles: true }}));
                }}
            ">
                {f1_logo}
                <p>MOTORSPORT</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col5:
            selected_class = "selected" if (st.session_state.selected_league == "ENTERTAINMENT" or st.session_state.selected_league == "Todas") else ""
            st.markdown(f"""
            <div class="league-card newera-card {selected_class}" onclick="
                const selectbox = window.parent.document.querySelector(\[data-testid=\\\"league_selector\\\"] select\);
                if (selectbox) {{
                    selectbox.value = ENTERTAINMENT;
                    selectbox.dispatchEvent(new Event(change, {{ bubbles: true }}));
                }}
            ">
                {ne_logo}
                <p>ENTERTAINMENT</p>
            </div>
            """, unsafe_allow_html=True)
        
        # Agregar espacio entre las tarjetas de ligas y las pestañas de países
        st.markdown("<div style='margin-bottom: 2rem;'></div>", unsafe_allow_html=True)
        
    
    def create_welcome_card(self, country_flag: str, country_name: str, description: str, stores_count: int, country_code: str = ""):
        """Crea una card de bienvenida profesional con colores personalizados por país"""
        
        # Determinar colores según el país
        if "Panamá" in country_name or "PANAMA" in country_name:
            # Rojo de la bandera de Panamá
            primary_color = "#dc2626"
            secondary_color = "#ef4444"
            accent_color = "#fecaca"
        elif "Honduras" in country_name or "HONDURAS" in country_name:
            # Azul 
            primary_color = "#1e40af"
            secondary_color = "#3b82f6"
            accent_color = "#bfdbfe"
        elif "El Salvador" in country_name or "EL SALVADOR" in country_name:
            # Azul oscuro de la bandera de El Salvador
            primary_color = "#1e3a8a"
            secondary_color = "#1d4ed8"
            accent_color = "#bfdbfe"
        elif "Costa Rica" in country_name or "COSTA RICA" in country_name:
            # Verde reciclaje de Costa Rica
            primary_color = "#16a34a"
            secondary_color = "#22c55e"
            accent_color = "#bbf7d0"
        elif "Guatemala" in country_name or "GUATEMALA" in country_name:
            # Celeste 
            primary_color = "#0ea5e9"
            secondary_color = "#38bdf8"
            accent_color = "#bae6fd"
        else:
            # Azul medio por defecto
            primary_color = "#1e40af"
            secondary_color = "#3b82f6"
            accent_color = "#bfdbfe"
        
        return f"""
<div class="welcome-card-{country_code.lower()}">
    <div style="position: absolute; top: 0; left: 0; right: 0; height: 4px; background: linear-gradient(90deg, {primary_color}, {secondary_color}); border-radius: 24px 24px 0 0; opacity: 0.8;"></div>
    
    <span class="country-flag">{country_flag}</span>
    
    <h3 style="color: {primary_color}; font-size: 1.75rem; font-weight: 700; margin-bottom: 1rem; transition: all 0.3s ease; text-shadow: 0 2px 4px rgba(0,0,0,0.02);">{country_name}</h3>
    
    <p style="color: #64748b; font-size: 1rem; font-weight: 500; line-height: 1.6; margin-bottom: 0; transition: all 0.3s ease; background: linear-gradient(135deg, rgba(255,255,255,0.7) 0%, {accent_color}20 100%); padding: 1rem; border-radius: 12px; border: 1px solid {accent_color}50;">
        {description}<br>
        <strong style="color: {primary_color};">{stores_count} tiendas</strong> en operación
    </p>
</div>"""
    
    def create_section_header(self, title: str, subtitle: str = "", icon: str = "📊"):
        """Crea un header de sección profesional con fondo degradado según el país"""
        subtitle_html = f"<p style='color: #d1d5db; margin: 0.5rem 0 0 0; font-size: 1.1rem;'>{subtitle}</p>" if subtitle else ""
        
        # Determinar colores según el país
        if "Panamá" in title or "PANAMA" in title or "Panama" in title:
            # Rojo de la bandera de Panamá
            background_gradient = "linear-gradient(135deg, #dc2626 0%, #ef4444 100%)"
            box_shadow = "0 10px 30px rgba(220, 38, 38, 0.3)"
        elif "Honduras" in title or "HONDURAS" in title:
            # Azul (anteriormente de Guatemala)
            background_gradient = "linear-gradient(135deg, #1e40af 0%, #3b82f6 100%)"
            box_shadow = "0 10px 30px rgba(30, 64, 175, 0.3)"
        elif "El Salvador" in title or "EL SALVADOR" in title:
            # Azul oscuro de la bandera de El Salvador
            background_gradient = "linear-gradient(135deg, #1e3a8a 0%, #1d4ed8 100%)"
            box_shadow = "0 10px 30px rgba(30, 58, 138, 0.3)"
        elif "Costa Rica" in title or "COSTA RICA" in title:
            # Verde reciclaje de Costa Rica
            background_gradient = "linear-gradient(135deg, #16a34a 0%, #22c55e 100%)"
            box_shadow = "0 10px 30px rgba(22, 163, 74, 0.3)"
        elif "Guatemala" in title or "GUATEMALA" in title:
            # Celeste (anteriormente de Honduras)
            background_gradient = "linear-gradient(135deg, #0ea5e9 0%, #38bdf8 100%)"
            box_shadow = "0 10px 30px rgba(14, 165, 233, 0.3)"
        elif "Puerto Rico" in title or "PUERTO RICO" in title:
            # Naranja para Puerto Rico
            background_gradient = "linear-gradient(135deg, #ea580c 0%, #f97316 100%)"
            box_shadow = "0 10px 30px rgba(234, 88, 12, 0.3)"
        else:
            # Azul medio por defecto
            background_gradient = "linear-gradient(135deg, #1e40af 0%, #3b82f6 100%)"
            box_shadow = "0 10px 30px rgba(30, 64, 175, 0.3)"
        
        st.markdown(f"""
        <div style="
            background: {background_gradient};
            padding: 2rem 3rem;
            border-radius: 15px;
            margin: 2rem 0;
            box-shadow: {box_shadow};
            animation: slideInFromLeft 0.8s ease-out;
            transition: all 0.3s ease;
        " onmouseover="this.style.transform=	ranslateY(-2px)" onmouseout="this.style.transform=	ranslateY(0)">
            <div style="display: flex; align-items: center;">
                <div>
                    <h2 style="color: white; margin: 0; font-size: 1.8rem; font-weight: 700;">
                        {title}
                    </h2>
                    {subtitle_html}
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

# Instancia del diseño profesional
professional_design = ProfessionalDesign()

@dataclass
class ProductClassification:
    """Clasificación de productos por silueta"""
    siluetas_planas: List[str]
    siluetas_curvas: List[str]
    
    def __post_init__(self):
        self.siluetas_planas = [
            "950", "5950", "5950 AF", "5950 DOGEAR", "5950 SPLIT PANEL",
            "950 AF", "950 AF TRUCKER", "950 SS", "950 TRUCKER",
            "LP 5950", "LP 950", "RC 5950", "RC 950"
        ]
        self.siluetas_curvas = [
            "920", "940", "970", "1920", "3930",
            "920 TRUCKER", "940 AF", "940 AF TRUCKER",
            "940 EF", "940 SS", "940 TRUCKER", "970 SS"
        ]
    
    def clasificar_silueta(self, silueta) -> Optional[str]:
        """Clasifica una silueta como plana o curva"""
        if pd.isna(silueta) or not isinstance(silueta, str):
            return None
        
        silueta_upper = str(silueta).strip().upper()
        if silueta_upper in self.siluetas_planas:
            return 'Planas'
        elif silueta_upper in self.siluetas_curvas:
            return 'Curvas'
        return None

# Instancia de clasificación
product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])

@dataclass
class CountryData:
    """Datos específicos de un país"""
    name: str
    bodegas: List[str]
    capacidades: Dict[str, int]
    tienda_mapping: Dict[str, str] = None
    
    def get_total_capacity(self) -> int:
        """Obtiene la capacidad total del país"""
        return sum(self.capacidades.values())

class CountryManager:
    """Gestor de datos de países"""
    
    def __init__(self):
        self.countries = {
            "Guatemala": CountryData(
                name="Guatemala",
                bodegas=[
                    "NE Cayala", "NE Chimaltenango", "NE Concepcion", "NE Interplaza Escuintla",
                    "NE InterXela", "NE Metrocentro Outlet", "NE Metronorte", "NE Metroplaza Jutiapa",
                    "NE Miraflores", "NE Naranjo", "NE Oakland", "NE Outlet Santa clara",
                    "NE Paseo Antigua", "NE Peri Roosvelt", "NE Plaza Magdalena", "NE Plaza Videre",
                    "NE Portales", "NE Pradera Chiquimula", "NE Pradera Escuintla", "NE Pradera Huehuetenango",
                    "NE Pradera Xela", "NE Puerto Barrios", "NE Vistares", "CENTRAL NEW ERA"
                ],
                capacidades={
                    "NE Miraflores": 6989, "NE Oakland": 5276, "NE Cayala": 5109,
                    "NE Plaza Videre": 2208, "NE Concepcion": 2754, "NE Portales": 6453,
                    "NE Naranjo": 2790, "NE Peri Roosvelt": 2476, "NE Vistares": 4890,
                    "NE Chimaltenango": 2357, "NE Pradera Escuintla": 4619, "NE Interplaza Escuintla": 6769,
                    "NE Pradera Xela": 3827, "NE InterXela": 3907, "NE Pradera Huehuetenango": 4835,
                    "NE Metroplaza Jutiapa": 2766, "NE Pradera Chiquimula": 4837, "NE Plaza Magdalena": 4710,
                    "NE Metronorte": 6735, "NE Metrocentro Outlet": 5184, "NE Outlet Santa clara": 4860,
                    "NE Paseo Antigua": 2952, "NE Puerto Barrios": 3024, "CENTRAL NEW ERA": 0
                }
            ),
            "El Salvador": CountryData(
                name="El Salvador",
                bodegas=[
                    "NE METROCENTRO LOURDES", "NE METROCENTRO SAN MIGUEL", "NE PLAZA MUNDO SOYAPANGO",
                    "NE USULUTÁN", "New Era Central", "NEW ERA EL PASEO", "NEW ERA METROCENTRO",
                    "NEW ERA METROCENTRO SANTA ANA", "NEW ERA MULTIPLAZA"
                ],
                capacidades={
                    "NEW ERA METROCENTRO": 4355, "NEW ERA MULTIPLAZA": 5443, "NEW ERA EL PASEO": 4436,
                    "NEW ERA METROCENTRO SANTA ANA": 5771, "NE USULUTÁN": 5760, "NE METROCENTRO SAN MIGUEL": 3600,
                    "NE PLAZA MUNDO SOYAPANGO": 3120, "NE METROCENTRO LOURDES": 6912, "New Era Central": 0
                },
                tienda_mapping={
                    # Mapeo: Bodega (Stock) -> Tienda (Ventas)
                    "NE METROCENTRO LOURDES": "NE METROCENTRO LOURDES",
                    "NEW ERA MULTIPLAZA": "NEW ERA MULTIPLAZA",
                    "NEW ERA METROCENTRO": "NE METROCENTRO",
                    "NEW ERA METROCENTRO SANTA ANA": "NE METROCENTRO SANTA ANA",
                    "NE PLAZA MUNDO SOYAPANGO": "NE PLAZA MUNDO SOYAPANGO",
                    "NE USULUTÁN": "NE PLAZA MUNDO USULUTÁN",
                    "NE METROCENTRO SAN MIGUEL": "NE METROCENTRO SAN MIGUEL",
                    "NEW ERA EL PASEO": "NEW ERA EL PASEO"
                    # Nota: "New Era Central" no tiene equivalente en ventas
                }
            ),
            "Honduras": CountryData(
                name="Honduras",
                bodegas=[
                    "NE – Cascadas Mall Tegucigalpa", "NE – City Mall Tegucigalpa", "NE – Mega Mall SPS",
                    "NE – Multiplaza Tegucigalpa", "NE –Multiplaza SPS"
                ],
                capacidades={
                    "NE – Mega Mall SPS": 2730, "NE –Multiplaza SPS": 6540, "NE – City Mall Tegucigalpa": 5190,
                    "NE – Cascadas Mall Tegucigalpa": 3816, "NE – Multiplaza Tegucigalpa": 0
                },
                tienda_mapping={
                    # Mapeo: Bodega (Stock) -> Tienda (Ventas)
                    "NE –Multiplaza SPS": "NE MULTIPLAZA SPS",
                    "NE – Cascadas Mall Tegucigalpa": "NEW ERA CASCADAS MALL TEGUCIGALPA",
                    "NE – Multiplaza Tegucigalpa": "NEW ERA MULTIPLAZA TEGUCIGALPA",
                    "NE – Mega Mall SPS": "NE MEGA  MALL SAN PEDRO SULA",
                    "NE – City Mall Tegucigalpa": "NE CITY MALL TEGUCIGALPA"
                }
            ),
            "Costa Rica": CountryData(
                name="Costa Rica",
                bodegas=[
                    "Bodega Central NEW ERA", "NE City Mall"
                ],
                capacidades={
                    "NE City Mall": 4260, "Bodega Central NEW ERA": 0
                },
                tienda_mapping={
                    # Mapeo: Bodega (Stock) -> Tienda (Ventas)
                    "NE City Mall": "NE CITY MAL"
                    # Nota: "Bodega Central NEW ERA" no tiene equivalente en ventas
                }
            ),
            "PANAMA": CountryData(
                name="PANAMA",
                bodegas=[
                    "Almacén general", "Bodega Central Albrook", "NE Albrookmall",
                    "NE Metromall", "NE Multiplaza Panamá", "NE Westland"
                ],
                capacidades={
                    "NE Multiplaza Panamá": 6318, "NE Westland": 2972, "NE Metromall": 4422,
                    "NE Albrookmall": 4224, "Almacén general": 0, "Bodega Central Albrook": 0
                },
                tienda_mapping={
                    # Mapeo: Bodega (Stock) -> Tienda (Ventas)
                    "NE Westland": "NE WESTLAND MALL",
                    "NE Multiplaza Panamá": "NE MULTIPLAZA PANAMÁ",
                    "NE Metromall": "NE METROMALL",
                    "NE Albrookmall": "NE ALBROOK MALL"
                    # Nota: "Bodega Central Albrook" y "Almacén general" no tienen equivalente en ventas
                }
            )
        }
    
    def get_country_data(self, country: str) -> Optional[CountryData]:
        """Obtiene los datos de un país"""
        return self.countries.get(country)
    
    def get_bodegas(self, country: str) -> List[str]:
        """Obtiene las bodegas de un país"""
        country_data = self.get_country_data(country)
        return country_data.bodegas if country_data else []
    
    def get_capacidades(self, country: str) -> Dict[str, int]:
        """Obtiene las capacidades de un país"""
        country_data = self.get_country_data(country)
        return country_data.capacidades if country_data else {}

# Instancia del gestor de países
country_manager = CountryManager()

@dataclass
class SalesProcessor:
    """Procesador de datos de ventas"""
    
    def __init__(self):
        # Inicializar country_manager
        self.country_manager = CountryManager()
        
        # Sistema de mapeo bidireccional para todos los países
        # Formato: nombre_canónico -> [variaciones_stock, variaciones_ventas]
        
        # GUATEMALA
        self.guatemala_mappings = {
            "NE_OAKLAND": ["NE Oakland", "NE OAKLAND"],
            "NE_CAYALA": ["NE Cayala", "NE CAYALA"],
            "NE_MIRAFLORES": ["NE Miraflores", "NE MIRAFLORES"],
            "NE_PORTALES": ["NE Portales", "NE PORTALES"],
            "NE_INTERXELA": ["NE InterXela", "NE INTERXELA"],
            "NE_METRONORTE": ["NE Metronorte", "NE METRONORTE"],
            "NE_CONCEPCION": ["NE Concepcion", "NE CONCEPCION"],
            "NE_INTERPLAZA_ESCUINTLA": ["NE Interplaza Escuintla", "NE INTERPLAZA ESCUINTLA"],
            "NE_PRADERA_HUEHUETENANGO": ["NE Pradera Huehuetenango", "NE PRADERA HUEHUETENANGO"],
            "NE_NARANJO": ["NE Naranjo", "NE NARANJO"],
            "NE_METROCENTRO_OUTLET": ["NE Metrocentro Outlet", "NEW ERA METROCENTRO VILLA NUEVA"],
            "NE_VISTARES": ["NE Vistares", "NE PRADERA VISTARES"],
            "NE_PERI_ROOSVELT": ["NE Peri Roosvelt", "NE PERIROOSVELT"],
            "NE_OUTLET_SANTA_CLARA": ["NE Outlet Santa clara", "NE SANTA CLARA"],
            "NE_PLAZA_MAGDALENA": ["NE Plaza Magdalena", "NE PLAZA MAGDALENA"],
            "NE_PRADERA_CHIQUIMULA": ["NE Pradera Chiquimula", "NE PRADERA CHIQUIMULA"],
            "NE_PRADERA_ESCUINTLA": ["NE Pradera Escuintla", "NE PRADERA ESCUINTLA"],
            "NE_PASEO_ANTIGUA": ["NE Paseo Antigua", "NE PASEO ANTIGUA"],
            "NE_PRADERA_XELA": ["NE Pradera Xela", "NE PRADERA XELA"],
            "NE_CHIMALTENANGO": ["NE Chimaltenango", "NE CHIMALTENANGO"],
            "NE_PLAZA_VIDERE": ["NE Plaza Videre", "NE PLAZA VIDERE"],
            "NE_METROPLAZA_JUTIAPA": ["NE Metroplaza Jutiapa", "NE METROPLAZA JUTIAPA"],
            "NE_PUERTO_BARRIOS": ["NE Puerto Barrios", "New Era Puerto Barrios"]
        }
        
        # EL SALVADOR
        self.el_salvador_mappings = {
            "NE_METROCENTRO_LOURDES": ["NE METROCENTRO LOURDES", "NE METROCENTRO LOURDES"],
            "NEW_ERA_MULTIPLAZA": ["NEW ERA MULTIPLAZA", "NEW ERA MULTIPLAZA"],
            "NEW_ERA_METROCENTRO": ["NEW ERA METROCENTRO", "NE METROCENTRO"],
            "NEW_ERA_METROCENTRO_SANTA_ANA": ["NEW ERA METROCENTRO SANTA ANA", "NE METROCENTRO SANTA ANA"],
            "NE_PLAZA_MUNDO_SOYAPANGO": ["NE PLAZA MUNDO SOYAPANGO", "NE PLAZA MUNDO SOYAPANGO"],
            "NE_USULUTN": ["NE USULUTÁN", "NE PLAZA MUNDO USULUTÁN"],
            "NE_METROCENTRO_SAN_MIGUEL": ["NE METROCENTRO SAN MIGUEL", "NE METROCENTRO SAN MIGUEL"],
            "NEW_ERA_EL_PASEO": ["NEW ERA EL PASEO", "NEW ERA EL PASEO"]
        }
        
        # HONDURAS
        self.honduras_mappings = {
            "NE_MULTIPLAZA_SPS": ["NE –Multiplaza SPS", "NE MULTIPLAZA SPS"],
            "NE_CASCADAS_MALL_TEGUCIGALPA": ["NE – Cascadas Mall Tegucigalpa", "NEW ERA CASCADAS MALL TEGUCIGALPA"],
            "NE_MULTIPLAZA_TEGUCIGALPA": ["NE – Multiplaza Tegucigalpa", "NEW ERA MULTIPLAZA TEGUCIGALPA"],
            "NE_MEGA_MALL_SPS": ["NE – Mega Mall SPS", "NE MEGA\xa0 MALL SAN PEDRO SULA"],
            "NE_CITY_MALL_TEGUCIGALPA": ["NE – City Mall Tegucigalpa", "NE CITY MALL TEGUCIGALPA"]
        }
        
        # COSTA RICA
        self.costa_rica_mappings = {
            "NE_CITY_MALL": ["NE City Mall", "NE CITY MAL"]
        }
        
        # PANAMA
        self.panama_mappings = {
            "NE_WESTLAND": ["NE Westland", "NE WESTLAND MALL"],
            "NE_MULTIPLAZA_PANAMA": ["NE Multiplaza Panamá", "NE MULTIPLAZA PANAMÁ"],
            "NE_METROMALL": ["NE Metromall", "NE METROMALL"],
            "NE_ALBROOKMALL": ["NE Albrookmall", "NE ALBROOK MALL"]
        }
        
        # Mapeo consolidado por país
        self.country_mappings = {
            "Guatemala": self.guatemala_mappings,
            "El Salvador": self.el_salvador_mappings,
            "Honduras": self.honduras_mappings,
            "Costa Rica": self.costa_rica_mappings,
            "PANAMA": self.panama_mappings
        }
        
        # Crear mapeos bidireccionales dinámicamente para todos los países
        self.nombre_to_canonico = {}  # cualquier_variacion -> (nombre_canonico, pais)
        self.canonico_to_stock = {}   # (nombre_canonico, pais) -> formato_stock
        self.canonico_to_ventas = {}  # (nombre_canonico, pais) -> formato_ventas
        
        for pais, mappings in self.country_mappings.items():
            for canonico, variaciones in mappings.items():
                formato_stock = variaciones[0]
                formato_ventas = variaciones[1]
                key = (canonico, pais)
                
                # Mapear todas las variaciones al canónico con país
                self.nombre_to_canonico[formato_stock] = key
                self.nombre_to_canonico[formato_ventas] = key
                self.nombre_to_canonico[self._normalize_text(formato_stock)] = key
                self.nombre_to_canonico[self._normalize_text(formato_ventas)] = key
                
                # Mapear canónico+país a formatos específicos
                self.canonico_to_stock[key] = formato_stock
                self.canonico_to_ventas[key] = formato_ventas

    def get_canonical_name(self, nombre_bodega, pais=None):
        """Obtiene el nombre canónico de una bodega desde cualquier variación"""
        if not nombre_bodega:
            return None
        
        # Intentar mapeo directo
        if nombre_bodega in self.nombre_to_canonico:
            canonical_key = self.nombre_to_canonico[nombre_bodega]
            # Si se especifica país, verificar que coincida
            if pais and canonical_key[1] != pais:
                return None
            return canonical_key
        
        # Intentar mapeo normalizado
        normalizado = self._normalize_text(nombre_bodega)
        if normalizado in self.nombre_to_canonico:
            canonical_key = self.nombre_to_canonico[normalizado]
            # Si se especifica país, verificar que coincida
            if pais and canonical_key[1] != pais:
                return None
            return canonical_key
        
        return None

    def normalize_bodega_name(self, nombre_bodega, target_format="stock", pais=None):
        """
        Normaliza nombre de bodega a formato específico.
        target_format: 'stock' o 'ventas'
        pais: país específico para filtrar mapeos
        """
        canonical_key = self.get_canonical_name(nombre_bodega, pais)
        if not canonical_key:
            return nombre_bodega  # Devolver original si no hay mapeo
        
        if target_format == "stock":
            return self.canonico_to_stock.get(canonical_key, nombre_bodega)
        elif target_format == "ventas":
            return self.canonico_to_ventas.get(canonical_key, nombre_bodega)
        else:
            return nombre_bodega

    def get_all_variations_for_country(self, pais):
        """Obtiene todas las variaciones de nombres para un país específico"""
        if pais not in self.country_mappings:
            return []
        
        variaciones = []
        for variaciones_lista in self.country_mappings[pais].values():
            variaciones.extend(variaciones_lista)
        return variaciones
            
    def _normalize_text(self, text):
        """Normaliza texto para comparación: mayúsculas, sin espacios extra, normalizar guiones"""
        if pd.isna(text):
            return ""
        # Normalizar diferentes tipos de guiones a guión estándar
        normalized = str(text).strip().upper()
        # Reemplazar diferentes tipos de guiones con guión estándar
        normalized = normalized.replace("–", "-").replace("—", "-").replace("−", "-")
        # Limpiar espacios múltiples
        normalized = normalized.replace("  ", " ")
        return normalized
    
    def procesar_ventas_guatemala(self, df_ventas: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, float]]]:
        """
        Procesa el archivo de ventas de Guatemala y retorna ventas desglosadas por bodega, liga y subcategoría
        Estructura: {bodega: {liga: {subcategoria: ventas}}}
        """
        if df_ventas is None or df_ventas.empty:
            return {}
        
        # Debug: Mostrar las columnas disponibles (solo en consola)
        print(f"Columnas disponibles en archivo de ventas: {list(df_ventas.columns)}")
        
        # Verificar columnas necesarias
        columnas_necesarias = ['U_Marca', 'U_Segmento', 'U_Liga', 'U_Silueta', 'USD_Total_SI_CD']
        columna_tienda = None
        
        # Buscar columna de tienda
        for col in df_ventas.columns:
            if 'tienda' in col.lower() or 'store' in col.lower() or 'bodega' in col.lower():
                columna_tienda = col
                break
        
        if columna_tienda is None:
            print("No se encontró columna de tienda")
            return {}
            
        columnas_necesarias.append(columna_tienda)
        
        # Verificar que existan todas las columnas
        for col in columnas_necesarias[:-1]:  # Excepto la columna_tienda que ya verificamos
            if col not in df_ventas.columns:
                print(f"No se encontró columna {col}")
                return {}
        
        print(f"Usando columna de tienda: {columna_tienda}")
        
        # Filtrar por marca NEW ERA
        df_new_era = df_ventas[df_ventas['U_Marca'].str.upper() == 'NEW ERA'].copy()
        
        # Filtrar tiendas usando el nuevo sistema de mapeo bidireccional
        # Obtener todas las variaciones posibles para Guatemala
        todas_variaciones = self.get_all_variations_for_country("Guatemala")
        
        df_mapeado = df_new_era[df_new_era[columna_tienda].isin(todas_variaciones)].copy()
        
        # Normalizar nombres de bodegas usando el nuevo sistema bidireccional
        df_mapeado['Bodega_Mapeada'] = df_mapeado[columna_tienda].apply(
            lambda x: self.normalize_bodega_name(x, target_format="stock", pais="Guatemala")
        )
        
        if len(df_mapeado) == 0:
            print("No hay registros mapeados para procesar")
            return {}
        
        # Definir categorías de ligas
        categorias_ligas = {
            "MLB": ["MLB", "MLB properties"],
            "NBA": ["NBA", "NBA Properties"],
            "NFL": ["NFL", "NFL Properties"], 
            "MOTORSPORT": ["MOTORSPORT"],
            "ENTERTAINMENT": [
                "NEW ERA BRANDED", "ENTERTAINMENT", "MARCA PAIS", "WARNER BROS",
                "DISNEY", "LOONEY TUNES", "MARVEL", "DC", "UNIVERSAL", "PARAMOUNT"
            ],
            "ACCESSORIES": ["ACCESSORIES"]
        }
        
        # Importar ProductClassification para clasificar siluetas
        product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])
        
        # Clasificar siluetas solo para HEADWEAR
        df_mapeado['Tipo'] = df_mapeado.apply(
            lambda row: product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        # Inicializar estructura de resultados
        ventas_desglosadas = {}
        
        # Procesar por bodega
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            df_bodega = df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega]
            ventas_desglosadas[bodega] = {}
            
            # Procesar por liga
            for categoria, ligas in categorias_ligas.items():
                ventas_desglosadas[bodega][categoria] = {}
                
                if categoria == 'ACCESSORIES':
                    # Para ACCESSORIES, filtrar directamente por segmento
                    df_accessories = df_bodega[df_bodega['U_Segmento'] == 'ACCESSORIES']
                    ventas_desglosadas[bodega][categoria]['Stock'] = df_accessories['USD_Total_SI_CD'].sum()
                    ventas_desglosadas[bodega][categoria]['Ventas'] = df_accessories['USD_Total_SI_CD'].sum()
                else:
                    # Lógica original para otras ligas
                    df_liga = df_bodega[df_bodega['U_Liga'].isin(ligas)]
                    
                    # Planas (HEADWEAR + Planas)
                    df_planas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Planas')]
                    ventas_desglosadas[bodega][categoria]['Planas'] = df_planas['USD_Total_SI_CD'].sum()
                    
                    # Curvas (HEADWEAR + Curvas)  
                    df_curvas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Curvas')]
                    ventas_desglosadas[bodega][categoria]['Curvas'] = df_curvas['USD_Total_SI_CD'].sum()
                    
                    # Apparel
                    df_apparel = df_liga[df_liga['U_Segmento'] == 'APPAREL']
                    ventas_desglosadas[bodega][categoria]['Apparel'] = df_apparel['USD_Total_SI_CD'].sum()
        
        print(f"Ventas desglosadas calculadas para {len(ventas_desglosadas)} bodegas")
        return ventas_desglosadas

    def procesar_ventas_el_salvador(self, df_ventas: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, float]]]:
        """
        Procesa el archivo de ventas de El Salvador y retorna ventas desglosadas por bodega, liga y subcategoría
        Estructura: {bodega: {liga: {subcategoria: ventas}}}
        """
        if df_ventas is None or df_ventas.empty:
            return {}
        
        # Debug: Mostrar las columnas disponibles (solo en consola)
        print(f"Columnas disponibles en archivo de ventas El Salvador: {list(df_ventas.columns)}")
        
        # Verificar columnas necesarias
        columnas_necesarias = ['U_Marca', 'U_Segmento', 'U_Liga', 'U_Silueta', 'USD_Total_SI_CD']
        columna_tienda = None
        
        # Buscar columna de tienda
        for col in df_ventas.columns:
            if 'tienda' in col.lower() or 'store' in col.lower() or 'bodega' in col.lower():
                columna_tienda = col
                break
        
        if columna_tienda is None:
            print("No se encontró columna de tienda en El Salvador")
            return {}
            
        columnas_necesarias.append(columna_tienda)
        
        # Verificar que existan todas las columnas
        for col in columnas_necesarias[:-1]:  # Excepto la columna_tienda que ya verificamos
            if col not in df_ventas.columns:
                print(f"No se encontró columna {col} en El Salvador")
                return {}
        
        print(f"Usando columna de tienda para El Salvador: {columna_tienda}")
        
        # Filtrar por marca NEW ERA
        df_new_era = df_ventas[df_ventas['U_Marca'].str.upper() == 'NEW ERA'].copy()
        
        # Filtrar tiendas usando el nuevo sistema de mapeo bidireccional
        # Obtener todas las variaciones posibles para El Salvador
        todas_variaciones = self.get_all_variations_for_country("El Salvador")
        
        df_mapeado = df_new_era[df_new_era[columna_tienda].isin(todas_variaciones)].copy()
        
        # Normalizar nombres de bodegas usando el nuevo sistema bidireccional
        df_mapeado['Bodega_Mapeada'] = df_mapeado[columna_tienda].apply(
            lambda x: self.normalize_bodega_name(x, target_format="stock", pais="El Salvador")
        )
        
        if len(df_mapeado) == 0:
            print("No hay registros mapeados para procesar en El Salvador")
            return {}
        
        # Definir categorías de ligas (mismas que Guatemala)
        categorias_ligas = {
            "MLB": ["MLB", "MLB properties"],
            "NBA": ["NBA", "NBA Properties"],
            "NFL": ["NFL", "NFL Properties"], 
            "MOTORSPORT": ["MOTORSPORT"],
            "ENTERTAINMENT": [
                "NEW ERA BRANDED", "ENTERTAINMENT", "MARCA PAIS", "WARNER BROS",
                "DISNEY", "LOONEY TUNES", "MARVEL", "DC", "UNIVERSAL", "PARAMOUNT"
            ],
            "ACCESSORIES": ["ACCESSORIES"]
        }
        
        # Importar ProductClassification para clasificar siluetas
        product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])
        
        # Clasificar siluetas solo para HEADWEAR
        df_mapeado['Tipo'] = df_mapeado.apply(
            lambda row: product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        # Inicializar estructura de resultados
        ventas_desglosadas = {}
        
        # Procesar por bodega
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            df_bodega = df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega]
            ventas_desglosadas[bodega] = {}
            
            # Procesar por liga
            for categoria, ligas in categorias_ligas.items():
                ventas_desglosadas[bodega][categoria] = {}
                
                if categoria == 'ACCESSORIES':
                    # Para ACCESSORIES, filtrar directamente por segmento
                    df_accessories = df_bodega[df_bodega['U_Segmento'] == 'ACCESSORIES']
                    ventas_desglosadas[bodega][categoria]['Stock'] = df_accessories['USD_Total_SI_CD'].sum()
                    ventas_desglosadas[bodega][categoria]['Ventas'] = df_accessories['USD_Total_SI_CD'].sum()
                else:
                    # Lógica original para otras ligas
                    df_liga = df_bodega[df_bodega['U_Liga'].isin(ligas)]
                    
                    # Planas (HEADWEAR + Planas)
                    df_planas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Planas')]
                    ventas_desglosadas[bodega][categoria]['Planas'] = df_planas['USD_Total_SI_CD'].sum()
                    
                    # Curvas (HEADWEAR + Curvas)  
                    df_curvas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Curvas')]
                    ventas_desglosadas[bodega][categoria]['Curvas'] = df_curvas['USD_Total_SI_CD'].sum()
                    
                    # Apparel
                    df_apparel = df_liga[df_liga['U_Segmento'] == 'APPAREL']
                    ventas_desglosadas[bodega][categoria]['Apparel'] = df_apparel['USD_Total_SI_CD'].sum()
        
        print(f"Ventas El Salvador desglosadas calculadas para {len(ventas_desglosadas)} bodegas")
        return ventas_desglosadas

    def procesar_ventas_costa_rica(self, df_ventas: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, float]]]:
        """
        Procesa el archivo de ventas de Costa Rica y retorna ventas desglosadas por bodega, liga y subcategoría
        Estructura: {bodega: {liga: {subcategoria: ventas}}}
        """
        if df_ventas is None or df_ventas.empty:
            return {}
        
        # Debug: Mostrar las columnas disponibles (solo en consola)
        print(f"Columnas disponibles en archivo de ventas Costa Rica: {list(df_ventas.columns)}")
        
        # Verificar columnas necesarias
        columnas_necesarias = ['U_Marca', 'U_Segmento', 'U_Liga', 'U_Silueta', 'USD_Total_SI_CD']
        columna_tienda = None
        
        # Buscar columna de tienda
        for col in df_ventas.columns:
            if 'tienda' in col.lower() or 'store' in col.lower() or 'bodega' in col.lower():
                columna_tienda = col
                break
        
        if columna_tienda is None:
            print("No se encontró columna de tienda en Costa Rica")
            return {}
            
        columnas_necesarias.append(columna_tienda)
        
        # Verificar que existan todas las columnas
        for col in columnas_necesarias[:-1]:  # Excepto la columna_tienda que ya verificamos
            if col not in df_ventas.columns:
                print(f"No se encontró columna {col} en Costa Rica")
                return {}
        
        print(f"Usando columna de tienda para Costa Rica: {columna_tienda}")
        
        # Filtrar por marca NEW ERA
        df_new_era = df_ventas[df_ventas['U_Marca'].str.upper() == 'NEW ERA'].copy()
        
        # Filtrar tiendas usando el nuevo sistema de mapeo bidireccional
        # Obtener todas las variaciones posibles para Costa Rica
        todas_variaciones = self.get_all_variations_for_country("Costa Rica")
        
        df_mapeado = df_new_era[df_new_era[columna_tienda].isin(todas_variaciones)].copy()
        
        # Normalizar nombres de bodegas usando el nuevo sistema bidireccional
        df_mapeado['Bodega_Mapeada'] = df_mapeado[columna_tienda].apply(
            lambda x: self.normalize_bodega_name(x, target_format="stock", pais="Costa Rica")
        )
        
        if len(df_mapeado) == 0:
            print("No hay registros mapeados para procesar en Costa Rica")
            return {}
        
        # Definir categorías de ligas (mismas que Guatemala y El Salvador)
        categorias_ligas = {
            "MLB": ["MLB", "MLB properties"],
            "NBA": ["NBA", "NBA Properties"],
            "NFL": ["NFL", "NFL Properties"], 
            "MOTORSPORT": ["MOTORSPORT"],
            "ENTERTAINMENT": [
                "NEW ERA BRANDED", "ENTERTAINMENT", "MARCA PAIS", "WARNER BROS",
                "DISNEY", "LOONEY TUNES", "MARVEL", "DC", "UNIVERSAL", "PARAMOUNT"
            ],
            "ACCESSORIES": ["ACCESSORIES"]
        }
        
        # Importar ProductClassification para clasificar siluetas
        product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])
        
        # Clasificar siluetas solo para HEADWEAR
        df_mapeado['Tipo'] = df_mapeado.apply(
            lambda row: product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        # Inicializar estructura de resultados
        ventas_desglosadas = {}
        
        # Procesar por bodega
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            df_bodega = df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega]
            ventas_desglosadas[bodega] = {}
            
            # Procesar por liga
            for categoria, ligas in categorias_ligas.items():
                ventas_desglosadas[bodega][categoria] = {}
                
                if categoria == 'ACCESSORIES':
                    # Para ACCESSORIES, filtrar directamente por segmento
                    df_accessories = df_bodega[df_bodega['U_Segmento'] == 'ACCESSORIES']
                    ventas_desglosadas[bodega][categoria]['Stock'] = df_accessories['USD_Total_SI_CD'].sum()
                    ventas_desglosadas[bodega][categoria]['Ventas'] = df_accessories['USD_Total_SI_CD'].sum()
                else:
                    # Lógica original para otras ligas
                    df_liga = df_bodega[df_bodega['U_Liga'].isin(ligas)]
                    
                    # Planas (HEADWEAR + Planas)
                    df_planas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Planas')]
                    ventas_desglosadas[bodega][categoria]['Planas'] = df_planas['USD_Total_SI_CD'].sum()
                    
                    # Curvas (HEADWEAR + Curvas)  
                    df_curvas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Curvas')]
                    ventas_desglosadas[bodega][categoria]['Curvas'] = df_curvas['USD_Total_SI_CD'].sum()
                    
                    # Apparel
                    df_apparel = df_liga[df_liga['U_Segmento'] == 'APPAREL']
                    ventas_desglosadas[bodega][categoria]['Apparel'] = df_apparel['USD_Total_SI_CD'].sum()
        
        print(f"Ventas Costa Rica desglosadas calculadas para {len(ventas_desglosadas)} bodegas")
        return ventas_desglosadas
    
    def procesar_ventas_honduras(self, df_ventas: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, float]]]:
        """
        Procesa el archivo de ventas de Honduras y retorna ventas desglosadas por bodega, liga y subcategoría
        Estructura: {bodega: {liga: {subcategoria: ventas}}}
        """
        print("🔥🔥🔥 EJECUTANDO procesar_ventas_honduras() 🔥🔥🔥")
        print(f"🔥 Archivo recibido: {df_ventas is not None}")
        if df_ventas is not None:
            print(f"🔥 Filas en archivo: {len(df_ventas)}")
            print(f"🔥 Columnas: {list(df_ventas.columns)}")
        
        if df_ventas is None or df_ventas.empty:
            print("🔥 ARCHIVO VACÍO O NULO - RETORNANDO {}")
            return {}
        
        print(f"Columnas disponibles en archivo de ventas Honduras: {list(df_ventas.columns)}")
        
        # Verificar columnas necesarias
        columnas_necesarias = ['U_Marca', 'U_Segmento', 'U_Liga', 'U_Silueta', 'USD_Total_SI_CD']
        columna_tienda = None
        
        # Buscar columna de tienda
        posibles_columnas_tienda = ['Tienda', 'Bodega', 'Store', 'Location']
        for col in posibles_columnas_tienda:
            if col in df_ventas.columns:
                columna_tienda = col
                break
        
        if columna_tienda is None:
            print("No se encontró columna de tienda en Honduras")
            return {}
            
        columnas_necesarias.append(columna_tienda)
        
        # Verificar que existan todas las columnas
        for col in columnas_necesarias:
            if col not in df_ventas.columns:
                print(f"No se encontró columna {col} en Honduras")
                return {}
        
        print(f"Usando columna de tienda para Honduras: {columna_tienda}")
        
        # Filtrar por marca NEW ERA
        df_new_era = df_ventas[df_ventas['U_Marca'].str.upper() == 'NEW ERA'].copy()
        
        # Filtrar tiendas usando el nuevo sistema de mapeo bidireccional
        # Obtener todas las variaciones posibles para Honduras
        todas_variaciones = self.get_all_variations_for_country("Honduras")
        
        df_mapeado = df_new_era[df_new_era[columna_tienda].isin(todas_variaciones)].copy()
        
        # Normalizar nombres de bodegas usando el nuevo sistema bidireccional
        df_mapeado['Bodega_Mapeada'] = df_mapeado[columna_tienda].apply(
            lambda x: self.normalize_bodega_name(x, target_format="stock", pais="Honduras")
        )
        
        # Importar ProductClassification para clasificar siluetas (IGUAL QUE OTROS PAÍSES)
        product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])
        
        # Clasificar siluetas solo para HEADWEAR (IGUAL QUE OTROS PAÍSES)
        df_mapeado['Tipo'] = df_mapeado.apply(
            lambda row: product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        print(f"Registros totales NEW ERA Honduras: {len(df_new_era)}")
        print(f"Registros con mapeo exitoso Honduras: {len(df_mapeado)}")
        print(f"Tiendas encontradas en archivo Honduras: {df_new_era[columna_tienda].unique()}")
        print(f"Bodegas mapeadas Honduras: {df_mapeado['Bodega_Mapeada'].unique()}")
        
        # DEBUG: Mostrar también tiendas que se mapearon exitosamente
        print(f"🎯 BODEGAS MAPEADAS EXITOSAMENTE:")
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            registros_bodega = len(df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega])
            print(f"  - {bodega}: {registros_bodega} registros")
        
        # DEBUG ESPECÍFICO PARA MEGA MALL SPS
        print(f"\n🔍 DEBUG ESPECÍFICO MEGA MALL SPS:")
        print(f"Todas las variaciones Honduras: {todas_variaciones}")
        mega_mall_variants = [v for v in todas_variaciones if 'MEGA' in v.upper() and 'MALL' in v.upper()]
        print(f"Variaciones que contienen MEGA MALL: {mega_mall_variants}")
        
        # Verificar si hay registros en el archivo original que contengan MEGA MALL
        tiendas_mega = df_new_era[df_new_era[columna_tienda].str.contains('MEGA.*MALL', case=False, na=False)]
        print(f"Registros con MEGA MALL en archivo original: {len(tiendas_mega)}")
        if len(tiendas_mega) > 0:
            print(f"Nombres exactos encontrados: {tiendas_mega[columna_tienda].unique()}")
            
        # Verificar mapeo específico
        for tienda_original in df_new_era[columna_tienda].unique():
            if 'MEGA' in str(tienda_original).upper() and 'MALL' in str(tienda_original).upper():
                mapeo_resultado = self.normalize_bodega_name(tienda_original, target_format="stock", pais="Honduras")
                print(f"Mapeo: '{tienda_original}' → '{mapeo_resultado}'")
        
        if df_mapeado.empty:
            print("No hay registros mapeados para procesar en Honduras")
            return {}
        
        # Definir categorías de ligas específicas para Honduras (sin las variantes "properties")
        categorias_ligas = {
            "MLB": ["MLB"],
            "NBA": ["NBA"],
            "NFL": ["NFL"],
            "MOTORSPORT": ["MOTORSPORT"],
            "ENTERTAINMENT": [
                "NEW ERA BRANDED", "ENTERTAINMENT", "MARCA PAIS", "WARNER BROS",
                "NONE LICENSED", "EUROPEAN SOCCER", "HONDURAS SOCCER"
            ],
            "ACCESSORIES": ["ACCESSORIES"]
        }
        
        # Debug crítico: Ver qué ligas están en el archivo  
        ligas_encontradas = sorted(df_mapeado['U_Liga'].unique())
        print(f"LIGAS ENCONTRADAS EN ARCHIVO HONDURAS: {ligas_encontradas}")
        print(f"LIGAS QUE ESPERAMOS:")
        for cat, ligas in categorias_ligas.items():
            if cat != 'ACCESSORIES':
                print(f"  {cat}: {ligas}")
        
        # Debug crítico: Ver qué siluetas están en el archivo
        siluetas_encontradas = sorted(df_mapeado[df_mapeado['U_Segmento'] == 'HEADWEAR']['U_Silueta'].unique())
        print(f"SILUETAS HEADWEAR ENCONTRADAS: {siluetas_encontradas}")
        
        
        print(f"Muestra de datos mapeados:")
        if not df_mapeado.empty:
            muestra = df_mapeado[['Bodega_Mapeada', columna_tienda, 'U_Liga', 'U_Segmento', 'U_Silueta', 'USD_Total_SI_CD']].head(10)
            print(muestra.to_string())
        
        ventas_desglosadas = {}
        
        # Procesar por bodega
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            df_bodega = df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega]
            ventas_desglosadas[bodega] = {}
            
            # Procesar por liga (IGUAL QUE GUATEMALA)
            for categoria, ligas in categorias_ligas.items():
                ventas_desglosadas[bodega][categoria] = {}
                
                if categoria == 'ACCESSORIES':
                    # Para ACCESSORIES, filtrar directamente por segmento
                    df_accessories = df_bodega[df_bodega['U_Segmento'] == 'ACCESSORIES']
                    ventas_desglosadas[bodega][categoria]['Stock'] = df_accessories['USD_Total_SI_CD'].sum()
                    ventas_desglosadas[bodega][categoria]['Ventas'] = df_accessories['USD_Total_SI_CD'].sum()
                else:
                    # Lógica original para otras ligas (IGUAL QUE GUATEMALA)
                    df_liga = df_bodega[df_bodega['U_Liga'].isin(ligas)]
                    
                    # Planas (HEADWEAR + Planas) - USANDO PRODUCT CLASSIFIER
                    df_planas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Planas')]
                    ventas_desglosadas[bodega][categoria]['Planas'] = df_planas['USD_Total_SI_CD'].sum()
                    
                    # Curvas (HEADWEAR + Curvas) - USANDO PRODUCT CLASSIFIER  
                    df_curvas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Curvas')]
                    ventas_desglosadas[bodega][categoria]['Curvas'] = df_curvas['USD_Total_SI_CD'].sum()
                    
                    # Apparel
                    df_apparel = df_liga[df_liga['U_Segmento'] == 'APPAREL']
                    ventas_desglosadas[bodega][categoria]['Apparel'] = df_apparel['USD_Total_SI_CD'].sum()
        
        print(f"Ventas Honduras desglosadas calculadas para {len(ventas_desglosadas)} bodegas")
        
        # Debug final: Mostrar resumen de ventas calculadas
        print("RESUMEN VENTAS HONDURAS:")
        for bodega, categorias in ventas_desglosadas.items():
            print(f"  {bodega}:")
            for categoria, subcategorias in categorias.items():
                total_cat = 0
                if categoria == 'ACCESSORIES':
                    total_cat = subcategorias.get('Ventas', 0)
                else:
                    total_cat = subcategorias.get('Planas', 0) + subcategorias.get('Curvas', 0) + subcategorias.get('Apparel', 0)
                if total_cat > 0:
                    print(f"    {categoria}: ${total_cat} ({subcategorias})")
                else:
                    print(f"    {categoria}: $0 (SIN DATOS)")
        
        return ventas_desglosadas

    def procesar_ventas_panama(self, df_ventas: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, float]]]:
        """
        Procesa el archivo de ventas de PANAMA y retorna ventas desglosadas por bodega, liga y subcategoría
        Estructura: {bodega: {liga: {subcategoria: ventas}}}
        """
        if df_ventas is None or df_ventas.empty:
            return {}
        
        print(f"Columnas disponibles en archivo de ventas PANAMA: {list(df_ventas.columns)}")
        
        # Verificar columnas necesarias
        columnas_necesarias = ['U_Marca', 'U_Segmento', 'U_Liga', 'U_Silueta', 'USD_Total_SI_CD']
        columna_tienda = None
        
        # Buscar columna de tienda
        posibles_columnas_tienda = ['Tienda', 'Bodega', 'Store', 'Location']
        for col in posibles_columnas_tienda:
            if col in df_ventas.columns:
                columna_tienda = col
                break
        
        if columna_tienda is None:
            print("No se encontró columna de tienda en PANAMA")
            return {}
            
        columnas_necesarias.append(columna_tienda)
        
        # Verificar que existan todas las columnas
        for col in columnas_necesarias:
            if col not in df_ventas.columns:
                print(f"No se encontró columna {col} en PANAMA")
                return {}
        
        print(f"Usando columna de tienda para PANAMA: {columna_tienda}")
        
        # Filtrar por marca NEW ERA
        df_new_era = df_ventas[df_ventas['U_Marca'].str.upper() == 'NEW ERA'].copy()
        
        # Filtrar tiendas usando el nuevo sistema de mapeo bidireccional
        # Obtener todas las variaciones posibles para PANAMA
        todas_variaciones = self.get_all_variations_for_country("PANAMA")
        
        df_mapeado = df_new_era[df_new_era[columna_tienda].isin(todas_variaciones)].copy()
        
        # Normalizar nombres de bodegas usando el nuevo sistema bidireccional
        df_mapeado['Bodega_Mapeada'] = df_mapeado[columna_tienda].apply(
            lambda x: self.normalize_bodega_name(x, target_format="stock", pais="PANAMA")
        )
        
        # Importar ProductClassification para clasificar siluetas (IGUAL QUE OTROS PAÍSES)
        product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])
        
        # Clasificar siluetas solo para HEADWEAR (IGUAL QUE OTROS PAÍSES)
        df_mapeado['Tipo'] = df_mapeado.apply(
            lambda row: product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        print(f"Registros totales NEW ERA PANAMA: {len(df_new_era)}")
        print(f"Registros con mapeo exitoso PANAMA: {len(df_mapeado)}")
        print(f"Tiendas encontradas en archivo PANAMA: {df_new_era[columna_tienda].unique()}")
        print(f"Bodegas mapeadas PANAMA: {df_mapeado['Bodega_Mapeada'].unique()}")
        print(f"Todas las variaciones PANAMA: {todas_variaciones}")
        print(f"Sistema de mapeo PANAMA: {self.country_mappings.get('PANAMA', {})}")
        
        # DEBUG: Mostrar bodegas que se mapearon exitosamente
        print(f"🎯 BODEGAS MAPEADAS EXITOSAMENTE EN PANAMA:")
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            registros_bodega = len(df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega])
            print(f"  - {bodega}: {registros_bodega} registros")
        
        if df_mapeado.empty:
            print("No hay registros mapeados para procesar en PANAMA")
            return {}
        
        # Definir categorías de ligas específicas para PANAMA (igual que Guatemala)
        categorias_ligas = {
            "MLB": ["MLB"],
            "NBA": ["NBA"],
            "NFL": ["NFL"],
            "MOTORSPORT": ["MOTORSPORT"],
            "ENTERTAINMENT": [
                "NEW ERA BRANDED", "ENTERTAINMENT", "MARCA PAIS", "WARNER BROS",
                "NONE LICENSED", "EUROPEAN SOCCER"
            ],
            "ACCESSORIES": ["ACCESSORIES"]
        }
        
        # Debug: Ver qué ligas están en el archivo  
        ligas_encontradas = sorted(df_mapeado['U_Liga'].unique())
        print(f"LIGAS ENCONTRADAS EN ARCHIVO PANAMA: {ligas_encontradas}")
        
        # Debug: Ver qué siluetas están en el archivo
        siluetas_encontradas = sorted(df_mapeado[df_mapeado['U_Segmento'] == 'HEADWEAR']['U_Silueta'].unique())
        print(f"SILUETAS HEADWEAR ENCONTRADAS EN PANAMA: {siluetas_encontradas}")
        
        print(f"Muestra de datos mapeados PANAMA:")
        if not df_mapeado.empty:
            muestra = df_mapeado[['Bodega_Mapeada', columna_tienda, 'U_Liga', 'U_Segmento', 'U_Silueta', 'USD_Total_SI_CD']].head(10)
            print(muestra.to_string())
        
        ventas_desglosadas = {}
        
        # Procesar por bodega
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            df_bodega = df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega]
            ventas_desglosadas[bodega] = {}
            
            # Procesar por liga (IGUAL QUE GUATEMALA)
            for categoria, ligas in categorias_ligas.items():
                ventas_desglosadas[bodega][categoria] = {}
                
                if categoria == 'ACCESSORIES':
                    # Para ACCESSORIES, filtrar directamente por segmento
                    df_accessories = df_bodega[df_bodega['U_Segmento'] == 'ACCESSORIES']
                    ventas_desglosadas[bodega][categoria]['Stock'] = df_accessories['USD_Total_SI_CD'].sum()
                    ventas_desglosadas[bodega][categoria]['Ventas'] = df_accessories['USD_Total_SI_CD'].sum()
                else:
                    # Lógica original para otras ligas (IGUAL QUE GUATEMALA)
                    df_liga = df_bodega[df_bodega['U_Liga'].isin(ligas)]
                    
                    # Planas (HEADWEAR + Planas) - USANDO PRODUCT CLASSIFIER
                    df_planas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Planas')]
                    ventas_desglosadas[bodega][categoria]['Planas'] = df_planas['USD_Total_SI_CD'].sum()
                    
                    # Curvas (HEADWEAR + Curvas) - USANDO PRODUCT CLASSIFIER  
                    df_curvas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Curvas')]
                    ventas_desglosadas[bodega][categoria]['Curvas'] = df_curvas['USD_Total_SI_CD'].sum()
                    
                    # Apparel
                    df_apparel = df_liga[df_liga['U_Segmento'] == 'APPAREL']
                    ventas_desglosadas[bodega][categoria]['Apparel'] = df_apparel['USD_Total_SI_CD'].sum()
        
        print(f"Ventas PANAMA desglosadas calculadas para {len(ventas_desglosadas)} bodegas")
        
        # Debug final: Mostrar resumen de ventas calculadas
        print("RESUMEN VENTAS PANAMA:")
        for bodega, categorias in ventas_desglosadas.items():
            print(f"  {bodega}:")
            for categoria, subcategorias in categorias.items():
                total_cat = 0
                if categoria == 'ACCESSORIES':
                    total_cat = subcategorias.get('Ventas', 0)
                else:
                    total_cat = subcategorias.get('Planas', 0) + subcategorias.get('Curvas', 0) + subcategorias.get('Apparel', 0)
                if total_cat > 0:
                    print(f"    {categoria}: ${total_cat} ({subcategorias})")
                else:
                    print(f"    {categoria}: $0 (SIN DATOS)")
        
        return ventas_desglosadas

    def procesar_cantidades_guatemala(self, df_ventas: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, float]]]:
        """
        Procesa el archivo de ventas de Guatemala basado en columna 'Cantidad' para tabla solo-ventas
        Estructura: {bodega: {liga: {subcategoria: cantidad}}}
        """
        if df_ventas is None or df_ventas.empty:
            return {}
        
        print(f"Procesando CANTIDADES Guatemala - Columnas disponibles: {list(df_ventas.columns)}")
        
        # Verificar columnas necesarias
        columnas_necesarias = ['U_Marca', 'U_Segmento', 'U_Liga', 'U_Silueta', 'Cantidad', 'USD_Total_SI_CD']
        columna_tienda = None
        
        # Buscar columna de tienda
        posibles_columnas_tienda = ['Tienda', 'Bodega', 'Store', 'Location']
        for col in posibles_columnas_tienda:
            if col in df_ventas.columns:
                columna_tienda = col
                break
        
        if columna_tienda is None:
            print("No se encontró columna de tienda en Guatemala (solo-ventas)")
            return {}
            
        columnas_necesarias.append(columna_tienda)
        
        # Verificar que existan todas las columnas
        for col in columnas_necesarias:
            if col not in df_ventas.columns:
                print(f"No se encontró columna {col} en Guatemala (solo-ventas)")
                return {}
        
        print(f"Usando columna de tienda para Guatemala (solo-ventas): {columna_tienda}")
        
        # Convertir columnas numéricas
        df_ventas['Cantidad'] = pd.to_numeric(df_ventas['Cantidad'], errors='coerce').fillna(0)
        df_ventas['USD_Total_SI_CD'] = pd.to_numeric(df_ventas['USD_Total_SI_CD'], errors='coerce').fillna(0)
        
        # Filtrar por marca NEW ERA
        df_new_era = df_ventas[df_ventas['U_Marca'].str.upper() == 'NEW ERA'].copy()
        
        # Usar listado exacto de tiendas especificado por el usuario
        tiendas_guatemala = [
            "NE CAYALA", "NE CHIMALTENANGO", "NE CONCEPCION", "NE INTERPLAZA ESCUINTLA",
            "NE INTERXELA", "NE METRONORTE", "NE METROPLAZA JUTIAPA", "NE MIRAFLORES",
            "NE NARANJO", "NE OAKLAND", "NE PASEO ANTIGUA", "NE PERIROOSVELT",
            "NE PLAZA MAGDALENA", "NE PLAZA VIDERE", "NE PORTALES", "NE PRADERA CHIQUIMULA",
            "NE PRADERA ESCUINTLA", "NE PRADERA HUEHUETENANGO", "NE PRADERA VISTARES",
            "NE PRADERA XELA", "NE SANTA CLARA", "NEW ERA METROCENTRO VILLA NUEVA",
            "New Era Puerto Barrios"
        ]
        
        # Filtrar solo las tiendas especificadas
        df_mapeado = df_new_era[df_new_era[columna_tienda].isin(tiendas_guatemala)].copy()
        
        # Renombrar columna para mantener consistencia con el resto del código
        df_mapeado['Bodega_Mapeada'] = df_mapeado[columna_tienda]
        
        # Importar ProductClassification para clasificar siluetas
        product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])
        
        # Clasificar siluetas solo para HEADWEAR
        df_mapeado['Tipo'] = df_mapeado.apply(
            lambda row: product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        print(f"Registros totales NEW ERA Guatemala (solo-ventas): {len(df_new_era)}")
        print(f"Registros con mapeo exitoso Guatemala (solo-ventas): {len(df_mapeado)}")
        print(f"Tiendas encontradas en archivo Guatemala: {df_new_era[columna_tienda].unique()}")
        print(f"Bodegas mapeadas Guatemala: {df_mapeado['Bodega_Mapeada'].unique()}")
        
        if df_mapeado.empty:
            print("No hay registros mapeados para procesar en Guatemala (solo-ventas)")
            return {}
        
        # Definir categorías de ligas específicas para Guatemala (IDÉNTICAS al modo stock+ventas)
        categorias_ligas = {
            "MLB": ["MLB", "MLB properties"],
            "NBA": ["NBA", "NBA Properties"],
            "NFL": ["NFL", "NFL Properties"],
            "MOTORSPORT": ["MOTORSPORT"],
            "ENTERTAINMENT": [
                "NEW ERA BRANDED", "ENTERTAINMENT", "MARCA PAIS", "WARNER BROS",
                "DISNEY", "LOONEY TUNES", "MARVEL", "DC", "UNIVERSAL", "PARAMOUNT"
            ],
            "ACCESSORIES": ["ACCESSORIES"]
        }
        
        # Debug: Ver qué ligas están en el archivo  
        ligas_encontradas = sorted(df_mapeado['U_Liga'].unique())
        print(f"LIGAS ENCONTRADAS EN ARCHIVO GUATEMALA (solo-ventas): {ligas_encontradas}")
        
        cantidades_desglosadas = {}
        
        # Procesar por bodega
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            df_bodega = df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega]
            cantidades_desglosadas[bodega] = {}
            
            # Procesar por liga
            for categoria, ligas in categorias_ligas.items():
                cantidades_desglosadas[bodega][categoria] = {}
                
                if categoria == 'ACCESSORIES':
                    # Para ACCESSORIES, filtrar directamente por segmento
                    df_accessories = df_bodega[df_bodega['U_Segmento'] == 'ACCESSORIES']
                    cantidades_desglosadas[bodega][categoria]['Stock'] = {
                        'Cantidad': df_accessories['Cantidad'].sum(),
                        'USD': df_accessories['USD_Total_SI_CD'].sum()
                    }
                else:
                    # Lógica para otras ligas
                    df_liga = df_bodega[df_bodega['U_Liga'].isin(ligas)]
                    
                    # Planas (HEADWEAR + Planas)
                    df_planas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Planas')]
                    cantidades_desglosadas[bodega][categoria]['Planas'] = {
                        'Cantidad': df_planas['Cantidad'].sum(),
                        'USD': df_planas['USD_Total_SI_CD'].sum()
                    }
                    
                    # Curvas (HEADWEAR + Curvas)  
                    df_curvas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Curvas')]
                    cantidades_desglosadas[bodega][categoria]['Curvas'] = {
                        'Cantidad': df_curvas['Cantidad'].sum(),
                        'USD': df_curvas['USD_Total_SI_CD'].sum()
                    }
                    
                    # Apparel
                    df_apparel = df_liga[df_liga['U_Segmento'] == 'APPAREL']
                    cantidades_desglosadas[bodega][categoria]['Apparel'] = {
                        'Cantidad': df_apparel['Cantidad'].sum(),
                        'USD': df_apparel['USD_Total_SI_CD'].sum()
                    }
        
        print(f"Cantidades Guatemala desglosadas calculadas para {len(cantidades_desglosadas)} bodegas")
        
        # Debug final: Mostrar resumen de cantidades calculadas
        print("RESUMEN CANTIDADES GUATEMALA (solo-ventas):")
        for bodega, categorias in cantidades_desglosadas.items():
            print(f"  {bodega}:")
            for categoria, subcategorias in categorias.items():
                total_cat = 0
                total_usd = 0
                if categoria == 'ACCESSORIES':
                    stock_data = subcategorias.get('Stock', {})
                    if isinstance(stock_data, dict):
                        total_cat = stock_data.get('Cantidad', 0)
                        total_usd = stock_data.get('USD', 0)
                else:
                    for subcat in ['Planas', 'Curvas', 'Apparel']:
                        subcat_data = subcategorias.get(subcat, {})
                        if isinstance(subcat_data, dict):
                            total_cat += subcat_data.get('Cantidad', 0)
                            total_usd += subcat_data.get('USD', 0)
                
                if total_cat > 0 or total_usd > 0:
                    print(f"    {categoria}: {total_cat} unidades, ${total_usd:.2f} USD")
                else:
                    print(f"    {categoria}: 0 unidades, $0.00 USD (SIN DATOS)")
        
        return cantidades_desglosadas

    def procesar_cantidades_el_salvador(self, df_ventas: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, float]]]:
        """
        Procesa el archivo de ventas de El Salvador basado en columna 'Cantidad' para tabla solo-ventas
        Estructura: {bodega: {liga: {subcategoria: cantidad}}}
        """
        if df_ventas is None or df_ventas.empty:
            return {}
        
        print(f"Procesando CANTIDADES El Salvador - Columnas disponibles: {list(df_ventas.columns)}")
        
        # Verificar columnas necesarias
        columnas_necesarias = ['U_Marca', 'U_Segmento', 'U_Liga', 'U_Silueta', 'Cantidad', 'USD_Total_SI_CD']
        columna_tienda = None
        
        # Buscar columna de tienda
        posibles_columnas_tienda = ['Tienda', 'Bodega', 'Store', 'Location']
        for col in posibles_columnas_tienda:
            if col in df_ventas.columns:
                columna_tienda = col
                break
        
        if columna_tienda is None:
            print("No se encontró columna de tienda en El Salvador (solo-ventas)")
            return {}
            
        columnas_necesarias.append(columna_tienda)
        
        # Verificar que existan todas las columnas
        for col in columnas_necesarias:
            if col not in df_ventas.columns:
                print(f"No se encontró columna {col} en El Salvador (solo-ventas)")
                return {}
        
        print(f"Usando columna de tienda para El Salvador (solo-ventas): {columna_tienda}")
        
        # Convertir columnas numéricas
        df_ventas['Cantidad'] = pd.to_numeric(df_ventas['Cantidad'], errors='coerce').fillna(0)
        df_ventas['USD_Total_SI_CD'] = pd.to_numeric(df_ventas['USD_Total_SI_CD'], errors='coerce').fillna(0)
        
        # Filtrar por marca NEW ERA
        df_new_era = df_ventas[df_ventas['U_Marca'].str.upper() == 'NEW ERA'].copy()
        
        # Usar listado de tiendas de El Salvador del diccionario bidireccional
        tiendas_el_salvador = [
            "NE METROCENTRO LOURDES", "NEW ERA MULTIPLAZA", "NE METROCENTRO", 
            "NE METROCENTRO SANTA ANA", "NE PLAZA MUNDO SOYAPANGO", "NE PLAZA MUNDO USULUTÁN",
            "NE METROCENTRO SAN MIGUEL", "NEW ERA EL PASEO"
        ]
        
        # Filtrar solo las tiendas especificadas
        df_mapeado = df_new_era[df_new_era[columna_tienda].isin(tiendas_el_salvador)].copy()
        
        # Renombrar columna para mantener consistencia con el resto del código
        df_mapeado['Bodega_Mapeada'] = df_mapeado[columna_tienda]
        
        # Importar ProductClassification para clasificar siluetas
        product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])
        
        # Clasificar siluetas solo para HEADWEAR
        df_mapeado['Tipo'] = df_mapeado.apply(
            lambda row: product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        print(f"Registros totales NEW ERA El Salvador (solo-ventas): {len(df_new_era)}")
        print(f"Registros con mapeo exitoso El Salvador (solo-ventas): {len(df_mapeado)}")
        print(f"Tiendas encontradas en archivo El Salvador: {df_new_era[columna_tienda].unique()}")
        print(f"Bodegas mapeadas El Salvador: {df_mapeado['Bodega_Mapeada'].unique()}")
        
        if df_mapeado.empty:
            print("No hay registros mapeados para procesar en El Salvador (solo-ventas)")
            return {}
        
        # Definir categorías de ligas específicas para El Salvador (IDÉNTICAS al modo stock+ventas)
        categorias_ligas = {
            "MLB": ["MLB", "MLB properties"],
            "NBA": ["NBA", "NBA Properties"],
            "NFL": ["NFL", "NFL Properties"],
            "MOTORSPORT": ["MOTORSPORT"],
            "ENTERTAINMENT": [
                "NEW ERA BRANDED", "ENTERTAINMENT", "MARCA PAIS", "WARNER BROS",
                "DISNEY", "LOONEY TUNES", "MARVEL", "DC", "UNIVERSAL", "PARAMOUNT"
            ],
            "ACCESSORIES": ["ACCESSORIES"]
        }
        
        # Debug: Ver qué ligas están en el archivo  
        ligas_encontradas = sorted(df_mapeado['U_Liga'].unique())
        print(f"LIGAS ENCONTRADAS EN ARCHIVO EL SALVADOR (solo-ventas): {ligas_encontradas}")
        
        cantidades_desglosadas = {}
        
        # Procesar por bodega
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            df_bodega = df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega]
            cantidades_desglosadas[bodega] = {}
            
            # Procesar por liga
            for categoria, ligas in categorias_ligas.items():
                cantidades_desglosadas[bodega][categoria] = {}
                
                if categoria == 'ACCESSORIES':
                    # Para ACCESSORIES, filtrar directamente por segmento
                    df_accessories = df_bodega[df_bodega['U_Segmento'] == 'ACCESSORIES']
                    cantidades_desglosadas[bodega][categoria]['Stock'] = {
                        'Cantidad': df_accessories['Cantidad'].sum(),
                        'USD': df_accessories['USD_Total_SI_CD'].sum()
                    }
                else:
                    # Lógica para otras ligas
                    df_liga = df_bodega[df_bodega['U_Liga'].isin(ligas)]
                    
                    # Planas (HEADWEAR + Planas)
                    df_planas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Planas')]
                    cantidades_desglosadas[bodega][categoria]['Planas'] = {
                        'Cantidad': df_planas['Cantidad'].sum(),
                        'USD': df_planas['USD_Total_SI_CD'].sum()
                    }
                    
                    # Curvas (HEADWEAR + Curvas)  
                    df_curvas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Curvas')]
                    cantidades_desglosadas[bodega][categoria]['Curvas'] = {
                        'Cantidad': df_curvas['Cantidad'].sum(),
                        'USD': df_curvas['USD_Total_SI_CD'].sum()
                    }
                    
                    # Apparel
                    df_apparel = df_liga[df_liga['U_Segmento'] == 'APPAREL']
                    cantidades_desglosadas[bodega][categoria]['Apparel'] = {
                        'Cantidad': df_apparel['Cantidad'].sum(),
                        'USD': df_apparel['USD_Total_SI_CD'].sum()
                    }
        
        print(f"Cantidades El Salvador desglosadas calculadas para {len(cantidades_desglosadas)} bodegas")
        
        # Debug final: Mostrar resumen de cantidades calculadas
        print("RESUMEN CANTIDADES EL SALVADOR (solo-ventas):")
        for bodega, categorias in cantidades_desglosadas.items():
            print(f"  {bodega}:")
            for categoria, subcategorias in categorias.items():
                total_cat = 0
                total_usd = 0
                if categoria == 'ACCESSORIES':
                    stock_data = subcategorias.get('Stock', {})
                    if isinstance(stock_data, dict):
                        total_cat = stock_data.get('Cantidad', 0)
                        total_usd = stock_data.get('USD', 0)
                else:
                    for subcat in ['Planas', 'Curvas', 'Apparel']:
                        subcat_data = subcategorias.get(subcat, {})
                        if isinstance(subcat_data, dict):
                            total_cat += subcat_data.get('Cantidad', 0)
                            total_usd += subcat_data.get('USD', 0)
                
                if total_cat > 0 or total_usd > 0:
                    print(f"    {categoria}: {total_cat} unidades, ${total_usd:.2f} USD")
                else:
                    print(f"    {categoria}: 0 unidades, $0.00 USD (SIN DATOS)")
        
        return cantidades_desglosadas

    def procesar_cantidades_honduras(self, df_ventas: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, float]]]:
        """
        Procesa el archivo de ventas de Honduras basado en columna 'Cantidad' para tabla solo-ventas
        Estructura: {bodega: {liga: {subcategoria: cantidad}}}
        """
        if df_ventas is None or df_ventas.empty:
            return {}
        
        print(f"Procesando CANTIDADES Honduras - Columnas disponibles: {list(df_ventas.columns)}")
        
        # Verificar columnas necesarias
        columnas_necesarias = ['U_Marca', 'U_Segmento', 'U_Liga', 'U_Silueta', 'Cantidad', 'USD_Total_SI_CD']
        columna_tienda = None
        
        # Buscar columna de tienda
        posibles_columnas_tienda = ['Tienda', 'Bodega', 'Store', 'Location']
        for col in posibles_columnas_tienda:
            if col in df_ventas.columns:
                columna_tienda = col
                break
        
        if columna_tienda is None:
            print("No se encontró columna de tienda en Honduras (solo-ventas)")
            return {}
            
        columnas_necesarias.append(columna_tienda)
        
        # Verificar que existan todas las columnas
        for col in columnas_necesarias:
            if col not in df_ventas.columns:
                print(f"No se encontró columna {col} en Honduras (solo-ventas)")
                return {}
        
        print(f"Usando columna de tienda para Honduras (solo-ventas): {columna_tienda}")
        
        # Convertir columnas numéricas
        df_ventas['Cantidad'] = pd.to_numeric(df_ventas['Cantidad'], errors='coerce').fillna(0)
        df_ventas['USD_Total_SI_CD'] = pd.to_numeric(df_ventas['USD_Total_SI_CD'], errors='coerce').fillna(0)
        
        # Filtrar por marca NEW ERA
        df_new_era = df_ventas[df_ventas['U_Marca'].str.upper() == 'NEW ERA'].copy()
        
        # Usar el mismo sistema bidireccional que procesar_ventas_honduras()
        # Obtener todas las variaciones posibles para Honduras
        todas_variaciones = self.get_all_variations_for_country("Honduras")
        
        df_mapeado = df_new_era[df_new_era[columna_tienda].isin(todas_variaciones)].copy()
        
        # Normalizar nombres de bodegas usando el mismo sistema bidireccional
        df_mapeado['Bodega_Mapeada'] = df_mapeado[columna_tienda].apply(
            lambda x: self.normalize_bodega_name(x, target_format="stock", pais="Honduras")
        )
        
        # Importar ProductClassification para clasificar siluetas
        product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])
        
        # Clasificar siluetas solo para HEADWEAR
        df_mapeado['Tipo'] = df_mapeado.apply(
            lambda row: product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        print(f"Registros totales NEW ERA Honduras (solo-ventas): {len(df_new_era)}")
        print(f"Registros con mapeo exitoso Honduras (solo-ventas): {len(df_mapeado)}")
        print(f"Tiendas encontradas en archivo Honduras: {df_new_era[columna_tienda].unique()}")
        print(f"Bodegas mapeadas Honduras: {df_mapeado['Bodega_Mapeada'].unique()}")
        
        
        if df_mapeado.empty:
            print("No hay registros mapeados para procesar en Honduras (solo-ventas)")
            return {}
        
        # Definir categorías de ligas específicas para Honduras (IDÉNTICAS al modo stock+ventas)
        categorias_ligas = {
            "MLB": ["MLB", "MLB properties"],
            "NBA": ["NBA", "NBA Properties"],
            "NFL": ["NFL", "NFL Properties"],
            "MOTORSPORT": ["MOTORSPORT"],
            "ENTERTAINMENT": [
                "NEW ERA BRANDED", "ENTERTAINMENT", "MARCA PAIS", "WARNER BROS",
                "DISNEY", "LOONEY TUNES", "MARVEL", "DC", "UNIVERSAL", "PARAMOUNT"
            ],
            "ACCESSORIES": ["ACCESSORIES"]
        }
        
        # Debug: Ver qué ligas están en el archivo  
        ligas_encontradas = sorted(df_mapeado['U_Liga'].unique())
        print(f"LIGAS ENCONTRADAS EN ARCHIVO HONDURAS (solo-ventas): {ligas_encontradas}")
        
        cantidades_desglosadas = {}
        
        # Procesar por bodega
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            df_bodega = df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega]
            cantidades_desglosadas[bodega] = {}
            
            # Procesar por liga
            for categoria, ligas in categorias_ligas.items():
                cantidades_desglosadas[bodega][categoria] = {}
                
                if categoria == 'ACCESSORIES':
                    # Para ACCESSORIES, filtrar directamente por segmento
                    df_accessories = df_bodega[df_bodega['U_Segmento'] == 'ACCESSORIES']
                    cantidades_desglosadas[bodega][categoria]['Stock'] = {
                        'Cantidad': df_accessories['Cantidad'].sum(),
                        'USD': df_accessories['USD_Total_SI_CD'].sum()
                    }
                    # También crear entrada para Ventas (USD) para consistencia con el sistema
                    cantidades_desglosadas[bodega][categoria]['Ventas'] = {
                        'Cantidad': df_accessories['Cantidad'].sum(),
                        'USD': df_accessories['USD_Total_SI_CD'].sum()
                    }
                else:
                    # Lógica para otras ligas
                    df_liga = df_bodega[df_bodega['U_Liga'].isin(ligas)]
                    
                    # Planas (HEADWEAR + Planas)
                    df_planas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Planas')]
                    cantidades_desglosadas[bodega][categoria]['Planas'] = {
                        'Cantidad': df_planas['Cantidad'].sum(),
                        'USD': df_planas['USD_Total_SI_CD'].sum()
                    }
                    
                    # Curvas (HEADWEAR + Curvas)  
                    df_curvas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Curvas')]
                    cantidades_desglosadas[bodega][categoria]['Curvas'] = {
                        'Cantidad': df_curvas['Cantidad'].sum(),
                        'USD': df_curvas['USD_Total_SI_CD'].sum()
                    }
                    
                    # Apparel
                    df_apparel = df_liga[df_liga['U_Segmento'] == 'APPAREL']
                    cantidades_desglosadas[bodega][categoria]['Apparel'] = {
                        'Cantidad': df_apparel['Cantidad'].sum(),
                        'USD': df_apparel['USD_Total_SI_CD'].sum()
                    }
        
        print(f"Cantidades Honduras desglosadas calculadas para {len(cantidades_desglosadas)} bodegas")
        
        
        # Debug final: Mostrar resumen de cantidades calculadas
        print("RESUMEN CANTIDADES HONDURAS (solo-ventas):")
        for bodega, categorias in cantidades_desglosadas.items():
            print(f"  {bodega}:")
            for categoria, subcategorias in categorias.items():
                total_cat = 0
                total_usd = 0
                if categoria == 'ACCESSORIES':
                    stock_data = subcategorias.get('Stock', {})
                    if isinstance(stock_data, dict):
                        total_cat = stock_data.get('Cantidad', 0)
                        total_usd = stock_data.get('USD', 0)
                else:
                    for subcat in ['Planas', 'Curvas', 'Apparel']:
                        subcat_data = subcategorias.get(subcat, {})
                        if isinstance(subcat_data, dict):
                            total_cat += subcat_data.get('Cantidad', 0)
                            total_usd += subcat_data.get('USD', 0)
                
                if total_cat > 0 or total_usd > 0:
                    print(f"    {categoria}: {total_cat} unidades, ${total_usd:.2f} USD")
                else:
                    print(f"    {categoria}: 0 unidades, $0.00 USD (SIN DATOS)")
        
        return cantidades_desglosadas

    def procesar_cantidades_costa_rica(self, df_ventas: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, float]]]:
        """
        Procesa el archivo de ventas de Costa Rica basado en columna 'Cantidad' para tabla solo-ventas
        Estructura: {bodega: {liga: {subcategoria: {'Cantidad': X, 'USD': Y}}}}
        """
        if df_ventas is None or df_ventas.empty:
            return {}
        
        print(f"Procesando CANTIDADES Costa Rica - Columnas disponibles: {list(df_ventas.columns)}")
        
        # Verificar columnas necesarias
        columnas_necesarias = ['U_Marca', 'U_Segmento', 'U_Liga', 'U_Silueta', 'Cantidad', 'USD_Total_SI_CD']
        columna_tienda = None
        
        # Buscar columna de tienda
        posibles_columnas_tienda = ['Tienda', 'Bodega', 'Store', 'Location']
        for col in posibles_columnas_tienda:
            if col in df_ventas.columns:
                columna_tienda = col
                break
        
        if columna_tienda is None:
            print("No se encontró columna de tienda en Costa Rica (solo-ventas)")
            return {}
            
        columnas_necesarias.append(columna_tienda)
        
        # Verificar que existan todas las columnas
        for col in columnas_necesarias:
            if col not in df_ventas.columns:
                print(f"No se encontró columna {col} en Costa Rica (solo-ventas)")
                return {}
        
        print(f"Usando columna de tienda para Costa Rica (solo-ventas): {columna_tienda}")
        
        # Convertir columnas numéricas
        df_ventas['Cantidad'] = pd.to_numeric(df_ventas['Cantidad'], errors='coerce').fillna(0)
        df_ventas['USD_Total_SI_CD'] = pd.to_numeric(df_ventas['USD_Total_SI_CD'], errors='coerce').fillna(0)
        
        # Filtrar por marca NEW ERA
        df_new_era = df_ventas[df_ventas['U_Marca'].str.upper() == 'NEW ERA'].copy()
        
        # Usar el mismo sistema bidireccional que otros países
        # Obtener todas las variaciones posibles para Costa Rica
        todas_variaciones = self.get_all_variations_for_country("Costa Rica")
        
        df_mapeado = df_new_era[df_new_era[columna_tienda].isin(todas_variaciones)].copy()
        
        # Normalizar nombres de bodegas usando el mismo sistema bidireccional
        df_mapeado['Bodega_Mapeada'] = df_mapeado[columna_tienda].apply(
            lambda x: self.normalize_bodega_name(x, target_format="stock", pais="Costa Rica")
        )
        
        # Importar ProductClassification para clasificar siluetas
        product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])
        
        # Clasificar siluetas solo para HEADWEAR
        df_mapeado['Tipo'] = df_mapeado.apply(
            lambda row: product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        print(f"Registros totales NEW ERA Costa Rica (solo-ventas): {len(df_new_era)}")
        print(f"Registros con mapeo exitoso Costa Rica (solo-ventas): {len(df_mapeado)}")
        print(f"Tiendas encontradas en archivo Costa Rica: {df_new_era[columna_tienda].unique()}")
        print(f"Bodegas mapeadas Costa Rica: {df_mapeado['Bodega_Mapeada'].unique()}")
        
        if df_mapeado.empty:
            print("No hay registros mapeados para procesar en Costa Rica (solo-ventas)")
            return {}
        
        # Definir categorías de ligas específicas para Costa Rica (IDÉNTICAS al modo stock+ventas)
        categorias_ligas = {
            "MLB": ["MLB", "MLB properties"],
            "NBA": ["NBA", "NBA Properties"],
            "NFL": ["NFL", "NFL Properties"],
            "MOTORSPORT": ["MOTORSPORT"],
            "ENTERTAINMENT": [
                "NEW ERA BRANDED", "ENTERTAINMENT", "MARCA PAIS", "WARNER BROS",
                "DISNEY", "LOONEY TUNES", "MARVEL", "DC", "UNIVERSAL", "PARAMOUNT"
            ],
            "ACCESSORIES": ["ACCESSORIES"]
        }
        
        # Debug: Ver qué ligas están en el archivo  
        ligas_encontradas = sorted(df_mapeado['U_Liga'].unique())
        print(f"LIGAS ENCONTRADAS EN ARCHIVO COSTA RICA (solo-ventas): {ligas_encontradas}")
        
        cantidades_desglosadas = {}
        
        # Procesar por bodega
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            df_bodega = df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega]
            cantidades_desglosadas[bodega] = {}
            
            # Procesar por liga
            for categoria, ligas in categorias_ligas.items():
                cantidades_desglosadas[bodega][categoria] = {}
                
                if categoria == 'ACCESSORIES':
                    # Para ACCESSORIES, filtrar directamente por segmento
                    df_accessories = df_bodega[df_bodega['U_Segmento'] == 'ACCESSORIES']
                    cantidades_desglosadas[bodega][categoria]['Stock'] = {
                        'Cantidad': df_accessories['Cantidad'].sum(),
                        'USD': df_accessories['USD_Total_SI_CD'].sum()
                    }
                    # También crear entrada para Ventas (USD) para consistencia con el sistema
                    cantidades_desglosadas[bodega][categoria]['Ventas'] = {
                        'Cantidad': df_accessories['Cantidad'].sum(),
                        'USD': df_accessories['USD_Total_SI_CD'].sum()
                    }
                else:
                    # Lógica para otras ligas
                    df_liga = df_bodega[df_bodega['U_Liga'].isin(ligas)]
                    
                    # Planas (HEADWEAR + Planas)
                    df_planas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Planas')]
                    cantidades_desglosadas[bodega][categoria]['Planas'] = {
                        'Cantidad': df_planas['Cantidad'].sum(),
                        'USD': df_planas['USD_Total_SI_CD'].sum()
                    }
                    
                    # Curvas (HEADWEAR + Curvas)  
                    df_curvas = df_liga[(df_liga['U_Segmento'] == 'HEADWEAR') & (df_liga['Tipo'] == 'Curvas')]
                    cantidades_desglosadas[bodega][categoria]['Curvas'] = {
                        'Cantidad': df_curvas['Cantidad'].sum(),
                        'USD': df_curvas['USD_Total_SI_CD'].sum()
                    }
                    
                    # Apparel
                    df_apparel = df_liga[df_liga['U_Segmento'] == 'APPAREL']
                    cantidades_desglosadas[bodega][categoria]['Apparel'] = {
                        'Cantidad': df_apparel['Cantidad'].sum(),
                        'USD': df_apparel['USD_Total_SI_CD'].sum()
                    }
        
        print(f"Cantidades Costa Rica desglosadas calculadas para {len(cantidades_desglosadas)} bodegas")
        
        # Debug final: Mostrar resumen de cantidades calculadas
        print("RESUMEN CANTIDADES COSTA RICA (solo-ventas):")
        for bodega, categorias in cantidades_desglosadas.items():
            print(f"  {bodega}:")
            for categoria, subcategorias in categorias.items():
                total_cat = 0
                total_usd = 0
                if categoria == 'ACCESSORIES':
                    stock_data = subcategorias.get('Stock', {})
                    if isinstance(stock_data, dict):
                        total_cat = stock_data.get('Cantidad', 0)
                        total_usd = stock_data.get('USD', 0)
                else:
                    for subcat in ['Planas', 'Curvas', 'Apparel']:
                        subcat_data = subcategorias.get(subcat, {})
                        if isinstance(subcat_data, dict):
                            total_cat += subcat_data.get('Cantidad', 0)
                            total_usd += subcat_data.get('USD', 0)
                
                if total_cat > 0 or total_usd > 0:
                    print(f"    {categoria}: {total_cat} unidades, ${total_usd:.2f} USD")
                else:
                    print(f"    {categoria}: 0 unidades, $0.00 USD (SIN DATOS)")
        
        return cantidades_desglosadas

    def procesar_cantidades_panama(self, df_ventas: pd.DataFrame) -> Dict[str, Dict[str, Dict[str, float]]]:
        """
        Procesa el archivo de ventas de Panama basado en columna 'Cantidad' para tabla solo-ventas
        Estructura: {bodega: {liga: {subcategoria: {'Cantidad': X, 'USD': Y}}}}
        """
        if df_ventas is None or df_ventas.empty:
            return {}
        
        print(f"Procesando CANTIDADES Panama - Columnas disponibles: {list(df_ventas.columns)}")
        
        # Verificar columnas necesarias
        columnas_necesarias = ['U_Marca', 'U_Segmento', 'U_Liga', 'U_Silueta', 'Cantidad', 'USD_Total_SI_CD']
        columna_tienda = None
        
        # Buscar columna de tienda
        posibles_columnas_tienda = ['Tienda', 'Bodega', 'Store', 'Location']
        for col in posibles_columnas_tienda:
            if col in df_ventas.columns:
                columna_tienda = col
                break
        
        if columna_tienda is None:
            print("No se encontró columna de tienda en Panama (solo-ventas)")
            return {}
            
        columnas_necesarias.append(columna_tienda)
        
        # Verificar que existan todas las columnas
        for col in columnas_necesarias:
            if col not in df_ventas.columns:
                print(f"No se encontró columna {col} en Panama (solo-ventas)")
                return {}
        
        print(f"Usando columna de tienda para Panama (solo-ventas): {columna_tienda}")
        
        # Convertir columnas numéricas
        df_ventas['Cantidad'] = pd.to_numeric(df_ventas['Cantidad'], errors='coerce').fillna(0)
        df_ventas['USD_Total_SI_CD'] = pd.to_numeric(df_ventas['USD_Total_SI_CD'], errors='coerce').fillna(0)
        
        # Filtrar por marca NEW ERA
        df_new_era = df_ventas[df_ventas['U_Marca'].str.upper() == 'NEW ERA'].copy()
        
        # Usar el mismo sistema bidireccional que otros países
        # Obtener todas las variaciones posibles para Panama
        todas_variaciones = self.get_all_variations_for_country("PANAMA")
        
        df_mapeado = df_new_era[df_new_era[columna_tienda].isin(todas_variaciones)].copy()
        
        # Normalizar nombres de bodegas usando el mismo sistema bidireccional
        df_mapeado['Bodega_Mapeada'] = df_mapeado[columna_tienda].apply(
            lambda x: self.normalize_bodega_name(x, target_format="stock", pais="PANAMA")
        )
        
        # Importar ProductClassification para clasificar siluetas
        product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])
        
        # Clasificar siluetas solo para HEADWEAR
        df_mapeado['Tipo'] = df_mapeado.apply(
            lambda row: product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        print(f"Registros totales NEW ERA Panama (solo-ventas): {len(df_new_era)}")
        print(f"Registros con mapeo exitoso Panama (solo-ventas): {len(df_mapeado)}")
        print(f"Tiendas encontradas en archivo Panama: {df_new_era[columna_tienda].unique()}")
        print(f"Bodegas mapeadas Panama: {df_mapeado['Bodega_Mapeada'].unique()}")
        
        if df_mapeado.empty:
            print("No hay registros mapeados para procesar en Panama (solo-ventas)")
            return {}
        
        # Definir categorías de ligas específicas para Panama (IDÉNTICAS al modo stock+ventas)
        categorias_ligas = {
            "MLB": ["MLB", "MLB properties"],
            "NBA": ["NBA", "NBA Properties"],
            "NFL": ["NFL", "NFL Properties"],
            "MOTORSPORT": ["MOTORSPORT"],
            "ENTERTAINMENT": [
                "NEW ERA BRANDED", "ENTERTAINMENT", "MARCA PAIS", "WARNER BROS",
                "DISNEY", "LOONEY TUNES", "MARVEL", "DC", "UNIVERSAL", "PARAMOUNT"
            ],
            "ACCESSORIES": ["ACCESSORIES"]
        }
        
        # Debug: Ver qué ligas están en el archivo  
        ligas_encontradas = sorted(df_mapeado['U_Liga'].unique())
        print(f"LIGAS ENCONTRADAS EN ARCHIVO PANAMA (solo-ventas): {ligas_encontradas}")
        
        cantidades_desglosadas = {}
        
        # Procesar por bodega
        for bodega in df_mapeado['Bodega_Mapeada'].unique():
            df_bodega = df_mapeado[df_mapeado['Bodega_Mapeada'] == bodega]
            cantidades_desglosadas[bodega] = {}
            
            # Procesar por liga
            for categoria, ligas in categorias_ligas.items():
                cantidades_desglosadas[bodega][categoria] = {}
                
                df_liga = df_bodega[df_bodega['U_Liga'].isin(ligas)]
                
                if categoria == "ACCESSORIES":
                    # Para ACCESSORIES no hay subdivisión
                    total_cantidad = df_liga['Cantidad'].sum()
                    total_usd = df_liga['USD_Total_SI_CD'].sum()
                    cantidades_desglosadas[bodega][categoria]['Stock'] = {
                        'Cantidad': total_cantidad,
                        'USD': total_usd
                    }
                else:
                    # Para otras ligas, separar por Planas, Curvas, Apparel
                    for subcategoria in ["Planas", "Curvas", "Apparel"]:
                        if subcategoria == "Apparel":
                            # Apparel
                            df_apparel = df_liga[df_liga['U_Segmento'] == 'APPAREL']
                            cantidades_desglosadas[bodega][categoria][subcategoria] = {
                                'Cantidad': df_apparel['Cantidad'].sum(),
                                'USD': df_apparel['USD_Total_SI_CD'].sum()
                            }
                        else:
                            # Headwear (Planas o Curvas)
                            df_headwear = df_liga[df_liga['U_Segmento'] == 'HEADWEAR']
                            df_tipo = df_headwear[df_headwear['Tipo'] == subcategoria]
                            cantidades_desglosadas[bodega][categoria][subcategoria] = {
                                'Cantidad': df_tipo['Cantidad'].sum(),
                                'USD': df_tipo['USD_Total_SI_CD'].sum()
                            }
        
        print(f"Cantidades PANAMA desglosadas calculadas para {len(cantidades_desglosadas)} bodegas")
        
        # Debug final: Mostrar resumen de cantidades calculadas
        print("RESUMEN CANTIDADES PANAMA:")
        for bodega, categorias in cantidades_desglosadas.items():
            print(f"  {bodega}:")
            for categoria, subcategorias in categorias.items():
                total_cat = 0
                total_usd = 0
                if categoria == 'ACCESSORIES':
                    stock_data = subcategorias.get('Stock', {})
                    if isinstance(stock_data, dict):
                        total_cat = stock_data.get('Cantidad', 0)
                        total_usd = stock_data.get('USD', 0)
                else:
                    for subcat in ['Planas', 'Curvas', 'Apparel']:
                        subcat_data = subcategorias.get(subcat, {})
                        if isinstance(subcat_data, dict):
                            total_cat += subcat_data.get('Cantidad', 0)
                            total_usd += subcat_data.get('USD', 0)
                
                if total_cat > 0 or total_usd > 0:
                    print(f"    {categoria}: {total_cat} unidades, ${total_usd:.2f} USD")
                else:
                    print(f"    {categoria}: 0 unidades, $0.00 USD (SIN DATOS)")
        
        return cantidades_desglosadas

    def procesar_usd_simple_guatemala(self, df_ventas: pd.DataFrame) -> Dict[str, float]:
        """
        Procesa archivo de ventas de Guatemala para extraer SOLO los USD por bodega
        Sigue el procedimiento simple de 4 pasos sin mapeos complejos
        """
        if df_ventas is None or df_ventas.empty:
            return {}
        
        print(f"Procesando USD simple Guatemala - Columnas: {list(df_ventas.columns)}")
        
        # Buscar columna de tienda/bodega
        columna_tienda = None
        for col in df_ventas.columns:
            if 'tienda' in col.lower() or 'store' in col.lower() or 'bodega' in col.lower():
                columna_tienda = col
                break
        
        if columna_tienda is None:
            print("No se encontró columna de tienda")
            return {}
        
        # Verificar columnas necesarias
        if 'USD_Total_SI_CD' not in df_ventas.columns:
            print("No se encontró columna USD_Total_SI_CD")
            return {}
        
        # Convertir USD a numérico
        df_ventas['USD_Total_SI_CD'] = pd.to_numeric(df_ventas['USD_Total_SI_CD'], errors='coerce').fillna(0)
        
        # PASO 1: Filtrar U_Marca con datos NEW ERA
        df_new_era = df_ventas[df_ventas['U_Marca'].str.upper() == 'NEW ERA'].copy()
        print(f"Registros después de filtrar NEW ERA: {len(df_new_era)}")
        
        if df_new_era.empty:
            return {}
        
        # Importar ProductClassification para clasificar siluetas
        product_classifier = ProductClassification(siluetas_planas=[], siluetas_curvas=[])
        
        # Clasificar siluetas para HEADWEAR
        df_new_era['Tipo'] = df_new_era.apply(
            lambda row: product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        # Resultado: USD por bodega
        usd_por_bodega = {}
        
        # Procesar por bodega
        for bodega in df_new_era[columna_tienda].unique():
            df_bodega = df_new_era[df_new_era[columna_tienda] == bodega]
            total_usd_bodega = 0
            
            # PASO 2: HEADWEAR - Filtrar U_Segmento con HEADWEAR y separar planas y curvas
            df_headwear = df_bodega[df_bodega['U_Segmento'] == 'HEADWEAR']
            df_planas = df_headwear[df_headwear['Tipo'] == 'Planas']
            df_curvas = df_headwear[df_headwear['Tipo'] == 'Curvas']
            
            # PASO 3: Sumar USD_TOTAL_SI_CD para planas y curvas
            total_usd_bodega += df_planas['USD_Total_SI_CD'].sum()
            total_usd_bodega += df_curvas['USD_Total_SI_CD'].sum()
            
            # PASO 4a: APPAREL - Filtrar U_Segmento con APPAREL
            df_apparel = df_bodega[df_bodega['U_Segmento'] == 'APPAREL']
            total_usd_bodega += df_apparel['USD_Total_SI_CD'].sum()
            
            # PASO 4b: ACCESSORIES - Filtrar U_Segmento con ACCESSORIES
            df_accessories = df_bodega[df_bodega['U_Segmento'] == 'ACCESSORIES']
            total_usd_bodega += df_accessories['USD_Total_SI_CD'].sum()
            
            # Guardar total por bodega
            usd_por_bodega[bodega] = total_usd_bodega
            
            if total_usd_bodega > 0:
                print(f"  {bodega}: ${total_usd_bodega:,.2f}")
        
        print(f"USD simple procesado para {len(usd_por_bodega)} bodegas")
        total_general = sum(usd_por_bodega.values())
        print(f"TOTAL USD SIMPLE: ${total_general:,.2f}")
        
        return usd_por_bodega

# Instancia del procesador de ventas
sales_processor = SalesProcessor()

@dataclass
class LeagueCategories:
    """Categorías de ligas deportivas"""
    categories: Dict[str, List[str]]
    
    def __post_init__(self):
        self.categories = {
            "MLB": ["MLB", "MLB properties"],
            "NBA": ["NBA", "NBA Properties"],
            "NFL": ["NFL", "NFL Properties"],
            "MOTORSPORT": ["MOTORSPORT"],
            "ENTERTAINMENT": [
                "NEW ERA BRANDED",
                "ENTERTAINMENT",
                "MARCA PAIS",
                "WARNER BROS",
                "NONE LICENSED",
                "MLS",
                "MILB",
                "GUATEMALA SOCCER LEAGUE",
                "WBC",
                "EUROPEAN SOCCER",
                "FEDERACION DE BASEBALL PUERTO RICO",
                "FEDERACION DOMINICANA DE BASEBALL",
                "BASEBALL FEDERATION"
            ],
            "ACCESSORIES": ["ACCESSORIES"]
        }
    
    def get_category_values(self, category: str) -> List[str]:
        """Obtiene los valores de una categoría"""
        return self.categories.get(category, [])
    
    def get_all_categories(self) -> Dict[str, List[str]]:
        """Obtiene todas las categorías"""
        return self.categories

# Instancia de categorías de liga
league_categories = LeagueCategories(categories={})

class StockAnalyzer:
    """Analizador de stock con métricas de cumplimiento"""
    
    def __init__(self, config: StockAnalysisConfig):
        self.config = config
    
    def obtener_color_semaforo(self, total_headwear: int, capacidad: int) -> str:
        """Determina el color del semáforo basado en el porcentaje de cumplimiento"""
        if capacidad == 0:
            return "rojo"
        
        # Aplicar nueva fórmula: (((TOTAL HEADWEAR/CAPACIDAD)*100%)-100%)
        porcentaje_cumplimiento = ((total_headwear / capacidad) * 100) - 100
        
        if porcentaje_cumplimiento < 0:  # Valores negativos
            return "rojo"
        elif 0 <= porcentaje_cumplimiento <= 15:  # 0% a 15%
            return "verde"
        else:  # Mayores al 15%
            return "amarillo"
    
    def calculate_performance_metrics(self, stock_data: List[Dict]) -> Dict[str, Any]:
        """Calcula métricas de rendimiento del stock"""
        if not stock_data:
            return {}
        
        total_stock = sum(item.get('stock', 0) for item in stock_data)
        total_capacity = sum(item.get('capacity', 0) for item in stock_data if item.get('capacity', 0) > 0)
        
        return {
            'total_stock': total_stock,
            'total_capacity': total_capacity,
            'capacity_utilization': total_stock / total_capacity if total_capacity > 0 else 0,
            'low_stock_stores': len([item for item in stock_data if self._is_low_stock(item)]),
            'overstock_stores': len([item for item in stock_data if self._is_overstock(item)])
        }
    
    def _is_low_stock(self, item: Dict) -> bool:
        """Verifica si una tienda tiene stock bajo"""
        capacity = item.get('capacity', 0)
        stock = item.get('stock', 0)
        if capacity == 0:
            return False
        return stock < (capacity * 0.95)
    
    def _is_overstock(self, item: Dict) -> bool:
        """Verifica si una tienda tiene sobrestock"""
        capacity = item.get('capacity', 0)
        stock = item.get('stock', 0)
        if capacity == 0:
            return False
        return stock > capacity

# Instancia del analizador
stock_analyzer = StockAnalyzer(config)

class DataLoader:
    """Cargador de datos con validación robusta"""
    
    def __init__(self, country_manager: CountryManager):
        self.country_manager = country_manager
        self.required_columns = ['U_Marca', 'U_Silueta', 'Stock_Actual', 'Bodega', 'U_Liga', 'U_Segmento']
        
        # Configuración centralizada de nombres de archivos permitidos
        self.nombres_permitidos = {
            'GUATEMALA': {
                'stock': 'GUATEMALA',
                'ventas': 'VENTAS_GUATEMALA'
            },
            'EL_SALVADOR': {
                'stock': 'EL_SALVADOR', 
                'ventas': 'VENTAS_EL_SALVADOR'
            },
            'PANAMA': {
                'stock': 'PANAMA',
                'ventas': 'VENTAS_PANAMA'
            },
            'HONDURAS': {
                'stock': 'HONDURAS',
                'ventas': 'VENTAS_HONDURAS'
            },
            'COSTA_RICA': {
                'stock': 'COSTA_RICA',
                'ventas': 'VENTAS_COSTA_RICA'
            },
            'GT': {
                'stock': 'GT',
                'ventas': 'GT'
            },
            # CONFIGURACIONES TEMPORALES - ESPACIO ADICIONAL
            'GUATEMALA_TEMP': {
                'stock': 'GUATEMALA',
                'ventas': 'VENTAS_GUATEMALA'
            },
            'EL_SALVADOR_TEMP': {
                'stock': 'EL_SALVADOR', 
                'ventas': 'VENTAS_EL_SALVADOR'
            },
            'HONDURAS_TEMP': {
                'stock': 'HONDURAS',
                'ventas': 'VENTAS_HONDURAS'
            },
            'COSTA_RICA_TEMP': {
                'stock': 'COSTA_RICA',
                'ventas': 'VENTAS_COSTA_RICA'
            },
            'PANAMA_TEMP': {
                'stock': 'PANAMA',
                'ventas': 'VENTAS_PANAMA'
            },
            'PUERTO_RICO_TEMP': {
                'stock': 'PUERTO_RICO',
                'ventas': 'VENTAS_PUERTO_RICO'
            },
            'GUATEMALA_MVP_TEMP': {
                'stock': 'GUATEMALA',
                'ventas': 'VENTAS_GUATEMALA'
            }
        }
    
    def _validar_nombre_archivo(self, archivo, pais: str, tipo: str) -> bool:
        """Valida que el nombre del archivo sea exactamente el esperado"""
        if not hasattr(archivo, 'name') or archivo.name is None:
            return False
            
        # Obtener el nombre sin extensión
        nombre_archivo = archivo.name.rsplit('.', 1)[0] if '.' in archivo.name else archivo.name
        
        # Obtener el nombre esperado para el país y tipo
        nombre_esperado = self.nombres_permitidos.get(pais, {}).get(tipo)
        
        if nombre_esperado is None:
            # Si no hay nombre definido para este tipo (ej: ventas no disponible)
            st.error(f"❌ **Error de seguridad:** No hay archivos de {tipo} configurados para {pais}")
            return False
            
        # Validación estricta de nombre
        if nombre_archivo != nombre_esperado:
            st.error(f"""
            ❌ **Error de seguridad:** Nombre de archivo incorrecto
            
            **Nombre recibido:** `{nombre_archivo}`
            **Nombre esperado:** `{nombre_esperado}`
            
            ⚠️ Por favor, renombra tu archivo exactamente como se indica: **{nombre_esperado}.csv**
            """)
            return False
            
        return True
    
    def cargar_archivo(self, label_texto: str, pais: str) -> Optional[pd.DataFrame]:
        """Carga y valida el archivo CSV con manejo robusto"""
        # Crear una etiqueta más elegante y minimalista
        with st.container():
            st.markdown(f"""
                <div style="
                    background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
                    border: 1px solid #e5e7eb;
                    border-radius: 16px;
                    padding: 1.5rem;
                    margin: 1rem 0;
                    text-align: center;
                ">
                    <h4 style="
                        color: #374151;
                        font-weight: 600;
                        margin: 0 0 0.5rem 0;
                        font-size: 1.1rem;
                    ">{label_texto}</h4>
                    <p style="
                        color: #6b7280;
                        font-size: 0.9rem;
                        margin: 0;
                    ">Arrastra tu archivo CSV aquí o haz clic para seleccionar</p>
                </div>
            """, unsafe_allow_html=True)
            
            archivo = st.file_uploader(
                label="",
                type=["csv"],
                key=f"uploader_{pais}",
                label_visibility="collapsed"
            )
        
        if archivo is None:
            return None
        
        # VALIDACIÓN DE SEGURIDAD: Verificar nombre del archivo
        if not self._validar_nombre_archivo(archivo, pais, 'stock'):
            return None
        
        try:
            return self._process_file(archivo, pais)
        except Exception as e:
            logger.error(f"Error al cargar archivo {pais}: {str(e)}")
            st.error(f"Error al cargar archivo {pais}: {str(e)}")
            return None
    
    def _process_file(self, archivo, pais: str) -> pd.DataFrame:
        """Procesa el archivo CSV"""
        with st.spinner(f"Cargando archivo {pais}..."):
            start_time = time.time()
            logger.info(f"Iniciando carga de archivo para {pais}")
            
            df = self._read_csv(archivo)
            
            # Para archivos de óptimos, procesamiento mínimo
            if pais == 'GT':
                # Solo limpiar nombres de columnas para archivo de óptimos
                df.columns = df.columns.str.strip()
            else:
                # Procesamiento completo para archivos de stock normales
                df = self._clean_data(df)
                df = self._filter_by_country(df, pais)
                self._validate_columns(df, pais)
                
                # Actualizar la fecha del último trabajo con stock
                current_date = datetime.now().strftime('%d/%m/%Y')
                st.session_state.last_stock_work_date = current_date
            
            elapsed_time = time.time() - start_time
            logger.info(f"Archivo {pais} cargado exitosamente en {elapsed_time:.2f}s - Registros: {len(df):,}")
            st.success(f"✅ Archivo {pais} cargado ({elapsed_time:.2f}s) | Registros: {len(df):,}")
            return df
    
    def cargar_archivo_ventas(self, label_texto: str, key: str, pais: str = None) -> Optional[pd.DataFrame]:
        """Carga archivo de ventas con validación de nombre"""
        with st.container():
            st.markdown(f"""
                <div style="
                    background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
                    border: 1px solid #e5e7eb;
                    border-radius: 16px;
                    padding: 1.5rem;
                    margin: 1rem 0;
                    text-align: center;
                ">
                    <h4 style="
                        color: #374151;
                        font-weight: 600;
                        margin: 0 0 0.5rem 0;
                        font-size: 1.1rem;
                    ">{label_texto}</h4>
                    <p style="
                        color: #6b7280;
                        font-size: 0.9rem;
                        margin: 0;
                    ">Arrastra tu archivo CSV aquí o haz clic para seleccionar</p>
                </div>
            """, unsafe_allow_html=True)
            
            archivo = st.file_uploader(
                "Seleccionar archivo",
                type=['csv'],
                key=key,
                label_visibility="collapsed"
            )
            
            if archivo is None:
                return None
            
            # VALIDACIÓN DE SEGURIDAD: Verificar nombre del archivo
            if pais and not self._validar_nombre_archivo(archivo, pais, 'ventas'):
                return None
            
            try:
                start_time = time.time()
                
                # Leer CSV sin validaciones específicas de stock
                df = pd.read_csv(
                    archivo,
                    encoding='utf-8',
                    delimiter=';',
                    low_memory=False,
                    on_bad_lines='skip'
                )
                
                # Limpieza básica sin columnas específicas
                df.columns = df.columns.str.strip()
                
                elapsed_time = time.time() - start_time
                st.success(f"✅ Archivo de ventas cargado ({elapsed_time:.2f}s) | Registros: {len(df):,}")
                return df
                
            except Exception as e:
                st.error(f"Error al cargar archivo de ventas: {str(e)}")
                return None
    
    def _read_csv(self, archivo) -> pd.DataFrame:
        """Lee el archivo CSV con configuración optimizada"""
        return pd.read_csv(
            archivo,
            encoding='utf-8',
            delimiter=';',
            dtype={'U_Silueta': str, 'Stock_Actual': str, 'Bodega': str, 'U_Liga': str, 'U_Segmento': str},
            low_memory=False,
            on_bad_lines='skip'
        )
    
    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Limpia y normaliza los datos"""
        df.columns = df.columns.str.strip()
        df['Bodega'] = df['Bodega'].astype(str).str.strip()
        df['U_Liga'] = df['U_Liga'].astype(str).str.strip().str.upper()
        df['U_Segmento'] = df['U_Segmento'].astype(str).str.strip().str.upper()
        
        # Conversión segura de stock a numérico
        df['Stock_Actual'] = pd.to_numeric(
            df['Stock_Actual'].str.replace(',', ''),
            errors='coerce'
        ).fillna(0)
        
        return df
    
    def _filter_by_country(self, df: pd.DataFrame, pais: str) -> pd.DataFrame:
        """Filtra datos por país y marca NEW ERA"""
        # Primero filtrar por marca NEW ERA
        df_new_era = df[df['U_Marca'].str.upper() == 'NEW ERA']
        
        # Luego filtrar por bodegas del país
        bodegas = self.country_manager.get_bodegas(pais)
        return df_new_era[df_new_era['Bodega'].isin(bodegas)] if bodegas else df_new_era
    
    def _validate_columns(self, df: pd.DataFrame, pais: str) -> None:
        """Valida que existan las columnas requeridas"""
        if not all(col in df.columns for col in self.required_columns):
            logger.error(f"Faltan columnas requeridas en el archivo de {pais}")
            st.error(f"❌ Faltan columnas requeridas en el archivo de {pais}")
            raise ValueError(f"Faltan columnas requeridas: {self.required_columns}")

# Instancia del cargador de datos
data_loader = DataLoader(country_manager)

class DataProcessor:
    """Procesador de datos consolidados"""
    
    def __init__(self, country_manager: CountryManager, league_categories: LeagueCategories, 
                 product_classifier: ProductClassification):
        self.country_manager = country_manager
        self.league_categories = league_categories
        self.product_classifier = product_classifier
    
    @st.cache_data(ttl=1)  # Cache con tiempo de vida muy corto para forzar actualización
    def procesar_datos_consolidados(_self, df_hash: List[Dict], pais: str, selected_league: str = None, df_ventas_hash: List[Dict] = None) -> Optional[pd.DataFrame]:
        """Procesa los datos para generar tabla con múltiples niveles de encabezados"""
        df = pd.DataFrame(df_hash)
        
        # Debug específico para Honduras
        if pais == "Honduras" and not df.empty:
            print(f"INICIO PROCESAMIENTO HONDURAS:")
            print(f"- Filas totales: {len(df)}")
            print(f"- Columnas: {list(df.columns)}")
            if 'Bodega' in df.columns:
                print(f"- Bodegas únicas en datos: {df['Bodega'].unique()}")
                city_mall_records = df[df['Bodega'].str.contains('City Mall Tegucigalpa', na=False)]
                print(f"- Registros con 'City Mall Tegucigalpa': {len(city_mall_records)}")
                if len(city_mall_records) > 0:
                    print(f"- Stock total bruto City Mall: {city_mall_records['Stock_Actual'].sum()}")
            print("="*50)
        
        if df is None or df.empty:
            return None
        
        with st.spinner(f"Generando tabla consolidada {pais}..."):
            logger.info(f"Iniciando procesamiento de datos consolidados para {pais}")
            
            df = _self._prepare_data(df)
            tabla_final = _self._create_base_table(pais)
            tabla_final = _self._process_categories(df, tabla_final, pais, selected_league, df_ventas_hash)
            tabla_final = _self._calculate_totals(tabla_final, pais, selected_league)
            
            # Agregar columna Ventas (USD) para ACCESSORIES solo si hay datos de ventas
            if (df_ventas_hash is not None and pais in ["Guatemala", "El Salvador", "Costa Rica", "Honduras", "PANAMA"] and 
                'ACCESSORIES - Stock' in tabla_final.columns and 
                'ACCESSORIES - Ventas (USD)' not in tabla_final.columns):
                tabla_final['ACCESSORIES - Ventas (USD)'] = 0.0
            
            # Agregar columnas de ventas si hay datos de ventas para Guatemala, El Salvador, Costa Rica, Honduras o PANAMA
            if df_ventas_hash is not None and pais in ["Guatemala", "El Salvador", "Costa Rica", "Honduras", "PANAMA"]:
                df_ventas = pd.DataFrame(df_ventas_hash)
                tabla_final = _self._add_sales_columns(tabla_final, df_ventas, selected_league, pais)
            
            # Calcular TOTAL (USD) SOLO si hay archivo de ventas cargado
            if df_ventas_hash is not None and pais in ["Guatemala", "El Salvador", "Costa Rica", "Honduras", "PANAMA"]:
                if selected_league:
                    # Para liga específica, solo sumar columnas de ventas de esa liga
                    columnas_usd = [col for col in tabla_final.columns if 
                                   ('Ventas (USD)' in col or ('Ventas' in col and 'Stock' not in col)) and 
                                   selected_league in col]
                    logger.info(f"TOTAL (USD) calculado solo para liga: {selected_league}")
                else:
                    # Para todas las ligas, sumar todas las columnas de ventas
                    columnas_usd = [col for col in tabla_final.columns if 'Ventas (USD)' in col or ('Ventas' in col and 'Stock' not in col)]
                    logger.info("TOTAL (USD) calculado para todas las ligas")
                
                if columnas_usd:
                    tabla_final['TOTAL (USD)'] = tabla_final[columnas_usd].sum(axis=1)
                else:
                    tabla_final['TOTAL (USD)'] = 0.0
            # Si no hay archivo de ventas, NO crear la columna TOTAL (USD)
            
            # Determinar si hay datos de ventas para pasarlo a _format_table
            hay_ventas = df_ventas_hash is not None and pais in ["Guatemala", "El Salvador", "Costa Rica", "Honduras", "PANAMA"]
            tabla_final = _self._format_table(tabla_final, selected_league, hay_ventas)
            
            logger.info(f"Procesamiento completado para {pais}")
            return tabla_final
    
    def _prepare_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Prepara los datos para el procesamiento"""
        # Asegurar que U_Silueta sea string y manejar valores NaN
        df['U_Silueta'] = df['U_Silueta'].astype(str).fillna('').str.strip().str.upper()
        
        # Normalizar nombres de bodegas para consistencia
        if 'Bodega' in df.columns:
            print(f"Bodegas originales encontradas: {df['Bodega'].unique()}")
            df['Bodega'] = df['Bodega'].astype(str).str.strip()
            
            # Mapeo específico para normalizar nombres conocidos
            mapeo_normalizacion = {
                # Costa Rica
                'BODEGA CENTRAL NEW ERA': 'Bodega Central NEW ERA',
                'Bodega Central New Era': 'Bodega Central NEW ERA',
                'bodega central new era': 'Bodega Central NEW ERA',
                'BODEGA CENTRAL NEW ERA ': 'Bodega Central NEW ERA',
                # Guatemala
                'central new era': 'CENTRAL NEW ERA',
                'Central New Era': 'CENTRAL NEW ERA',
                'Central NEW ERA': 'CENTRAL NEW ERA',
                'CENTRAL NEW ERA ': 'CENTRAL NEW ERA',
                # El Salvador
                'new era central': 'New Era Central',
                'NEW ERA CENTRAL': 'New Era Central',
                'New era central': 'New Era Central',
                'NEW ERA CENTRAL ': 'New Era Central',
                # Tiendas comunes
                'NE CITY MALL': 'NE City Mall',
                'Ne City Mall': 'NE City Mall',
                'ne city mall': 'NE City Mall',
                # Honduras - normalizar guiones y espacios
                'NE - City Mall Tegucigalpa': 'NE – City Mall Tegucigalpa',
                'NE -City Mall Tegucigalpa': 'NE – City Mall Tegucigalpa',
                'NE- City Mall Tegucigalpa': 'NE – City Mall Tegucigalpa',
                'NE-City Mall Tegucigalpa': 'NE – City Mall Tegucigalpa',
                'NE - CIty Mall Tegucigalpa': 'NE – City Mall Tegucigalpa',  # Con I mayúscula
                'NE -CIty Mall Tegucigalpa': 'NE – City Mall Tegucigalpa',
                'NE- CIty Mall Tegucigalpa': 'NE – City Mall Tegucigalpa',
                'NE-CIty Mall Tegucigalpa': 'NE – City Mall Tegucigalpa',
                'NE - Cascadas Mall Tegucigalpa': 'NE – Cascadas Mall Tegucigalpa',
                'NE -Cascadas Mall Tegucigalpa': 'NE – Cascadas Mall Tegucigalpa',
                'NE- Cascadas Mall Tegucigalpa': 'NE – Cascadas Mall Tegucigalpa',
                'NE-Cascadas Mall Tegucigalpa': 'NE – Cascadas Mall Tegucigalpa',
                'NE - Multiplaza Tegucigalpa': 'NE – Multiplaza Tegucigalpa',
                'NE -Multiplaza Tegucigalpa': 'NE – Multiplaza Tegucigalpa',
                'NE- Multiplaza Tegucigalpa': 'NE – Multiplaza Tegucigalpa',
                'NE-Multiplaza Tegucigalpa': 'NE – Multiplaza Tegucigalpa',
                'NE - Mega Mall SPS': 'NE – Mega Mall SPS',
                'NE -Mega Mall SPS': 'NE – Mega Mall SPS',
                'NE- Mega Mall SPS': 'NE – Mega Mall SPS',
                'NE-Mega Mall SPS': 'NE – Mega Mall SPS',
                'NE -Multiplaza SPS': 'NE –Multiplaza SPS',
                'NE- Multiplaza SPS': 'NE –Multiplaza SPS',
                'NE-Multiplaza SPS': 'NE –Multiplaza SPS',
                'NE - Multiplaza SPS': 'NE –Multiplaza SPS'
            }
            
            # Aplicar normalización exacta primero
            df['Bodega'] = df['Bodega'].replace(mapeo_normalizacion)
            print(f"Bodegas después de normalización: {df['Bodega'].unique()}")
            
            # Normalización adicional por texto similar
            print(f"Iniciando normalización adicional...")
            for idx, bodega_actual in df['Bodega'].items():
                if 'central' in bodega_actual.lower() and 'new era' in bodega_actual.lower():
                    # Determinar país basado en el formato del nombre
                    if 'bodega' in bodega_actual.lower():
                        df.at[idx, 'Bodega'] = 'Bodega Central NEW ERA'  # Costa Rica
                    elif bodega_actual.lower().startswith('new era'):
                        df.at[idx, 'Bodega'] = 'New Era Central'  # El Salvador
                    else:
                        df.at[idx, 'Bodega'] = 'CENTRAL NEW ERA'  # Guatemala
                elif 'city mall' in bodega_actual.lower() and 'ne' in bodega_actual.lower():
                    if 'tegucigalpa' in bodega_actual.lower():
                        print(f"NORMALIZANDO HONDURAS: '{bodega_actual}' -> 'NE – City Mall Tegucigalpa'")
                        df.at[idx, 'Bodega'] = 'NE – City Mall Tegucigalpa'  # Honduras
                    else:
                        df.at[idx, 'Bodega'] = 'NE City Mall'  # Otros países
                # Normalización adicional para otras bodegas de Honduras
                elif 'cascadas mall' in bodega_actual.lower() and 'ne' in bodega_actual.lower():
                    df.at[idx, 'Bodega'] = 'NE – Cascadas Mall Tegucigalpa'
                elif 'multiplaza' in bodega_actual.lower() and 'tegucigalpa' in bodega_actual.lower() and 'ne' in bodega_actual.lower():
                    df.at[idx, 'Bodega'] = 'NE – Multiplaza Tegucigalpa'
                elif 'mega mall' in bodega_actual.lower() and 'sps' in bodega_actual.lower() and 'ne' in bodega_actual.lower():
                    df.at[idx, 'Bodega'] = 'NE – Mega Mall SPS'
                elif 'multiplaza' in bodega_actual.lower() and 'sps' in bodega_actual.lower() and 'ne' in bodega_actual.lower():
                    df.at[idx, 'Bodega'] = 'NE –Multiplaza SPS'
        
        # Clasificar solo productos HEADWEAR por silueta
        df['Tipo'] = df.apply(
            lambda row: self.product_classifier.clasificar_silueta(row['U_Silueta']) 
            if row['U_Segmento'] == 'HEADWEAR' else None, 
            axis=1
        )
        
        # Filtrar solo siluetas válidas de HEADWEAR, Apparel y Accessories
        df_filtrado = df[(df['Tipo'].notna()) | (df['U_Segmento'] == 'APPAREL') | (df['U_Segmento'] == 'ACCESSORIES')].copy()
        
        # Debug específico para Honduras después del filtrado
        if 'Bodega' in df_filtrado.columns:
            bodegas_en_datos = df_filtrado['Bodega'].unique()
            if any('City Mall Tegucigalpa' in bodega for bodega in bodegas_en_datos):
                print(f"DEBUG: Después del filtrado, bodegas con 'City Mall Tegucigalpa': {[b for b in bodegas_en_datos if 'City Mall Tegucigalpa' in b]}")
                city_mall_data = df_filtrado[df_filtrado['Bodega'].str.contains('City Mall Tegucigalpa', na=False)]
                if not city_mall_data.empty:
                    print(f"Stock total NE City Mall Tegucigalpa: {city_mall_data['Stock_Actual'].sum()}")
                    print(f"Registros por segmento: {city_mall_data['U_Segmento'].value_counts().to_dict()}")
        
        return df_filtrado
    
    def _create_base_table(self, pais: str) -> pd.DataFrame:
        """Crea la tabla base con las bodegas del país"""
        bodegas = self.country_manager.get_bodegas(pais)
        if pais == "Honduras":
            print(f"Creando tabla base para Honduras con bodegas: {bodegas}")
        return pd.DataFrame(index=bodegas)
    
    def _process_categories(self, df: pd.DataFrame, tabla_final: pd.DataFrame, pais: str, selected_league: str = None, df_ventas_hash: List[Dict] = None) -> pd.DataFrame:
        """Procesa cada categoría de liga"""
        # Usar el parámetro pasado en lugar de session_state para compatibilidad con cache
        
        # SIEMPRE procesar todas las categorías para generar tabla completa
        categorias_a_procesar = self.league_categories.get_all_categories()
        logger.info("Procesando todas las categorías para tabla completa")
        
        for categoria, valores in categorias_a_procesar.items():
            if categoria == 'ACCESSORIES':
                # Para ACCESSORIES, filtrar por segmento en lugar de liga
                df_cat = df[df['U_Segmento'].str.upper() == 'ACCESSORIES']
                logger.info(f"Categoría: {categoria}, Registros filtrados por segmento: {len(df_cat)}")
                
                # Debug específico para Honduras
                if pais == "Honduras":
                    print(f"ACCESSORIES Honduras - Registros encontrados: {len(df_cat)}")
                    if len(df_cat) > 0:
                        print(f"Bodegas en ACCESSORIES: {df_cat['Bodega'].unique()}")
                        print(f"Stock por bodega en ACCESSORIES:")
                        for bodega in df_cat['Bodega'].unique():
                            stock = df_cat[df_cat['Bodega'] == bodega]['Stock_Actual'].sum()
                            print(f"  {bodega}: {stock}")
                
                if len(df_cat) == 0:
                    logger.warning(f"No se encontraron datos para la categoría {categoria}")
                    continue
                
                # Para ACCESSORIES, crear columnas Stock y opcionalmente Ventas
                accessories_stock = self._process_accessories_stock(df_cat)
                
                # Agregar las columnas al DataFrame final
                for col in accessories_stock.columns:
                    tabla_final[f"{categoria} - {col}"] = accessories_stock[col]
            else:
                # Lógica original para otras ligas
                df_cat = df[df['U_Liga'].str.upper().isin([v.upper() for v in valores])]
                logger.info(f"Categoría: {categoria}, Registros filtrados: {len(df_cat)}")
                
                # Debug específico para Honduras
                if pais == "Honduras":
                    print(f"{categoria} Honduras - Registros encontrados: {len(df_cat)}")
                    if len(df_cat) > 0:
                        print(f"Bodegas en {categoria}: {df_cat['Bodega'].unique()}")
                        if categoria == "MLB":  # Solo para MLB para no saturar logs
                            print(f"Stock por bodega en {categoria}:")
                            for bodega in df_cat['Bodega'].unique():
                                stock_planas = df_cat[(df_cat['Bodega'] == bodega) & (df_cat['Tipo'] == 'Planas')]['Stock_Actual'].sum()
                                stock_curvas = df_cat[(df_cat['Bodega'] == bodega) & (df_cat['Tipo'] == 'Curvas')]['Stock_Actual'].sum()
                                stock_apparel = df_cat[(df_cat['Bodega'] == bodega) & (df_cat['U_Segmento'] == 'APPAREL')]['Stock_Actual'].sum()
                                print(f"  {bodega}: Planas={stock_planas}, Curvas={stock_curvas}, Apparel={stock_apparel}")
                
                if len(df_cat) == 0:
                    logger.warning(f"No se encontraron datos para la categoría {categoria}")
                    continue
                
                # Procesar Planas y Curvas
                pivot = self._process_headwear_types(df_cat)
                
                # Procesar Apparel
                apparel = self._process_apparel(df_cat)
            
            # Combinar resultados
            pivot = pivot.join(apparel, how='left').fillna(0)
            
            # SIEMPRE usar el formato original para tabla completa
            pivot.columns = [f"{categoria} - {col}" for col in pivot.columns]
            logger.info(f"Columnas generadas: {list(pivot.columns)}")
            
            tabla_final = tabla_final.join(pivot, how='left')
        
        # SIEMPRE procesar Accessories para tabla completa
        accessories = self._process_accessories(df)
        tabla_final = tabla_final.join(accessories, how='left').fillna(0)
        
        # Asegurar que todas las bodegas del país aparezcan en la tabla final
        todas_bodegas = self.country_manager.get_bodegas(pais)
        for bodega in todas_bodegas:
            if bodega not in tabla_final.index:
                # Agregar bodega faltante con ceros
                nueva_fila = pd.Series(0, index=tabla_final.columns, name=bodega)
                tabla_final = pd.concat([tabla_final, nueva_fila.to_frame().T])
        
        return tabla_final.fillna(0).astype(int)
    
    def _process_headwear_types(self, df_cat: pd.DataFrame) -> pd.DataFrame:
        """Procesa tipos de headwear (planas y curvas)"""
        return df_cat[df_cat['Tipo'].notna()].pivot_table(
            index='Bodega',
            columns='Tipo',
            values='Stock_Actual',
            aggfunc='sum',
            fill_value=0
        )
    
    def _process_apparel(self, df_cat: pd.DataFrame) -> pd.Series:
        """Procesa datos de apparel"""
        return df_cat[df_cat['U_Segmento'] == 'APPAREL'].groupby('Bodega')['Stock_Actual'].sum().rename('Apparel')
    
    def _process_accessories(self, df: pd.DataFrame) -> pd.Series:
        """Procesa datos de accessories"""
        return df[df['U_Segmento'] == 'ACCESSORIES'].groupby('Bodega')['Stock_Actual'].sum().rename('TOTAL ACCESSORIES')
    
    def _process_accessories_stock(self, df_cat: pd.DataFrame) -> pd.DataFrame:
        """Procesa datos de stock para ACCESSORIES como columnas independientes"""
        # Crear DataFrame con Stock para ACCESSORIES
        accessories_data = df_cat.groupby('Bodega')['Stock_Actual'].sum()
        
        # Crear DataFrame con columnas Stock y Ventas (USD)
        result = pd.DataFrame(index=accessories_data.index)
        result['Stock'] = accessories_data
        result['Ventas (USD)'] = 0.0  # Inicializar con 0, se llenará con datos reales si están disponibles
        
        return result
    
    def _calculate_totals(self, tabla_final: pd.DataFrame, pais: str, selected_league: str = None) -> pd.DataFrame:
        """Calcula totales y métricas"""
        
        
        # Calcular totales según el filtro aplicado
        if selected_league:
            # Para liga específica, calcular totales SOLO de esa liga y SOLO Stock
            tabla_final['TOTAL PLANAS'] = tabla_final[[col for col in tabla_final.columns if str(col) == f'{selected_league} - Planas']].sum(axis=1)
            tabla_final['TOTAL CURVAS'] = tabla_final[[col for col in tabla_final.columns if str(col) == f'{selected_league} - Curvas']].sum(axis=1)
            tabla_final['TOTAL APPAREL'] = tabla_final[[col for col in tabla_final.columns if str(col) == f'{selected_league} - Apparel']].sum(axis=1)
            logger.info(f"Calculando totales solo para liga: {selected_league}")
        else:
            # Columnas PLANAS: Buscar solo las que terminan en "Planas" (excluyendo ACCESSORIES y Ventas)
            columnas_planas = [col for col in tabla_final.columns if str(col).endswith('- Planas') and 'ACCESSORIES' not in str(col) and 'Ventas' not in str(col)]
            tabla_final['TOTAL PLANAS'] = tabla_final[columnas_planas].sum(axis=1) if columnas_planas else 0
            
            # Columnas CURVAS: Buscar solo las que terminan en "Curvas" (excluyendo ACCESSORIES y Ventas)
            columnas_curvas = [col for col in tabla_final.columns if str(col).endswith('- Curvas') and 'ACCESSORIES' not in str(col) and 'Ventas' not in str(col)]
            tabla_final['TOTAL CURVAS'] = tabla_final[columnas_curvas].sum(axis=1) if columnas_curvas else 0
            
            # Columnas APPAREL: Buscar solo las que terminan en "Apparel" (excluyendo ACCESSORIES y Ventas)
            columnas_apparel = [col for col in tabla_final.columns if str(col).endswith('- Apparel') and 'ACCESSORIES' not in str(col) and 'Ventas' not in str(col)]
            tabla_final['TOTAL APPAREL'] = tabla_final[columnas_apparel].sum(axis=1) if columnas_apparel else 0
        
        # TOTAL HEADWEAR y TOTAL STOCK se calculan igual en ambos casos
        tabla_final['TOTAL HEADWEAR'] = tabla_final['TOTAL PLANAS'] + tabla_final['TOTAL CURVAS']
        tabla_final['TOTAL STOCK'] = tabla_final[['TOTAL HEADWEAR', 'TOTAL APPAREL']].sum(axis=1)
        
        # Calcular TOTAL USD sumando todas las columnas de ventas USD
        # Buscar todas las columnas que contengan 'Ventas (USD)' para sumarlas
        columnas_usd = [col for col in tabla_final.columns if 'Ventas (USD)' in str(col)]
        if columnas_usd:
            tabla_final['TOTAL USD'] = tabla_final[columnas_usd].sum(axis=1)
        else:
            # Si no hay columnas de ventas USD, inicializar en 0
            tabla_final['TOTAL USD'] = 0
        
        # Solo calcular capacidad y % cumplimiento cuando NO hay selected_league
        if not selected_league:
            # Agregar capacidad
            capacidades = self.country_manager.get_capacidades(pais)
            tabla_final['CAPACIDAD EN TIENDA'] = tabla_final.index.map(lambda x: capacidades.get(x, 0))
            
            # Calcular % DE CUMPLIMIENTO
            tabla_final['% DE CUMPLIMIENTO'] = (((tabla_final['TOTAL HEADWEAR'] / tabla_final['CAPACIDAD EN TIENDA']) * 100) - 100).replace([np.inf, -np.inf], 0).fillna(0)
            tabla_final['% DE CUMPLIMIENTO'] = tabla_final['% DE CUMPLIMIENTO'].apply(lambda x: f"{x:.2f}%" if pd.notnull(x) and x != 0 else "N/A")
            
        
        # Ordenar por TOTAL STOCK
        tabla_final = tabla_final.sort_values('TOTAL STOCK', ascending=False)
        
        # Agregar fila de TOTALES
        return self._add_totals_row(tabla_final, pais)
    
    def _add_totals_row(self, tabla_final: pd.DataFrame, pais: str) -> pd.DataFrame:
        """Agrega fila de totales"""
        fila_totales = tabla_final.sum()
        
        capacidad_total = self.country_manager.get_country_data(pais).get_total_capacity()
        total_headwear_suma = fila_totales['TOTAL HEADWEAR']
        
        if capacidad_total > 0:
            porcentaje_total = ((total_headwear_suma / capacidad_total) * 100) - 100
            fila_totales['% DE CUMPLIMIENTO'] = f"{porcentaje_total:.2f}%"
        else:
            fila_totales['% DE CUMPLIMIENTO'] = "N/A"
        
        fila_totales['CAPACIDAD EN TIENDA'] = capacidad_total
        tabla_final.loc['TOTAL'] = fila_totales
        
        return tabla_final
    
    def _add_sales_columns(self, tabla_final: pd.DataFrame, df_ventas: pd.DataFrame, selected_league: str = None, pais: str = "Guatemala") -> pd.DataFrame:
        """Agrega las columnas de ventas desglosadas por liga y subcategoría"""
        import streamlit as st
        
        # Procesar datos de ventas usando el SalesProcessor según el país
        if pais == "Guatemala":
            ventas_desglosadas = sales_processor.procesar_ventas_guatemala(df_ventas)
        elif pais == "El Salvador":
            ventas_desglosadas = sales_processor.procesar_ventas_el_salvador(df_ventas)
        elif pais == "Costa Rica":
            ventas_desglosadas = sales_processor.procesar_ventas_costa_rica(df_ventas)
        elif pais == "Honduras":
            # Para la verificación, usar la misma función que usa la tabla (cantidades en lugar de ventas USD)
            ventas_desglosadas = sales_processor.procesar_cantidades_honduras(df_ventas)
            
            # Verificar si hay bodegas que no coinciden usando mapeo bidireccional
            if ventas_desglosadas:
                bodegas_tabla = set(tabla_final.index) - {'TOTAL'}
                bodegas_sin_ventas = []
                
                for bodega_stock in bodegas_tabla:
                    # Usar el mapeo bidireccional para encontrar correspondencia
                    found = False
                    for bodega_ventas in ventas_desglosadas.keys():
                        canonical_stock = sales_processor.get_canonical_name(bodega_stock, pais)
                        canonical_ventas = sales_processor.get_canonical_name(bodega_ventas, pais)
                        if canonical_stock == canonical_ventas and canonical_stock is not None:
                            found = True
                            break
                    
                    if not found:
                        bodegas_sin_ventas.append(bodega_stock)
                
        elif pais == "PANAMA":
            print(f"Procesando ventas PANAMA - Archivo recibido: {df_ventas is not None}")
            if df_ventas is not None:
                print(f"Filas en archivo ventas PANAMA: {len(df_ventas)}")
            ventas_desglosadas = sales_processor.procesar_ventas_panama(df_ventas)
            
            # Verificar si hay bodegas que no coinciden (solo mostrar si hay problema)
            if ventas_desglosadas:
                bodegas_tabla = set(tabla_final.index) - {'TOTAL'}
                # Para PANAMA, excluir bodegas centrales de la verificación (es normal que no tengan ventas)
                bodegas_centrales_panama = {'Almacén general', 'Bodega Central Albrook'}
                bodegas_tabla = bodegas_tabla - bodegas_centrales_panama
                
                bodegas_ventas = set(ventas_desglosadas.keys())
                if not bodegas_tabla.issubset(bodegas_ventas):
                    import streamlit as st
                    st.warning("⚠️ **Algunas bodegas no tienen datos de ventas**")
                    st.write(f"❌ **Sin datos de ventas**: {list(bodegas_tabla - bodegas_ventas)}")
        else:
            ventas_desglosadas = {}
        
        # Función auxiliar para encontrar bodega en ventas_desglosadas
        def encontrar_bodega_ventas(bodega_tabla, ventas_desglosadas, pais):
            """Encuentra la bodega correspondiente en ventas_desglosadas usando mapeo bidireccional"""
            if bodega_tabla in ventas_desglosadas:
                return bodega_tabla
            
            if pais in ["Guatemala", "El Salvador", "Honduras", "Costa Rica", "PANAMA"]:
                # Buscar usando mapeo bidireccional
                for venta_bodega in ventas_desglosadas.keys():
                    canonical_venta = sales_processor.get_canonical_name(venta_bodega, pais)
                    canonical_stock = sales_processor.get_canonical_name(bodega_tabla, pais)
                    
                    if canonical_stock == canonical_venta:
                        return venta_bodega
            
            return None
        
        # SIEMPRE procesar todas las categorías para generar tabla completa
        categorias_ligas = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT", "ACCESSORIES"]
        logger.info("Agregando columnas de ventas para todas las ligas (tabla completa)")
        subcategorias = ["Planas", "Curvas", "Apparel"]
        
        # Agregar columnas de ventas para cada combinación liga-subcategoría
        for categoria in categorias_ligas:
                
            if categoria == 'ACCESSORIES':
                # Para ACCESSORIES, solo agregar columnas Stock y Ventas (USD)
                for subcategoria in ['Stock', 'Ventas (USD)']:
                    # SIEMPRE usar formato original para tabla completa
                    col_name = f"{categoria} - {subcategoria}"
                    # NO sobrescribir la columna Stock si ya existe con datos
                    if col_name not in tabla_final.columns:
                        tabla_final[col_name] = 0.0
                    
                    # Llenar valores por bodega
                    for bodega in tabla_final.index:
                        bodega_ventas = encontrar_bodega_ventas(bodega, ventas_desglosadas, pais)
                        if bodega != 'TOTAL' and bodega_ventas:
                            ventas_bodega = ventas_desglosadas[bodega_ventas]
                            if categoria in ventas_bodega:
                                # Para ACCESSORIES - Stock: NO sobrescribir si ya existe (viene del archivo de stock)
                                # Solo llenar si la celda está en 0 (inicializada pero sin datos de stock)
                                if subcategoria == 'Stock' and 'Stock' in ventas_bodega[categoria]:
                                    # Solo sobrescribir si no hay datos de stock (valor es 0)
                                    if tabla_final.loc[bodega, col_name] == 0:
                                        datos_stock = ventas_bodega[categoria]['Stock']
                                        if isinstance(datos_stock, dict) and 'Cantidad' in datos_stock:
                                            tabla_final.loc[bodega, col_name] = datos_stock['Cantidad']  # Para stock usar cantidad
                                        else:
                                            tabla_final.loc[bodega, col_name] = datos_stock
                                elif subcategoria == 'Ventas (USD)' and 'Ventas' in ventas_bodega[categoria]:
                                    datos_ventas = ventas_bodega[categoria]['Ventas']
                                    if isinstance(datos_ventas, dict) and 'USD' in datos_ventas:
                                        tabla_final.loc[bodega, col_name] = datos_ventas['USD']  # Para ventas usar USD
                                    else:
                                        tabla_final.loc[bodega, col_name] = datos_ventas
                    
                    # Calcular total para la fila TOTAL
                    if 'TOTAL' in tabla_final.index:
                        total_categoria_subcategoria = 0
                        if subcategoria == 'Stock':
                            # Para Stock de ACCESSORIES, NUNCA sobrescribir datos existentes
                            # Los datos de stock original ya están calculados correctamente
                            # Solo calcular desde ventas si NO HAY datos de stock en absoluto
                            current_stock_total = tabla_final.loc['TOTAL', col_name]
                            if current_stock_total == 0:
                                # Verificar si hay datos de stock reales en otras filas
                                has_real_stock_data = any(
                                    tabla_final.loc[bodega, col_name] > 0 
                                    for bodega in tabla_final.index 
                                    if bodega != 'TOTAL'
                                )
                                # Solo usar datos de ventas si no hay datos de stock reales
                                if not has_real_stock_data:
                                    for bodega, ventas_bodega in ventas_desglosadas.items():
                                        if categoria in ventas_bodega and 'Stock' in ventas_bodega[categoria]:
                                            datos_stock = ventas_bodega[categoria]['Stock']
                                            if isinstance(datos_stock, dict) and 'Cantidad' in datos_stock:
                                                total_categoria_subcategoria += datos_stock['Cantidad']
                                            else:
                                                total_categoria_subcategoria += datos_stock
                                    tabla_final.loc['TOTAL', col_name] = total_categoria_subcategoria
                        else:
                            # Para Ventas (USD), siempre usar datos de ventas
                            for bodega, ventas_bodega in ventas_desglosadas.items():
                                if categoria in ventas_bodega and 'Ventas' in ventas_bodega[categoria]:
                                    datos_ventas = ventas_bodega[categoria]['Ventas']
                                    if isinstance(datos_ventas, dict) and 'USD' in datos_ventas:
                                        total_categoria_subcategoria += datos_ventas['USD']
                                    else:
                                        total_categoria_subcategoria += datos_ventas
                            tabla_final.loc['TOTAL', col_name] = total_categoria_subcategoria
            else:
                # Lógica para otras ligas - SIEMPRE generar tabla completa
                for subcategoria in subcategorias:
                    # SIEMPRE usar formato original para tabla completa
                    col_name = f"{categoria} - {subcategoria} - Ventas"
                    tabla_final[col_name] = 0.0
                    
                    # Llenar valores por bodega
                    for bodega in tabla_final.index:
                        bodega_ventas = encontrar_bodega_ventas(bodega, ventas_desglosadas, pais)
                        if bodega != 'TOTAL' and bodega_ventas:
                            ventas_bodega = ventas_desglosadas[bodega_ventas]
                            if categoria in ventas_bodega and subcategoria in ventas_bodega[categoria]:
                                datos_venta = ventas_bodega[categoria][subcategoria]
                                # Para Honduras, los datos vienen como diccionario {'Cantidad': X, 'USD': Y}
                                # Para las columnas de ventas, usar el valor USD
                                if isinstance(datos_venta, dict) and 'USD' in datos_venta:
                                    valor_venta = datos_venta['USD']
                                else:
                                    valor_venta = datos_venta  # Para otros países que ya devuelven valor escalar
                                tabla_final.loc[bodega, col_name] = valor_venta
                    
                    # Calcular total para la fila TOTAL
                    if 'TOTAL' in tabla_final.index:
                        total_categoria_subcategoria = 0
                        for bodega, ventas_bodega in ventas_desglosadas.items():
                            if categoria in ventas_bodega and subcategoria in ventas_bodega[categoria]:
                                datos_venta = ventas_bodega[categoria][subcategoria]
                                # Para Honduras, los datos vienen como diccionario {'Cantidad': X, 'USD': Y}
                                if isinstance(datos_venta, dict) and 'USD' in datos_venta:
                                    total_categoria_subcategoria += datos_venta['USD']
                                else:
                                    total_categoria_subcategoria += datos_venta
                        tabla_final.loc['TOTAL', col_name] = total_categoria_subcategoria
        
        return tabla_final
    
    def _format_table(self, tabla_final: pd.DataFrame, selected_league: str = None, hay_ventas: bool = False) -> pd.DataFrame:
        """Formatea la tabla final con MultiIndex de 3 niveles: Liga → Subcategoría → Stock/Ventas"""
        tabla_final.reset_index(inplace=True)
        tabla_final.rename(columns={'index': 'Bodega'}, inplace=True)
        
        # Crear MultiIndex para columnas con 3 niveles
        columnas_multi = [('INFO', 'INFO', 'Bodega')]
        
        # Definir qué categorías incluir según el filtro
        if selected_league:
            # Solo incluir la liga seleccionada
            categorias_para_multiindex = {selected_league: self.league_categories.get_category_values(selected_league)}
        else:
            # Incluir todas las categorías
            categorias_para_multiindex = self.league_categories.get_all_categories()
            
        # Para cada liga y subcategoría, crear columnas Stock y Ventas
        for categoria in categorias_para_multiindex.keys():
            if categoria == 'ACCESSORIES':
                # Para ACCESSORIES, crear columna Stock siempre, Ventas solo si hay datos de ventas
                columnas_multi.append((categoria, 'Accessories', 'Stock'))
                if hay_ventas:
                    columnas_multi.append((categoria, 'Accessories', 'Ventas (USD)'))
            else:
                for subcategoria in ['Planas', 'Curvas', 'Apparel']:
                    # Stock y Ventas para cada subcategoría
                    columnas_multi.extend([
                        (categoria, subcategoria, 'Stock'),
                        (categoria, subcategoria, 'Ventas')
                    ])
        
        # Columnas de totales 
        if selected_league:
            # Para liga específica, NO incluir CAPACIDAD EN TIENDA ni % DE CUMPLIMIENTO
            columnas_totales = [
                ('TOTALES', 'RESUMEN', 'TOTAL PLANAS'),
                ('TOTALES', 'RESUMEN', 'TOTAL CURVAS'),
                ('TOTALES', 'RESUMEN', 'TOTAL APPAREL'),
                ('TOTALES', 'RESUMEN', 'TOTAL HEADWEAR'),
                ('TOTALES', 'RESUMEN', 'TOTAL STOCK')
            ]
            # Solo agregar TOTAL (USD) si hay datos de ventas
            if hay_ventas:
                columnas_totales.append(('TOTALES', 'RESUMEN', 'TOTAL (USD)'))
            columnas_multi.extend(columnas_totales)
        else:
            # Para todas las ligas, incluir columnas según disponibilidad
            columnas_totales = [
                ('TOTALES', 'RESUMEN', 'TOTAL PLANAS'),
                ('TOTALES', 'RESUMEN', 'TOTAL CURVAS'),
                ('TOTALES', 'RESUMEN', 'TOTAL APPAREL'),
                ('TOTALES', 'RESUMEN', 'TOTAL HEADWEAR'),
                ('TOTALES', 'RESUMEN', 'CAPACIDAD EN TIENDA'),
                ('TOTALES', 'RESUMEN', '% DE CUMPLIMIENTO'),
                ('TOTALES', 'RESUMEN', 'TOTAL STOCK')
            ]
            # Solo agregar TOTAL (USD) si hay datos de ventas
            if hay_ventas:
                columnas_totales.append(('TOTALES', 'RESUMEN', 'TOTAL (USD)'))
            columnas_multi.extend(columnas_totales)
        
        # Crear diccionario de mapeo de nombres de columnas
        mapeo_columnas = {'Bodega': ('INFO', 'INFO', 'Bodega')}
        
        # Definir qué categorías incluir según el filtro
        if selected_league:
            # Solo incluir la liga seleccionada
            categorias_a_incluir = {selected_league: self.league_categories.get_category_values(selected_league)}
            logger.info(f"Filtrando tabla para mostrar solo: {selected_league}")
        else:
            # Incluir todas las categorías
            categorias_a_incluir = self.league_categories.get_all_categories()
            logger.info("Mostrando tabla completa con todas las ligas")
        
        # Mapear columnas de stock y ventas existentes
        for categoria in categorias_a_incluir.keys():
            if categoria == 'ACCESSORIES':
                # Para ACCESSORIES, mapear columnas Stock y Ventas (USD)
                nombre_stock = f"{categoria} - Stock"
                nombre_ventas = f"{categoria} - Ventas (USD)"
                
                if nombre_stock in tabla_final.columns:
                    mapeo_columnas[nombre_stock] = (categoria, 'Accessories', 'Stock')
                if nombre_ventas in tabla_final.columns:
                    mapeo_columnas[nombre_ventas] = (categoria, 'Accessories', 'Ventas (USD)')
            else:
                for subcategoria in ['Planas', 'Curvas', 'Apparel']:
                    nombre_stock = f"{categoria} - {subcategoria}"
                    nombre_ventas = f"{categoria} - {subcategoria} - Ventas"
                    
                    if nombre_stock in tabla_final.columns:
                        mapeo_columnas[nombre_stock] = (categoria, subcategoria, 'Stock')
                    if nombre_ventas in tabla_final.columns:
                        mapeo_columnas[nombre_ventas] = (categoria, subcategoria, 'Ventas')
        
        # Mapear columnas de totales
        totales_mapping = {
            'TOTAL PLANAS': ('TOTALES', 'RESUMEN', 'TOTAL PLANAS'),
            'TOTAL CURVAS': ('TOTALES', 'RESUMEN', 'TOTAL CURVAS'),
            'TOTAL APPAREL': ('TOTALES', 'RESUMEN', 'TOTAL APPAREL'),
            'TOTAL HEADWEAR': ('TOTALES', 'RESUMEN', 'TOTAL HEADWEAR'),
            'CAPACIDAD EN TIENDA': ('TOTALES', 'RESUMEN', 'CAPACIDAD EN TIENDA'),
            '% DE CUMPLIMIENTO': ('TOTALES', 'RESUMEN', '% DE CUMPLIMIENTO'),
            'TOTAL STOCK': ('TOTALES', 'RESUMEN', 'TOTAL STOCK'),
            'TOTAL (USD)': ('TOTALES', 'RESUMEN', 'TOTAL (USD)')
        }
        
        for col_original, col_multi in totales_mapping.items():
            if col_original in tabla_final.columns:
                mapeo_columnas[col_original] = col_multi
        
        # Reordenar columnas según el orden esperado
        columnas_ordenadas = []
        tuples_ordenadas = []
        
        for col_multi in columnas_multi:
            for col_original, col_mapped in mapeo_columnas.items():
                if col_mapped == col_multi and col_original in tabla_final.columns:
                    columnas_ordenadas.append(col_original)
                    tuples_ordenadas.append(col_multi)
                    break
        
        # Aplicar reordenamiento
        tabla_final = tabla_final[columnas_ordenadas]
        
        # Crear MultiIndex con 3 niveles explícitamente nombrados
        multi_index = pd.MultiIndex.from_tuples(
            tuples_ordenadas,
            names=['Liga', 'Subcategoría', 'Tipo']
        )
        tabla_final.columns = multi_index
        
        return tabla_final

    @st.cache_data(ttl=1)
    def procesar_solo_ventas_guatemala(_self, df_ventas_hash: List[Dict], selected_league: str = None) -> Optional[pd.DataFrame]:
        """Procesa datos solo de ventas (cantidades) para Guatemala sin requerir archivo de stock"""
        df_ventas = pd.DataFrame(df_ventas_hash)
        pais = "Guatemala"
        
        print(f"Procesando SOLO VENTAS Guatemala - Archivo recibido con {len(df_ventas)} filas")
        
        # Procesar cantidades usando el nuevo método
        cantidades_desglosadas = sales_processor.procesar_cantidades_guatemala(df_ventas)
        
        if not cantidades_desglosadas:
            print("No se pudieron procesar las cantidades de Guatemala")
            return None
        
        
        # Debug: Verificar qué bodegas están en los datos
        print("DEBUG - Bodegas disponibles en cantidades_desglosadas:")
        for bodega in cantidades_desglosadas.keys():
            print(f"  - {bodega}")
        
        # Usar directamente las bodegas de los datos procesados en lugar de country_data
        # Crear estructura de tabla similar a procesar_datos_consolidados pero sin capacidades
        categorias_ligas = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT", "ACCESSORIES"]
        tabla_final = pd.DataFrame()
        
        # Inicializar tabla con bodegas de los datos procesados
        for bodega in cantidades_desglosadas.keys():
            if bodega == "CENTRAL NEW ERA":  # Excluir bodega central
                continue
                
            fila_bodega = {'Bodega': bodega}
            
            # Procesar cada liga
            for categoria in categorias_ligas:
                if categoria == "ACCESSORIES":
                    # Para ACCESSORIES, dos columnas: Cantidad y TOTAL USD
                    col_cantidad = f"{categoria} - Cantidad"
                    col_usd = f"{categoria} - TOTAL USD"
                    if (bodega in cantidades_desglosadas and 
                        categoria in cantidades_desglosadas[bodega] and
                        'Stock' in cantidades_desglosadas[bodega][categoria]):
                        stock_data = cantidades_desglosadas[bodega][categoria]['Stock']
                        if isinstance(stock_data, dict):
                            fila_bodega[col_cantidad] = stock_data.get('Cantidad', 0)
                            fila_bodega[col_usd] = stock_data.get('USD', 0)
                        else:
                            fila_bodega[col_cantidad] = 0
                            fila_bodega[col_usd] = 0
                    else:
                        fila_bodega[col_cantidad] = 0
                        fila_bodega[col_usd] = 0
                else:
                    # Para otras ligas, subcategorías con dos columnas cada una
                    subcategorias = ["Planas", "Curvas", "Apparel"]
                    for subcat in subcategorias:
                        col_cantidad = f"{categoria} - {subcat} - Cantidad"
                        col_usd = f"{categoria} - {subcat} - TOTAL USD"
                        if (bodega in cantidades_desglosadas and 
                            categoria in cantidades_desglosadas[bodega] and
                            subcat in cantidades_desglosadas[bodega][categoria]):
                            subcat_data = cantidades_desglosadas[bodega][categoria][subcat]
                            if isinstance(subcat_data, dict):
                                fila_bodega[col_cantidad] = subcat_data.get('Cantidad', 0)
                                fila_bodega[col_usd] = subcat_data.get('USD', 0)
                            else:
                                fila_bodega[col_cantidad] = 0
                                fila_bodega[col_usd] = 0
                        else:
                            fila_bodega[col_cantidad] = 0
                            fila_bodega[col_usd] = 0
            
            # Calcular totales por bodega (solo cantidades, sin USD para totales)
            categorias_principales = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT"]
            total_planas = sum([fila_bodega.get(f"{cat} - Planas - Cantidad", 0) for cat in categorias_principales])
            total_curvas = sum([fila_bodega.get(f"{cat} - Curvas - Cantidad", 0) for cat in categorias_principales])
            total_apparel = sum([fila_bodega.get(f"{cat} - Apparel - Cantidad", 0) for cat in categorias_principales])
            
            fila_bodega['TOTAL PLANAS'] = total_planas
            fila_bodega['TOTAL CURVAS'] = total_curvas
            fila_bodega['TOTAL APPAREL'] = total_apparel
            fila_bodega['TOTAL HEADWEAR'] = total_planas + total_curvas
            fila_bodega['TOTAL STOCK'] = total_planas + total_curvas + total_apparel + fila_bodega.get('ACCESSORIES - Cantidad', 0)
            
            # Calcular TOTAL USD como suma horizontal de todas las celdas TOTAL USD de la fila
            total_usd = 0
            # Sumar USD de todas las ligas principales (Planas, Curvas, Apparel)
            for categoria in categorias_principales:
                total_usd += fila_bodega.get(f"{categoria} - Planas - TOTAL USD", 0)
                total_usd += fila_bodega.get(f"{categoria} - Curvas - TOTAL USD", 0)
                total_usd += fila_bodega.get(f"{categoria} - Apparel - TOTAL USD", 0)
            # Sumar USD de ACCESSORIES
            total_usd += fila_bodega.get('ACCESSORIES - TOTAL USD', 0)
            
            fila_bodega['TOTAL USD'] = total_usd
            
            # Agregar fila a la tabla
            tabla_final = pd.concat([tabla_final, pd.DataFrame([fila_bodega], index=[bodega])])
        
        # Calcular fila de totales
        if not tabla_final.empty:
            fila_totales = {}
            for col in tabla_final.columns:
                if col != 'Bodega':  # Excluir columna de texto
                    fila_totales[col] = tabla_final[col].sum()
                else:
                    fila_totales[col] = 'TOTAL'  # Poner etiqueta 'TOTAL' en lugar de sumar
            
            # Agregar fila TOTAL
            tabla_final.loc['TOTAL'] = fila_totales
        
        # Llenar valores NaN con 0 y convertir columnas apropiadamente
        tabla_final = tabla_final.fillna(0)
        
        # Convertir columnas de cantidad a enteros, mantener USD como float
        for col in tabla_final.columns:
            if 'USD' in col:
                tabla_final[col] = tabla_final[col].astype(float)
            elif col != 'Bodega':
                tabla_final[col] = tabla_final[col].astype(int)
        
        # IMPORTANTE: Agregar columna 'Bodega' desde el índice
        tabla_final['Bodega'] = tabla_final.index
        
        # Aplicar filtro de liga si es necesario
        if selected_league and selected_league != "Todas":
            columnas_filtradas = ['TOTAL PLANAS', 'TOTAL CURVAS', 'TOTAL APPAREL', 'TOTAL HEADWEAR', 'TOTAL STOCK']
            # Agregar columnas de la liga específica
            for col in tabla_final.columns:
                if selected_league in col:
                    columnas_filtradas.append(col)
            
            tabla_final = tabla_final[columnas_filtradas]
        
        # Formatear tabla con MultiIndex (sin capacidades ni % cumplimiento)
        tabla_final = _self._format_table_solo_ventas(tabla_final, selected_league)
        
        print(f"Tabla solo-ventas Guatemala generada con {len(tabla_final)} filas y {len(tabla_final.columns)} columnas")
        return tabla_final

    def procesar_solo_ventas_el_salvador(_self, df_ventas_hash: List[Dict], selected_league: str = None) -> Optional[pd.DataFrame]:
        """Procesa datos solo de ventas (cantidades) para El Salvador sin requerir archivo de stock"""
        df_ventas = pd.DataFrame(df_ventas_hash)
        pais = "El Salvador"
        
        print(f"Procesando SOLO VENTAS El Salvador - Archivo recibido con {len(df_ventas)} filas")
        
        # Procesar cantidades usando el nuevo método
        cantidades_desglosadas = sales_processor.procesar_cantidades_el_salvador(df_ventas)
        
        if not cantidades_desglosadas:
            print("No se pudieron procesar las cantidades de El Salvador")
            return None
        
        
        # Debug: Verificar qué bodegas están en los datos
        print("DEBUG - Bodegas disponibles en cantidades_desglosadas:")
        for bodega in cantidades_desglosadas.keys():
            print(f"  - {bodega}")
        
        # Usar directamente las bodegas de los datos procesados en lugar de country_data
        # Crear estructura de tabla similar a procesar_datos_consolidados pero sin capacidades
        categorias_ligas = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT", "ACCESSORIES"]
        tabla_final = pd.DataFrame()
        
        # Inicializar tabla con bodegas de los datos procesados
        for bodega in cantidades_desglosadas.keys():
            if bodega == "CENTRAL NEW ERA":  # Excluir bodega central
                continue
                
            fila_bodega = {'Bodega': bodega}
            
            # Procesar cada liga
            for categoria in categorias_ligas:
                if categoria == "ACCESSORIES":
                    # Para ACCESSORIES, dos columnas: Cantidad y TOTAL USD
                    col_cantidad = f"{categoria} - Cantidad"
                    col_usd = f"{categoria} - TOTAL USD"
                    if (bodega in cantidades_desglosadas and 
                        categoria in cantidades_desglosadas[bodega] and
                        'Stock' in cantidades_desglosadas[bodega][categoria]):
                        stock_data = cantidades_desglosadas[bodega][categoria]['Stock']
                        if isinstance(stock_data, dict):
                            fila_bodega[col_cantidad] = stock_data.get('Cantidad', 0)
                            fila_bodega[col_usd] = stock_data.get('USD', 0)
                        else:
                            fila_bodega[col_cantidad] = 0
                            fila_bodega[col_usd] = 0
                    else:
                        fila_bodega[col_cantidad] = 0
                        fila_bodega[col_usd] = 0
                else:
                    # Para otras ligas, subcategorías con dos columnas cada una
                    subcategorias = ["Planas", "Curvas", "Apparel"]
                    for subcat in subcategorias:
                        col_cantidad = f"{categoria} - {subcat} - Cantidad"
                        col_usd = f"{categoria} - {subcat} - TOTAL USD"
                        if (bodega in cantidades_desglosadas and 
                            categoria in cantidades_desglosadas[bodega] and
                            subcat in cantidades_desglosadas[bodega][categoria]):
                            subcat_data = cantidades_desglosadas[bodega][categoria][subcat]
                            if isinstance(subcat_data, dict):
                                fila_bodega[col_cantidad] = subcat_data.get('Cantidad', 0)
                                fila_bodega[col_usd] = subcat_data.get('USD', 0)
                            else:
                                fila_bodega[col_cantidad] = 0
                                fila_bodega[col_usd] = 0
                        else:
                            fila_bodega[col_cantidad] = 0
                            fila_bodega[col_usd] = 0
            
            # Calcular totales por bodega (solo cantidades, sin USD para totales)
            categorias_principales = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT"]
            total_planas = sum([fila_bodega.get(f"{cat} - Planas - Cantidad", 0) for cat in categorias_principales])
            total_curvas = sum([fila_bodega.get(f"{cat} - Curvas - Cantidad", 0) for cat in categorias_principales])
            total_apparel = sum([fila_bodega.get(f"{cat} - Apparel - Cantidad", 0) for cat in categorias_principales])
            
            fila_bodega['TOTAL PLANAS'] = total_planas
            fila_bodega['TOTAL CURVAS'] = total_curvas
            fila_bodega['TOTAL APPAREL'] = total_apparel
            fila_bodega['TOTAL HEADWEAR'] = total_planas + total_curvas
            fila_bodega['TOTAL STOCK'] = total_planas + total_curvas + total_apparel + fila_bodega.get('ACCESSORIES - Cantidad', 0)
            
            # Calcular TOTAL USD como suma horizontal de todas las celdas TOTAL USD de la fila
            total_usd = 0
            # Sumar USD de todas las ligas principales (Planas, Curvas, Apparel)
            for categoria in categorias_principales:
                total_usd += fila_bodega.get(f"{categoria} - Planas - TOTAL USD", 0)
                total_usd += fila_bodega.get(f"{categoria} - Curvas - TOTAL USD", 0)
                total_usd += fila_bodega.get(f"{categoria} - Apparel - TOTAL USD", 0)
            # Sumar USD de ACCESSORIES
            total_usd += fila_bodega.get('ACCESSORIES - TOTAL USD', 0)
            
            fila_bodega['TOTAL USD'] = total_usd
            
            # Agregar fila a la tabla
            tabla_final = pd.concat([tabla_final, pd.DataFrame([fila_bodega], index=[bodega])])
        
        # Calcular fila de totales
        if not tabla_final.empty:
            fila_totales = {}
            for col in tabla_final.columns:
                if col != 'Bodega':  # Excluir columna de texto
                    fila_totales[col] = tabla_final[col].sum()
                else:
                    fila_totales[col] = 'TOTAL'  # Poner etiqueta 'TOTAL' en lugar de sumar
            
            # Agregar fila TOTAL
            tabla_final.loc['TOTAL'] = fila_totales
        
        # Llenar valores NaN con 0 y convertir columnas apropiadamente
        tabla_final = tabla_final.fillna(0)
        
        # Convertir columnas de cantidad a enteros, mantener USD como float
        for col in tabla_final.columns:
            if 'USD' in col:
                tabla_final[col] = tabla_final[col].astype(float)
            elif col != 'Bodega':
                tabla_final[col] = tabla_final[col].astype(int)
        
        # IMPORTANTE: Agregar columna 'Bodega' desde el índice
        tabla_final['Bodega'] = tabla_final.index
        
        # Aplicar filtro de liga si es necesario
        if selected_league and selected_league != "Todas":
            columnas_filtradas = ['TOTAL PLANAS', 'TOTAL CURVAS', 'TOTAL APPAREL', 'TOTAL HEADWEAR', 'TOTAL STOCK']
            # Agregar columnas de la liga específica
            for col in tabla_final.columns:
                if selected_league in col:
                    columnas_filtradas.append(col)
            
            tabla_final = tabla_final[columnas_filtradas]
        
        # Formatear tabla con MultiIndex (sin capacidades ni % cumplimiento)
        tabla_final = _self._format_table_solo_ventas(tabla_final, selected_league)
        
        print(f"Tabla solo-ventas El Salvador generada con {len(tabla_final)} filas y {len(tabla_final.columns)} columnas")
        return tabla_final

    def procesar_solo_ventas_honduras(_self, df_ventas_hash: List[Dict], selected_league: str = None) -> Optional[pd.DataFrame]:
        """Procesa datos solo de ventas (cantidades) para Honduras sin requerir archivo de stock"""
        df_ventas = pd.DataFrame(df_ventas_hash)
        pais = "Honduras"
        
        print(f"Procesando SOLO VENTAS Honduras - Archivo recibido con {len(df_ventas)} filas")
        
        # Procesar cantidades usando el nuevo método
        cantidades_desglosadas = sales_processor.procesar_cantidades_honduras(df_ventas)
        
        if not cantidades_desglosadas:
            print("No se pudieron procesar las cantidades de Honduras")
            return None
        
        
        # Debug: Verificar qué bodegas están en los datos
        print("DEBUG - Bodegas disponibles en cantidades_desglosadas:")
        for bodega in cantidades_desglosadas.keys():
            print(f"  - {bodega}")
        
        # Usar directamente las bodegas de los datos procesados en lugar de country_data
        # Crear estructura de tabla similar a procesar_datos_consolidados pero sin capacidades
        categorias_ligas = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT", "ACCESSORIES"]
        tabla_final = pd.DataFrame()
        
        # Inicializar tabla con bodegas de los datos procesados
        for bodega in cantidades_desglosadas.keys():
            if bodega == "CENTRAL NEW ERA":  # Excluir bodega central
                continue
                
            fila_bodega = {'Bodega': bodega}
            
            # Procesar cada liga
            for categoria in categorias_ligas:
                if categoria == "ACCESSORIES":
                    # Para ACCESSORIES, dos columnas: Cantidad y TOTAL USD
                    col_cantidad = f"{categoria} - Cantidad"
                    col_usd = f"{categoria} - TOTAL USD"
                    if (bodega in cantidades_desglosadas and 
                        categoria in cantidades_desglosadas[bodega] and
                        'Stock' in cantidades_desglosadas[bodega][categoria]):
                        stock_data = cantidades_desglosadas[bodega][categoria]['Stock']
                        if isinstance(stock_data, dict):
                            fila_bodega[col_cantidad] = stock_data.get('Cantidad', 0)
                            fila_bodega[col_usd] = stock_data.get('USD', 0)
                        else:
                            fila_bodega[col_cantidad] = 0
                            fila_bodega[col_usd] = 0
                    else:
                        fila_bodega[col_cantidad] = 0
                        fila_bodega[col_usd] = 0
                else:
                    # Para otras ligas, subcategorías con dos columnas cada una
                    subcategorias = ["Planas", "Curvas", "Apparel"]
                    for subcat in subcategorias:
                        col_cantidad = f"{categoria} - {subcat} - Cantidad"
                        col_usd = f"{categoria} - {subcat} - TOTAL USD"
                        if (bodega in cantidades_desglosadas and 
                            categoria in cantidades_desglosadas[bodega] and
                            subcat in cantidades_desglosadas[bodega][categoria]):
                            subcat_data = cantidades_desglosadas[bodega][categoria][subcat]
                            if isinstance(subcat_data, dict):
                                fila_bodega[col_cantidad] = subcat_data.get('Cantidad', 0)
                                fila_bodega[col_usd] = subcat_data.get('USD', 0)
                            else:
                                fila_bodega[col_cantidad] = 0
                                fila_bodega[col_usd] = 0
                        else:
                            fila_bodega[col_cantidad] = 0
                            fila_bodega[col_usd] = 0
            
            # Calcular totales por bodega (solo cantidades, sin USD para totales)
            categorias_principales = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT"]
            total_planas = sum([fila_bodega.get(f"{cat} - Planas - Cantidad", 0) for cat in categorias_principales])
            total_curvas = sum([fila_bodega.get(f"{cat} - Curvas - Cantidad", 0) for cat in categorias_principales])
            total_apparel = sum([fila_bodega.get(f"{cat} - Apparel - Cantidad", 0) for cat in categorias_principales])
            
            fila_bodega['TOTAL PLANAS'] = total_planas
            fila_bodega['TOTAL CURVAS'] = total_curvas
            fila_bodega['TOTAL APPAREL'] = total_apparel
            fila_bodega['TOTAL HEADWEAR'] = total_planas + total_curvas
            fila_bodega['TOTAL STOCK'] = total_planas + total_curvas + total_apparel + fila_bodega.get('ACCESSORIES - Cantidad', 0)
            
            # Calcular TOTAL USD como suma horizontal de todas las celdas TOTAL USD de la fila
            total_usd = 0
            # Sumar USD de todas las ligas principales (Planas, Curvas, Apparel)
            for categoria in categorias_principales:
                total_usd += fila_bodega.get(f"{categoria} - Planas - TOTAL USD", 0)
                total_usd += fila_bodega.get(f"{categoria} - Curvas - TOTAL USD", 0)
                total_usd += fila_bodega.get(f"{categoria} - Apparel - TOTAL USD", 0)
            # Sumar USD de ACCESSORIES
            total_usd += fila_bodega.get('ACCESSORIES - TOTAL USD', 0)
            
            fila_bodega['TOTAL USD'] = total_usd
            
            # Agregar fila a la tabla
            tabla_final = pd.concat([tabla_final, pd.DataFrame([fila_bodega], index=[bodega])])
        
        # Calcular fila de totales
        if not tabla_final.empty:
            fila_totales = {}
            for col in tabla_final.columns:
                if col != 'Bodega':  # Excluir columna de texto
                    fila_totales[col] = tabla_final[col].sum()
                else:
                    fila_totales[col] = 'TOTAL'  # Poner etiqueta 'TOTAL' en lugar de sumar
            
            # Agregar fila TOTAL
            tabla_final.loc['TOTAL'] = fila_totales
        
        # Llenar valores NaN con 0 y convertir columnas apropiadamente
        tabla_final = tabla_final.fillna(0)
        
        # Convertir columnas de cantidad a enteros, mantener USD como float
        for col in tabla_final.columns:
            if 'USD' in col:
                tabla_final[col] = tabla_final[col].astype(float)
            elif col != 'Bodega':
                tabla_final[col] = tabla_final[col].astype(int)
        
        # IMPORTANTE: Agregar columna 'Bodega' desde el índice
        tabla_final['Bodega'] = tabla_final.index
        
        # Aplicar filtro de liga si es necesario
        if selected_league and selected_league != "Todas":
            columnas_filtradas = ['TOTAL PLANAS', 'TOTAL CURVAS', 'TOTAL APPAREL', 'TOTAL HEADWEAR', 'TOTAL STOCK']
            # Agregar columnas de la liga específica
            for col in tabla_final.columns:
                if selected_league in col:
                    columnas_filtradas.append(col)
            
            tabla_final = tabla_final[columnas_filtradas]
        
        # Formatear tabla con MultiIndex (sin capacidades ni % cumplimiento)
        tabla_final = _self._format_table_solo_ventas(tabla_final, selected_league)
        
        print(f"Tabla solo-ventas Honduras generada con {len(tabla_final)} filas y {len(tabla_final.columns)} columnas")
        return tabla_final

    def procesar_solo_ventas_costa_rica(_self, df_ventas_hash: List[Dict], selected_league: str = None) -> Optional[pd.DataFrame]:
        """Procesa datos solo de ventas (cantidades) para Costa Rica sin requerir archivo de stock"""
        df_ventas = pd.DataFrame(df_ventas_hash)
        pais = "Costa Rica"
        
        print(f"Procesando SOLO VENTAS Costa Rica - Archivo recibido con {len(df_ventas)} filas")
        
        # Procesar cantidades usando el nuevo método
        cantidades_desglosadas = sales_processor.procesar_cantidades_costa_rica(df_ventas)
        
        if not cantidades_desglosadas:
            print("No se pudieron procesar las cantidades de Costa Rica")
            return None
        
        
        # Debug: Verificar qué bodegas están en los datos
        print("DEBUG - Bodegas disponibles en cantidades_desglosadas:")
        for bodega in cantidades_desglosadas.keys():
            print(f"  - {bodega}")
        
        # Usar directamente las bodegas de los datos procesados en lugar de country_data
        bodegas_disponibles = list(cantidades_desglosadas.keys())
        print(f"Bodegas que se incluirán en la tabla: {bodegas_disponibles}")
        
        # Crear tabla vacía con las bodegas disponibles
        tabla_final = pd.DataFrame(index=bodegas_disponibles)
        
        # Crear estructura de tabla similar a Guatemala
        categorias_ligas = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT", "ACCESSORIES"]
        tabla_final = pd.DataFrame()
        
        # Inicializar tabla con bodegas de los datos procesados
        for bodega in cantidades_desglosadas.keys():
            if "Central" in bodega or "central" in bodega:  # Excluir bodega central
                continue
                
            fila_bodega = {'Bodega': bodega}
            
            # Procesar cada liga
            for categoria in categorias_ligas:
                if categoria == "ACCESSORIES":
                    # Para ACCESSORIES, dos columnas: Cantidad y TOTAL USD
                    col_cantidad = f"{categoria} - Cantidad"
                    col_usd = f"{categoria} - TOTAL USD"
                    if (bodega in cantidades_desglosadas and 
                        categoria in cantidades_desglosadas[bodega] and
                        'Stock' in cantidades_desglosadas[bodega][categoria]):
                        subcat_data = cantidades_desglosadas[bodega][categoria]['Stock']
                        if isinstance(subcat_data, dict):
                            fila_bodega[col_cantidad] = subcat_data.get('Cantidad', 0)
                            fila_bodega[col_usd] = subcat_data.get('USD', 0)
                        else:
                            fila_bodega[col_cantidad] = 0
                            fila_bodega[col_usd] = 0
                    else:
                        fila_bodega[col_cantidad] = 0
                        fila_bodega[col_usd] = 0
                else:
                    # Para otras ligas: Planas, Curvas, Apparel (cada una con Cantidad y TOTAL USD)
                    for subcategoria in ["Planas", "Curvas", "Apparel"]:
                        col_cantidad = f"{categoria} - {subcategoria} - Cantidad"
                        col_usd = f"{categoria} - {subcategoria} - TOTAL USD"
                        
                        if (bodega in cantidades_desglosadas and 
                            categoria in cantidades_desglosadas[bodega] and
                            subcategoria in cantidades_desglosadas[bodega][categoria]):
                            subcat_data = cantidades_desglosadas[bodega][categoria][subcategoria]
                            if isinstance(subcat_data, dict):
                                fila_bodega[col_cantidad] = subcat_data.get('Cantidad', 0)
                                fila_bodega[col_usd] = subcat_data.get('USD', 0)
                            else:
                                fila_bodega[col_cantidad] = 0
                                fila_bodega[col_usd] = 0
                        else:
                            fila_bodega[col_cantidad] = 0
                            fila_bodega[col_usd] = 0
            
            # Calcular totales por bodega (solo cantidades, sin USD para totales)
            categorias_principales = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT"]
            total_planas = sum([fila_bodega.get(f"{cat} - Planas - Cantidad", 0) for cat in categorias_principales])
            total_curvas = sum([fila_bodega.get(f"{cat} - Curvas - Cantidad", 0) for cat in categorias_principales])
            total_apparel = sum([fila_bodega.get(f"{cat} - Apparel - Cantidad", 0) for cat in categorias_principales])
            
            fila_bodega['TOTAL PLANAS'] = total_planas
            fila_bodega['TOTAL CURVAS'] = total_curvas
            fila_bodega['TOTAL APPAREL'] = total_apparel
            fila_bodega['TOTAL HEADWEAR'] = total_planas + total_curvas
            fila_bodega['TOTAL STOCK'] = total_planas + total_curvas + total_apparel + fila_bodega.get('ACCESSORIES - Cantidad', 0)
            
            # Calcular TOTAL USD como suma horizontal de todas las celdas TOTAL USD de la fila
            total_usd = 0
            # Sumar USD de todas las ligas principales (Planas, Curvas, Apparel)
            for categoria in categorias_principales:
                total_usd += fila_bodega.get(f"{categoria} - Planas - TOTAL USD", 0)
                total_usd += fila_bodega.get(f"{categoria} - Curvas - TOTAL USD", 0)
                total_usd += fila_bodega.get(f"{categoria} - Apparel - TOTAL USD", 0)
            # Sumar USD de ACCESSORIES
            total_usd += fila_bodega.get('ACCESSORIES - TOTAL USD', 0)
            
            fila_bodega['TOTAL USD'] = total_usd
            
            # Agregar fila a la tabla
            tabla_final = pd.concat([tabla_final, pd.DataFrame([fila_bodega], index=[bodega])])
        
        # Calcular fila de totales
        if not tabla_final.empty:
            fila_totales = {}
            fila_totales['Bodega'] = 'TOTAL'
            
            # Sumar todas las columnas numéricas
            for col in tabla_final.columns:
                if col != 'Bodega':
                    try:
                        fila_totales[col] = tabla_final[col].sum()
                    except:
                        fila_totales[col] = 0
            
            # Agregar fila de totales
            tabla_final = pd.concat([tabla_final, pd.DataFrame([fila_totales], index=['TOTAL'])])
        
        # Aplicar filtro de liga si se especifica
        if selected_league and selected_league != "Todas":
            columnas_a_mantener = ['Bodega']
            for col in tabla_final.columns:
                if col != 'Bodega':
                    if col.startswith(selected_league) or col.startswith('TOTAL'):
                        columnas_a_mantener.append(col)
            tabla_final = tabla_final[columnas_a_mantener]
        
        # Crear mapeo para MultiIndex (igual que Guatemala)
        mapeo_columnas = {}
        
        # Mapeo de columnas regulares a formato MultiIndex
        for categoria in categorias_ligas:
            if categoria == "ACCESSORIES":
                mapeo_columnas[f"{categoria} - Cantidad"] = (categoria, 'Accessories', 'Cantidad')
                mapeo_columnas[f"{categoria} - TOTAL USD"] = (categoria, 'Accessories', 'TOTAL USD')
            else:
                for subcategoria in ["Planas", "Curvas", "Apparel"]:
                    mapeo_columnas[f"{categoria} - {subcategoria} - Cantidad"] = (categoria, subcategoria, 'Cantidad')
                    mapeo_columnas[f"{categoria} - {subcategoria} - TOTAL USD"] = (categoria, subcategoria, 'TOTAL USD')
        
        # Mapeo de totales
        totales_mapping = {
            'Bodega': ('INFO', 'INFO', 'Bodega'),
            'TOTAL PLANAS': ('TOTAL', 'TOTAL PLANAS', 'Cantidad'),
            'TOTAL CURVAS': ('TOTAL', 'TOTAL CURVAS', 'Cantidad'),
            'TOTAL APPAREL': ('TOTAL', 'TOTAL APPAREL', 'Cantidad'),
            'TOTAL HEADWEAR': ('TOTAL', 'TOTAL HEADWEAR', 'Cantidad'),
            'TOTAL STOCK': ('TOTAL', 'TOTAL STOCK', 'Cantidad'),
            'TOTAL USD': ('TOTAL', 'TOTAL USD', 'TOTAL USD')
        }
        
        # Combinar mapeos
        for col_original, col_multi in totales_mapping.items():
            if col_original in tabla_final.columns:
                mapeo_columnas[col_original] = col_multi
        
        # Crear MultiIndex
        columnas_multi = []
        
        # Agregar Bodega primero si existe
        if 'Bodega' in tabla_final.columns:
            columnas_multi.append(mapeo_columnas['Bodega'])
        
        # Luego agregar las demás columnas
        for col in tabla_final.columns:
            if col != 'Bodega':  # Excluir Bodega ya que la agregamos primero
                if col in mapeo_columnas:
                    columnas_multi.append(mapeo_columnas[col])
                else:
                    columnas_multi.append(('OTROS', 'OTROS', col))
        
        # Reordenar DataFrame para que Bodega esté primero
        if 'Bodega' in tabla_final.columns:
            cols = ['Bodega'] + [col for col in tabla_final.columns if col != 'Bodega']
            tabla_final = tabla_final[cols]
        
        multi_index = pd.MultiIndex.from_tuples(
            columnas_multi,
            names=['Liga', 'Subcategoría', 'Tipo']
        )
        tabla_final.columns = multi_index
        
        return tabla_final

    def procesar_solo_ventas_panama(_self, df_ventas_hash: List[Dict], selected_league: str = None) -> Optional[pd.DataFrame]:
        """Procesa datos solo de ventas (cantidades) para Panama sin requerir archivo de stock"""
        df_ventas = pd.DataFrame(df_ventas_hash)
        pais = "PANAMA"
        
        print(f"Procesando SOLO VENTAS Panama - Archivo recibido con {len(df_ventas)} filas")
        
        # Procesar cantidades usando el nuevo método
        cantidades_desglosadas = sales_processor.procesar_cantidades_panama(df_ventas)
        
        if not cantidades_desglosadas:
            print("No se pudieron procesar las cantidades de Panama")
            return None
        
        
        # Debug: Verificar qué bodegas están en los datos
        print("DEBUG - Bodegas disponibles en cantidades_desglosadas:")
        for bodega in cantidades_desglosadas.keys():
            print(f"  - {bodega}")
        
        # Usar directamente las bodegas de los datos procesados en lugar de country_data
        bodegas_disponibles = list(cantidades_desglosadas.keys())
        print(f"Bodegas que se incluirán en la tabla: {bodegas_disponibles}")
        
        # Crear tabla vacía con las bodegas disponibles
        tabla_final = pd.DataFrame(index=bodegas_disponibles)
        
        # Crear estructura de tabla similar a Guatemala
        categorias_ligas = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT", "ACCESSORIES"]
        tabla_final = pd.DataFrame()
        
        # Inicializar tabla con bodegas de los datos procesados
        for bodega in cantidades_desglosadas.keys():
            if "Central" in bodega or "central" in bodega or "Almacén" in bodega or "Bodega Central" in bodega:  # Excluir bodegas centrales
                continue
                
            fila_bodega = {'Bodega': bodega}
            
            # Procesar cada liga
            for categoria in categorias_ligas:
                if categoria == "ACCESSORIES":
                    # Para ACCESSORIES, dos columnas: Cantidad y TOTAL USD
                    col_cantidad = f"{categoria} - Cantidad"
                    col_usd = f"{categoria} - TOTAL USD"
                    if (bodega in cantidades_desglosadas and 
                        categoria in cantidades_desglosadas[bodega] and
                        'Stock' in cantidades_desglosadas[bodega][categoria]):
                        subcat_data = cantidades_desglosadas[bodega][categoria]['Stock']
                        if isinstance(subcat_data, dict):
                            fila_bodega[col_cantidad] = subcat_data.get('Cantidad', 0)
                            fila_bodega[col_usd] = subcat_data.get('USD', 0)
                        else:
                            fila_bodega[col_cantidad] = 0
                            fila_bodega[col_usd] = 0
                    else:
                        fila_bodega[col_cantidad] = 0
                        fila_bodega[col_usd] = 0
                else:
                    # Para otras ligas: Planas, Curvas, Apparel (cada una con Cantidad y TOTAL USD)
                    for subcategoria in ["Planas", "Curvas", "Apparel"]:
                        col_cantidad = f"{categoria} - {subcategoria} - Cantidad"
                        col_usd = f"{categoria} - {subcategoria} - TOTAL USD"
                        
                        if (bodega in cantidades_desglosadas and 
                            categoria in cantidades_desglosadas[bodega] and
                            subcategoria in cantidades_desglosadas[bodega][categoria]):
                            subcat_data = cantidades_desglosadas[bodega][categoria][subcategoria]
                            if isinstance(subcat_data, dict):
                                fila_bodega[col_cantidad] = subcat_data.get('Cantidad', 0)
                                fila_bodega[col_usd] = subcat_data.get('USD', 0)
                            else:
                                fila_bodega[col_cantidad] = 0
                                fila_bodega[col_usd] = 0
                        else:
                            fila_bodega[col_cantidad] = 0
                            fila_bodega[col_usd] = 0
            
            # Calcular totales por bodega (solo cantidades, sin USD para totales)
            categorias_principales = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT"]
            total_planas = sum([fila_bodega.get(f"{cat} - Planas - Cantidad", 0) for cat in categorias_principales])
            total_curvas = sum([fila_bodega.get(f"{cat} - Curvas - Cantidad", 0) for cat in categorias_principales])
            total_apparel = sum([fila_bodega.get(f"{cat} - Apparel - Cantidad", 0) for cat in categorias_principales])
            
            fila_bodega['TOTAL PLANAS'] = total_planas
            fila_bodega['TOTAL CURVAS'] = total_curvas
            fila_bodega['TOTAL APPAREL'] = total_apparel
            fila_bodega['TOTAL HEADWEAR'] = total_planas + total_curvas
            fila_bodega['TOTAL STOCK'] = total_planas + total_curvas + total_apparel + fila_bodega.get('ACCESSORIES - Cantidad', 0)
            
            # Calcular TOTAL USD como suma horizontal de todas las celdas TOTAL USD de la fila
            total_usd = 0
            # Sumar USD de todas las ligas principales (Planas, Curvas, Apparel)
            for categoria in categorias_principales:
                total_usd += fila_bodega.get(f"{categoria} - Planas - TOTAL USD", 0)
                total_usd += fila_bodega.get(f"{categoria} - Curvas - TOTAL USD", 0)
                total_usd += fila_bodega.get(f"{categoria} - Apparel - TOTAL USD", 0)
            # Sumar USD de ACCESSORIES
            total_usd += fila_bodega.get('ACCESSORIES - TOTAL USD', 0)
            
            fila_bodega['TOTAL USD'] = total_usd
            
            # Agregar fila a la tabla
            tabla_final = pd.concat([tabla_final, pd.DataFrame([fila_bodega], index=[bodega])])
        
        # Calcular fila de totales
        if not tabla_final.empty:
            fila_totales = {}
            fila_totales['Bodega'] = 'TOTAL'
            
            # Sumar todas las columnas numéricas
            for col in tabla_final.columns:
                if col != 'Bodega':
                    try:
                        fila_totales[col] = tabla_final[col].sum()
                    except:
                        fila_totales[col] = 0
            
            # Agregar fila de totales
            tabla_final = pd.concat([tabla_final, pd.DataFrame([fila_totales], index=['TOTAL'])])
        
        # Aplicar filtro de liga si se especifica
        if selected_league and selected_league != "Todas":
            columnas_a_mantener = ['Bodega']
            for col in tabla_final.columns:
                if col != 'Bodega':
                    if col.startswith(selected_league) or col.startswith('TOTAL'):
                        columnas_a_mantener.append(col)
            tabla_final = tabla_final[columnas_a_mantener]
        
        # Crear mapeo para MultiIndex (igual que Guatemala)
        mapeo_columnas = {}
        
        # Mapeo de columnas regulares a formato MultiIndex
        for categoria in categorias_ligas:
            if categoria == "ACCESSORIES":
                mapeo_columnas[f"{categoria} - Cantidad"] = (categoria, 'Accessories', 'Cantidad')
                mapeo_columnas[f"{categoria} - TOTAL USD"] = (categoria, 'Accessories', 'TOTAL USD')
            else:
                for subcategoria in ["Planas", "Curvas", "Apparel"]:
                    mapeo_columnas[f"{categoria} - {subcategoria} - Cantidad"] = (categoria, subcategoria, 'Cantidad')
                    mapeo_columnas[f"{categoria} - {subcategoria} - TOTAL USD"] = (categoria, subcategoria, 'TOTAL USD')
        
        # Mapeo de totales
        totales_mapping = {
            'Bodega': ('INFO', 'INFO', 'Bodega'),
            'TOTAL PLANAS': ('TOTAL', 'TOTAL PLANAS', 'Cantidad'),
            'TOTAL CURVAS': ('TOTAL', 'TOTAL CURVAS', 'Cantidad'),
            'TOTAL APPAREL': ('TOTAL', 'TOTAL APPAREL', 'Cantidad'),
            'TOTAL HEADWEAR': ('TOTAL', 'TOTAL HEADWEAR', 'Cantidad'),
            'TOTAL STOCK': ('TOTAL', 'TOTAL STOCK', 'Cantidad'),
            'TOTAL USD': ('TOTAL', 'TOTAL USD', 'TOTAL USD')
        }
        
        # Combinar mapeos
        for col_original, col_multi in totales_mapping.items():
            if col_original in tabla_final.columns:
                mapeo_columnas[col_original] = col_multi
        
        # Crear MultiIndex
        columnas_multi = []
        
        # Agregar Bodega primero si existe
        if 'Bodega' in tabla_final.columns:
            columnas_multi.append(mapeo_columnas['Bodega'])
        
        # Luego agregar las demás columnas
        for col in tabla_final.columns:
            if col != 'Bodega':  # Excluir Bodega ya que la agregamos primero
                if col in mapeo_columnas:
                    columnas_multi.append(mapeo_columnas[col])
                else:
                    columnas_multi.append(('OTROS', 'OTROS', col))
        
        # Reordenar DataFrame para que Bodega esté primero
        if 'Bodega' in tabla_final.columns:
            cols = ['Bodega'] + [col for col in tabla_final.columns if col != 'Bodega']
            tabla_final = tabla_final[cols]
        
        multi_index = pd.MultiIndex.from_tuples(
            columnas_multi,
            names=['Liga', 'Subcategoría', 'Tipo']
        )
        tabla_final.columns = multi_index
        
        return tabla_final

    def _format_table_solo_ventas(self, tabla_final: pd.DataFrame, selected_league: str = None) -> pd.DataFrame:
        """Formatea la tabla solo-ventas con MultiIndex sin capacidades ni % cumplimiento"""
        
        # Crear mapeo de columnas para MultiIndex
        mapeo_columnas = {'Bodega': ('INFO', 'INFO', 'Bodega')}
        
        # Mapear columnas de ligas y subcategorías
        categorias_ligas = ["MLB", "NBA", "NFL", "MOTORSPORT", "ENTERTAINMENT", "ACCESSORIES"]
        for categoria in categorias_ligas:
            if categoria == "ACCESSORIES":
                # Para ACCESSORIES, mapear ambas columnas
                col_cantidad = f"{categoria} - Cantidad"
                col_usd = f"{categoria} - TOTAL USD"
                if col_cantidad in tabla_final.columns:
                    mapeo_columnas[col_cantidad] = (categoria, 'Accessories', 'Cantidad')
                if col_usd in tabla_final.columns:
                    mapeo_columnas[col_usd] = (categoria, 'Accessories', 'TOTAL USD')
            else:
                subcategorias = ["Planas", "Curvas", "Apparel"]
                for subcat in subcategorias:
                    # Para cada subcategoría, mapear ambas columnas
                    col_cantidad = f"{categoria} - {subcat} - Cantidad"
                    col_usd = f"{categoria} - {subcat} - TOTAL USD"
                    if col_cantidad in tabla_final.columns:
                        mapeo_columnas[col_cantidad] = (categoria, subcat, 'Cantidad')
                    if col_usd in tabla_final.columns:
                        mapeo_columnas[col_usd] = (categoria, subcat, 'TOTAL USD')
        
        # Mapear columnas de totales
        totales_mapping = {
            'TOTAL PLANAS': ('TOTALES', 'RESUMEN', 'TOTAL PLANAS'),
            'TOTAL CURVAS': ('TOTALES', 'RESUMEN', 'TOTAL CURVAS'),
            'TOTAL APPAREL': ('TOTALES', 'RESUMEN', 'TOTAL APPAREL'),
            'TOTAL HEADWEAR': ('TOTALES', 'RESUMEN', 'TOTAL HEADWEAR'),
            'TOTAL STOCK': ('TOTALES', 'RESUMEN', 'TOTAL STOCK'),
            'TOTAL USD': ('TOTALES', 'RESUMEN', 'TOTAL USD')
        }
        
        for col_original, col_multi in totales_mapping.items():
            if col_original in tabla_final.columns:
                mapeo_columnas[col_original] = col_multi
        
        # Crear MultiIndex - IMPORTANTE: Asegurar que Bodega aparezca primero
        columnas_multi = []
        
        # Agregar Bodega primero si existe
        if 'Bodega' in tabla_final.columns:
            columnas_multi.append(mapeo_columnas['Bodega'])
        
        # Luego agregar las demás columnas
        for col in tabla_final.columns:
            if col != 'Bodega':  # Excluir Bodega ya que la agregamos primero
                if col in mapeo_columnas:
                    columnas_multi.append(mapeo_columnas[col])
                else:
                    columnas_multi.append(('OTROS', 'OTROS', col))
        
        # Reordenar DataFrame para que Bodega esté primero
        if 'Bodega' in tabla_final.columns:
            cols = ['Bodega'] + [col for col in tabla_final.columns if col != 'Bodega']
            tabla_final = tabla_final[cols]
        
        multi_index = pd.MultiIndex.from_tuples(
            columnas_multi,
            names=['Liga', 'Subcategoría', 'Tipo']
        )
        tabla_final.columns = multi_index
        
        return tabla_final

# Instancia del procesador de datos
data_processor = DataProcessor(country_manager, league_categories, product_classifier)

class ChartVisualizer:
    """Visualizador de gráficas con Plotly"""
    
    def __init__(self, stock_analyzer: StockAnalyzer, country_manager: CountryManager):
        self.stock_analyzer = stock_analyzer
        self.country_manager = country_manager
    
    def mostrar_grafica_comparativa(self, tabla: pd.DataFrame, pais: str) -> None:
        """Muestra gráfica comparativa de Stock vs Capacidad por bodega"""
        if tabla is None:
            return
        
        logger.info(f"Generando gráfica comparativa para {pais}")
        
        selected_league = st.session_state.get('selected_league', None)
        # Convertir "Todas" a None para mostrar todas las ligas
        if selected_league == "Todas":
            selected_league = None
        
        if selected_league:
            st.markdown(f"#### 📊 Stock por Bodega - {selected_league} - {pais}")
        else:
            st.markdown(f"#### 📊 Comparativa Stock vs Capacidad - {pais}")
        
        df_grafica = self._prepare_chart_data(tabla, pais)
        fig = self._create_chart(df_grafica)
        
        st.plotly_chart(fig, use_container_width=True)
        
        # Solo mostrar alertas y análisis de performance si no hay liga específica
        if not selected_league:
            self._show_alerts(df_grafica, pais)
            self._show_performance_analysis(df_grafica, pais)
        else:
            self._show_performance_analysis(df_grafica, pais)
    
    def _prepare_chart_data(self, tabla: pd.DataFrame, pais: str) -> pd.DataFrame:
        """Prepara los datos para la gráfica"""
        # Obtener liga seleccionada para determinar qué columna usar
        selected_league = st.session_state.get('selected_league', None)
        if selected_league == "Todas":
            selected_league = None
            
        # Buscar columnas en la nueva estructura MultiIndex
        bodega_col = None
        total_headwear_col = None
        capacidad_col = None
        
        for col in tabla.columns:
            if len(col) == 3 and col[2] == 'Bodega':
                bodega_col = col
            elif len(col) == 3 and col[2] == 'CAPACIDAD EN TIENDA':
                capacidad_col = col
            elif len(col) == 3 and col[2] == 'TOTAL HEADWEAR':
                total_headwear_col = col
        
        # Si hay liga específica seleccionada, buscar las columnas de esa liga específica
        if selected_league and total_headwear_col is None:
            planas_col = None
            curvas_col = None
            for col in tabla.columns:
                if len(col) == 3 and col[0] == selected_league and col[1] == 'Planas' and col[2] == 'Stock':
                    planas_col = col
                elif len(col) == 3 and col[0] == selected_league and col[1] == 'Curvas' and col[2] == 'Stock':
                    curvas_col = col
            
            # Si encontramos las columnas específicas de la liga, calcular el total
            if planas_col and curvas_col and bodega_col:
                # Crear un DataFrame temporal para calcular el stock de la liga específica
                df_temp = tabla[[bodega_col, planas_col, curvas_col]].copy()
                df_temp['Stock_Liga'] = df_temp[planas_col] + df_temp[curvas_col]
                
                # Crear DataFrame final con las columnas correctas
                df_grafica = pd.DataFrame({
                    'Bodega': df_temp[bodega_col].tolist(),
                    'Stock': df_temp['Stock_Liga'].tolist(),
                    'Capacidad': [0] * len(df_temp)  # Para liga específica, capacidad es 0
                })
                
                # Aplicar filtros por país para excluir bodegas centrales
                if pais == "Guatemala":
                    df_grafica = df_grafica[df_grafica['Bodega'] != 'CENTRAL NEW ERA'].copy()
                elif pais == "El Salvador":
                    df_grafica = df_grafica[df_grafica['Bodega'] != 'New Era Central'].copy()
                elif pais == "PANAMA":
                    df_grafica = df_grafica[
                        (~df_grafica['Bodega'].isin(['Almacén general', 'Bodega Central Albrook']))
                    ].copy()
                elif pais == "Costa Rica":
                    df_grafica = df_grafica[df_grafica['Bodega'] != 'Bodega Central NEW ERA'].copy()
                
                return df_grafica.sort_values('Stock', ascending=True)
        
        if bodega_col is None or total_headwear_col is None:
            return pd.DataFrame()  # Return empty if can	 find required columns
            
        datos_grafica = tabla[tabla[bodega_col] != 'TOTAL'].copy()
        
        # Excluir bodegas específicas de las gráficas por país
        if pais == "Guatemala":
            datos_grafica = datos_grafica[datos_grafica[bodega_col] != 'CENTRAL NEW ERA'].copy()
        elif pais == "El Salvador":
            datos_grafica = datos_grafica[datos_grafica[bodega_col] != 'New Era Central'].copy()
        elif pais == "PANAMA":
            datos_grafica = datos_grafica[
                (~datos_grafica[bodega_col].isin(['Almacén general', 'Bodega Central Albrook']))
            ].copy()
        elif pais == "Costa Rica":
            datos_grafica = datos_grafica[datos_grafica[bodega_col] != 'Bodega Central NEW ERA'].copy()
        
        # Verificar si existe la columna CAPACIDAD EN TIENDA
        if capacidad_col is not None and capacidad_col in datos_grafica.columns:
            capacidad_data = datos_grafica[capacidad_col].tolist()
        else:
            # Para liga específica, no hay columna de capacidad
            capacidad_data = [0] * len(datos_grafica)
        
        df_grafica = pd.DataFrame({
            'Bodega': datos_grafica[bodega_col].tolist(),
            'Stock': datos_grafica[total_headwear_col].tolist(),
            'Capacidad': capacidad_data
        })
        
        return df_grafica.sort_values('Stock', ascending=True)
    
    def _create_chart(self, df_grafica: pd.DataFrame) -> go.Figure:
        """Crea una gráfica ultra minimalista y limpia"""
        fig = go.Figure()
        
        selected_league = st.session_state.get('selected_league', None)
        # Convertir "Todas" a None para mostrar todas las ligas
        if selected_league == "Todas":
            selected_league = None
        
        # Barras de Capacidad - diseño minimalista (solo si no hay liga específica)
        if not selected_league and any(cap > 0 for cap in df_grafica['Capacidad']):
            fig.add_trace(go.Bar(
                y=df_grafica['Bodega'],
                x=df_grafica['Capacidad'],
                name='Capacidad Máxima',
                orientation='h',
                marker=dict(
                    color='rgba(0, 0, 0, 0.8)',
                    line=dict(width=0)  # Sin bordes para look minimalista
                ),
                text=[f"{x:,}" if x > 0 else "N/A" for x in df_grafica['Capacidad']],
                textposition='outside',
                textfont=dict(size=12, color='rgba(0, 0, 0, 0.8)'),
                hovertemplate='<b>%{y}</b><br>Capacidad: %{x:,}<extra></extra>',
                hoverlabel=dict(
                    bgcolor="white",
                    bordercolor="rgba(0, 0, 0, 0.1)",
                    font=dict(color="#000", size=12)
                )
            ))
        
        # Barras de Stock - diseño minimalista
        fig.add_trace(go.Bar(
            y=df_grafica['Bodega'],
            x=df_grafica['Stock'],
            name='Stock Actual',
            orientation='h',
            marker=dict(
                color='rgba(107, 114, 128, 0.8)',
                line=dict(width=0)  # Sin bordes
            ),
            text=[f"{x:,}" for x in df_grafica['Stock']],
            textposition='outside',
            textfont=dict(size=12, color='rgba(107, 114, 128, 0.8)'),
            hovertemplate='<b>%{y}</b><br>Stock: %{x:,}<extra></extra>',
            hoverlabel=dict(
                bgcolor="white",
                bordercolor="rgba(107, 114, 128, 0.1)",
                font=dict(color="#6b7280", size=12)
            )
        ))
        
        # Anotaciones minimalistas
        self._add_overstock_annotations(fig, df_grafica)
        
        # Layout ultra minimalista
        fig.update_layout(
            # Sin títulos de ejes para máximo minimalismo
            xaxis_title="",
            yaxis_title="",
            barmode='group',
            height=max(600, len(df_grafica) * 45),
            margin=dict(l=20, r=20, t=20, b=20),
            
            # Leyenda minimalista
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                xanchor="center",
                x=0.5,
                font=dict(size=12, color='#6b7280'),
                bgcolor='rgba(0,0,0,0)',  # Transparente
                borderwidth=0  # Sin borde
            ),
            
            # Ejes minimalistas
            xaxis=dict(
                showgrid=True,
                gridwidth=0.5,
                gridcolor='rgba(0, 0, 0, 0.05)',  # Grillas casi invisibles
                tickformat=',',
                tickfont=dict(size=11, color='#9ca3af'),
                showline=False,  # Sin líneas de ejes
                zeroline=False,  # Sin línea de cero
                ticks=""  # Sin marcas de tick
            ),
            yaxis=dict(
                showgrid=False,
                tickfont=dict(size=12, color='#374151'),
                showline=False,
                zeroline=False,
                ticks=""
            ),
            
            # Fondo completamente limpio
            plot_bgcolor='white',
            paper_bgcolor='white',
            
            # Sin marcos ni decoraciones
            showlegend=True,
            hovermode='closest'
        )
        
        return fig
    
    def _add_overstock_annotations(self, fig: go.Figure, df_grafica: pd.DataFrame) -> None:
        """Agrega anotaciones de sobrestock y faltante de stock"""
        if not any(cap > 0 for cap in df_grafica['Capacidad']):
            return
        
        # Calcular distancia de referencia
        distancia_referencia = self._calculate_reference_distance(df_grafica)
        
        for _, row in df_grafica.iterrows():
            if row['Capacidad'] > 0:
                max_value_bodega = max(row['Stock'], row['Capacidad'])
                annotation_position = max_value_bodega + (distancia_referencia * 0.3)
                
                # SOBRESTOCK (Stock > Capacidad) - Color amarillo dorado
                if row['Stock'] > row['Capacidad']:
                    fig.add_annotation(
                        x=annotation_position,
                        y=row['Bodega'],
                        text="SOBRESTOCK",
                        showarrow=False,
                        font=dict(size=9, color='#f59e0b', family='Arial Black'),  # Amarillo dorado
                        bgcolor='rgba(255,255,255,0.9)',
                        bordercolor='#f59e0b',  # Amarillo dorado
                        borderwidth=1
                    )
                
                # FALTANTE DE STOCK (Stock < Capacidad) - Color rojo
                elif row['Stock'] < row['Capacidad']:
                    fig.add_annotation(
                        x=annotation_position,
                        y=row['Bodega'],
                        text="FALTANTE DE STOCK",
                        showarrow=False,
                        font=dict(size=9, color='#ef4444', family='Arial Black'),  # Rojo
                        bgcolor='rgba(255,255,255,0.9)',
                        bordercolor='#ef4444',  # Rojo
                        borderwidth=1
                    )
    
    def _calculate_reference_distance(self, df_grafica: pd.DataFrame) -> float:
        """Calcula la distancia de referencia para anotaciones"""
        referencia_bodega = "NE Cayala"
        
        for _, row in df_grafica.iterrows():
            if row['Bodega'] == referencia_bodega and row['Capacidad'] > 0:
                max_value_cayala = max(row['Stock'], row['Capacidad'])
                return (max_value_cayala * 0.40) + (max_value_cayala * 0.10)
        
        return 1000  # Valor por defecto
    
    def _show_alerts(self, df_grafica: pd.DataFrame, pais: str) -> None:
        """Muestra alertas de stock con diseño compacto"""
        st.markdown("---")
        st.markdown("### 🚨 ALERTAS DE STOCK")
        
        alertas = []
        
        # Buscar todas las bodegas que tienen FALTANTE DE STOCK (Stock < Capacidad)
        for _, row in df_grafica.iterrows():
            bodega = row['Bodega']
            stock_actual = row['Stock']
            capacidad = row['Capacidad']
            
            # Condición para FALTANTE DE STOCK: Stock < Capacidad y Capacidad > 0
            if capacidad > 0 and stock_actual < capacidad:
                faltante = capacidad - stock_actual
                alertas.append({
                    'bodega': bodega,
                    'stock': stock_actual,
                    'capacidad': capacidad,
                    'faltante': faltante
                })
        
        if alertas:
            alertas.sort(key=lambda x: x['faltante'], reverse=True)
            
            # Crear recuadros compactos - hasta 3 por fila
            for i in range(0, len(alertas), 3):
                cols = st.columns(3)
                batch_alertas = alertas[i:i+3]
                
                for j, alerta in enumerate(batch_alertas):
                    with cols[j]:
                        self._create_compact_alert_card(alerta)
        else:
            # Mensaje moderno y amigable cuando no hay alertas usando componentes nativos
            st.markdown("""
            <div style="
                background: linear-gradient(135deg, #ecfdf5 0%, #d1fae5 100%);
                border: 2px solid #10b981;
                border-radius: 20px;
                padding: 30px;
                margin: 20px 0;
                text-align: center;
                box-shadow: 0 8px 25px rgba(16, 185, 129, 0.2);
            ">
            """, unsafe_allow_html=True)
            
            # Icono grande centrado
            st.markdown("""
            <div style="text-align: center; margin-bottom: 20px;">
                <span style="font-size: 4rem;">✅</span>
            </div>
            """, unsafe_allow_html=True)
            
            # Mensaje principal
            st.markdown("""
            <h2 style="
                color: #065f46;
                text-align: center;
                font-size: 1.8rem;
                font-weight: 700;
                margin-bottom: 15px;
            ">
                ¡Excelente gestión de inventario!
            </h2>
            """, unsafe_allow_html=True)
            
            # Mensaje clave
            st.markdown("""
            <h3 style="
                color: #047857;
                text-align: center;
                font-size: 1.3rem;
                font-weight: 600;
                margin-bottom: 10px;
            ">
                Capacidades abastecidas con suficiente stock
            </h3>
            """, unsafe_allow_html=True)
            
            # Descripción
            st.markdown("""
            <p style="
                color: #059669;
                text-align: center;
                font-size: 1rem;
                margin-bottom: 20px;
            ">
                Todas las bodegas mantienen niveles óptimos de inventario
            </p>
            """, unsafe_allow_html=True)
            
            # Badges usando columnas
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.markdown("""
                <div style="display: flex; justify-content: center; gap: 15px; flex-wrap: wrap;">
                    <span style="
                        background: rgba(16, 185, 129, 0.2);
                        color: #065f46;
                        padding: 8px 16px;
                        border-radius: 20px;
                        font-size: 0.9rem;
                        font-weight: 600;
                        border: 1px solid rgba(16, 185, 129, 0.3);
                        display: inline-block;
                        margin: 5px;
                    ">
                        📊 Stock Óptimo
                    </span>
                    <span style="
                        background: rgba(16, 185, 129, 0.2);
                        color: #065f46;
                        padding: 8px 16px;
                        border-radius: 20px;
                        font-size: 0.9rem;
                        font-weight: 600;
                        border: 1px solid rgba(16, 185, 129, 0.3);
                        display: inline-block;
                        margin: 5px;
                    ">
                        🎯 Capacidades Completas
                    </span>
                </div>
                """, unsafe_allow_html=True)
            
            # Cerrar container
            st.markdown("</div>", unsafe_allow_html=True)
    
    def _create_compact_alert_card(self, alerta: dict) -> None:
        """Crea una card moderna usando componentes mixtos"""
        # Determinar severidad y colores
        porcentaje_faltante = (alerta['faltante'] / alerta['capacidad']) * 100
        
        if porcentaje_faltante >= 50:
            color_principal = "#dc2626"
            severidad = "CRÍTICO"
            icono = "🚨"
            bg_color = "#fee2e2"
        elif porcentaje_faltante >= 25:
            color_principal = "#ef4444"
            severidad = "MODERADO"
            icono = "⚠️"
            bg_color = "#fecaca"
        else:
            color_principal = "#f87171"
            severidad = "LEVE"
            icono = "📋"
            bg_color = "#fed7d7"
        
        # Container principal con fondo de color
        st.markdown(f"""
        <div style="
            background: {bg_color};
            border: 2px solid {color_principal};
            border-radius: 20px;
            padding: 20px;
            margin: 15px 0;
            box-shadow: 0 8px 25px rgba(0,0,0,0.15);
            position: relative;
        ">
        """, unsafe_allow_html=True)
        
        # Badge de severidad
        st.markdown(f"""
        <div style="
            position: absolute;
            top: -8px;
            right: 20px;
            background: {color_principal};
            color: white;
            padding: 6px 16px;
            border-radius: 15px;
            font-size: 0.75rem;
            font-weight: 700;
            text-transform: uppercase;
        ">
            {severidad}
        </div>
        """, unsafe_allow_html=True)
        
        # Header con nombre de bodega
        st.markdown(f"""
        <div style="text-align: center; padding: 15px 0;">
            <div style="
                background: white;
                border-radius: 15px;
                padding: 12px 24px;
                display: inline-block;
                box-shadow: 0 4px 15px rgba(0,0,0,0.1);
                border: 1px solid rgba(0,0,0,0.1);
            ">
                <span style="font-size: 1.3rem; margin-right: 8px;">{icono}</span>
                <span style="
                    color: #000000; 
                    font-size: 1.2rem; 
                    font-weight: 700;
                ">{alerta['bodega']}</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Métricas usando columnas de Streamlit
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown(f"""
            <div style="
                background: white;
                border-radius: 15px;
                padding: 20px;
                text-align: center;
                box-shadow: 0 4px 15px rgba(0,0,0,0.1);
                margin: 5px;
            ">
                <div style="color: {color_principal}; font-size: 2rem; font-weight: 800; margin-bottom: 8px;">
                    {alerta['faltante']:,}
                </div>
                <div style="color: #6b7280; font-size: 0.9rem; font-weight: 600; text-transform: uppercase;">
                    Faltante
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div style="
                background: white;
                border-radius: 15px;
                padding: 20px;
                text-align: center;
                box-shadow: 0 4px 15px rgba(0,0,0,0.1);
                margin: 5px;
            ">
                <div style="color: #374151; font-size: 2rem; font-weight: 800; margin-bottom: 8px;">
                    {alerta['stock']:,}
                </div>
                <div style="color: #6b7280; font-size: 0.9rem; font-weight: 600; text-transform: uppercase;">
                    Stock
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div style="
                background: white;
                border-radius: 15px;
                padding: 20px;
                text-align: center;
                box-shadow: 0 4px 15px rgba(0,0,0,0.1);
                margin: 5px;
            ">
                <div style="color: #374151; font-size: 2rem; font-weight: 800; margin-bottom: 8px;">
                    {alerta['capacidad']:,}
                </div>
                <div style="color: #6b7280; font-size: 0.9rem; font-weight: 600; text-transform: uppercase;">
                    Capacidad
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        # Barra de progreso
        capacidad_utilizada = 100 - porcentaje_faltante
        st.markdown(f"""
        <div style="margin: 20px 10px 10px 10px;">
            <div style="
                background: rgba(255,255,255,0.8);
                border-radius: 10px;
                padding: 4px;
                overflow: hidden;
            ">
                <div style="
                    background: {color_principal};
                    height: 10px;
                    border-radius: 6px;
                    width: {capacidad_utilizada:.1f}%;
                    transition: width 0.8s ease;
                "></div>
            </div>
            <div style="
                text-align: center;
                margin-top: 8px;
                font-size: 0.85rem;
                color: #6b7280;
                font-weight: 600;
            ">
                {capacidad_utilizada:.1f}% de capacidad utilizada
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Cerrar container principal
        st.markdown("</div>", unsafe_allow_html=True)
    
    def _show_performance_analysis(self, df_grafica: pd.DataFrame, pais: str) -> None:
        """Muestra análisis de performance con diseño profesional"""
        selected_league = st.session_state.get('selected_league', None)
        # Convertir "Todas" a None para mostrar todas las ligas
        if selected_league == "Todas":
            selected_league = None
        
        # Header profesional con colores según el país
        st.markdown("---")
        
        # Determinar colores según el país
        if pais == "PANAMA":
            # Rojo de la bandera de Panamá
            background_gradient = "linear-gradient(135deg, #dc2626 0%, #ef4444 100%)"
            box_shadow = "0 10px 30px rgba(220, 38, 38, 0.3)"
        elif pais == "Honduras":
            # Azul (anteriormente de Guatemala)
            background_gradient = "linear-gradient(135deg, #1e40af 0%, #3b82f6 100%)"
            box_shadow = "0 10px 30px rgba(30, 64, 175, 0.3)"
        elif pais == "El Salvador":
            # Azul oscuro de la bandera de El Salvador
            background_gradient = "linear-gradient(135deg, #1e3a8a 0%, #1d4ed8 100%)"
            box_shadow = "0 10px 30px rgba(30, 58, 138, 0.3)"
        elif pais == "Costa Rica":
            # Verde reciclaje de Costa Rica
            background_gradient = "linear-gradient(135deg, #16a34a 0%, #22c55e 100%)"
            box_shadow = "0 10px 30px rgba(22, 163, 74, 0.3)"
        elif pais == "Guatemala":
            # Celeste (anteriormente de Honduras)
            background_gradient = "linear-gradient(135deg, #0ea5e9 0%, #38bdf8 100%)"
            box_shadow = "0 10px 30px rgba(14, 165, 233, 0.3)"
        else:
            # Azul medio por defecto
            background_gradient = "linear-gradient(135deg, #1e40af 0%, #3b82f6 100%)"
            box_shadow = "0 10px 30px rgba(30, 64, 175, 0.3)"
        
        st.markdown(f"""
        <div style="
            background: {background_gradient};
            padding: 2rem 3rem;
            border-radius: 15px;
            margin-bottom: 2rem;
            box-shadow: {box_shadow};
            animation: slideInFromRight 0.8s ease-out;
            transition: all 0.3s ease;
        " onmouseover="this.style.transform=	ranslateY(-2px)" onmouseout="this.style.transform=	ranslateY(0)">
            <div style="display: flex; align-items: center;">
                <div>
                    <h2 style="color: white; margin: 0; font-size: 1.8rem; font-weight: 700;">
                        Análisis de Performance{f" - {selected_league}" if selected_league else ""}
                    </h2>
                    <p style="color: #d1d5db; margin: 0.5rem 0 0 0; font-size: 1.1rem;">
                        Indicadores clave de rendimiento por bodega{f" para {selected_league}" if selected_league else ""}
                    </p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        selected_league = st.session_state.get('selected_league', None)
        # Convertir "Todas" a None para mostrar todas las ligas
        if selected_league == "Todas":
            selected_league = None
        
        if selected_league:
            # Para liga específica, usar todos los datos sin filtrar por capacidad
            df_analisis = df_grafica.copy()
            # No calcular porcentaje de cumplimiento para liga específica
        else:
            # Para vista completa, usar todos los datos del gráfico (ya excluye bodegas centrales)
            # Solo calcular porcentaje de cumplimiento para bodegas con capacidad > 0
            df_analisis = df_grafica.copy()
            
            # Crear columna de porcentaje solo para bodegas con capacidad > 0
            df_analisis['Porcentaje_Cumplimiento'] = df_analisis.apply(
                lambda row: (row['Stock'] / row['Capacidad']) * 100 if row['Capacidad'] > 0 else 0, 
                axis=1
            )
        
        # Calcular métricas siempre que haya datos
        if len(df_analisis) > 0:
            max_stock = df_analisis.loc[df_analisis['Stock'].idxmax()]
            min_stock = df_analisis.loc[df_analisis['Stock'].idxmin()]
            promedio_stock = df_analisis['Stock'].mean()
        else:
            return
        
        # Métricas con diseño profesional igual que Métricas Generales
        cols = st.columns(3)
        
        if selected_league:
            metricas_performance = [
                (max_stock['Bodega'], f"{max_stock['Stock']:,}", f"Mayor Stock {selected_league}", "🏆", "#10b981"),
                (min_stock['Bodega'], f"{min_stock['Stock']:,}", f"Menor Stock {selected_league}", "📊", "#ef4444"),
                (f"{promedio_stock:,.0f}", "unidades", f"Promedio de Headwear {selected_league}", "📈", "#6b7280")
            ]
        else:
            metricas_performance = [
                (max_stock['Bodega'], f"{max_stock['Stock']:,}", "Mayor Stock", "🏆", "#10b981"),
                (min_stock['Bodega'], f"{min_stock['Stock']:,}", "Menor Stock", "📊", "#ef4444"),
                (f"{promedio_stock:,.0f}", "unidades", "Promedio de Headwear", "📈", "#6b7280")
            ]
        
        for i, (valor_principal, valor_secundario, nombre, emoji, color) in enumerate(metricas_performance):
            with cols[i]:
                st.markdown(f"""
                <div class="metric-card" style="border-left: 4px solid {color};">
                    <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 0.5rem;">
                        <span style="font-size: 1.5rem; color: #000000;">{emoji}</span>
                        <span style="color: {color}; font-weight: 600; font-size: 0.9rem;">{nombre.upper()}</span>
                    </div>
                    <div style="font-size: 1.4rem; font-weight: 700; color: #374151; margin-bottom: 0.25rem;">
                        {valor_principal}
                    </div>
                    <div style="color: #6b7280; font-size: 1.2rem; font-weight: 600;">
                        {valor_secundario}
                    </div>
                </div>
                """, unsafe_allow_html=True)

def mostrar_distribucion_ligas_por_bodega(tabla: pd.DataFrame, pais: str) -> None:
    """Muestra la distribución porcentual de ligas por bodega en gráfica de barras verticales"""
    if tabla is None or len(tabla) == 0:
        return
    
    logger.info(f"Generando distribución de ligas por bodega para {pais}")
    
    # DEBUG: Mostrar información de la tabla
    logger.info(f"Columnas disponibles: {list(tabla.columns)}")
    logger.info(f"Índices (bodegas): {list(tabla.index)}")
    
    # Buscar la columna que contiene los nombres de las bodegas
    nombre_columna_bodega = None
    if isinstance(tabla.columns, pd.MultiIndex):
        # Para MultiIndex, buscar columna con nombres de bodegas
        for col in tabla.columns:
            if len(col) >= 3 and col[2] == 'Bodega':
                nombre_columna_bodega = col
                break
        # Si no encuentra "Bodega" en nivel 2, buscar en otros niveles
        if nombre_columna_bodega is None:
            for col in tabla.columns:
                if 'INFO' in str(col) and 'Bodega' in str(col):
                    nombre_columna_bodega = col
                    break
    
    logger.info(f"Columna de nombres de bodega encontrada: {nombre_columna_bodega}")
    
    # Filtrar solo las bodegas (excluir fila TOTAL)
    df_bodegas = tabla[tabla.index != 'TOTAL'].copy()
    
    # Excluir bodegas centrales de las distribuciones
    if pais == "Costa Rica" and "Bodega Central NEW ERA" in df_bodegas.index:
        df_bodegas = df_bodegas.drop("Bodega Central NEW ERA")
    elif pais == "PANAMA":
        bodegas_panama_excluir = ['Almacén general', 'Bodega Central Albrook']
        for bodega in bodegas_panama_excluir:
            if bodega in df_bodegas.index:
                df_bodegas = df_bodegas.drop(bodega)
    
    if len(df_bodegas) == 0:
        st.warning("No se encontraron bodegas en los datos")
        return
    
    # Obtener nombres reales de bodegas
    if nombre_columna_bodega is not None:
        nombres_reales_bodegas = df_bodegas[nombre_columna_bodega].tolist()
        logger.info(f"Nombres reales de bodegas: {nombres_reales_bodegas}")
    else:
        # Fallback: usar índices si no encuentra la columna de nombres
        nombres_reales_bodegas = list(df_bodegas.index)
        logger.info(f"Usando índices como nombres de bodegas: {nombres_reales_bodegas}")
        st.warning("No se pudo encontrar la columna de nombres de bodegas, usando índices")
        st.write("Estructura de columnas:")
        st.write(tabla.columns.tolist()[:5])
    
    # Definir las ligas a analizar
    ligas = ['MLB', 'NBA', 'NFL', 'MOTORSPORT', 'ENTERTAINMENT']
    
    # Verificar si la tabla tiene columnas MultiIndex
    es_multiindex = isinstance(df_bodegas.columns, pd.MultiIndex)
    logger.info(f"Es MultiIndex: {es_multiindex}")
    
    # DEBUG: Verificar qué columnas de ligas existen (para MultiIndex)
    columnas_encontradas = []
    if es_multiindex:
        for liga in ligas:
            # Buscar columnas con estructura (Liga, Tipo, 'Stock')
            col_planas = (liga, 'Planas', 'Stock')
            col_curvas = (liga, 'Curvas', 'Stock')
            
            if col_planas in df_bodegas.columns:
                columnas_encontradas.append(col_planas)
            if col_curvas in df_bodegas.columns:
                columnas_encontradas.append(col_curvas)
    else:
        for liga in ligas:
            col_planas = f"{liga} - Planas"
            col_curvas = f"{liga} - Curvas"
            if col_planas in df_bodegas.columns:
                columnas_encontradas.append(col_planas)
            if col_curvas in df_bodegas.columns:
                columnas_encontradas.append(col_curvas)
    
    logger.info(f"Columnas de ligas encontradas: {columnas_encontradas}")
    
    if not columnas_encontradas:
        st.warning("No se encontraron columnas de stock por liga")
        # Mostrar algunas columnas de ejemplo
        st.write("Columnas disponibles en la tabla:")
        st.write(list(tabla.columns)[:10])  # Mostrar primeras 10 columnas
        return
    
    # Calcular stock de planas + curvas por liga para cada bodega
    distribucion_data = []
    
    for i, bodega_idx in enumerate(df_bodegas.index):
        # Usar nombre real de bodega si está disponible
        nombre_bodega = nombres_reales_bodegas[i] if i < len(nombres_reales_bodegas) else bodega_idx
        bodega_data = {'Bodega': nombre_bodega}
        total_stock_bodega = 0
        
        # Calcular stock por liga (planas + curvas)
        for liga in ligas:
            if es_multiindex:
                # Para columnas MultiIndex: (Liga, Tipo, 'Stock')
                col_planas = (liga, 'Planas', 'Stock')
                col_curvas = (liga, 'Curvas', 'Stock')
                
                stock_planas = df_bodegas.loc[bodega_idx, col_planas] if col_planas in df_bodegas.columns else 0
                stock_curvas = df_bodegas.loc[bodega_idx, col_curvas] if col_curvas in df_bodegas.columns else 0
            else:
                # Para columnas simples: "LIGA - Tipo"
                col_planas = f"{liga} - Planas"
                col_curvas = f"{liga} - Curvas"
                
                stock_planas = df_bodegas.loc[bodega_idx, col_planas] if col_planas in df_bodegas.columns else 0
                stock_curvas = df_bodegas.loc[bodega_idx, col_curvas] if col_curvas in df_bodegas.columns else 0
            
            # Asegurar que son números
            try:
                stock_planas = float(stock_planas) if stock_planas != 0 else 0
                stock_curvas = float(stock_curvas) if stock_curvas != 0 else 0
            except:
                stock_planas = 0
                stock_curvas = 0
            
            stock_liga = stock_planas + stock_curvas
            bodega_data[liga] = stock_liga
            total_stock_bodega += stock_liga
            
            # DEBUG: Mostrar stock por liga y bodega
            if stock_liga > 0:
                logger.info(f"Bodega {nombre_bodega}, Liga {liga}: Planas={stock_planas}, Curvas={stock_curvas}, Total={stock_liga}")
        
        # DEBUG: Mostrar totales por bodega
        logger.info(f"Bodega {nombre_bodega}: Total stock = {total_stock_bodega}")
        
        # Calcular porcentajes
        if total_stock_bodega > 0:
            for liga in ligas:
                bodega_data[f"{liga}_porcentaje"] = (bodega_data[liga] / total_stock_bodega) * 100
        else:
            for liga in ligas:
                bodega_data[f"{liga}_porcentaje"] = 0
        
        bodega_data['Total'] = total_stock_bodega
        distribucion_data.append(bodega_data)
    
    # Convertir a DataFrame
    df_distribucion = pd.DataFrame(distribucion_data)
    
    if len(df_distribucion) == 0:
        return
    
    # Filtrar CENTRAL NEW ERA, New Era Central, Bodega Central NEW ERA y TOTAL del gráfico
    bodegas_excluir = ['CENTRAL NEW ERA', 'New Era Central', 'TOTAL']
    
    # Para Costa Rica, también excluir "Bodega Central NEW ERA"
    if pais == "Costa Rica":
        bodegas_excluir.append('Bodega Central NEW ERA')
    # Para PANAMA, excluir bodegas centrales
    elif pais == "PANAMA":
        bodegas_excluir.extend(['Almacén general', 'Bodega Central Albrook'])
    
    df_distribucion = df_distribucion[
        ~df_distribucion['Bodega'].isin(bodegas_excluir)
    ].copy()
    
    if len(df_distribucion) == 0:
        return
    
    # DEBUG: Verificar contenido del DataFrame
    logger.info(f"DataFrame de distribución creado con {len(df_distribucion)} filas (sin CENTRAL NEW ERA y TOTAL)")
    logger.info(f"Bodegas encontradas: {df_distribucion['Bodega'].tolist()}")
    
    # Definir nombres dinámicos según el país
    if pais == "Guatemala":
        nombre_tiendas_secundarias = "Tiendas Departamentales"
        nombre_tiendas_principales = "Tiendas de Ciudad"
    elif pais == "Costa Rica":
        nombre_tiendas_secundarias = "Tiendas Departamentales"
        nombre_tiendas_principales = "Tiendas Franquicia"
    elif pais == "Honduras":
        nombre_tiendas_secundarias = "Tiendas Departamentales"
        nombre_tiendas_principales = "Tiendas Franquicia"
    else:
        nombre_tiendas_secundarias = "Tiendas Franquicia"
        nombre_tiendas_principales = "Tiendas de Ciudad"
    
    # Definir tiendas de ciudad, outlets y secundarias
    bodegas_principales = [
        'NE Oakland', 'NE Cayala', 'NE Miraflores', 'NE Portales', 'NE Concepcion', 
        'NE Naranjo', 'NE Vistares', 'NE Peri Roosvelt', 'NE Plaza Videre'
    ]
    
    bodegas_outlets = [
        'NE Metronorte', 'NE Metrocentro Outlet', 'NE Outlet Santa clara'
    ]
    
    # Para El Salvador, separar NE METROCENTRO LOURDES como tienda outlet especial
    if pais == "El Salvador":
        bodega_outlet_especial = ['NE METROCENTRO LOURDES']
        df_outlet_especial = df_distribucion[df_distribucion['Bodega'].isin(bodega_outlet_especial)].copy()
        
        # Separar los datos en cuatro grupos para El Salvador
        df_principales = df_distribucion[df_distribucion['Bodega'].isin(bodegas_principales)].copy()
        df_outlets = df_distribucion[df_distribucion['Bodega'].isin(bodegas_outlets)].copy()
        df_secundarias = df_distribucion[
            ~df_distribucion['Bodega'].isin(bodegas_principales + bodegas_outlets + bodega_outlet_especial)
        ].copy()
    else:
        # Separar los datos en tres grupos para otros países
        df_principales = df_distribucion[df_distribucion['Bodega'].isin(bodegas_principales)].copy()
        df_outlets = df_distribucion[df_distribucion['Bodega'].isin(bodegas_outlets)].copy()
        df_secundarias = df_distribucion[
            ~df_distribucion['Bodega'].isin(bodegas_principales + bodegas_outlets)
        ].copy()
        
        # Excluir bodegas centrales específicamente de todas las categorías
        if pais == "Costa Rica":
            df_secundarias = df_secundarias[df_secundarias['Bodega'] != 'Bodega Central NEW ERA'].copy()
        elif pais == "PANAMA":
            df_secundarias = df_secundarias[
                ~df_secundarias['Bodega'].isin(['Almacén general', 'Bodega Central Albrook'])
            ].copy()
        
        df_outlet_especial = pd.DataFrame()  # DataFrame vacío para otros países
    
    # DEBUG: Verificar separación de datos
    logger.info(f"Bodegas principales encontradas: {df_principales['Bodega'].tolist() if len(df_principales) > 0 else 'NINGUNA'}")
    logger.info(f"Total tiendas de ciudad: {len(df_principales)}")
    logger.info(f"Bodegas outlets encontradas: {df_outlets['Bodega'].tolist() if len(df_outlets) > 0 else 'NINGUNA'}")
    logger.info(f"Total bodegas outlets: {len(df_outlets)}")
    logger.info(f"Bodegas secundarias encontradas: {df_secundarias['Bodega'].tolist() if len(df_secundarias) > 0 else 'NINGUNA'}")
    logger.info(f"Total tiendas departamentales: {len(df_secundarias)}")
    
    # Crear header de sección
    professional_design.create_section_header(
        f"Distribución de Stock por Bodega - {pais}",
        "Porcentaje de stock (planas + curvas) por liga en cada bodega",
        "📊"
    )
    
    # Función auxiliar para crear leyenda de ligas
    def crear_leyenda_ligas():
        st.markdown("""
        <div style="display: flex; justify-content: center; align-items: center; margin: 10px 0; padding: 15px; background: #f8fafc; border-radius: 8px; border: 1px solid #e2e8f0;">
            <div style="display: flex; flex-wrap: wrap; gap: 20px; justify-content: center; align-items: center;">
                <div style="display: flex; align-items: center; gap: 8px;">
                    <div style="width: 16px; height: 16px; background: #1f77b4; border-radius: 3px;"></div>
                    <span style="font-size: 12px; font-weight: 600; color: #374151;">MLB</span>
                </div>
                <div style="display: flex; align-items: center; gap: 8px;">
                    <div style="width: 16px; height: 16px; background: #ff7f0e; border-radius: 3px;"></div>
                    <span style="font-size: 12px; font-weight: 600; color: #374151;">NBA</span>
                </div>
                <div style="display: flex; align-items: center; gap: 8px;">
                    <div style="width: 16px; height: 16px; background: #2ca02c; border-radius: 3px;"></div>
                    <span style="font-size: 12px; font-weight: 600; color: #374151;">NFL</span>
                </div>
                <div style="display: flex; align-items: center; gap: 8px;">
                    <div style="width: 16px; height: 16px; background: #d62728; border-radius: 3px;"></div>
                    <span style="font-size: 12px; font-weight: 600; color: #374151;">MOTORSPORT</span>
                </div>
                <div style="display: flex; align-items: center; gap: 8px;">
                    <div style="width: 16px; height: 16px; background: #9467bd; border-radius: 3px;"></div>
                    <span style="font-size: 12px; font-weight: 600; color: #374151;">ENTERTAINMENT</span>
                </div>
                <div style="width: 2px; height: 20px; background: #d1d5db; margin: 0 10px;"></div>
                <div style="display: flex; align-items: center; gap: 8px;">
                    <div style="width: 20px; height: 3px; background: #374151; border-radius: 1px;"></div>
                    <span style="font-size: 12px; font-weight: 600; color: #374151;">Línea Sólida (Ventas)</span>
                </div>
                <div style="display: flex; align-items: center; gap: 8px;">
                    <div style="width: 20px; height: 2px; background: transparent; border-top: 2px dashed #374151;"></div>
                    <span style="font-size: 12px; font-weight: 600; color: #374151;">Línea Punteada (Stock)</span>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    # Función auxiliar para crear gráfico
    def crear_grafico_distribucion(df_data, titulo_grafico, ligas):
        if len(df_data) == 0:
            return None
            
        fig = go.Figure()
        
        # Colores para cada liga
        colores_ligas = {
            'MLB': '#1f77b4',      # Azul
            'NBA': '#ff7f0e',      # Naranja
            'NFL': '#2ca02c',      # Verde
            'MOTORSPORT': '#d62728', # Rojo
            'ENTERTAINMENT': '#9467bd' # Púrpura
        }
        
        # Obtener nombres de bodegas para el eje X
        nombres_bodegas = df_data['Bodega'].tolist()
        
        # Agregar barras para cada liga
        for liga in ligas:
            fig.add_trace(go.Bar(
                name=liga,
                x=nombres_bodegas,
                y=df_data[f'{liga}_porcentaje'],
                marker_color=colores_ligas[liga],
                text=[f'{val:.1f}%' for val in df_data[f'{liga}_porcentaje']],
                textposition='outside',
                textfont=dict(
                    size=16,
                    color='black',
                    family='Inter, sans-serif',
                    weight='bold'
                )
            ))
        
        # Configurar layout
        fig.update_layout(
            title=titulo_grafico,
            xaxis_title='Bodegas/Tiendas',
            yaxis_title='Porcentaje (%)',
            barmode='group',
            height=600,
            showlegend=False,
            xaxis=dict(
                categoryorder='array',
                categoryarray=nombres_bodegas
            ),
            margin=dict(l=60, r=60, t=100, b=80)
        )
        
        # Configurar ejes
        fig.update_xaxes(
            tickangle=45,
            tickmode='array',
            tickvals=list(range(len(nombres_bodegas))),
            ticktext=nombres_bodegas
        )
        fig.update_yaxes(range=[0, 100])
        
        # Agregar líneas de cuadrícula
        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='rgba(0,0,0,0.1)')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='rgba(0,0,0,0.1)')
        
        return fig
    
    # Función auxiliar para crear tabla resumen
    def crear_tabla_resumen(df_data, titulo_tabla, ligas):
        if len(df_data) == 0:
            return
        
        st.markdown(f"#### {titulo_tabla}")
        
        # Crear tabla para mostrar con índice de nombres de bodegas
        tabla_resumen = df_data[['Bodega'] + [f'{liga}_porcentaje' for liga in ligas] + ['Total']].copy()
        
        # Usar nombres de bodegas como índice para mejor visualización
        tabla_resumen = tabla_resumen.set_index('Bodega')
        
        # Renombrar columnas para mejor presentación
        columnas_rename = {'Total': 'Total Stock'}
        for liga in ligas:
            columnas_rename[f'{liga}_porcentaje'] = f'{liga}'
        
        tabla_resumen = tabla_resumen.rename(columns=columnas_rename)
        
        # Formatear porcentajes
        for liga in ligas:
            tabla_resumen[liga] = tabla_resumen[liga].apply(lambda x: f'{x:.1f}%')
        
        # Formatear total con comas
        tabla_resumen['Total Stock'] = tabla_resumen['Total Stock'].apply(lambda x: f'{x:,}')
        
        # Crear HTML personalizado con el mismo estilo que la tabla consolidada
        def crear_html_tabla_resumen(df):
            html = '<table style="border-collapse: collapse; text-align: center; font-size: 11px; width: 100%;">'
            
            # Header con estilo profesional
            html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
            html += '<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; min-width: 120px;">Bodega</td>'
            
            for liga in ligas:
                html += f'<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; min-width: 80px;">{liga}</td>'
            
            html += '<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; min-width: 100px;">Total Stock</td>'
            html += '</tr>'
            
            # Filas de datos
            for idx, row in df.iterrows():
                html += '<tr style="background-color: #f9f9f9;" onmouseover="this.style.backgroundColor=\'#e8f4f8\'" onmouseout="this.style.backgroundColor=\'#f9f9f9\'">'
                html += f'<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; text-align: left; font-weight: 600;">{idx}</td>'
                
                for liga in ligas:
                    valor = row[liga]
                    html += f'<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; text-align: center;">{valor}</td>'
                
                total_valor = row['Total Stock']
                html += f'<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; text-align: center; font-weight: 600; background-color: #f0f0f0;">{total_valor}</td>'
                html += '</tr>'
            
            html += '</table>'
            return html
        
        # Aplicar estilos CSS
        st.markdown("""
        <style>
            .tabla-resumen-container {
                overflow-x: auto;
                width: 100%;
                border: 1px solid #ddd;
                background: white;
                border-radius: 8px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                margin: 10px 0;
            }
        </style>
        """, unsafe_allow_html=True)
        
        # Mostrar tabla con estilo profesional
        st.markdown('<div class="tabla-resumen-container">', unsafe_allow_html=True)
        tabla_html = crear_html_tabla_resumen(tabla_resumen)
        st.markdown(tabla_html, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # Crear y mostrar gráfico de tiendas principales con su tabla
    if len(df_principales) > 0:
        st.markdown(f"#### 🏪 {nombre_tiendas_principales}")
        fig_principales = crear_grafico_distribucion(
            df_principales, 
            f'Distribución por Ligas - {nombre_tiendas_principales} ({pais})', 
            ligas
        )
        if fig_principales:
            st.plotly_chart(fig_principales, use_container_width=True)
            
            # Mostrar leyenda de ligas justo después del gráfico
            crear_leyenda_ligas()
        
        # Mostrar tabla de tiendas principales después de la leyenda
        crear_tabla_resumen(df_principales, f"📋 Resumen - {nombre_tiendas_principales}", ligas)
    
    # Crear y mostrar gráfico de outlets con su tabla
    if len(df_outlets) > 0:
        st.markdown("#### 🛒 Outlets")
        fig_outlets = crear_grafico_distribucion(
            df_outlets, 
            f'Distribución por Ligas - Outlets ({pais})', 
            ligas
        )
        if fig_outlets:
            st.plotly_chart(fig_outlets, use_container_width=True)
            
            # Mostrar leyenda de ligas justo después del gráfico
            crear_leyenda_ligas()
        
        # Mostrar tabla de outlets después de la leyenda
        crear_tabla_resumen(df_outlets, "📋 Resumen - Outlets", ligas)
    
    # Crear y mostrar gráfico de tiendas departamentales con su tabla
    if len(df_secundarias) > 0:
        st.markdown(f"#### 🏬 {nombre_tiendas_secundarias}")
        fig_secundarias = crear_grafico_distribucion(
            df_secundarias, 
            f'Distribución por Ligas - {nombre_tiendas_secundarias} ({pais})', 
            ligas
        )
        if fig_secundarias:
            st.plotly_chart(fig_secundarias, use_container_width=True)
            
            # Mostrar leyenda de ligas justo después del gráfico
            crear_leyenda_ligas()
        
        # Mostrar tabla de tiendas departamentales después de la leyenda
        crear_tabla_resumen(df_secundarias, f"📋 Resumen - {nombre_tiendas_secundarias}", ligas)
    
    # Crear y mostrar gráfico de tienda outlet especial (solo para El Salvador)
    if pais == "El Salvador" and len(df_outlet_especial) > 0:
        st.markdown("#### 🏪 Tienda Outlet")
        fig_outlet_especial = crear_grafico_distribucion(
            df_outlet_especial, 
            f'Distribución por Ligas - Tienda Outlet ({pais})', 
            ligas
        )
        if fig_outlet_especial:
            st.plotly_chart(fig_outlet_especial, use_container_width=True)
            
            # Mostrar leyenda de ligas justo después del gráfico
            crear_leyenda_ligas()
        
        # Mostrar tabla de tienda outlet después de la leyenda
        crear_tabla_resumen(df_outlet_especial, "📋 Resumen - Tienda Outlet", ligas)
    
    # ==================== NUEVA SECCIÓN: DISTRIBUCIÓN DE VENTAS POR BODEGA ====================
    
    # Función auxiliar para crear gráfico de distribución de ventas
    def crear_grafico_distribucion_ventas(df_data, titulo_grafico, ligas):
        if len(df_data) == 0:
            return None
            
        fig = go.Figure()
        
        # Colores para cada liga
        colores_ligas = {
            'MLB': '#1f77b4',      # Azul
            'NBA': '#ff7f0e',      # Naranja
            'NFL': '#2ca02c',      # Verde
            'MOTORSPORT': '#d62728', # Rojo
            'ENTERTAINMENT': '#9467bd' # Púrpura
        }
        
        # Obtener nombres de bodegas para el eje X
        nombres_bodegas = df_data['Bodega'].tolist()
        
        # Agregar barras para cada liga
        for liga in ligas:
            fig.add_trace(go.Bar(
                name=liga,
                x=nombres_bodegas,
                y=df_data[f'{liga}_porcentaje_ventas'],
                marker_color=colores_ligas[liga],
                text=[f'{val:.1f}%' for val in df_data[f'{liga}_porcentaje_ventas']],
                textposition='outside',
                textfont=dict(
                    size=16,
                    color='black',
                    family='Inter, sans-serif',
                    weight='bold'
                )
            ))
        
        # Configurar layout
        fig.update_layout(
            title=titulo_grafico,
            xaxis_title='Bodegas/Tiendas',
            yaxis_title='Porcentaje (%)',
            barmode='group',
            height=600,
            showlegend=False,
            xaxis=dict(
                categoryorder='array',
                categoryarray=nombres_bodegas
            ),
            margin=dict(l=60, r=60, t=100, b=80)
        )
        
        # Configurar ejes
        fig.update_xaxes(
            tickangle=45,
            tickmode='array',
            tickvals=list(range(len(nombres_bodegas))),
            ticktext=nombres_bodegas
        )
        fig.update_yaxes(range=[0, 100])
        
        # Agregar líneas de cuadrícula
        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='rgba(0,0,0,0.1)')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='rgba(0,0,0,0.1)')
        
        return fig
    
    # Función auxiliar para crear tabla resumen de ventas
    def crear_tabla_resumen_ventas(df_data, titulo_tabla, ligas):
        if len(df_data) == 0:
            return
        
        st.markdown(f"#### {titulo_tabla}")
        
        # Crear tabla para mostrar con índice de nombres de bodegas
        tabla_resumen = df_data[['Bodega'] + [f'{liga}_porcentaje_ventas' for liga in ligas] + ['Total_Ventas']].copy()
        
        # Usar nombres de bodegas como índice para mejor visualización
        tabla_resumen = tabla_resumen.set_index('Bodega')
        
        # Renombrar columnas para mejor presentación
        columnas_rename = {'Total_Ventas': 'Total Ventas (USD)'}
        for liga in ligas:
            columnas_rename[f'{liga}_porcentaje_ventas'] = f'{liga}'
        
        tabla_resumen = tabla_resumen.rename(columns=columnas_rename)
        
        # Formatear porcentajes
        for liga in ligas:
            tabla_resumen[liga] = tabla_resumen[liga].apply(lambda x: f'{x:.1f}%')
        
        # Formatear total con comas y símbolo de dólar
        tabla_resumen['Total Ventas (USD)'] = tabla_resumen['Total Ventas (USD)'].apply(lambda x: f'${x:,.2f}')
        
        # Crear HTML personalizado con el mismo estilo que la tabla consolidada
        def crear_html_tabla_ventas(df):
            html = '<table style="border-collapse: collapse; text-align: center; font-size: 11px; width: 100%;">'
            
            # Header con estilo profesional (color diferente para ventas)
            html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
            html += '<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; min-width: 120px;">Bodega</td>'
            
            for liga in ligas:
                html += f'<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; min-width: 80px;">{liga}</td>'
            
            html += '<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; min-width: 120px;">Total Ventas (USD)</td>'
            html += '</tr>'
            
            # Filas de datos
            for idx, row in df.iterrows():
                html += '<tr style="background-color: #f9f9f9;" onmouseover="this.style.backgroundColor=\'#e8f8f2\'" onmouseout="this.style.backgroundColor=\'#f9f9f9\'">'
                html += f'<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; text-align: left; font-weight: 600;">{idx}</td>'
                
                for liga in ligas:
                    valor = row[liga]
                    html += f'<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; text-align: center;">{valor}</td>'
                
                total_valor = row['Total Ventas (USD)']
                html += f'<td style="border: 1px solid #ddd; padding: 6px; font-size: 11px; text-align: center; font-weight: 600; background-color: #ecfdf5; color: #059669;">{total_valor}</td>'
                html += '</tr>'
            
            html += '</table>'
            return html
        
        # Aplicar estilos CSS
        st.markdown("""
        <style>
            .tabla-ventas-container {
                overflow-x: auto;
                width: 100%;
                border: 1px solid #ddd;
                background: white;
                border-radius: 8px;
                box-shadow: 0 2px 8px rgba(0,0,0,0.1);
                margin: 10px 0;
            }
        </style>
        """, unsafe_allow_html=True)
        
        # Mostrar tabla con estilo profesional
        st.markdown('<div class="tabla-ventas-container">', unsafe_allow_html=True)
        tabla_html = crear_html_tabla_ventas(tabla_resumen)
        st.markdown(tabla_html, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # Verificar si hay datos de ventas disponibles (para Guatemala, El Salvador, Costa Rica y Honduras)
    if pais in ["Guatemala", "El Salvador", "Costa Rica", "Honduras", "PANAMA"] and any('Ventas' in str(col) for col in df_bodegas.columns):
        # Crear header de sección para ventas
        professional_design.create_section_header(
            f"Distribución de Ventas por Bodega - {pais}",
            "Porcentaje de ventas (USD) por liga en cada bodega",
            "💰"
        )
        
        # Procesar datos de distribución de ventas
        distribucion_ventas_data = []
        
        for i, bodega_idx in enumerate(df_bodegas.index):
            # Usar nombre real de bodega si está disponible
            nombre_bodega = nombres_reales_bodegas[i] if i < len(nombres_reales_bodegas) else bodega_idx
            
            # Excluir bodegas centrales de las distribuciones de ventas
            if pais == "Costa Rica" and nombre_bodega == "Bodega Central NEW ERA":
                continue
            elif pais == "PANAMA" and nombre_bodega in ['Almacén general', 'Bodega Central Albrook']:
                continue
            bodega_data_ventas = {'Bodega': nombre_bodega}
            total_ventas_bodega = 0
            
            # Calcular ventas por liga (planas + curvas)
            for liga in ligas:
                if es_multiindex:
                    # Para columnas MultiIndex: (Liga, Tipo, 'Ventas')
                    col_planas_ventas = (liga, 'Planas', 'Ventas')
                    col_curvas_ventas = (liga, 'Curvas', 'Ventas')
                    
                    ventas_planas = df_bodegas.loc[bodega_idx, col_planas_ventas] if col_planas_ventas in df_bodegas.columns else 0
                    ventas_curvas = df_bodegas.loc[bodega_idx, col_curvas_ventas] if col_curvas_ventas in df_bodegas.columns else 0
                else:
                    # Para columnas simples: "LIGA - Tipo - Ventas"  
                    col_planas_ventas = f"{liga} - Planas - Ventas"
                    col_curvas_ventas = f"{liga} - Curvas - Ventas"
                    
                    ventas_planas = df_bodegas.loc[bodega_idx, col_planas_ventas] if col_planas_ventas in df_bodegas.columns else 0
                    ventas_curvas = df_bodegas.loc[bodega_idx, col_curvas_ventas] if col_curvas_ventas in df_bodegas.columns else 0
                
                # Asegurar que son números
                try:
                    ventas_planas = float(ventas_planas) if ventas_planas != 0 else 0
                    ventas_curvas = float(ventas_curvas) if ventas_curvas != 0 else 0
                except:
                    ventas_planas = 0
                    ventas_curvas = 0
                
                ventas_liga = ventas_planas + ventas_curvas
                bodega_data_ventas[liga] = ventas_liga
                total_ventas_bodega += ventas_liga
            
            # Calcular porcentajes de ventas
            if total_ventas_bodega > 0:
                for liga in ligas:
                    bodega_data_ventas[f"{liga}_porcentaje_ventas"] = (bodega_data_ventas[liga] / total_ventas_bodega) * 100
            else:
                for liga in ligas:
                    bodega_data_ventas[f"{liga}_porcentaje_ventas"] = 0
            
            bodega_data_ventas['Total_Ventas'] = total_ventas_bodega
            distribucion_ventas_data.append(bodega_data_ventas)
        
        # Convertir a DataFrame  
        df_distribucion_ventas = pd.DataFrame(distribucion_ventas_data)
        
        if len(df_distribucion_ventas) == 0:
            st.warning("No hay datos de ventas disponibles para mostrar gráficos.")
        else:
            # Filtrar CENTRAL NEW ERA, New Era Central y TOTAL del gráfico
            bodegas_excluir_ventas = ['CENTRAL NEW ERA', 'New Era Central', 'TOTAL']
            # Para PANAMA, excluir bodegas centrales
            if pais == "PANAMA":
                bodegas_excluir_ventas.extend(['Almacén general', 'Bodega Central Albrook'])
            
            df_distribucion_ventas = df_distribucion_ventas[
                ~df_distribucion_ventas['Bodega'].isin(bodegas_excluir_ventas)
            ].copy()
            
            if len(df_distribucion_ventas) > 0:
                # Separar los datos para ventas
                df_principales_ventas = df_distribucion_ventas[df_distribucion_ventas['Bodega'].isin(bodegas_principales)].copy()
                df_outlets_ventas = df_distribucion_ventas[df_distribucion_ventas['Bodega'].isin(bodegas_outlets)].copy()
                
                # Para El Salvador, separar NE METROCENTRO LOURDES como tienda outlet especial
                if pais == "El Salvador":
                    df_outlet_especial_ventas = df_distribucion_ventas[df_distribucion_ventas['Bodega'].isin(bodega_outlet_especial)].copy()
                    df_secundarias_ventas = df_distribucion_ventas[
                        ~df_distribucion_ventas['Bodega'].isin(bodegas_principales + bodegas_outlets + bodega_outlet_especial)
                    ].copy()
                else:
                    df_outlet_especial_ventas = pd.DataFrame()
                    df_secundarias_ventas = df_distribucion_ventas[
                        ~df_distribucion_ventas['Bodega'].isin(bodegas_principales + bodegas_outlets)
                    ].copy()
                    
                    # Excluir bodegas centrales específicamente de ventas
                    if pais == "Costa Rica":
                        df_secundarias_ventas = df_secundarias_ventas[df_secundarias_ventas['Bodega'] != 'Bodega Central NEW ERA'].copy()
                    elif pais == "PANAMA":
                        df_secundarias_ventas = df_secundarias_ventas[
                            ~df_secundarias_ventas['Bodega'].isin(['Almacén general', 'Bodega Central Albrook'])
                        ].copy()
                
                # Crear y mostrar gráfico de tiendas principales con ventas
                if len(df_principales_ventas) > 0:
                    st.markdown(f"#### 🏪 {nombre_tiendas_principales} - Ventas")
                    fig_principales_ventas = crear_grafico_distribucion_ventas(
                        df_principales_ventas, 
                        f'Distribución por Ligas - {nombre_tiendas_principales} - Ventas ({pais})', 
                        ligas
                    )
                    if fig_principales_ventas:
                        st.plotly_chart(fig_principales_ventas, use_container_width=True)
                        
                        # Mostrar leyenda de ligas justo después del gráfico
                        crear_leyenda_ligas()
                    
                    # Mostrar tabla de tiendas principales de ventas después de la leyenda
                    crear_tabla_resumen_ventas(df_principales_ventas, f"📋 Resumen - {nombre_tiendas_principales} - Ventas", ligas)
                
                # Crear y mostrar gráfico de outlets con ventas
                if len(df_outlets_ventas) > 0:
                    st.markdown("#### 🛒 Outlets - Ventas")
                    fig_outlets_ventas = crear_grafico_distribucion_ventas(
                        df_outlets_ventas, 
                        f'Distribución por Ligas - Outlets - Ventas ({pais})', 
                        ligas
                    )
                    if fig_outlets_ventas:
                        st.plotly_chart(fig_outlets_ventas, use_container_width=True)
                        
                        # Mostrar leyenda de ligas justo después del gráfico
                        crear_leyenda_ligas()
                    
                    # Mostrar tabla de outlets de ventas después de la leyenda
                    crear_tabla_resumen_ventas(df_outlets_ventas, "📋 Resumen - Outlets - Ventas", ligas)
                
                # Crear y mostrar gráfico de tiendas departamentales con ventas
                if len(df_secundarias_ventas) > 0:
                    st.markdown(f"#### 🏬 {nombre_tiendas_secundarias} - Ventas")
                    fig_secundarias_ventas = crear_grafico_distribucion_ventas(
                        df_secundarias_ventas, 
                        f'Distribución por Ligas - {nombre_tiendas_secundarias} - Ventas ({pais})', 
                        ligas
                    )
                    if fig_secundarias_ventas:
                        st.plotly_chart(fig_secundarias_ventas, use_container_width=True)
                        
                        # Mostrar leyenda de ligas justo después del gráfico
                        crear_leyenda_ligas()
                    
                    # Mostrar tabla de tiendas departamentales de ventas después de la leyenda
                    crear_tabla_resumen_ventas(df_secundarias_ventas, f"📋 Resumen - {nombre_tiendas_secundarias} - Ventas", ligas)
                
                # Crear y mostrar gráfico de tienda outlet especial con ventas (solo para El Salvador)
                if pais == "El Salvador" and len(df_outlet_especial_ventas) > 0:
                    st.markdown("#### 🏪 Tienda Outlet - Ventas")
                    fig_outlet_especial_ventas = crear_grafico_distribucion_ventas(
                        df_outlet_especial_ventas, 
                        f'Distribución por Ligas - Tienda Outlet - Ventas ({pais})', 
                        ligas
                    )
                    if fig_outlet_especial_ventas:
                        st.plotly_chart(fig_outlet_especial_ventas, use_container_width=True)
                        
                        # Mostrar leyenda de ligas justo después del gráfico
                        crear_leyenda_ligas()
                    
                    # Mostrar tabla de tienda outlet de ventas después de la leyenda
                    crear_tabla_resumen_ventas(df_outlet_especial_ventas, "📋 Resumen - Tienda Outlet - Ventas", ligas)
        
        # ==================== NUEVA SECCIÓN: COMPARACIÓN STOCK VS VENTAS ====================
        
        # Función para crear gráfico comparativo - LÍNEAS VERTICALES DELGADAS AGRUPADAS
        def crear_grafico_comparativo_stock_ventas(df_data_stock, df_data_ventas, titulo_grafico, ligas):
            if len(df_data_stock) == 0 or len(df_data_ventas) == 0:
                return None
                
            fig = go.Figure()
            
            # Colores para cada liga
            colores_ligas = {
                'MLB': '#1f77b4',      # Azul  
                'NBA': '#ff7f0e',      # Naranja
                'NFL': '#2ca02c',      # Verde
                'MOTORSPORT': '#d62728', # Rojo
                'ENTERTAINMENT': '#9467bd' # Púrpura
            }
            
            # Obtener nombres de bodegas para el eje X
            nombres_bodegas = df_data_ventas['Bodega'].tolist()
            num_ligas = len(ligas)
            
            # Crear líneas verticales para cada liga y bodega
            for i, liga in enumerate(ligas):
                for j, bodega in enumerate(nombres_bodegas):
                    # Calcular posiciones X para agrupar las líneas por bodega
                    # Cada bodega tendrá 10 líneas: 5 ventas + 5 stock
                    base_x = j * 0.8  # Reducir espaciado entre bodegas (era j)
                    offset_ventas = (i - (num_ligas - 1) / 2) * 0.12  # Aumentar espaciado entre líneas (era 0.08)
                    offset_stock = offset_ventas + 0.05  # Aumentar separación stock-ventas (era 0.03)
                    
                    x_ventas = base_x + offset_ventas
                    x_stock = base_x + offset_stock
                    
                    # Obtener valores
                    valor_ventas = df_data_ventas.iloc[j][f'{liga}_porcentaje_ventas']
                    valor_stock = df_data_stock.iloc[j][f'{liga}_porcentaje']
                    
                    # Posiciones de texto con rotación para evitar sobreposición
                    text_pos_ventas = 'top center'
                    text_pos_stock = 'top center'
                    text_y_offset_ventas = 2
                    text_y_offset_stock = 2
                    
                    # LÍNEA VERTICAL PARA VENTAS (sólida, delgada)
                    fig.add_trace(go.Scatter(
                        x=[x_ventas, x_ventas],
                        y=[0, valor_ventas],
                        mode='lines',
                        line=dict(
                            color=colores_ligas[liga],
                            width=4,  # Línea delgada pero visible
                            dash='solid'
                        ),
                        showlegend=False,
                        hovertemplate=f'<b>{liga} - Ventas</b><br>{bodega}<br>{valor_ventas:.1f}%<extra></extra>',
                        name=f'{liga} - Ventas'
                    ))
                    
                    # MARCA CIRCULAR PARA VENTAS (extremo de línea sólida)
                    fig.add_trace(go.Scatter(
                        x=[x_ventas],
                        y=[valor_ventas],
                        mode='markers',
                        marker=dict(
                            color=colores_ligas[liga],
                            size=6,
                            symbol='circle',
                            line=dict(width=1, color='white')
                        ),
                        showlegend=False,
                        hoverinfo='skip'
                    ))
                    
                    # ANOTACIÓN PARA VENTAS (rotada -90 grados)
                    fig.add_annotation(
                        x=x_ventas,
                        y=valor_ventas + text_y_offset_ventas,
                        text=f'{valor_ventas:.1f}%',
                        textangle=-90,
                        showarrow=False,
                        font=dict(
                            size=9,
                            color=colores_ligas[liga],
                            family='Inter, sans-serif'
                        ),
                        xanchor='center',
                        yanchor='bottom'
                    )
                    
                    # LÍNEA VERTICAL PARA STOCK (punteada, delgada)
                    fig.add_trace(go.Scatter(
                        x=[x_stock, x_stock],
                        y=[0, valor_stock],
                        mode='lines',
                        line=dict(
                            color=colores_ligas[liga],
                            width=3,  # Más delgada que ventas
                            dash='dash'  # Punteada
                        ),
                        showlegend=False,
                        hovertemplate=f'<b>{liga} - Stock</b><br>{bodega}<br>{valor_stock:.1f}%<extra></extra>',
                        name=f'{liga} - Stock'
                    ))
                    
                    # MARCA CUADRADA PARA STOCK (extremo de línea punteada)
                    fig.add_trace(go.Scatter(
                        x=[x_stock],
                        y=[valor_stock],
                        mode='markers',
                        marker=dict(
                            color=colores_ligas[liga],
                            size=6,
                            symbol='square',
                            line=dict(width=1, color='white')
                        ),
                        showlegend=False,
                        hoverinfo='skip'
                    ))
                    
                    # ANOTACIÓN PARA STOCK (rotada -90 grados)
                    fig.add_annotation(
                        x=x_stock,
                        y=valor_stock + text_y_offset_stock,
                        text=f'{valor_stock:.1f}%',
                        textangle=-90,
                        showarrow=False,
                        font=dict(
                            size=8,
                            color=colores_ligas[liga],
                            family='Inter, sans-serif'
                        ),
                        xanchor='center',
                        yanchor='bottom'
                    )
            
            # Configurar layout
            fig.update_layout(
                title=titulo_grafico,
                xaxis_title='Bodegas/Tiendas',
                yaxis_title='Porcentaje (%)',
                height=600,
                showlegend=False,
                xaxis=dict(
                    tickmode='array',
                    tickvals=[i * 0.8 for i in range(len(nombres_bodegas))],  # Ajustar posiciones de etiquetas
                    ticktext=nombres_bodegas
                ),
                margin=dict(l=60, r=60, t=100, b=80)
            )
            
            # Configurar ejes
            fig.update_xaxes(
                tickangle=45,
                range=[-0.4, (len(nombres_bodegas) - 1) * 0.8 + 0.4]  # Ajustar rango para el nuevo espaciado
            )
            fig.update_yaxes(range=[0, 100])
            
            # Agregar líneas de cuadrícula
            fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='rgba(0,0,0,0.1)')
            fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='rgba(0,0,0,0.1)')
            
            return fig
        
        # Crear header de sección para comparación
        professional_design.create_section_header(
            f"Comparación Stock vs Ventas por Bodega - {pais}",
            "Comparación visual entre distribución de stock (transparente) y ventas (sólido) por liga",
            "📊"
        )
        
        # Crear gráfico comparativo para tiendas principales
        if len(df_principales) > 0 and len(df_principales_ventas) > 0:
            st.markdown(f"#### 🏪 {nombre_tiendas_principales} - Comparación Stock vs Ventas")
            fig_comparativo_principales = crear_grafico_comparativo_stock_ventas(
                df_principales,
                df_principales_ventas, 
                f'Stock vs Ventas - {nombre_tiendas_principales} ({pais})', 
                ligas
            )
            if fig_comparativo_principales:
                st.plotly_chart(fig_comparativo_principales, use_container_width=True)
                
                # Mostrar leyenda de ligas justo después del gráfico
                crear_leyenda_ligas()
        
        # Crear gráfico comparativo para outlets
        if len(df_outlets) > 0 and len(df_outlets_ventas) > 0:
            st.markdown("#### 🛒 Outlets - Comparación Stock vs Ventas")
            fig_comparativo_outlets = crear_grafico_comparativo_stock_ventas(
                df_outlets,
                df_outlets_ventas, 
                f'Stock vs Ventas - Outlets ({pais})', 
                ligas
            )
            if fig_comparativo_outlets:
                st.plotly_chart(fig_comparativo_outlets, use_container_width=True)
                
                # Mostrar leyenda de ligas justo después del gráfico
                crear_leyenda_ligas()
        
        # Crear gráfico comparativo para tiendas departamentales
        if len(df_secundarias) > 0 and len(df_secundarias_ventas) > 0:
            st.markdown(f"#### 🏬 {nombre_tiendas_secundarias} - Comparación Stock vs Ventas")
            fig_comparativo_secundarias = crear_grafico_comparativo_stock_ventas(
                df_secundarias,
                df_secundarias_ventas, 
                f'Stock vs Ventas - {nombre_tiendas_secundarias} ({pais})', 
                ligas
            )
            if fig_comparativo_secundarias:
                st.plotly_chart(fig_comparativo_secundarias, use_container_width=True)
                
                # Mostrar leyenda de ligas justo después del gráfico
                crear_leyenda_ligas()
        
        # Crear y mostrar gráfico comparativo de tienda outlet especial (solo para El Salvador)
        if pais == "El Salvador" and len(df_outlet_especial) > 0 and len(df_outlet_especial_ventas) > 0:
            st.markdown("#### 🏪 Tienda Outlet - Comparación Stock vs Ventas")
            fig_comparativo_outlet_especial = crear_grafico_comparativo_stock_ventas(
                df_outlet_especial,
                df_outlet_especial_ventas, 
                f'Stock vs Ventas - Tienda Outlet ({pais})', 
                ligas
            )
            if fig_comparativo_outlet_especial:
                st.plotly_chart(fig_comparativo_outlet_especial, use_container_width=True)
                
                # Mostrar leyenda de ligas justo después del gráfico
                crear_leyenda_ligas()
        
        # NUEVA SECCIÓN: Exportar Distribuciones cuando hay ventas (al final de todo)
        if pais in ["Guatemala", "El Salvador", "Costa Rica", "Honduras", "PANAMA"]:
            tiene_ventas_final = any('Ventas' in str(col) for col in df_bodegas.columns)
            if tiene_ventas_final:
                # Crear diccionario completo con todas las tablas (stock y ventas)
                tablas_completas = {
                    'df_principales': df_principales if len(df_principales) > 0 else pd.DataFrame(),
                    'df_outlets': df_outlets if len(df_outlets) > 0 else pd.DataFrame(),
                    'df_secundarias': df_secundarias if len(df_secundarias) > 0 else pd.DataFrame(),
                    'df_principales_ventas': df_principales_ventas if len(df_principales_ventas) > 0 else pd.DataFrame(),
                    'df_outlets_ventas': df_outlets_ventas if len(df_outlets_ventas) > 0 else pd.DataFrame(),
                    'df_secundarias_ventas': df_secundarias_ventas if len(df_secundarias_ventas) > 0 else pd.DataFrame()
                }
                
                # Agregar Tienda Outlet especial para El Salvador
                if pais == "El Salvador":
                    tablas_completas['df_outlet_especial'] = df_outlet_especial if len(df_outlet_especial) > 0 else pd.DataFrame()
                    tablas_completas['df_outlet_especial_ventas'] = df_outlet_especial_ventas if len(df_outlet_especial_ventas) > 0 else pd.DataFrame()
                agregar_seccion_exportar_distribuciones(tablas_completas, pais, tiene_ventas_final)
        
    else:
        # Mostrar mensaje informativo para otros países
        st.info(f"📊 Los gráficos de distribución de ventas solo están disponibles para Guatemala cuando se cargan datos de ventas.")
    
    # NUEVA SECCIÓN: Exportar Distribuciones (para Guatemala, El Salvador, Costa Rica y Honduras)
    # Recolectar las tablas reales para exportación
    if pais in ["Guatemala", "El Salvador", "Costa Rica", "Honduras"]:
        tiene_ventas = any('Ventas' in str(col) for col in df_bodegas.columns)
        
        # Crear diccionario con las tablas de stock
        tablas_reales = {
            'df_principales': df_principales if len(df_principales) > 0 else pd.DataFrame(),
            'df_outlets': df_outlets if len(df_outlets) > 0 else pd.DataFrame(),
            'df_secundarias': df_secundarias if len(df_secundarias) > 0 else pd.DataFrame()
        }
        
        # Agregar Tienda Outlet especial para El Salvador
        if pais == "El Salvador":
            tablas_reales['df_outlet_especial'] = df_outlet_especial if len(df_outlet_especial) > 0 else pd.DataFrame()
        
        # Si NO hay ventas, mostrar la sección aquí (después de distribución de stock)
        # Si SÍ hay ventas, la sección se mostrará al final de la función (después de comparación)
        if not tiene_ventas:
            agregar_seccion_exportar_distribuciones(tablas_reales, pais, tiene_ventas)
    
    # CSS aplicado de forma más simple y compatible
    st.markdown("""
    <style>
    /* Estilos específicos para la tabla de distribución */
    [data-testid="stDataFrame"] {
        background: white;
        border-radius: 8px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        overflow: hidden;
        border: 1px solid #e5e7eb;
        font-family: 'Inter', sans-serif;
    }
    
    [data-testid="stDataFrame"] thead th {
        background: #3b82f6;
        color: white;
        font-weight: bold;
        text-align: center;
        padding: 8px;
        font-size: 11px;
        border: none;
    }
    
    [data-testid="stDataFrame"] tbody td {
        text-align: center;
        padding: 6px 8px;
        font-size: 11px;
        border-bottom: 1px solid #f3f4f6;
    }
    
    [data-testid="stDataFrame"] tbody td:first-child {
        text-align: left;
        font-weight: bold;
        background: #f8fafc;
        color: #374151;
        padding-left: 12px;
    }
    
    [data-testid="stDataFrame"] tbody tr:nth-child(even) {
        background: #f9fafb;
    }
    
    [data-testid="stDataFrame"] tbody tr:hover {
        background: #eff6ff;
    }
    </style>
    """, unsafe_allow_html=True)

# Instancia del visualizador de gráficas
chart_visualizer = ChartVisualizer(stock_analyzer, country_manager)

def mostrar_grafica_comparativa(tabla, pais):
    """Wrapper para compatibilidad"""
    chart_visualizer.mostrar_grafica_comparativa(tabla, pais)

def mostrar_tabla_consolidada(tabla, pais):
    """Muestra la tabla con múltiples niveles de encabezados"""
    if tabla is None:
        return
    
    logger.info(f"Mostrando tabla consolidada para {pais}")
    
    professional_design.create_section_header(
        f"Tabla Consolidada - {pais}", 
        "Detalle completo de inventario por bodega y categoría",
        "📊"
    )
    
    # Generar estilos dinámicos para el semáforo
    estilos_semaforo = []
    capacidades = country_manager.get_capacidades(pais)
    
    # Encontrar la posición de la columna "% DE CUMPLIMIENTO"
    col_cumplimiento_index = None
    for idx, col in enumerate(tabla.columns):
        # Manejar estructura de 3 niveles para totales
        if len(col) == 3 and col[2] == '% DE CUMPLIMIENTO':
            col_cumplimiento_index = idx + 1  # +1 porque CSS es 1-based
            break
    
    if col_cumplimiento_index is None:
        logger.warning("No se encontró la columna % DE CUMPLIMIENTO")
        col_cumplimiento_index = 17  # Fallback al valor original
    
    for i, fila in tabla.iterrows():
        # Buscar columna Bodega en estructura nueva o antigua
        bodega = None
        total_headwear = None
        
        for col in tabla.columns:
            if len(col) == 3 and col[2] == 'Bodega':
                bodega = fila[col]
            elif len(col) == 3 and col[2] == 'TOTAL HEADWEAR':
                total_headwear = fila[col]
        
        if bodega is None or total_headwear is None:
            continue
        
        if bodega == 'TOTAL':
            capacidad = country_manager.get_country_data(pais).get_total_capacity()
        else:
            capacidad = capacidades.get(bodega, 0)
        
        if capacidad > 0:
            color = stock_analyzer.obtener_color_semaforo(total_headwear, capacidad)
            
            if color == "verde":
                color_css = "#28a745"  # Verde
            elif color == "amarillo":
                color_css = "#ffc107"  # Amarillo
            else:
                color_css = "#dc3545"  # Rojo
        else:
            color_css = "#6c757d"  # Gris para N/A
        
        # Crear selector CSS para la celda específica
        estilos_semaforo.append(f"""
            .stDataFrame tbody tr:nth-child({i+1}) td:nth-child({col_cumplimiento_index}) {{
                background-color: {color_css} !important;
                color: white !important;
                font-weight: bold !important;
            }}
        """)
    
    # Formatear números con separadores de miles
    tabla_formateada = tabla.copy()
    for col in tabla_formateada.columns:
        # Formato especial para columnas de ventas (MultiIndex con 3 niveles) - SIN USD
        if len(col) == 3 and col[2] == 'Ventas':  # Nueva estructura: (Liga, Subcategoría, 'Ventas')
            tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) and x > 0 else "0")
        elif len(col) >= 2 and col[-1] not in ['Bodega', '% DE CUMPLIMIENTO']:  # Stock y otros números
            tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) else "0")
        elif len(col) >= 2 and col[-1] == '% DE CUMPLIMIENTO':  # Porcentajes
            tabla_formateada[col] = tabla_formateada[col].astype(str)  # Ya está formateado
    
    # Convertir tabla a HTML con celdas combinadas en MultiIndex
    def crear_tabla_html_con_celdas_combinadas(df):
        # Obtener información del MultiIndex
        liga_counts = {}
        subcategoria_counts = {}
        
        # Contar cuántas columnas tiene cada liga y subcategoría
        for col in df.columns:
            if len(col) == 3:  # (Liga, Subcategoría, Tipo)
                liga, subcategoria, tipo = col
                
                # Contar por liga
                if liga not in liga_counts:
                    liga_counts[liga] = 0
                liga_counts[liga] += 1
                
                # Contar por combinación liga-subcategoría
                key = (liga, subcategoria)
                if key not in subcategoria_counts:
                    subcategoria_counts[key] = 0
                subcategoria_counts[key] += 1
        
        # Crear HTML de la tabla - Tamaño más compacto
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 7px; width: 100%;">'
        
        # Fila 1: Ligas (con colspan)
        html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
        html += '<td rowspan="3" style="border: 1px solid #ddd; padding: 2px; vertical-align: middle; font-size: 7px; width: 50px;">Bodega</td>'
        
        for liga, count in liga_counts.items():
            if liga != 'INFO':  # Skip INFO column
                html += f'<td colspan="{count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{liga}</td>'
        
        html += '</tr>'
        
        # Fila 2: Subcategorías (con colspan)
        html += '<tr style="background-color: #f0f0f0; font-weight: bold;">'
        
        processed_subcategorias = set()
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO':
                liga, subcategoria, tipo = col
                key = (liga, subcategoria)
                
                # Solo procesar cada subcategoría una vez
                if key not in processed_subcategorias:
                    processed_subcategorias.add(key)
                    sub_count = subcategoria_counts.get(key, 1)
                    html += f'<td colspan="{sub_count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{subcategoria}</td>'
        
        html += '</tr>'
        
        # Fila 3: Tipos (Stock/Ventas)
        html += '<tr style="background-color: #e8e8e8; font-weight: bold;">'
        
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO':
                liga, subcategoria, tipo = col
                # Cambiar "Ventas" por "Ventas (USD)"
                tipo_display = "Ventas (USD)" if tipo == "Ventas" else tipo
                html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; width: 30px;">{tipo_display}</td>'
        
        html += '</tr>'
        
        # Encontrar la columna "% DE CUMPLIMIENTO" para aplicar semáforo
        col_cumplimiento_index = None
        for idx, col in enumerate(df.columns):
            if len(col) == 3 and col[2] == '% DE CUMPLIMIENTO':
                col_cumplimiento_index = idx
                break
        
        # Filas de datos
        for idx, row in df.iterrows():
            if idx == len(df) - 1:  # Fila TOTAL
                html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
            else:
                html += '<tr>'
            
            for col_idx, col in enumerate(df.columns):
                value = row[col]
                
                # Aplicar semáforo solo a la columna % DE CUMPLIMIENTO (y no a fila TOTAL)
                if col_idx == col_cumplimiento_index and idx < len(df) - 1:
                    # Obtener capacidades para calcular color del semáforo
                    bodega = None
                    total_headwear = None
                    
                    # Buscar valores necesarios para el cálculo del semáforo
                    for search_col in df.columns:
                        if len(search_col) == 3 and search_col[2] == 'Bodega':
                            bodega = row[search_col]
                        elif len(search_col) == 3 and search_col[2] == 'TOTAL HEADWEAR':
                            total_headwear = row[search_col]
                    
                    # Calcular color del semáforo
                    if bodega and total_headwear is not None:
                        capacidades = country_manager.get_capacidades(pais)
                        capacidad = capacidades.get(bodega, 0)
                        
                        if capacidad > 0:
                            # Convertir total_headwear a número si está como string
                            try:
                                if isinstance(total_headwear, str):
                                    # Remover comas y convertir a float
                                    total_headwear_num = float(total_headwear.replace(',', ''))
                                else:
                                    total_headwear_num = float(total_headwear)
                                
                                color = stock_analyzer.obtener_color_semaforo(total_headwear_num, capacidad)
                                
                                if color == "verde":
                                    color_css = "#d4edda"  # Verde (mismo que MVP)
                                elif color == "amarillo":
                                    color_css = "#fff3cd"  # Amarillo (mismo que MVP)
                                else:
                                    color_css = "#f8d7da"  # Rojo (mismo que MVP)
                            except (ValueError, TypeError):
                                color_css = "#f8f9fa"  # Gris para errores de conversión (mismo que MVP)
                        else:
                            color_css = "#f8f9fa"  # Gris para N/A (mismo que MVP)
                        
                        html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; background-color: {color_css}; color: black; font-weight: bold;">{value}</td>'
                    else:
                        html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{value}</td>'
                else:
                    html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{value}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        return html
    
    # CSS para tabla con ancho igual a títulos y scroll interno
    st.markdown("""
    <style>
        .tabla-solo-ventas {
            overflow-x: auto !important;
            overflow-y: auto !important;
            width: 100% !important;
            max-width: 100% !important;
            max-height: 500px !important;
            border: 1px solid #ddd !important;
            background: white !important;
            margin: 0 !important;
            border-radius: 8px !important;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1) !important;
        }
        
        .tabla-solo-ventas table {
            font-size: 7px !important;
            border-collapse: collapse !important;
            width: max-content !important;
            min-width: 100% !important;
        }
        
        .tabla-solo-ventas td, .tabla-solo-ventas th {
            padding: 4px 6px !important;
            border: 1px solid #ddd !important;
            font-size: 7px !important;
            white-space: nowrap !important;
            text-align: center !important;
            min-width: 70px !important;
            max-width: 70px !important;
        }
        
        .tabla-solo-ventas td:first-child, .tabla-solo-ventas th:first-child {
            min-width: 100px !important;
            max-width: 100px !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Crear contenedor con estilos inline - REVERTIR a dimensiones originales
    container_style = """
        overflow-x: auto; 
        overflow-y: auto; 
        width: 100%; 
        max-width: 100%; 
        max-height: 400px; 
        border: 2px solid #333; 
        background: white; 
        margin: 0; 
        border-radius: 8px; 
        box-shadow: 0 2px 8px rgba(0,0,0,0.2);
        font-size: 12px;
    """
    
    st.markdown(f'<div style="{container_style}">', unsafe_allow_html=True)
    tabla_html = crear_tabla_html_con_celdas_combinadas(tabla_formateada)
    st.markdown(tabla_html, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Sección de exportación (para Guatemala, El Salvador, Costa Rica, Honduras y PANAMA)
    if pais in ["Guatemala", "El Salvador", "Costa Rica", "Honduras", "PANAMA"]:
        # Configurar header según el país
        if pais == "Guatemala":
            codigo_pais = "GT"
        elif pais == "Costa Rica":
            codigo_pais = "CR"
        elif pais == "Honduras":
            codigo_pais = "HN"
        elif pais == "PANAMA":
            codigo_pais = "PA"
        else:  # El Salvador
            codigo_pais = "SV"
        professional_design.create_section_header(
            f"Exportar Reporte - {pais}", 
            "Generar archivo Excel con formato profesional",
            codigo_pais
        )
        
        col1, col2 = st.columns([3, 2])
        
        with col1:
            # Obtener nombre del archivo desde session state o usar valor por defecto
            if pais == "Guatemala":
                archivo_default = "GUATEMALA.csv"
                session_key = 'archivo_guatemala_name'
                export_key = "nombre_gt_export"
            elif pais == "Costa Rica":
                archivo_default = "COSTA_RICA.csv"
                session_key = 'archivo_costa_rica_name'
                export_key = "nombre_cr_export"
            elif pais == "Honduras":
                archivo_default = "HONDURAS.csv"
                session_key = 'archivo_honduras_name'
                export_key = "nombre_hn_export"
            elif pais == "PANAMA":
                archivo_default = "PANAMA.csv"
                session_key = 'archivo_panama_name'
                export_key = "nombre_pa_export"
            else:  # El Salvador
                archivo_default = "EL_SALVADOR.csv"
                session_key = 'archivo_el_salvador_name'
                export_key = "nombre_sv_export"
            
            archivo_nombre = st.session_state.get(session_key, archivo_default)
            nombre_archivo = st.text_input("📝 Nombre del archivo origen", archivo_nombre, key=export_key)
        
        with col2:
            st.markdown("<br>", unsafe_allow_html=True)  # Espaciado
            if pais == "Guatemala":
                button_key = "excel_gt_export"
            elif pais == "Costa Rica":
                button_key = "excel_cr_export"
            elif pais == "Honduras":
                button_key = "excel_hn_export"
            elif pais == "PANAMA":
                button_key = "excel_pa_export"
            else:  # El Salvador
                button_key = "excel_sv_export"
            if st.button(f"🚀 Generar Excel {pais}", key=button_key, use_container_width=True):
                exportar_excel_consolidado(tabla, nombre_archivo, pais)
    
    # Mostrar métricas resumidas mejoradas
    selected_league = st.session_state.get('selected_league', None)
    # Convertir "Todas" a None para mostrar todas las ligas
    if selected_league == "Todas":
        selected_league = None
    
    if selected_league:
        professional_design.create_section_header(
            f"Métricas {selected_league} - {pais}", 
            f"Resumen ejecutivo de inventario específico para {selected_league}",
            "📈"
        )
    else:
        professional_design.create_section_header(
            f"Métricas Generales - {pais}", 
            "Resumen ejecutivo de inventario por categoría",
            "📈"
        )
    
    # Verificar si hay datos de ventas para incluir TOTAL VENTAS
    hay_total_usd = False
    total_ventas_valor = 0
    for tabla_col in tabla.columns:
        if len(tabla_col) == 3 and tabla_col[2] == 'TOTAL (USD)':
            hay_total_usd = True
            total_ventas_valor = tabla[tabla_col].iloc[-1]
            break
    
    # Definir métricas según disponibilidad de datos de ventas
    if hay_total_usd:
        cols = st.columns(5)
        metricas = [
            ('TOTAL HEADWEAR', 'Headwear Total', "🧢", "#000000"),
            ('TOTAL APPAREL', 'Apparel Total', "👕", "#000000"),
            ('ACCESSORIES', 'Accessories Total', "🧦", "#000000"),
            ('TOTAL STOCK', 'Inventario Total', "📦", "#000000"),
            ('TOTAL (USD)', 'Total Ventas', "💰", "#000000")
        ]
    else:
        cols = st.columns(4)
        metricas = [
            ('TOTAL HEADWEAR', 'Headwear Total', "🧢", "#000000"),
            ('TOTAL APPAREL', 'Apparel Total', "👕", "#000000"),
            ('ACCESSORIES', 'Accessories Total', "🧦", "#000000"),
            ('TOTAL STOCK', 'Inventario Total', "📦", "#000000")
        ]
    
    for i, (col, nombre, emoji, color) in enumerate(metricas):
        with cols[i]:
            # Buscar la columna en la estructura de 3 niveles (totales)
            valor = 0
            if col == 'ACCESSORIES':
                # Para ACCESSORIES, buscar en la estructura específica
                try:
                    # Buscar columna con estructura (ACCESSORIES, Accessories, Stock)
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'ACCESSORIES' and 
                            tabla_col[1] == 'Accessories' and 
                            tabla_col[2] == 'Stock'):
                            # Tomar solo la fila TOTAL (última fila)
                            valor = tabla[tabla_col].iloc[-1]
                            break
                except:
                    valor = 0
            else:
                # Para las demás métricas, usar la lógica original
                for tabla_col in tabla.columns:
                    if len(tabla_col) == 3 and tabla_col[2] == col:
                        valor = tabla[tabla_col].iloc[-1]
                        break
            
            # Determinar el texto de descripción según el tipo de métrica
            if col == 'TOTAL (USD)':
                descripcion = f"USD en ventas{f' - {selected_league}' if selected_league else ''}"
                valor_formato = f"${valor:,.2f}"
            else:
                descripcion = f"unidades en stock{f' - {selected_league}' if selected_league else ''}"
                # Convertir a entero para eliminar decimales
                valor_entero = int(valor) if valor else 0
                valor_formato = f"{valor_entero:,}"
            
            st.markdown(f"""
            <div class="metric-card" style="border-left: 4px solid {color};">
                <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 0.5rem;">
                    <span style="font-size: 2rem; color: #000000;">{emoji}</span>
                    <span style="color: {color}; font-weight: 600; font-size: 1.1rem;">{nombre.upper()}</span>
                </div>
                <div style="font-size: 2rem; font-weight: 700; color: #374151; margin-bottom: 0.25rem;">
                    {valor_formato}
                </div>
                <div style="color: #6b7280; font-size: 0.85rem;">
                    {descripcion}
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    # AGREGAR GRÁFICA COMPARATIVA
    if selected_league:
        professional_design.create_section_header(
            f"Análisis Visual {selected_league} - {pais}", 
            f"Visualización interactiva de stock específico para {selected_league}",
            "📊"
        )
    else:
        professional_design.create_section_header(
            f"Análisis Visual - {pais}", 
            "Comparativa interactiva de stock vs capacidad por bodega",
            "📊"
        )
    mostrar_grafica_comparativa(tabla, pais)
    
    # AGREGAR NUEVA SECCIÓN: Distribución de Ligas por Bodega (para Guatemala, El Salvador, Costa Rica, Honduras y PANAMA)
    if pais in ["Guatemala", "El Salvador", "Costa Rica", "Honduras", "PANAMA"]:
        mostrar_distribucion_ligas_por_bodega(tabla, pais)

def mostrar_tabla_solo_ventas_guatemala(tabla):
    """Muestra la tabla consolidada para solo-ventas de Guatemala (sin capacidades ni % cumplimiento)"""
    if tabla is None:
        return
    
    logger.info("Mostrando tabla solo-ventas para Guatemala")
    
    professional_design.create_section_header(
        "Tabla Consolidada - Guatemala (Solo Ventas)", 
        "Detalle completo de cantidades vendidas por bodega y categoría",
        "📊"
    )
    
    # Formatear números con separadores de miles
    tabla_formateada = tabla.copy()
    for col in tabla_formateada.columns:
        # Formato para columnas según tipo
        if len(col) == 3 and col[2] not in ['Bodega']:
            if col[2] == 'TOTAL USD':
                # Formato para columnas USD sin símbolo de dólar
                tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{float(x):,.2f}" if pd.notnull(x) else "0.00")
            else:
                # Formato para columnas de cantidades
                tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) else "0")
        # También formatear columnas de totales que no tienen MultiIndex de 3 niveles
        elif col == 'TOTAL USD':
            tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{float(x):,.2f}" if pd.notnull(x) else "0.00")
    
    # Convertir tabla a HTML con celdas combinadas en MultiIndex
    def crear_tabla_html_solo_ventas(df):
        # Obtener información del MultiIndex
        liga_counts = {}
        subcategoria_counts = {}
        
        # Contar cuántas columnas tiene cada liga y subcategoría
        for col in df.columns:
            if len(col) == 3:  # (Liga, Subcategoría, Tipo)
                liga, subcategoria, tipo = col
                
                # Solo contar las columnas que no son INFO ni Bodega
                if liga != 'INFO' and tipo != 'Bodega':
                    # Contar por liga
                    if liga not in liga_counts:
                        liga_counts[liga] = 0
                    liga_counts[liga] += 1
                    
                    # Contar por combinación liga-subcategoría
                    key = (liga, subcategoria)
                    if key not in subcategoria_counts:
                        subcategoria_counts[key] = 0
                    subcategoria_counts[key] += 1
        
        # Crear HTML de la tabla
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 7px; width: 100%; table-layout: fixed;">'
        
        # Fila 1: Ligas (con colspan)
        html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
        html += '<td rowspan="3" style="border: 1px solid #ddd; padding: 2px; vertical-align: middle; font-size: 7px; width: 50px;">Bodega</td>'
        
        for liga, count in liga_counts.items():
            if liga != 'INFO':  # Skip INFO column
                html += f'<td colspan="{count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{liga}</td>'
        
        html += '</tr>'
        
        # Fila 2: Subcategorías (con colspan)
        html += '<tr style="background-color: #f0f0f0; font-weight: bold;">'
        
        processed_subcategorias = set()
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO':
                liga, subcategoria, tipo = col
                key = (liga, subcategoria)
                
                # Solo procesar cada subcategoría una vez
                if key not in processed_subcategorias:
                    processed_subcategorias.add(key)
                    sub_count = subcategoria_counts.get(key, 1)
                    html += f'<td colspan="{sub_count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{subcategoria}</td>'
        
        html += '</tr>'
        
        # Fila 3: Tipos (Cantidad para reflejar el origen de los datos)
        html += '<tr style="background-color: #e8e8e8; font-weight: bold;">'
        
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO' and col[2] != 'Bodega':
                liga, subcategoria, tipo = col
                # Mostrar etiquetas apropiadas
                if tipo == "Cantidad":
                    tipo_display = "Cantidad"
                elif tipo == "TOTAL USD":
                    tipo_display = "TOTAL USD"
                else:
                    tipo_display = tipo
                html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; width: 30px;">{tipo_display}</td>'
        
        html += '</tr>'
        
        # Filas de datos
        for idx, row in df.iterrows():
            if idx == 'TOTAL':  # Fila TOTAL (verificar por nombre del índice)
                html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
            else:
                html += '<tr>'
            
            # Primero agregar la columna Bodega
            for col in df.columns:
                if len(col) == 3 and col[2] == 'Bodega':
                    value = row[col]
                    html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; text-align: left;">{value}</td>'
                    break
            
            # Luego agregar las demás columnas (cantidades)
            for col in df.columns:
                if len(col) == 3 and col[0] != 'INFO' and col[2] != 'Bodega':
                    value = row[col]
                    html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{value}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        return html
    
    # CSS para tabla con ancho igual a títulos y scroll interno
    st.markdown("""
    <style>
        .tabla-solo-ventas {
            overflow-x: auto !important;
            overflow-y: auto !important;
            width: 100% !important;
            max-width: 100% !important;
            max-height: 500px !important;
            border: 1px solid #ddd !important;
            background: white !important;
            margin: 0 !important;
            border-radius: 8px !important;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1) !important;
        }
        
        .tabla-solo-ventas table {
            font-size: 7px !important;
            border-collapse: collapse !important;
            width: max-content !important;
            min-width: 100% !important;
        }
        
        .tabla-solo-ventas td, .tabla-solo-ventas th {
            padding: 4px 6px !important;
            border: 1px solid #ddd !important;
            font-size: 7px !important;
            white-space: nowrap !important;
            text-align: center !important;
            min-width: 70px !important;
            max-width: 70px !important;
        }
        
        .tabla-solo-ventas td:first-child, .tabla-solo-ventas th:first-child {
            min-width: 100px !important;
            max-width: 100px !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    
    st.markdown('<div class="tabla-solo-ventas">', unsafe_allow_html=True)
    tabla_html = crear_tabla_html_solo_ventas(tabla_formateada)
    st.markdown(tabla_html, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Sección de exportación para Guatemala solo-ventas
    professional_design.create_section_header(
        "Exportar Reporte - Guatemala (Solo Ventas)", 
        "Generar archivo Excel con formato profesional",
        "GT"
    )
    
    col1, col2 = st.columns([3, 2])
    
    with col1:
        archivo_default = "GUATEMALA_SOLO_VENTAS.csv"
        session_key = 'archivo_guatemala_solo_ventas_name'
        export_key = "nombre_gt_solo_ventas_export"
        
        archivo_nombre = st.session_state.get(session_key, archivo_default)
        nombre_archivo = st.text_input("📝 Nombre del archivo origen", archivo_nombre, key=export_key)
    
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)  # Espaciado
        button_key = "excel_gt_solo_ventas_export"
        if st.button("🚀 Generar Excel Guatemala (Solo Ventas)", key=button_key, use_container_width=True):
            exportar_excel_consolidado(tabla, nombre_archivo, "Guatemala")
    
    # Mostrar métricas resumidas adaptadas para cantidades
    professional_design.create_section_header(
        "Métricas de Ventas - Guatemala", 
        "Resumen ejecutivo de cantidades vendidas por categoría",
        "📈"
    )
    
    # Definir métricas para cantidades
    cols = st.columns(5)
    metricas = [
        ('TOTAL HEADWEAR', 'Headwear Total', "🧢", "#000000"),
        ('TOTAL APPAREL', 'Apparel Total', "👕", "#000000"),
        ('ACCESSORIES', 'Accessories Total', "🧦", "#000000"),
        ('TOTAL CANTIDADES', 'Total Cantidades Vendidas', "📊", "#000000"),
        ('TOTAL USD', 'Total USD Vendidos', "💰", "#000000")
    ]
    
    for i, (col, nombre, emoji, color) in enumerate(metricas):
        with cols[i]:
            # Buscar la columna en la estructura de 3 niveles (totales)
            valor = 0
            if col == 'ACCESSORIES':
                # Para ACCESSORIES, buscar en la estructura específica
                try:
                    # Buscar columna con estructura (ACCESSORIES, Accessories, Cantidad)
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'ACCESSORIES' and 
                            tabla_col[1] == 'Accessories' and 
                            tabla_col[2] == 'Cantidad'):
                            # Tomar solo la fila TOTAL (última fila)
                            valor = tabla[tabla_col].iloc[-1]
                            break
                except:
                    valor = 0
            elif col == 'TOTAL CANTIDADES':
                # Para TOTAL CANTIDADES, sumar TOTAL HEADWEAR + TOTAL APPAREL + ACCESSORIES
                try:
                    total_headwear = 0
                    total_apparel = 0
                    total_accessories = 0
                    
                    # Buscar TOTAL HEADWEAR
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'HEADWEAR' and 
                            tabla_col[2] == 'Cantidad'):
                            total_headwear = tabla[tabla_col].iloc[-1]
                            break
                    
                    # Buscar TOTAL APPAREL
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'APPAREL' and 
                            tabla_col[2] == 'Cantidad'):
                            total_apparel = tabla[tabla_col].iloc[-1]
                            break
                    
                    # Buscar ACCESSORIES Cantidad
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'ACCESSORIES' and 
                            tabla_col[1] == 'Accessories' and 
                            tabla_col[2] == 'Cantidad'):
                            total_accessories = tabla[tabla_col].iloc[-1]
                            break
                    
                    valor = total_headwear + total_apparel + total_accessories
                except:
                    valor = 0
            elif col == 'TOTAL USD':
                # Para TOTAL USD, usar la columna consolidada 'TOTAL USD' que ya existe
                try:
                    # Buscar la columna consolidada TOTAL USD
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTALES' and 
                            tabla_col[1] == 'RESUMEN' and 
                            tabla_col[2] == 'TOTAL USD'):
                            valor = tabla[tabla_col].iloc[-1]
                            break
                    else:
                        valor = 0
                except:
                    valor = 0
            else:
                # Para las demás métricas (TOTAL HEADWEAR, TOTAL APPAREL)
                try:
                    for tabla_col in tabla.columns:
                        if col == 'TOTAL HEADWEAR':
                            if (len(tabla_col) == 3 and 
                                tabla_col[0] == 'TOTAL' and 
                                tabla_col[1] == 'HEADWEAR' and 
                                tabla_col[2] == 'Cantidad'):
                                valor = tabla[tabla_col].iloc[-1]
                                break
                        elif col == 'TOTAL APPAREL':
                            if (len(tabla_col) == 3 and 
                                tabla_col[0] == 'TOTAL' and 
                                tabla_col[1] == 'APPAREL' and 
                                tabla_col[2] == 'Cantidad'):
                                valor = tabla[tabla_col].iloc[-1]
                                break
                except:
                    valor = 0
            
            # Formato según tipo de métrica
            if col == 'TOTAL USD':
                descripcion = "USD en ventas"
                valor_formato = f"${valor:,.2f}" if valor else "$0.00"
            else:
                descripcion = "unidades vendidas"
                valor_entero = int(valor) if valor else 0
                valor_formato = f"{valor_entero:,}"
            
            st.markdown(f"""
            <div class="metric-card" style="border-left: 4px solid {color};">
                <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 0.5rem;">
                    <span style="font-size: 2rem; color: #000000;">{emoji}</span>
                    <span style="color: {color}; font-weight: 600; font-size: 1.1rem;">{nombre.upper()}</span>
                </div>
                <div style="font-size: 2rem; font-weight: 700; color: #374151; margin-bottom: 0.25rem;">
                    {valor_formato}
                </div>
                <div style="color: #6b7280; font-size: 0.85rem;">
                    {descripcion}
                </div>
            </div>
            """, unsafe_allow_html=True)

def mostrar_tabla_solo_ventas_el_salvador(tabla):
    """Muestra la tabla consolidada para solo-ventas de El Salvador (sin capacidades ni % cumplimiento)"""
    if tabla is None:
        return
    
    logger.info("Mostrando tabla solo-ventas para El Salvador")
    
    professional_design.create_section_header(
        "Tabla Consolidada - El Salvador (Solo Ventas)", 
        "Detalle completo de cantidades vendidas por bodega y categoría",
        "📊"
    )
    
    # Formatear números con separadores de miles
    tabla_formateada = tabla.copy()
    for col in tabla_formateada.columns:
        # Formato para columnas según tipo
        if len(col) == 3 and col[2] not in ['Bodega']:
            if col[2] == 'TOTAL USD':
                # Formato para columnas USD sin símbolo de dólar
                tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{float(x):,.2f}" if pd.notnull(x) else "0.00")
            else:
                # Formato para columnas de cantidades
                tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) else "0")
        # También formatear columnas de totales que no tienen MultiIndex de 3 niveles
        elif col == 'TOTAL USD':
            tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{float(x):,.2f}" if pd.notnull(x) else "0.00")
    
    # Convertir tabla a HTML con celdas combinadas en MultiIndex
    def crear_tabla_html_solo_ventas(df):
        # Obtener información del MultiIndex
        liga_counts = {}
        subcategoria_counts = {}
        
        # Contar cuántas columnas tiene cada liga y subcategoría
        for col in df.columns:
            if len(col) == 3:  # (Liga, Subcategoría, Tipo)
                liga, subcategoria, tipo = col
                
                # Solo contar las columnas que no son INFO ni Bodega
                if liga != 'INFO' and tipo != 'Bodega':
                    # Contar por liga
                    if liga not in liga_counts:
                        liga_counts[liga] = 0
                    liga_counts[liga] += 1
                    
                    # Contar por combinación liga-subcategoría
                    key = (liga, subcategoria)
                    if key not in subcategoria_counts:
                        subcategoria_counts[key] = 0
                    subcategoria_counts[key] += 1
        
        # Crear HTML de la tabla
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 7px; width: 100%; table-layout: fixed;">'
        
        # Fila 1: Ligas (con colspan)
        html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
        html += '<td rowspan="3" style="border: 1px solid #ddd; padding: 2px; vertical-align: middle; font-size: 7px; width: 50px;">Bodega</td>'
        
        for liga, count in liga_counts.items():
            if liga != 'INFO':  # Skip INFO column
                html += f'<td colspan="{count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{liga}</td>'
        
        html += '</tr>'
        
        # Fila 2: Subcategorías (con colspan)
        html += '<tr style="background-color: #f0f0f0; font-weight: bold;">'
        
        processed_subcategorias = set()
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO':
                liga, subcategoria, tipo = col
                key = (liga, subcategoria)
                
                # Solo procesar cada subcategoría una vez
                if key not in processed_subcategorias:
                    processed_subcategorias.add(key)
                    sub_count = subcategoria_counts.get(key, 1)
                    html += f'<td colspan="{sub_count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{subcategoria}</td>'
        
        html += '</tr>'
        
        # Fila 3: Tipos (Cantidad para reflejar el origen de los datos)
        html += '<tr style="background-color: #e8e8e8; font-weight: bold;">'
        
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO' and col[2] != 'Bodega':
                liga, subcategoria, tipo = col
                # Mostrar etiquetas apropiadas
                if tipo == "Cantidad":
                    tipo_display = "Cantidad"
                elif tipo == "TOTAL USD":
                    tipo_display = "TOTAL USD"
                else:
                    tipo_display = tipo
                html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; width: 30px;">{tipo_display}</td>'
        
        html += '</tr>'
        
        # Filas de datos
        for idx, row in df.iterrows():
            if idx == 'TOTAL':  # Fila TOTAL (verificar por nombre del índice)
                html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
            else:
                html += '<tr>'
            
            # Primero agregar la columna Bodega
            for col in df.columns:
                if len(col) == 3 and col[2] == 'Bodega':
                    value = row[col]
                    html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; text-align: left;">{value}</td>'
                    break
            
            # Luego agregar las demás columnas (cantidades)
            for col in df.columns:
                if len(col) == 3 and col[0] != 'INFO' and col[2] != 'Bodega':
                    value = row[col]
                    html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{value}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        return html
    
    # CSS para tabla con ancho igual a títulos y scroll interno
    st.markdown("""
    <style>
        .tabla-solo-ventas-el-salvador {
            overflow-x: auto !important;
            overflow-y: auto !important;
            width: 100% !important;
            max-width: 100% !important;
            max-height: 500px !important;
            border: 1px solid #ddd !important;
            background: white !important;
            margin: 0 !important;
            border-radius: 8px !important;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1) !important;
        }
        
        .tabla-solo-ventas-el-salvador table {
            font-size: 7px !important;
            border-collapse: collapse !important;
            width: max-content !important;
            min-width: 100% !important;
        }
        
        .tabla-solo-ventas-el-salvador td, .tabla-solo-ventas-el-salvador th {
            padding: 4px 6px !important;
            border: 1px solid #ddd !important;
            font-size: 7px !important;
            white-space: nowrap !important;
            text-align: center !important;
            min-width: 70px !important;
            max-width: 70px !important;
        }
        
        .tabla-solo-ventas-el-salvador td:first-child, .tabla-solo-ventas-el-salvador th:first-child {
            min-width: 100px !important;
            max-width: 100px !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    
    st.markdown('<div class="tabla-solo-ventas-el-salvador">', unsafe_allow_html=True)
    tabla_html = crear_tabla_html_solo_ventas(tabla_formateada)
    st.markdown(tabla_html, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Sección de exportación para El Salvador solo-ventas
    professional_design.create_section_header(
        "Exportar Reporte - El Salvador (Solo Ventas)", 
        "Generar archivo Excel con formato profesional",
        "SV"
    )
    
    col1, col2 = st.columns([3, 2])
    
    with col1:
        archivo_default = "EL_SALVADOR_SOLO_VENTAS.csv"
        session_key = 'archivo_el_salvador_solo_ventas_name'
        export_key = "nombre_sv_solo_ventas_export"
        
        archivo_nombre = st.session_state.get(session_key, archivo_default)
        nombre_archivo = st.text_input("📝 Nombre del archivo origen", archivo_nombre, key=export_key)
    
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)  # Espaciado
        button_key = "excel_sv_solo_ventas_export"
        if st.button("🚀 Generar Excel El Salvador (Solo Ventas)", key=button_key, use_container_width=True):
            exportar_excel_consolidado(tabla, nombre_archivo, "El Salvador")
    
    # Mostrar métricas resumidas adaptadas para cantidades
    professional_design.create_section_header(
        "Métricas de Ventas - El Salvador", 
        "Resumen ejecutivo de cantidades vendidas por categoría",
        "📈"
    )
    
    # Definir métricas para cantidades
    cols = st.columns(5)
    metricas = [
        ('TOTAL HEADWEAR', 'Headwear Total', "🧢", "#000000"),
        ('TOTAL APPAREL', 'Apparel Total', "👕", "#000000"),
        ('ACCESSORIES', 'Accessories Total', "🧦", "#000000"),
        ('TOTAL CANTIDADES', 'Total Cantidades Vendidas', "📊", "#000000"),
        ('TOTAL USD', 'Total USD Vendidos', "💰", "#000000")
    ]
    
    for i, (col, nombre, emoji, color) in enumerate(metricas):
        with cols[i]:
            # Buscar la columna en la estructura de 3 niveles (totales)
            valor = 0
            if col == 'ACCESSORIES':
                # Para ACCESSORIES, buscar en la estructura específica
                try:
                    # Buscar columna con estructura (ACCESSORIES, Accessories, Cantidad)
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'ACCESSORIES' and 
                            tabla_col[1] == 'Accessories' and 
                            tabla_col[2] == 'Cantidad'):
                            # Tomar solo la fila TOTAL (última fila)
                            valor = tabla[tabla_col].iloc[-1]
                            break
                except:
                    valor = 0
            elif col == 'TOTAL CANTIDADES':
                # Para TOTAL CANTIDADES, sumar TOTAL HEADWEAR + TOTAL APPAREL + ACCESSORIES
                try:
                    total_headwear = 0
                    total_apparel = 0
                    total_accessories = 0
                    
                    # Buscar TOTAL HEADWEAR
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'HEADWEAR' and 
                            tabla_col[2] == 'Cantidad'):
                            total_headwear = tabla[tabla_col].iloc[-1]
                            break
                    
                    # Buscar TOTAL APPAREL
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'APPAREL' and 
                            tabla_col[2] == 'Cantidad'):
                            total_apparel = tabla[tabla_col].iloc[-1]
                            break
                    
                    # Buscar ACCESSORIES Cantidad
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'ACCESSORIES' and 
                            tabla_col[1] == 'Accessories' and 
                            tabla_col[2] == 'Cantidad'):
                            total_accessories = tabla[tabla_col].iloc[-1]
                            break
                    
                    valor = total_headwear + total_apparel + total_accessories
                except:
                    valor = 0
            elif col == 'TOTAL USD':
                # Para TOTAL USD, usar la columna consolidada 'TOTAL USD' que ya existe
                try:
                    # Buscar la columna consolidada TOTAL USD
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTALES' and 
                            tabla_col[1] == 'RESUMEN' and 
                            tabla_col[2] == 'TOTAL USD'):
                            valor = tabla[tabla_col].iloc[-1]
                            break
                    else:
                        valor = 0
                except:
                    valor = 0
            else:
                # Para las demás métricas (TOTAL HEADWEAR, TOTAL APPAREL)
                try:
                    for tabla_col in tabla.columns:
                        if col == 'TOTAL HEADWEAR':
                            if (len(tabla_col) == 3 and 
                                tabla_col[0] == 'TOTAL' and 
                                tabla_col[1] == 'HEADWEAR' and 
                                tabla_col[2] == 'Cantidad'):
                                valor = tabla[tabla_col].iloc[-1]
                                break
                        elif col == 'TOTAL APPAREL':
                            if (len(tabla_col) == 3 and 
                                tabla_col[0] == 'TOTAL' and 
                                tabla_col[1] == 'APPAREL' and 
                                tabla_col[2] == 'Cantidad'):
                                valor = tabla[tabla_col].iloc[-1]
                                break
                except:
                    valor = 0
            
            # Formato según tipo de métrica
            if col == 'TOTAL USD':
                descripcion = "USD en ventas"
                valor_formato = f"${valor:,.2f}" if valor else "$0.00"
            else:
                descripcion = "unidades vendidas"
                valor_entero = int(valor) if valor else 0
                valor_formato = f"{valor_entero:,}"
            
            st.markdown(f"""
            <div class="metric-card" style="border-left: 4px solid {color};">
                <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 0.5rem;">
                    <span style="font-size: 2rem; color: #000000;">{emoji}</span>
                    <span style="color: {color}; font-weight: 600; font-size: 1.1rem;">{nombre.upper()}</span>
                </div>
                <div style="font-size: 2rem; font-weight: 700; color: #374151; margin-bottom: 0.25rem;">
                    {valor_formato}
                </div>
                <div style="color: #6b7280; font-size: 0.85rem;">
                    {descripcion}
                </div>
            </div>
            """, unsafe_allow_html=True)

def mostrar_tabla_solo_ventas_honduras(tabla):
    """Muestra la tabla consolidada para solo-ventas de Honduras (sin capacidades ni % cumplimiento)"""
    if tabla is None:
        return
    
    logger.info("Mostrando tabla solo-ventas para Honduras")
    
    professional_design.create_section_header(
        "Tabla Consolidada - Honduras (Solo Ventas)", 
        "Detalle completo de cantidades vendidas por bodega y categoría",
        "📊"
    )
    
    # Formatear números con separadores de miles
    tabla_formateada = tabla.copy()
    for col in tabla_formateada.columns:
        # Formato para columnas según tipo
        if len(col) == 3 and col[2] not in ['Bodega']:
            if col[2] == 'TOTAL USD':
                # Formato para columnas USD sin símbolo de dólar
                tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{float(x):,.2f}" if pd.notnull(x) else "0.00")
            else:
                # Formato para columnas de cantidades
                tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) else "0")
        # También formatear columnas de totales que no tienen MultiIndex de 3 niveles
        elif col == 'TOTAL USD':
            tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{float(x):,.2f}" if pd.notnull(x) else "0.00")
    
    # Convertir tabla a HTML con celdas combinadas en MultiIndex
    def crear_tabla_html_solo_ventas(df):
        # Obtener información del MultiIndex
        liga_counts = {}
        subcategoria_counts = {}
        
        # Contar cuántas columnas tiene cada liga y subcategoría
        for col in df.columns:
            if len(col) == 3:  # (Liga, Subcategoría, Tipo)
                liga, subcategoria, tipo = col
                
                # Solo contar las columnas que no son INFO ni Bodega
                if liga != 'INFO' and tipo != 'Bodega':
                    # Contar por liga
                    if liga not in liga_counts:
                        liga_counts[liga] = 0
                    liga_counts[liga] += 1
                    
                    # Contar por combinación liga-subcategoría
                    key = (liga, subcategoria)
                    if key not in subcategoria_counts:
                        subcategoria_counts[key] = 0
                    subcategoria_counts[key] += 1
        
        # Crear HTML de la tabla
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 7px; width: 100%; table-layout: fixed;">'
        
        # Fila 1: Ligas (con colspan)
        html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
        html += '<td rowspan="3" style="border: 1px solid #ddd; padding: 2px; vertical-align: middle; font-size: 7px; width: 50px;">Bodega</td>'
        
        for liga, count in liga_counts.items():
            if liga != 'INFO':  # Skip INFO column
                html += f'<td colspan="{count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{liga}</td>'
        
        html += '</tr>'
        
        # Fila 2: Subcategorías (con colspan)
        html += '<tr style="background-color: #f0f0f0; font-weight: bold;">'
        
        processed_subcategorias = set()
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO':
                liga, subcategoria, tipo = col
                key = (liga, subcategoria)
                
                # Solo procesar cada subcategoría una vez
                if key not in processed_subcategorias:
                    processed_subcategorias.add(key)
                    sub_count = subcategoria_counts.get(key, 1)
                    html += f'<td colspan="{sub_count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{subcategoria}</td>'
        
        html += '</tr>'
        
        # Fila 3: Tipos (Cantidad para reflejar el origen de los datos)
        html += '<tr style="background-color: #e8e8e8; font-weight: bold;">'
        
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO' and col[2] != 'Bodega':
                liga, subcategoria, tipo = col
                # Mostrar etiquetas apropiadas
                if tipo == "Cantidad":
                    tipo_display = "Cantidad"
                elif tipo == "TOTAL USD":
                    tipo_display = "TOTAL USD"
                else:
                    tipo_display = tipo
                html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; width: 30px;">{tipo_display}</td>'
        
        html += '</tr>'
        
        # Filas de datos
        for idx, row in df.iterrows():
            if idx == 'TOTAL':  # Fila TOTAL (verificar por nombre del índice)
                html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
            else:
                html += '<tr>'
            
            # Primero agregar la columna Bodega
            for col in df.columns:
                if len(col) == 3 and col[2] == 'Bodega':
                    value = row[col]
                    html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; text-align: left;">{value}</td>'
                    break
            
            # Luego agregar las demás columnas (cantidades)
            for col in df.columns:
                if len(col) == 3 and col[0] != 'INFO' and col[2] != 'Bodega':
                    value = row[col]
                    html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{value}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        return html
    
    # CSS para tabla con ancho igual a títulos y scroll interno
    st.markdown("""
    <style>
        .tabla-solo-ventas-honduras {
            overflow-x: auto !important;
            overflow-y: auto !important;
            width: 100% !important;
            max-width: 100% !important;
            max-height: 500px !important;
            border: 1px solid #ddd !important;
            background: white !important;
            margin: 0 !important;
            border-radius: 8px !important;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1) !important;
        }
        
        .tabla-solo-ventas-honduras table {
            font-size: 7px !important;
            border-collapse: collapse !important;
            width: max-content !important;
            min-width: 100% !important;
        }
        
        .tabla-solo-ventas-honduras td, .tabla-solo-ventas-honduras th {
            padding: 4px 6px !important;
            border: 1px solid #ddd !important;
            font-size: 7px !important;
            white-space: nowrap !important;
            text-align: center !important;
            min-width: 70px !important;
            max-width: 70px !important;
        }
        
        .tabla-solo-ventas-honduras td:first-child, .tabla-solo-ventas-honduras th:first-child {
            min-width: 100px !important;
            max-width: 100px !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    
    st.markdown('<div class="tabla-solo-ventas-honduras">', unsafe_allow_html=True)
    tabla_html = crear_tabla_html_solo_ventas(tabla_formateada)
    st.markdown(tabla_html, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Sección de exportación para Honduras solo-ventas
    professional_design.create_section_header(
        "Exportar Reporte - Honduras (Solo Ventas)", 
        "Generar archivo Excel con formato profesional",
        "HN"
    )
    
    col1, col2 = st.columns([3, 2])
    
    with col1:
        archivo_default = "HONDURAS_SOLO_VENTAS.csv"
        session_key = 'archivo_honduras_solo_ventas_name'
        export_key = "nombre_hn_solo_ventas_export"
        
        archivo_nombre = st.session_state.get(session_key, archivo_default)
        nombre_archivo = st.text_input("📝 Nombre del archivo origen", archivo_nombre, key=export_key)
    
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)  # Espaciado
        button_key = "excel_hn_solo_ventas_export"
        if st.button("🚀 Generar Excel Honduras (Solo Ventas)", key=button_key, use_container_width=True):
            exportar_excel_consolidado(tabla, nombre_archivo, "Honduras")
    
    # Mostrar métricas resumidas adaptadas para cantidades
    professional_design.create_section_header(
        "Métricas de Ventas - Honduras", 
        "Resumen ejecutivo de cantidades vendidas por categoría",
        "📈"
    )
    
    # Definir métricas para cantidades
    cols = st.columns(5)
    metricas = [
        ('TOTAL HEADWEAR', 'Headwear Total', "🧢", "#000000"),
        ('TOTAL APPAREL', 'Apparel Total', "👕", "#000000"),
        ('ACCESSORIES', 'Accessories Total', "🧦", "#000000"),
        ('TOTAL CANTIDADES', 'Total Cantidades Vendidas', "📊", "#000000"),
        ('TOTAL USD', 'Total USD Vendidos', "💰", "#000000")
    ]
    
    for i, (col, nombre, emoji, color) in enumerate(metricas):
        with cols[i]:
            # Buscar la columna en la estructura de 3 niveles (totales)
            valor = 0
            if col == 'ACCESSORIES':
                # Para ACCESSORIES, buscar en la estructura específica
                try:
                    # Buscar columna con estructura (ACCESSORIES, Accessories, Cantidad)
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'ACCESSORIES' and 
                            tabla_col[1] == 'Accessories' and 
                            tabla_col[2] == 'Cantidad'):
                            # Tomar solo la fila TOTAL (última fila)
                            valor = tabla[tabla_col].iloc[-1]
                            break
                except:
                    valor = 0
            elif col == 'TOTAL CANTIDADES':
                # Para TOTAL CANTIDADES, sumar TOTAL HEADWEAR + TOTAL APPAREL + ACCESSORIES
                try:
                    total_headwear = 0
                    total_apparel = 0
                    total_accessories = 0
                    
                    # Buscar TOTAL HEADWEAR
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'HEADWEAR' and 
                            tabla_col[2] == 'Cantidad'):
                            total_headwear = tabla[tabla_col].iloc[-1]
                            break
                    
                    # Buscar TOTAL APPAREL
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'APPAREL' and 
                            tabla_col[2] == 'Cantidad'):
                            total_apparel = tabla[tabla_col].iloc[-1]
                            break
                    
                    # Buscar ACCESSORIES Cantidad
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'ACCESSORIES' and 
                            tabla_col[1] == 'Accessories' and 
                            tabla_col[2] == 'Cantidad'):
                            total_accessories = tabla[tabla_col].iloc[-1]
                            break
                    
                    valor = total_headwear + total_apparel + total_accessories
                except:
                    valor = 0
            elif col == 'TOTAL USD':
                # Para TOTAL USD, usar la columna consolidada 'TOTAL USD' que ya existe
                try:
                    # Buscar la columna consolidada TOTAL USD
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTALES' and 
                            tabla_col[1] == 'RESUMEN' and 
                            tabla_col[2] == 'TOTAL USD'):
                            valor = tabla[tabla_col].iloc[-1]
                            break
                    else:
                        valor = 0
                except:
                    valor = 0
            else:
                # Para las demás métricas (TOTAL HEADWEAR, TOTAL APPAREL)
                try:
                    for tabla_col in tabla.columns:
                        if col == 'TOTAL HEADWEAR':
                            if (len(tabla_col) == 3 and 
                                tabla_col[0] == 'TOTAL' and 
                                tabla_col[1] == 'HEADWEAR' and 
                                tabla_col[2] == 'Cantidad'):
                                valor = tabla[tabla_col].iloc[-1]
                                break
                        elif col == 'TOTAL APPAREL':
                            if (len(tabla_col) == 3 and 
                                tabla_col[0] == 'TOTAL' and 
                                tabla_col[1] == 'APPAREL' and 
                                tabla_col[2] == 'Cantidad'):
                                valor = tabla[tabla_col].iloc[-1]
                                break
                except:
                    valor = 0
            
            # Formato según tipo de métrica
            if col == 'TOTAL USD':
                descripcion = "USD en ventas"
                valor_formato = f"${valor:,.2f}" if valor else "$0.00"
            else:
                descripcion = "unidades vendidas"
                valor_entero = int(valor) if valor else 0
                valor_formato = f"{valor_entero:,}"
            
            st.markdown(f"""
            <div class="metric-card" style="border-left: 4px solid {color};">
                <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 0.5rem;">
                    <span style="font-size: 2rem; color: #000000;">{emoji}</span>
                    <span style="color: {color}; font-weight: 600; font-size: 1.1rem;">{nombre.upper()}</span>
                </div>
                <div style="font-size: 2rem; font-weight: 700; color: #374151; margin-bottom: 0.25rem;">
                    {valor_formato}
                </div>
                <div style="color: #6b7280; font-size: 0.85rem;">
                    {descripcion}
                </div>
            </div>
            """, unsafe_allow_html=True)

def mostrar_tabla_solo_ventas_costa_rica(tabla):
    """Muestra la tabla consolidada para solo-ventas de Costa Rica (sin capacidades ni % cumplimiento)"""
    if tabla is None:
        return
    
    logger.info("Mostrando tabla solo-ventas para Costa Rica")
    
    professional_design.create_section_header(
        "Tabla Consolidada - Costa Rica (Solo Ventas)", 
        "Detalle completo de cantidades vendidas por bodega y categoría",
        "📊"
    )
    
    # Formatear números con separadores de miles
    tabla_formateada = tabla.copy()
    for col in tabla_formateada.columns:
        # Formato para columnas según tipo
        if len(col) == 3 and col[2] not in ['Bodega']:
            if col[2] == 'TOTAL USD':
                # Formato para columnas USD sin símbolo de dólar
                tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{float(x):,.2f}" if pd.notnull(x) else "0.00")
            else:
                # Formato para columnas de cantidades
                tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) else "0")
        # También formatear columnas de totales que no tienen MultiIndex de 3 niveles
        elif col == 'TOTAL USD':
            tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{float(x):,.2f}" if pd.notnull(x) else "0.00")
    
    # Convertir tabla a HTML con celdas combinadas en MultiIndex
    def crear_tabla_html_solo_ventas(df):
        # Obtener información del MultiIndex
        liga_counts = {}
        subcategoria_counts = {}
        
        # Contar cuántas columnas tiene cada liga y subcategoría
        for col in df.columns:
            if len(col) == 3:  # (Liga, Subcategoría, Tipo)
                liga, subcategoria, tipo = col
                
                # Solo contar las columnas que no son INFO ni Bodega
                if liga != 'INFO' and tipo != 'Bodega':
                    # Contar por liga
                    if liga not in liga_counts:
                        liga_counts[liga] = 0
                    liga_counts[liga] += 1
                    
                    # Contar por combinación liga-subcategoría
                    key = (liga, subcategoria)
                    if key not in subcategoria_counts:
                        subcategoria_counts[key] = 0
                    subcategoria_counts[key] += 1
        
        # Crear HTML de la tabla
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 7px; width: 100%; table-layout: fixed;">'
        
        # Fila 1: Ligas (con colspan)
        html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
        html += '<td rowspan="3" style="border: 1px solid #ddd; padding: 2px; vertical-align: middle; font-size: 7px; width: 50px;">Bodega</td>'
        
        for liga, count in liga_counts.items():
            if liga != 'INFO':  # Skip INFO column
                html += f'<td colspan="{count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{liga}</td>'
        
        html += '</tr>'
        
        # Fila 2: Subcategorías (con colspan)
        html += '<tr style="background-color: #f0f0f0; font-weight: bold;">'
        
        processed_subcategorias = set()
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO':
                liga, subcategoria, tipo = col
                key = (liga, subcategoria)
                
                # Solo procesar cada subcategoría una vez
                if key not in processed_subcategorias:
                    processed_subcategorias.add(key)
                    sub_count = subcategoria_counts.get(key, 1)
                    html += f'<td colspan="{sub_count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{subcategoria}</td>'
        
        html += '</tr>'
        
        # Fila 3: Tipos (Cantidad para reflejar el origen de los datos)
        html += '<tr style="background-color: #e8e8e8; font-weight: bold;">'
        
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO' and col[2] != 'Bodega':
                liga, subcategoria, tipo = col
                # Mostrar etiquetas apropiadas
                if tipo == "Cantidad":
                    tipo_display = "Cantidad"
                elif tipo == "TOTAL USD":
                    tipo_display = "TOTAL USD"
                else:
                    tipo_display = tipo
                html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; width: 30px;">{tipo_display}</td>'
        
        html += '</tr>'
        
        # Filas de datos
        for idx, row in df.iterrows():
            if idx == 'TOTAL':  # Fila TOTAL (verificar por nombre del índice)
                html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
            else:
                html += '<tr>'
            
            # Primero agregar la columna Bodega
            for col in df.columns:
                if len(col) == 3 and col[2] == 'Bodega':
                    value = row[col]
                    html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; text-align: left;">{value}</td>'
                    break
            
            # Luego agregar las demás columnas (cantidades)
            for col in df.columns:
                if len(col) == 3 and col[0] != 'INFO' and col[2] != 'Bodega':
                    value = row[col]
                    html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{value}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        return html
    
    # CSS para tabla con ancho igual a títulos y scroll interno
    st.markdown("""
    <style>
        .tabla-solo-ventas-costa-rica {
            overflow-x: auto !important;
            overflow-y: auto !important;
            width: 100% !important;
            max-width: 100% !important;
            max-height: 500px !important;
            border: 1px solid #ddd !important;
            background: white !important;
            margin: 0 !important;
            border-radius: 8px !important;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1) !important;
        }
        
        .tabla-solo-ventas-costa-rica table {
            font-size: 7px !important;
            border-collapse: collapse !important;
            width: max-content !important;
            min-width: 100% !important;
        }
        
        .tabla-solo-ventas-costa-rica td, .tabla-solo-ventas-costa-rica th {
            padding: 4px 6px !important;
            border: 1px solid #ddd !important;
            text-align: center !important;
            white-space: nowrap !important;
            overflow: hidden !important;
            text-overflow: ellipsis !important;
            font-size: 7px !important;
        }
        
        .tabla-solo-ventas-costa-rica td:first-child {
            text-align: left !important;
            position: sticky !important;
            left: 0 !important;
            background-color: white !important;
            z-index: 1 !important;
            min-width: 100px !important;
        }
        
        .tabla-solo-ventas-costa-rica tr:nth-child(odd) {
            background-color: #f8f9fa !important;
        }
        
        .tabla-solo-ventas-costa-rica tr:nth-child(even) {
            background-color: white !important;
        }
        
        .tabla-solo-ventas-costa-rica tr:hover {
            background-color: #e3f2fd !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Mostrar la tabla
    st.markdown('<div class="tabla-solo-ventas-costa-rica">', unsafe_allow_html=True)
    tabla_html = crear_tabla_html_solo_ventas(tabla_formateada)
    st.markdown(tabla_html, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Mostrar métricas resumidas adaptadas para cantidades
    professional_design.create_section_header(
        "Métricas de Ventas - Costa Rica", 
        "Resumen ejecutivo de cantidades vendidas por categoría",
        "📈"
    )
    
    # Definir métricas para cantidades
    cols = st.columns(5)
    metricas = [
        ('TOTAL HEADWEAR', 'Headwear Total', "🧢", "#000000"),
        ('TOTAL APPAREL', 'Apparel Total', "👕", "#000000"),
        ('ACCESSORIES', 'Accessories Total', "🧦", "#000000"),
        ('TOTAL CANTIDADES', 'Total Cantidades Vendidas', "📊", "#000000"),
        ('TOTAL USD', 'Total USD Vendidos', "💰", "#000000")
    ]
    
    for i, (col, nombre, emoji, color) in enumerate(metricas):
        with cols[i]:
            # Buscar la columna en la estructura de 3 niveles (totales)
            valor = 0
            if col == 'ACCESSORIES':
                # Para ACCESSORIES, buscar en la estructura específica
                try:
                    # Buscar columna con estructura (ACCESSORIES, Accessories, Cantidad)
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'ACCESSORIES' and 
                            tabla_col[1] == 'Accessories' and 
                            tabla_col[2] == 'Cantidad'):
                            # Tomar solo la fila TOTAL (última fila)
                            valor = tabla[tabla_col].iloc[-1]
                            break
                except:
                    valor = 0
            elif col == 'TOTAL CANTIDADES':
                # Para TOTAL CANTIDADES, sumar TOTAL HEADWEAR + TOTAL APPAREL + ACCESSORIES
                try:
                    total_headwear = 0
                    total_apparel = 0
                    total_accessories = 0
                    
                    # Buscar TOTAL HEADWEAR
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'HEADWEAR' and 
                            tabla_col[2] == 'Cantidad'):
                            total_headwear = tabla[tabla_col].iloc[-1]
                            break
                    
                    # Buscar TOTAL APPAREL
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'APPAREL' and 
                            tabla_col[2] == 'Cantidad'):
                            total_apparel = tabla[tabla_col].iloc[-1]
                            break
                    
                    # Buscar ACCESSORIES Cantidad
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'ACCESSORIES' and 
                            tabla_col[1] == 'Accessories' and 
                            tabla_col[2] == 'Cantidad'):
                            total_accessories = tabla[tabla_col].iloc[-1]
                            break
                    
                    valor = total_headwear + total_apparel + total_accessories
                except:
                    valor = 0
            elif col == 'TOTAL USD':
                # Para TOTAL USD, usar la columna consolidada 'TOTAL USD' que ya existe
                try:
                    # Buscar la columna consolidada TOTAL USD
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'USD' and 
                            tabla_col[2] == 'TOTAL USD'):
                            valor = tabla[tabla_col].iloc[-1]
                            break
                    else:
                        valor = 0
                except:
                    valor = 0
            else:
                # Para las demás métricas (TOTAL HEADWEAR, TOTAL APPAREL)
                try:
                    for tabla_col in tabla.columns:
                        if col == 'TOTAL HEADWEAR':
                            if (len(tabla_col) == 3 and 
                                tabla_col[0] == 'TOTAL' and 
                                tabla_col[1] == 'HEADWEAR' and 
                                tabla_col[2] == 'Cantidad'):
                                valor = tabla[tabla_col].iloc[-1]
                                break
                        elif col == 'TOTAL APPAREL':
                            if (len(tabla_col) == 3 and 
                                tabla_col[0] == 'TOTAL' and 
                                tabla_col[1] == 'APPAREL' and 
                                tabla_col[2] == 'Cantidad'):
                                valor = tabla[tabla_col].iloc[-1]
                                break
                except:
                    valor = 0
            
            # Formato según tipo de métrica
            if col == 'TOTAL USD':
                descripcion = "USD en ventas"
                valor_formato = f"${valor:,.2f}" if valor else "$0.00"
            else:
                descripcion = "unidades vendidas"
                valor_entero = int(valor) if valor else 0
                valor_formato = f"{valor_entero:,}"
            
            st.markdown(f"""
            <div class="metric-card" style="border-left: 4px solid {color};">
                <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 0.5rem;">
                    <span style="font-size: 2rem; color: #000000;">{emoji}</span>
                    <span style="color: {color}; font-weight: 600; font-size: 1.1rem;">{nombre.upper()}</span>
                </div>
                <div style="font-size: 2rem; font-weight: 700; color: #374151; margin-bottom: 0.25rem;">
                    {valor_formato}
                </div>
                <div style="color: #6b7280; font-size: 0.85rem;">
                    {descripcion}
                </div>
            </div>
            """, unsafe_allow_html=True)

def mostrar_tabla_solo_ventas_panama(tabla):
    """Muestra la tabla consolidada para solo-ventas de Panama (sin capacidades ni % cumplimiento)"""
    if tabla is None:
        return
    
    logger.info("Mostrando tabla solo-ventas para Panama")
    
    professional_design.create_section_header(
        "Tabla Consolidada - Panamá (Solo Ventas)", 
        "Detalle completo de cantidades vendidas por bodega y categoría",
        "📊"
    )
    
    # Formatear números con separadores de miles
    tabla_formateada = tabla.copy()
    for col in tabla_formateada.columns:
        # Formato para columnas según tipo
        if len(col) == 3 and col[2] not in ['Bodega']:
            if col[2] == 'TOTAL USD':
                # Formato para columnas USD sin símbolo de dólar
                tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{float(x):,.2f}" if pd.notnull(x) else "0.00")
            else:
                # Formato para columnas de cantidades
                tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) else "0")
        # También formatear columnas de totales que no tienen MultiIndex de 3 niveles
        elif col == 'TOTAL USD':
            tabla_formateada[col] = tabla_formateada[col].apply(lambda x: f"{float(x):,.2f}" if pd.notnull(x) else "0.00")
    
    # Convertir tabla a HTML con celdas combinadas en MultiIndex
    def crear_tabla_html_solo_ventas(df):
        # Obtener información del MultiIndex
        liga_counts = {}
        subcategoria_counts = {}
        
        # Contar cuántas columnas tiene cada liga y subcategoría
        for col in df.columns:
            if len(col) == 3:  # (Liga, Subcategoría, Tipo)
                liga, subcategoria, tipo = col
                
                # Solo contar las columnas que no son INFO ni Bodega
                if liga != 'INFO' and tipo != 'Bodega':
                    # Contar por liga
                    if liga not in liga_counts:
                        liga_counts[liga] = 0
                    liga_counts[liga] += 1
                    
                    # Contar por combinación liga-subcategoría
                    key = (liga, subcategoria)
                    if key not in subcategoria_counts:
                        subcategoria_counts[key] = 0
                    subcategoria_counts[key] += 1
        
        # Crear HTML de la tabla
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 7px; width: 100%; table-layout: fixed;">'
        
        # Fila 1: Ligas (con colspan)
        html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
        html += '<td rowspan="3" style="border: 1px solid #ddd; padding: 2px; vertical-align: middle; font-size: 7px; width: 50px;">Bodega</td>'
        
        for liga, count in liga_counts.items():
            if liga != 'INFO':  # Skip INFO column
                html += f'<td colspan="{count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{liga}</td>'
        
        html += '</tr>'
        
        # Fila 2: Subcategorías (con colspan)
        html += '<tr style="background-color: #f0f0f0; font-weight: bold;">'
        
        processed_subcategorias = set()
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO':
                liga, subcategoria, tipo = col
                key = (liga, subcategoria)
                
                # Solo procesar cada subcategoría una vez
                if key not in processed_subcategorias:
                    processed_subcategorias.add(key)
                    sub_count = subcategoria_counts.get(key, 1)
                    html += f'<td colspan="{sub_count}" style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{subcategoria}</td>'
        
        html += '</tr>'
        
        # Fila 3: Tipos (Cantidad para reflejar el origen de los datos)
        html += '<tr style="background-color: #e8e8e8; font-weight: bold;">'
        
        for col in df.columns:
            if len(col) == 3 and col[0] != 'INFO' and col[2] != 'Bodega':
                liga, subcategoria, tipo = col
                # Mostrar etiquetas apropiadas
                if tipo == "Cantidad":
                    tipo_display = "Cantidad"
                elif tipo == "TOTAL USD":
                    tipo_display = "TOTAL USD"
                else:
                    tipo_display = tipo
                html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; width: 30px;">{tipo_display}</td>'
        
        html += '</tr>'
        
        # Filas de datos
        for idx, row in df.iterrows():
            if idx == 'TOTAL':  # Fila TOTAL (verificar por nombre del índice)
                html += '<tr style="background-color: #000000; color: white; font-weight: bold;">'
            else:
                html += '<tr>'
            
            # Primero agregar la columna Bodega
            for col in df.columns:
                if len(col) == 3 and col[2] == 'Bodega':
                    value = row[col]
                    html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px; text-align: left;">{value}</td>'
                    break
            
            # Luego agregar las demás columnas (cantidades)
            for col in df.columns:
                if len(col) == 3 and col[0] != 'INFO' and col[2] != 'Bodega':
                    value = row[col]
                    html += f'<td style="border: 1px solid #ddd; padding: 2px; font-size: 7px;">{value}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        return html
    
    # CSS para tabla con ancho igual a títulos y scroll interno
    st.markdown("""
    <style>
        .tabla-solo-ventas-panama {
            overflow-x: auto !important;
            overflow-y: auto !important;
            width: 100% !important;
            max-width: 100% !important;
            max-height: 500px !important;
            border: 1px solid #ddd !important;
            background: white !important;
            margin: 0 !important;
            border-radius: 8px !important;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1) !important;
        }
        
        .tabla-solo-ventas-panama table {
            font-size: 7px !important;
            border-collapse: collapse !important;
            width: max-content !important;
            min-width: 100% !important;
        }
        
        .tabla-solo-ventas-panama td, .tabla-solo-ventas-panama th {
            padding: 4px 6px !important;
            border: 1px solid #ddd !important;
            text-align: center !important;
            white-space: nowrap !important;
            overflow: hidden !important;
            text-overflow: ellipsis !important;
            font-size: 7px !important;
        }
        
        .tabla-solo-ventas-panama td:first-child {
            text-align: left !important;
            position: sticky !important;
            left: 0 !important;
            background-color: white !important;
            z-index: 1 !important;
            min-width: 100px !important;
        }
        
        .tabla-solo-ventas-panama tr:nth-child(odd) {
            background-color: #f8f9fa !important;
        }
        
        .tabla-solo-ventas-panama tr:nth-child(even) {
            background-color: white !important;
        }
        
        .tabla-solo-ventas-panama tr:hover {
            background-color: #e3f2fd !important;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Mostrar la tabla
    st.markdown('<div class="tabla-solo-ventas-panama">', unsafe_allow_html=True)
    tabla_html = crear_tabla_html_solo_ventas(tabla_formateada)
    st.markdown(tabla_html, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Mostrar métricas resumidas adaptadas para cantidades
    professional_design.create_section_header(
        "Métricas de Ventas - Panamá", 
        "Resumen ejecutivo de cantidades vendidas por categoría",
        "📈"
    )
    
    # Definir métricas para cantidades
    cols = st.columns(5)
    metricas = [
        ('TOTAL HEADWEAR', 'Headwear Total', "🧢", "#000000"),
        ('TOTAL APPAREL', 'Apparel Total', "👕", "#000000"),
        ('ACCESSORIES', 'Accessories Total', "🧦", "#000000"),
        ('TOTAL CANTIDADES', 'Total Cantidades Vendidas', "📊", "#000000"),
        ('TOTAL USD', 'Total USD Vendidos', "💰", "#000000")
    ]
    
    for i, (col, nombre, emoji, color) in enumerate(metricas):
        with cols[i]:
            # Buscar la columna en la estructura de 3 niveles (totales)
            valor = 0
            if col == 'ACCESSORIES':
                # Para ACCESSORIES, buscar en la estructura específica
                try:
                    # Buscar columna con estructura (ACCESSORIES, Accessories, Cantidad)
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'ACCESSORIES' and 
                            tabla_col[1] == 'Accessories' and 
                            tabla_col[2] == 'Cantidad'):
                            # Tomar solo la fila TOTAL (última fila)
                            valor = tabla[tabla_col].iloc[-1]
                            break
                except:
                    valor = 0
            elif col == 'TOTAL CANTIDADES':
                # Para TOTAL CANTIDADES, sumar TOTAL HEADWEAR + TOTAL APPAREL + ACCESSORIES
                try:
                    total_headwear = 0
                    total_apparel = 0
                    total_accessories = 0
                    
                    # Buscar TOTAL HEADWEAR
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'HEADWEAR' and 
                            tabla_col[2] == 'Cantidad'):
                            total_headwear = tabla[tabla_col].iloc[-1]
                            break
                    
                    # Buscar TOTAL APPAREL
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'APPAREL' and 
                            tabla_col[2] == 'Cantidad'):
                            total_apparel = tabla[tabla_col].iloc[-1]
                            break
                    
                    # Buscar ACCESSORIES Cantidad
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'ACCESSORIES' and 
                            tabla_col[1] == 'Accessories' and 
                            tabla_col[2] == 'Cantidad'):
                            total_accessories = tabla[tabla_col].iloc[-1]
                            break
                    
                    valor = total_headwear + total_apparel + total_accessories
                except:
                    valor = 0
            elif col == 'TOTAL USD':
                # Para TOTAL USD, usar la columna consolidada 'TOTAL USD' que ya existe
                try:
                    # Buscar la columna consolidada TOTAL USD
                    for tabla_col in tabla.columns:
                        if (len(tabla_col) == 3 and 
                            tabla_col[0] == 'TOTAL' and 
                            tabla_col[1] == 'USD' and 
                            tabla_col[2] == 'TOTAL USD'):
                            valor = tabla[tabla_col].iloc[-1]
                            break
                    else:
                        valor = 0
                except:
                    valor = 0
            else:
                # Para las demás métricas (TOTAL HEADWEAR, TOTAL APPAREL)
                try:
                    for tabla_col in tabla.columns:
                        if col == 'TOTAL HEADWEAR':
                            if (len(tabla_col) == 3 and 
                                tabla_col[0] == 'TOTAL' and 
                                tabla_col[1] == 'HEADWEAR' and 
                                tabla_col[2] == 'Cantidad'):
                                valor = tabla[tabla_col].iloc[-1]
                                break
                        elif col == 'TOTAL APPAREL':
                            if (len(tabla_col) == 3 and 
                                tabla_col[0] == 'TOTAL' and 
                                tabla_col[1] == 'APPAREL' and 
                                tabla_col[2] == 'Cantidad'):
                                valor = tabla[tabla_col].iloc[-1]
                                break
                except:
                    valor = 0
            
            # Formato según tipo de métrica
            if col == 'TOTAL USD':
                descripcion = "USD en ventas"
                valor_formato = f"${valor:,.2f}" if valor else "$0.00"
            else:
                descripcion = "unidades vendidas"
                valor_entero = int(valor) if valor else 0
                valor_formato = f"{valor_entero:,}"
            
            st.markdown(f"""
            <div class="metric-card" style="border-left: 4px solid {color};">
                <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 0.5rem;">
                    <span style="font-size: 2rem; color: #000000;">{emoji}</span>
                    <span style="color: {color}; font-weight: 600; font-size: 1.1rem;">{nombre.upper()}</span>
                </div>
                <div style="font-size: 2rem; font-weight: 700; color: #374151; margin-bottom: 0.25rem;">
                    {valor_formato}
                </div>
                <div style="color: #6b7280; font-size: 0.85rem;">
                    {descripcion}
                </div>
            </div>
            """, unsafe_allow_html=True)

def agregar_seccion_exportar_distribuciones(tablas_reales, pais, tiene_ventas):
    """Agrega la sección de exportación de distribuciones idéntica a la sección existente"""
    # Crear header de sección idéntico a la sección de exportación existente
    professional_design = ProfessionalDesign()
    if pais == "Guatemala":
        codigo_pais = "GT"
    elif pais == "Costa Rica":
        codigo_pais = "CR"
    elif pais == "Honduras":
        codigo_pais = "HN"
    else:  # El Salvador
        codigo_pais = "SV"
    professional_design.create_section_header(
        f"Exportar Distribuciones - {pais}", 
        "Generar archivo Excel con distribuciones por bodega",
        codigo_pais
    )
    
    col1, col2 = st.columns([3, 2])
    
    with col1:
        # Input idéntico al existente
        if pais == "Guatemala":
            archivo_default = "GUATEMALA.csv"
            session_key = 'archivo_guatemala_name'
            dist_key = "nombre_dist_export"
        elif pais == "Costa Rica":
            archivo_default = "COSTA_RICA.csv"
            session_key = 'archivo_costa_rica_name'
            dist_key = "nombre_dist_export_cr"
        elif pais == "Honduras":
            archivo_default = "HONDURAS.csv"
            session_key = 'archivo_honduras_name'
            dist_key = "nombre_dist_export_hn"
        elif pais == "PANAMA":
            archivo_default = "PANAMA.csv"
            session_key = 'archivo_panama_name'
            dist_key = "nombre_dist_export_pa"
        else:  # El Salvador
            archivo_default = "EL_SALVADOR.csv"
            session_key = 'archivo_el_salvador_name'
            dist_key = "nombre_dist_export_sv"
        
        archivo_nombre = st.session_state.get(session_key, archivo_default)
        nombre_archivo_dist = st.text_input("📝 Nombre del archivo origen", archivo_nombre, key=dist_key)
    
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)  # Espaciado
        if pais == "Guatemala":
            button_dist_key = "excel_dist_export"
        elif pais == "Costa Rica":
            button_dist_key = "excel_dist_export_cr"
        elif pais == "Honduras":
            button_dist_key = "excel_dist_export_hn"
        elif pais == "PANAMA":
            button_dist_key = "excel_dist_export_pa"
        else:  # El Salvador
            button_dist_key = "excel_dist_export_sv"
        if st.button("🚀 Generar Excel Distribuciones", key=button_dist_key, use_container_width=True):
            exportar_excel_distribuciones_reales(tablas_reales, pais, tiene_ventas)

def exportar_excel_distribuciones_reales(tablas_reales, pais, tiene_ventas):
    """Exporta las tablas reales de distribución tal como aparecen en Streamlit"""
    if not tablas_reales:
        st.warning(f"No hay tablas de distribución para exportar de {pais}")
        return
    
    try:
        logger.info(f"Iniciando exportación de distribuciones reales para {pais}")
        
        # Definir nombres dinámicos según el país
        if pais == "Guatemala":
            nombre_tiendas_secundarias = "Tiendas Departamentales"
            nombre_tiendas_principales = "Tiendas de Ciudad"
        elif pais == "Costa Rica":
            nombre_tiendas_secundarias = "Tiendas Departamentales"
            nombre_tiendas_principales = "Tiendas Franquicia"
        elif pais == "Honduras":
            nombre_tiendas_secundarias = "Tiendas Departamentales"
            nombre_tiendas_principales = "Tiendas Franquicia"
        else:
            nombre_tiendas_secundarias = "Tiendas Franquicia"
            nombre_tiendas_principales = "Tiendas de Ciudad"
        
        # Crear archivo Excel
        nombre_excel = f"distribucion_bodegas_{pais.lower().replace(' ', '_')}_{config.fecha_reporte}.xlsx"
        output = pd.ExcelWriter(nombre_excel, engine='openpyxl')
        
        if not tiene_ventas:
            # Solo hay stock - crear una pestaña con las 3 tablas
            sheet_name = "Distribución Stock"
            row_offset = 0
            
            # Escribir tabla de Tiendas Principales
            if 'df_principales' in tablas_reales and len(tablas_reales['df_principales']) > 0:
                # Agregar título
                titulo_principales = pd.DataFrame([[f'🏪 {nombre_tiendas_principales.upper()}']], columns=[''])
                titulo_principales.to_excel(output, sheet_name=sheet_name, startrow=row_offset, index=False, header=False)
                row_offset += 2
                
                tablas_reales['df_principales'].to_excel(output, sheet_name=sheet_name, startrow=row_offset, index=False)
                row_offset += len(tablas_reales['df_principales']) + 3
            
            # Escribir tabla de Outlets
            if 'df_outlets' in tablas_reales and len(tablas_reales['df_outlets']) > 0:
                titulo_outlets = pd.DataFrame([['🛒 OUTLETS']], columns=[''])
                titulo_outlets.to_excel(output, sheet_name=sheet_name, startrow=row_offset, index=False, header=False)
                row_offset += 2
                
                tablas_reales['df_outlets'].to_excel(output, sheet_name=sheet_name, startrow=row_offset, index=False)
                row_offset += len(tablas_reales['df_outlets']) + 3
            
            # Escribir tabla de Tiendas Departamentales
            if 'df_secundarias' in tablas_reales and len(tablas_reales['df_secundarias']) > 0:
                titulo_secundarias = pd.DataFrame([[f'🏬 {nombre_tiendas_secundarias.upper()}']], columns=[''])
                titulo_secundarias.to_excel(output, sheet_name=sheet_name, startrow=row_offset, index=False, header=False)
                row_offset += 2
                
                tablas_reales['df_secundarias'].to_excel(output, sheet_name=sheet_name, startrow=row_offset, index=False)
                row_offset += len(tablas_reales['df_secundarias']) + 3
            
            # Escribir tabla de Tienda Outlet (solo para El Salvador)
            if 'df_outlet_especial' in tablas_reales and len(tablas_reales['df_outlet_especial']) > 0:
                titulo_outlet_especial = pd.DataFrame([['🏪 TIENDA OUTLET']], columns=[''])
                titulo_outlet_especial.to_excel(output, sheet_name=sheet_name, startrow=row_offset, index=False, header=False)
                row_offset += 2
                
                tablas_reales['df_outlet_especial'].to_excel(output, sheet_name=sheet_name, startrow=row_offset, index=False)
        
        else:
            # Hay stock y ventas - crear 3 pestañas
            
            # PESTAÑA 1: Distribución Stock
            sheet_name_stock = "Distribución Stock"
            row_offset = 0
            
            if 'df_principales' in tablas_reales and len(tablas_reales['df_principales']) > 0:
                titulo = pd.DataFrame([[f'🏪 {nombre_tiendas_principales.upper()}']], columns=[''])
                titulo.to_excel(output, sheet_name=sheet_name_stock, startrow=row_offset, index=False, header=False)
                row_offset += 2
                tablas_reales['df_principales'].to_excel(output, sheet_name=sheet_name_stock, startrow=row_offset, index=False)
                row_offset += len(tablas_reales['df_principales']) + 3
            
            if 'df_outlets' in tablas_reales and len(tablas_reales['df_outlets']) > 0:
                titulo = pd.DataFrame([['🛒 OUTLETS']], columns=[''])
                titulo.to_excel(output, sheet_name=sheet_name_stock, startrow=row_offset, index=False, header=False)
                row_offset += 2
                tablas_reales['df_outlets'].to_excel(output, sheet_name=sheet_name_stock, startrow=row_offset, index=False)
                row_offset += len(tablas_reales['df_outlets']) + 3
            
            if 'df_secundarias' in tablas_reales and len(tablas_reales['df_secundarias']) > 0:
                titulo = pd.DataFrame([[f'🏬 {nombre_tiendas_secundarias.upper()}']], columns=[''])
                titulo.to_excel(output, sheet_name=sheet_name_stock, startrow=row_offset, index=False, header=False)
                row_offset += 2
                tablas_reales['df_secundarias'].to_excel(output, sheet_name=sheet_name_stock, startrow=row_offset, index=False)
                row_offset += len(tablas_reales['df_secundarias']) + 3
            
            # Escribir tabla de Tienda Outlet en pestaña de stock (solo para El Salvador)
            if 'df_outlet_especial' in tablas_reales and len(tablas_reales['df_outlet_especial']) > 0:
                titulo = pd.DataFrame([['🏪 TIENDA OUTLET']], columns=[''])
                titulo.to_excel(output, sheet_name=sheet_name_stock, startrow=row_offset, index=False, header=False)
                row_offset += 2
                tablas_reales['df_outlet_especial'].to_excel(output, sheet_name=sheet_name_stock, startrow=row_offset, index=False)
            
            # PESTAÑA 2: Distribución Ventas
            sheet_name_ventas = "Distribución Ventas"
            row_offset = 0
            
            if 'df_principales_ventas' in tablas_reales and len(tablas_reales['df_principales_ventas']) > 0:
                titulo = pd.DataFrame([[f'🏪 {nombre_tiendas_principales.upper()} - VENTAS']], columns=[''])
                titulo.to_excel(output, sheet_name=sheet_name_ventas, startrow=row_offset, index=False, header=False)
                row_offset += 2
                tablas_reales['df_principales_ventas'].to_excel(output, sheet_name=sheet_name_ventas, startrow=row_offset, index=False)
                row_offset += len(tablas_reales['df_principales_ventas']) + 3
            
            if 'df_outlets_ventas' in tablas_reales and len(tablas_reales['df_outlets_ventas']) > 0:
                titulo = pd.DataFrame([['🛒 OUTLETS - VENTAS']], columns=[''])
                titulo.to_excel(output, sheet_name=sheet_name_ventas, startrow=row_offset, index=False, header=False)
                row_offset += 2
                tablas_reales['df_outlets_ventas'].to_excel(output, sheet_name=sheet_name_ventas, startrow=row_offset, index=False)
                row_offset += len(tablas_reales['df_outlets_ventas']) + 3
            
            if 'df_secundarias_ventas' in tablas_reales and len(tablas_reales['df_secundarias_ventas']) > 0:
                titulo = pd.DataFrame([[f'🏬 {nombre_tiendas_secundarias.upper()} - VENTAS']], columns=[''])
                titulo.to_excel(output, sheet_name=sheet_name_ventas, startrow=row_offset, index=False, header=False)
                row_offset += 2
                tablas_reales['df_secundarias_ventas'].to_excel(output, sheet_name=sheet_name_ventas, startrow=row_offset, index=False)
                row_offset += len(tablas_reales['df_secundarias_ventas']) + 3
            
            # Escribir tabla de Tienda Outlet - Ventas (solo para El Salvador)
            if 'df_outlet_especial_ventas' in tablas_reales and len(tablas_reales['df_outlet_especial_ventas']) > 0:
                titulo = pd.DataFrame([['🏪 TIENDA OUTLET - VENTAS']], columns=[''])
                titulo.to_excel(output, sheet_name=sheet_name_ventas, startrow=row_offset, index=False, header=False)
                row_offset += 2
                tablas_reales['df_outlet_especial_ventas'].to_excel(output, sheet_name=sheet_name_ventas, startrow=row_offset, index=False)
            
            # Nota: Se removió la pestaña de comparación como se solicitó
        
        # Aplicar formato básico a todas las pestañas
        workbook = output.book
        header_fill = PatternFill(start_color='4a7a8c', end_color='4a7a8c', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        titulo_fill = PatternFill(start_color='2d3748', end_color='2d3748', fill_type='solid')
        titulo_font = Font(color='FFFFFF', bold=True, size=14)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Aplicar formato a todas las celdas
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    # Formato numérico con 2 decimales
                    if isinstance(cell.value, (int, float)) and cell.value != 0:
                        cell.number_format = '0.00'
                    
                    # Formato para títulos de secciones (🏪, 🛒, 🏬)
                    if cell.value and isinstance(cell.value, str) and any(emoji in str(cell.value) for emoji in ['🏪', '🛒', '🏬']):
                        cell.fill = titulo_fill
                        cell.font = titulo_font
                    # Formato para headers de tablas
                    elif cell.row > 1 and cell.value and isinstance(cell.value, str) and 'Bodega' in str(cell.value):
                        cell.fill = header_fill
                        cell.font = header_font
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.close()
        
        # Descargar archivo
        with open(nombre_excel, "rb") as f:
            if tiene_ventas:
                label_text = f"Descargar Distribuciones Completas {pais}"
            else:
                label_text = f"Descargar Distribución Stock {pais}"
            
            st.download_button(
                label=label_text,
                data=f,
                file_name=f"DISTRIBUCION_BODEGAS_{pais.upper().replace(' ', '_')}_{config.fecha_reporte}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_distribucion_real_{pais}"
            )
        
        # Limpiar archivo temporal
        os.remove(nombre_excel)
        logger.info(f"Exportación de distribuciones reales completada para {pais}")
        
    except Exception as e:
        logger.error(f"Error al exportar distribuciones reales {pais}: {str(e)}")
        st.error(f"Error al exportar distribuciones reales {pais}: {str(e)}")

def exportar_excel_distribuciones(df_bodegas, nombres_reales_bodegas, pais):
    """Exporta las tablas de distribución por bodega a Excel con pestañas según los datos disponibles"""
    if df_bodegas is None:
        st.warning(f"No hay datos de distribución para exportar de {pais}")
        return
    
    try:
        logger.info(f"Iniciando exportación de distribuciones para {pais}")
        
        # Definir nombres dinámicos según el país
        if pais == "Guatemala":
            nombre_tiendas_secundarias = "Tiendas Departamentales"
        else:
            nombre_tiendas_secundarias = "Tiendas Franquicia"
        
        # Detectar si hay datos de ventas
        tiene_ventas = any('Ventas' in str(col) for col in df_bodegas.columns)
        
        # Crear archivo Excel
        nombre_excel = f"distribucion_bodegas_{pais.lower().replace(' ', '_')}_{config.fecha_reporte}.xlsx"
        output = pd.ExcelWriter(nombre_excel, engine='openpyxl')
        
        # Configurar ligas
        ligas = ['MLB', 'NBA', 'NFL', 'MOTORSPORT', 'ENTERTAINMENT']
        es_multiindex = isinstance(df_bodegas.columns, pd.MultiIndex)
        
        # Función auxiliar para crear tablas de distribución
        def crear_tablas_distribucion(df_bodegas, nombres_bodegas, tipo_datos="Stock"):
            # Procesar datos de distribución
            distribucion_data = []
            
            for i, bodega_idx in enumerate(df_bodegas.index):
                nombre_bodega = nombres_bodegas[i] if i < len(nombres_bodegas) else bodega_idx
                
                # Excluir bodegas centrales de las distribuciones
                if pais == "Costa Rica" and nombre_bodega == "Bodega Central NEW ERA":
                    continue
                elif pais == "PANAMA" and nombre_bodega in ['Almacén general', 'Bodega Central Albrook']:
                    continue
                    
                bodega_data = {'Bodega': nombre_bodega}
                total_bodega = 0
                
                # Calcular totales por liga
                for liga in ligas:
                    if es_multiindex:
                        if tipo_datos == "Stock":
                            col_planas = (liga, 'Planas', 'Stock')
                            col_curvas = (liga, 'Curvas', 'Stock')
                        else:  # Ventas
                            col_planas = (liga, 'Planas', 'Ventas')
                            col_curvas = (liga, 'Curvas', 'Ventas')
                    else:
                        if tipo_datos == "Stock":
                            col_planas = f"{liga} - Planas - Stock"
                            col_curvas = f"{liga} - Curvas - Stock"
                        else:  # Ventas
                            col_planas = f"{liga} - Planas - Ventas"
                            col_curvas = f"{liga} - Curvas - Ventas"
                    
                    valor_planas = df_bodegas.loc[bodega_idx, col_planas] if col_planas in df_bodegas.columns else 0
                    valor_curvas = df_bodegas.loc[bodega_idx, col_curvas] if col_curvas in df_bodegas.columns else 0
                    
                    try:
                        valor_planas = float(valor_planas) if valor_planas != 0 else 0
                        valor_curvas = float(valor_curvas) if valor_curvas != 0 else 0
                    except:
                        valor_planas = 0
                        valor_curvas = 0
                    
                    valor_liga = valor_planas + valor_curvas
                    bodega_data[liga] = valor_liga
                    total_bodega += valor_liga
                
                # Calcular porcentajes
                if total_bodega > 0:
                    for liga in ligas:
                        pct = (bodega_data[liga] / total_bodega) * 100
                        bodega_data[f'{liga} %'] = round(pct, 1)
                else:
                    for liga in ligas:
                        bodega_data[f'{liga} %'] = 0.0
                
                bodega_data['Total'] = total_bodega
                distribucion_data.append(bodega_data)
            
            df_distribucion = pd.DataFrame(distribucion_data)
            
            # Separar en tres tipos de bodegas
            df_principales = df_distribucion[df_distribucion['Bodega'].str.contains('Principal|Ciudad', case=False, na=False)]
            df_outlets = df_distribucion[df_distribucion['Bodega'].str.contains('Outlet', case=False, na=False)]
            df_secundarias = df_distribucion[~df_distribucion['Bodega'].str.contains('Principal|Ciudad|Outlet', case=False, na=False)]
            
            return df_principales, df_outlets, df_secundarias
        
        # Exportar datos de stock
        df_principales_stock, df_outlets_stock, df_secundarias_stock = crear_tablas_distribucion(df_bodegas, nombres_reales_bodegas, "Stock")
        
        # Crear pestaña de Stock
        sheet_name_stock = "Distribución Stock"
        
        # Escribir las tres tablas en la pestaña de stock
        row_offset = 0
        
        # Tiendas de Ciudad
        if len(df_principales_stock) > 0:
            df_principales_stock.to_excel(output, sheet_name=sheet_name_stock, startrow=row_offset, index=False)
            row_offset += len(df_principales_stock) + 3  # Espacio entre tablas
        
        # Outlets
        if len(df_outlets_stock) > 0:
            df_outlets_stock.to_excel(output, sheet_name=sheet_name_stock, startrow=row_offset, index=False)
            row_offset += len(df_outlets_stock) + 3
        
        # Tiendas Departamentales
        if len(df_secundarias_stock) > 0:
            df_secundarias_stock.to_excel(output, sheet_name=sheet_name_stock, startrow=row_offset, index=False)
        
        # Si hay datos de ventas, crear pestañas adicionales
        if tiene_ventas:
            # Exportar datos de ventas
            df_principales_ventas, df_outlets_ventas, df_secundarias_ventas = crear_tablas_distribucion(df_bodegas, nombres_reales_bodegas, "Ventas")
            
            # Crear pestaña de Ventas
            sheet_name_ventas = "Distribución Ventas"
            row_offset = 0
            
            if len(df_principales_ventas) > 0:
                df_principales_ventas.to_excel(output, sheet_name=sheet_name_ventas, startrow=row_offset, index=False)
                row_offset += len(df_principales_ventas) + 3
            
            if len(df_outlets_ventas) > 0:
                df_outlets_ventas.to_excel(output, sheet_name=sheet_name_ventas, startrow=row_offset, index=False)
                row_offset += len(df_outlets_ventas) + 3
            
            if len(df_secundarias_ventas) > 0:
                df_secundarias_ventas.to_excel(output, sheet_name=sheet_name_ventas, startrow=row_offset, index=False)
            
            # Crear pestaña de Comparación (datos combinados)
            sheet_name_comparacion = "Comparación Stock vs Ventas"
            row_offset = 0
            
            # Combinar datos de stock y ventas para comparación
            if len(df_principales_stock) > 0 and len(df_principales_ventas) > 0:
                df_comparacion_principales = df_principales_stock.merge(df_principales_ventas, on='Bodega', suffixes=(' (Stock)', ' (Ventas)'))
                df_comparacion_principales.to_excel(output, sheet_name=sheet_name_comparacion, startrow=row_offset, index=False)
                row_offset += len(df_comparacion_principales) + 3
            
            if len(df_outlets_stock) > 0 and len(df_outlets_ventas) > 0:
                df_comparacion_outlets = df_outlets_stock.merge(df_outlets_ventas, on='Bodega', suffixes=(' (Stock)', ' (Ventas)'))
                df_comparacion_outlets.to_excel(output, sheet_name=sheet_name_comparacion, startrow=row_offset, index=False)
                row_offset += len(df_comparacion_outlets) + 3
            
            if len(df_secundarias_stock) > 0 and len(df_secundarias_ventas) > 0:
                df_comparacion_secundarias = df_secundarias_stock.merge(df_secundarias_ventas, on='Bodega', suffixes=(' (Stock)', ' (Ventas)'))
                df_comparacion_secundarias.to_excel(output, sheet_name=sheet_name_comparacion, startrow=row_offset, index=False)
        
        # Aplicar formato básico a todas las pestañas
        workbook = output.book
        header_fill = PatternFill(start_color='4a7a8c', end_color='4a7a8c', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Aplicar formato a todas las celdas
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    # Formato para headers
                    if cell.row == 1 or (cell.value and isinstance(cell.value, str) and 'Bodega' in str(cell.value)):
                        cell.fill = header_fill
                        cell.font = header_font
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.close()
        
        # Descargar archivo
        with open(nombre_excel, "rb") as f:
            if tiene_ventas:
                label_text = f"Descargar Distribuciones Completas {pais}"
            else:
                label_text = f"Descargar Distribución Stock {pais}"
            
            st.download_button(
                label=label_text,
                data=f,
                file_name=f"DISTRIBUCION_BODEGAS_{pais.upper().replace(' ', '_')}_{config.fecha_reporte}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_distribucion_{pais}"
            )
        
        # Limpiar archivo temporal
        os.remove(nombre_excel)
        logger.info(f"Exportación de distribuciones completada para {pais}")
        
    except Exception as e:
        logger.error(f"Error al exportar distribuciones {pais}: {str(e)}")
        st.error(f"Error al exportar distribuciones {pais}: {str(e)}")

def exportar_excel_consolidado(tabla, nombre_archivo, pais):
    """Exporta la tabla consolidada a Excel con formato profesional"""
    if tabla is None:
        st.warning(f"No hay datos para exportar de {pais}")
        return
    
    try:
        selected_league = st.session_state.get('selected_league', None)
        # Convertir "Todas" a None para mostrar todas las ligas
        if selected_league == "Todas":
            selected_league = None
        
        if selected_league:
            logger.info(f"Iniciando exportación a Excel para {selected_league} - {pais}")
        else:
            logger.info(f"Iniciando exportación a Excel para {pais}")
        
        # Crear copia del DataFrame para exportación
        df_export = tabla.copy()
        df_export.columns = [' - '.join(col).strip(' - ') for col in df_export.columns.values]
        
        # Crear archivo Excel
        if selected_league:
            nombre_excel = f"stock_{selected_league.lower()}_{pais.lower().replace(' ', '_')}.xlsx"
        else:
            nombre_excel = f"stock_consolidado_{pais.lower().replace(' ', '_')}.xlsx"
        output = pd.ExcelWriter(nombre_excel, engine='openpyxl')
        
        if selected_league:
            sheet_name = f"{selected_league} {pais}"
        else:
            sheet_name = f"Stock {pais}"
        
        df_export.to_excel(output, sheet_name=sheet_name, index=False)
        
        # Aplicar formato
        workbook = output.book
        worksheet = output.sheets[sheet_name]
        
        # Estilos
        header_fill = PatternFill(start_color='4a7a8c', end_color='4a7a8c', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=14)
        
        total_fill = PatternFill(start_color='d35400', end_color='d35400', fill_type='solid')
        total_font = Font(color='FFFFFF', bold=True, size=14)
        
        normal_font = Font(color='000000', size=10)
        
        # Colores para el semáforo
        verde_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
        amarillo_fill = PatternFill(start_color='ffc107', end_color='ffc107', fill_type='solid')
        rojo_fill = PatternFill(start_color='dc3545', end_color='dc3545', fill_type='solid')
        gris_fill = PatternFill(start_color='6c757d', end_color='6c757d', fill_type='solid')
        semaforo_font = Font(color='FFFFFF', bold=True, size=10)
        
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Aplicar formatos
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = center_alignment
                
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                elif cell.row == worksheet.max_row:
                    cell.fill = total_fill
                    cell.font = total_font
                else:
                    cell.font = normal_font
        
        # Aplicar semáforo a la columna "% DE CUMPLIMIENTO"
        col_cumplimiento = None
        col_total_headwear = None
        
        # Buscar columnas por nombre que contenga las palabras clave
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value:
                if "% DE CUMPLIMIENTO" in str(cell_value):
                    col_cumplimiento = col
                elif "TOTAL HEADWEAR" in str(cell_value):
                    col_total_headwear = col
        
        if col_cumplimiento and col_total_headwear:
            logger.info(f"Aplicando semáforo - Col cumplimiento: {col_cumplimiento}, Col total headwear: {col_total_headwear}")
            capacidades = country_manager.get_capacidades(pais)
            
            for row in range(2, worksheet.max_row + 1):
                bodega = worksheet.cell(row=row, column=1).value
                
                # Obtener total_headwear de la columna encontrada
                total_headwear = worksheet.cell(row=row, column=col_total_headwear).value or 0
                
                if bodega == 'TOTAL':
                    capacidad = country_manager.get_country_data(pais).get_total_capacity()
                else:
                    capacidad = capacidades.get(bodega, 0)
                
                cell = worksheet.cell(row=row, column=col_cumplimiento)
                
                if capacidad > 0:
                    color = stock_analyzer.obtener_color_semaforo(total_headwear, capacidad)
                    if color == "verde":
                        cell.fill = verde_fill
                    elif color == "amarillo":
                        cell.fill = amarillo_fill
                    else:
                        cell.fill = rojo_fill
                else:
                    cell.fill = gris_fill
                
                if row == worksheet.max_row:
                    cell.font = Font(color='FFFFFF', bold=True, size=14)
                else:
                    cell.font = semaforo_font
        else:
            logger.warning(f"No se pudieron encontrar las columnas para el semáforo - Col cumplimiento: {col_cumplimiento}, Col total headwear: {col_total_headwear}")
            logger.info("Columnas disponibles en Excel:")
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=1, column=col).value
                logger.info(f"  Columna {col}: {cell_value}")
        
        # Autoajustar columnas
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.1
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
        # Agregar información adicional
        info_row = worksheet.max_row + 2
        worksheet.cell(row=info_row, column=1, value="Fecha:").font = Font(bold=True)
        worksheet.cell(row=info_row, column=2, value=datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        
        worksheet.cell(row=info_row+1, column=1, value="Archivo origen:").font = Font(bold=True)
        worksheet.cell(row=info_row+1, column=2, value=nombre_archivo)
        
        worksheet.cell(row=info_row+2, column=1, value="País:").font = Font(bold=True)
        worksheet.cell(row=info_row+2, column=2, value=pais)
        
        # Agregar leyenda del semáforo
        worksheet.cell(row=info_row+4, column=1, value="Leyenda Semáforo:").font = Font(bold=True)
        worksheet.cell(row=info_row+5, column=1, value="Verde: 0%-15%").fill = verde_fill
        worksheet.cell(row=info_row+5, column=1).font = semaforo_font
        worksheet.cell(row=info_row+6, column=1, value="Amarillo: >15%").fill = amarillo_fill
        worksheet.cell(row=info_row+6, column=1).font = semaforo_font
        worksheet.cell(row=info_row+7, column=1, value="Rojo: <0%").fill = rojo_fill
        worksheet.cell(row=info_row+7, column=1).font = semaforo_font
        worksheet.cell(row=info_row+8, column=1, value="Gris: Sin capacidad definida").fill = gris_fill
        worksheet.cell(row=info_row+8, column=1).font = semaforo_font
        
        output.close()
        
        # Descargar archivo
        with open(nombre_excel, "rb") as f:
            st.download_button(
                label=f"Descargar Reporte {pais}",
                data=f,
                file_name=f"STOCK_CONSOLIDADO_{pais.upper().replace(' ', '_')}_{config.fecha_reporte}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_{pais}"
            )
        
        # Limpiar archivo temporal
        os.remove(nombre_excel)
        logger.info(f"Exportación a Excel completada para {pais}")
        
    except Exception as e:
        logger.error(f"Error al exportar {pais}: {str(e)}")
        st.error(f"Error al exportar {pais}: {str(e)}")

def obtener_optimos_mvp() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por código y bodega
    {codigo: {bodega: cantidad_optima}}
    """
    # Mapeo de nombres de tiendas: Stock Real (base) -> Stock Códigos -> Stock Tallas
    mapeo_tiendas = {
        "NE Miraflores": "NEW ERA MIRAFLORES",
        "NE Oakland": "NEW ERA OAKLAND", 
        "NE Portales": "NEW ERA PORTALES",
        "NE InterXela": "NEW ERA INT XELA",
        "NE Cayala": "NEW ERA CAYALA",
        "NE Metronorte": "NEW ERA METRONORTE",
        "NE Concepcion": "NEW ERA CONCEPCION",
        "NE Interplaza Escuintla": "NE INT ESCUINTLA",
        "NE Pradera Huehuetenango": "NEW ERA HUEHUETENANGO",
        "NE Naranjo": "NEW ERA NARANJO",
        "NE Metrocentro Outlet": "NEW ERA METROCENTRO OUTLET",
        "NE Vistares": "NEW ERA VISTARES",
        "NE Peri Roosvelt": "NEW ERA PERI-ROOSELVET",
        "NE Outlet Santa clara": "NEW ERA SANTA CLARA",
        "NE Pradera Chiquimula": "PRADERA CHIQUIMULA",
        "NE Pradera Escuintla": "NE PRADERA ESCUINTLA",
        "NE Paseo Antigua": "NEW ERA ANTIGUA",
        "NE Pradera Xela": "PRADERA XELA",
        "NE Puerto Barrios": "NEW ERA PUERTO BARRIOS",
        "NE Metroplaza Jutiapa": "PRADERA JUTIAPA",
        "NE Chimaltenango": "NEW ERA CHIMALTENANGO",
        "NE Plaza Magdalena": "NEW ERA COBAN"
    }
    
    # Stock óptimo por códigos (nuevos datos cargados)
    optimos_data = {
        "10030709": {"NEW ERA MIRAFLORES": 20, "NEW ERA OAKLAND": 20, "NEW ERA PERI-ROOSELVET": 18, "NE INT ESCUINTLA": 12, "NEW ERA CONCEPCION": 12, "NEW ERA NARANJO": 18, "NEW ERA PORTALES": 18, "NEW ERA CHIMALTENANGO": 12, "NEW ERA INT XELA": 18, "NEW ERA CAYALA": 12, "NEW ERA METRONORTE": 12, "NEW ERA HUEHUETENANGO": 18, "NE PRADERA ESCUINTLA": 12, "PRADERA CHIQUIMULA": 12, "PRADERA XELA": 12, "PRADERA JUTIAPA": 12, "NEW ERA VISTARES": 18, "NEW ERA SANTA CLARA": 12, "NEW ERA COBAN": 12, "NEW ERA METROCENTRO OUTLET": 12, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 12},
        "10030708": {"NEW ERA MIRAFLORES": 10, "NEW ERA OAKLAND": 10, "NEW ERA PERI-ROOSELVET": 8, "NE INT ESCUINTLA": 6, "NEW ERA CONCEPCION": 6, "NEW ERA NARANJO": 8, "NEW ERA PORTALES": 8, "NEW ERA CHIMALTENANGO": 6, "NEW ERA INT XELA": 8, "NEW ERA CAYALA": 6, "NEW ERA METRONORTE": 6, "NEW ERA HUEHUETENANGO": 8, "NE PRADERA ESCUINTLA": 6, "PRADERA CHIQUIMULA": 6, "PRADERA XELA": 6, "PRADERA JUTIAPA": 6, "NEW ERA VISTARES": 8, "NEW ERA SANTA CLARA": 6, "NEW ERA COBAN": 6, "NEW ERA METROCENTRO OUTLET": 6, "NEW ERA ANTIGUA": 8, "NEW ERA PUERTO BARRIOS": 6},
        "10112874": {"NEW ERA MIRAFLORES": 6, "NEW ERA OAKLAND": 6, "NEW ERA PERI-ROOSELVET": 3, "NE INT ESCUINTLA": 3, "NEW ERA CONCEPCION": 3, "NEW ERA NARANJO": 3, "NEW ERA PORTALES": 3, "NEW ERA CHIMALTENANGO": 3, "NEW ERA INT XELA": 3, "NEW ERA CAYALA": 3, "NEW ERA METRONORTE": 2, "NEW ERA HUEHUETENANGO": 3, "NE PRADERA ESCUINTLA": 2, "PRADERA CHIQUIMULA": 2, "PRADERA XELA": 3, "PRADERA JUTIAPA": 2, "NEW ERA VISTARES": 3, "NEW ERA SANTA CLARA": 2, "NEW ERA COBAN": 3, "NEW ERA METROCENTRO OUTLET": 2, "NEW ERA ANTIGUA": 3, "NEW ERA PUERTO BARRIOS": 3},
        "11591122": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11591128": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11591150": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11591175": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70331909": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70331911": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70331962": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "10975804": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "10975815": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "10975835": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70192970": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70353249": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70353266": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70360899": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70360903": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70428987": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70430338": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70457634": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11591024": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11591025": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11591026": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11591043": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11591046": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11591047": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11591077": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11591078": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11941921": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70556851": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70556867": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70556869": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "70558225": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "10047511": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "10047531": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "10047538": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11405605": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "11405614": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "12650335": {"NEW ERA MIRAFLORES": 120, "NEW ERA OAKLAND": 108, "NEW ERA PERI-ROOSELVET": 60, "NE INT ESCUINTLA": 48, "NEW ERA CONCEPCION": 60, "NEW ERA NARANJO": 60, "NEW ERA PORTALES": 60, "NEW ERA CHIMALTENANGO": 48, "NEW ERA INT XELA": 60, "NEW ERA CAYALA": 48, "NEW ERA METRONORTE": 60, "NEW ERA HUEHUETENANGO": 48, "NE PRADERA ESCUINTLA": 48, "PRADERA CHIQUIMULA": 48, "PRADERA XELA": 60, "PRADERA JUTIAPA": 48, "NEW ERA VISTARES": 48, "NEW ERA SANTA CLARA": 60, "NEW ERA COBAN": 48, "NEW ERA METROCENTRO OUTLET": 60, "NEW ERA ANTIGUA": 48, "NEW ERA PUERTO BARRIOS": 48},
        "12650337": {"NEW ERA MIRAFLORES": 36, "NEW ERA OAKLAND": 36, "NEW ERA PERI-ROOSELVET": 18, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 18, "NEW ERA NARANJO": 18, "NEW ERA PORTALES": 18, "NEW ERA CHIMALTENANGO": 18, "NEW ERA INT XELA": 18, "NEW ERA CAYALA": 18, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 18, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 18, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 18, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "12650340": {"NEW ERA MIRAFLORES": 72, "NEW ERA OAKLAND": 60, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 24, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 36, "NEW ERA PORTALES": 36, "NEW ERA CHIMALTENANGO": 24, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 24, "NEW ERA METRONORTE": 36, "NEW ERA HUEHUETENANGO": 24, "NE PRADERA ESCUINTLA": 24, "PRADERA CHIQUIMULA": 24, "PRADERA XELA": 36, "PRADERA JUTIAPA": 24, "NEW ERA VISTARES": 24, "NEW ERA SANTA CLARA": 36, "NEW ERA COBAN": 24, "NEW ERA METROCENTRO OUTLET": 36, "NEW ERA ANTIGUA": 36, "NEW ERA PUERTO BARRIOS": 24},
        "12650342": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "12650343": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18},
        "12650344": {"NEW ERA MIRAFLORES": 48, "NEW ERA OAKLAND": 48, "NEW ERA PERI-ROOSELVET": 36, "NE INT ESCUINTLA": 18, "NEW ERA CONCEPCION": 36, "NEW ERA NARANJO": 48, "NEW ERA PORTALES": 48, "NEW ERA CHIMALTENANGO": 36, "NEW ERA INT XELA": 36, "NEW ERA CAYALA": 36, "NEW ERA METRONORTE": 18, "NEW ERA HUEHUETENANGO": 36, "NE PRADERA ESCUINTLA": 18, "PRADERA CHIQUIMULA": 18, "PRADERA XELA": 36, "PRADERA JUTIAPA": 18, "NEW ERA VISTARES": 36, "NEW ERA SANTA CLARA": 18, "NEW ERA COBAN": 18, "NEW ERA METROCENTRO OUTLET": 18, "NEW ERA ANTIGUA": 18, "NEW ERA PUERTO BARRIOS": 18}
    }
    
    # Convertir nombres de tiendas de códigos a nombres del stock real
    optimos_dict = {}
    for codigo, tiendas_codigo in optimos_data.items():
        optimos_dict[codigo] = {}
        for tienda_real, tienda_codigo in mapeo_tiendas.items():
            if tienda_codigo in tiendas_codigo:
                optimos_dict[codigo][tienda_real] = tiendas_codigo[tienda_codigo]
    
    return optimos_dict


def obtener_optimos_mvp_elsalvador() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por código y bodega para El Salvador
    {codigo: {bodega: cantidad_optima}}
    """
    # Mapeo de nombres de tiendas: Stock Real (dashboard) -> Stock Códigos
    mapeo_tiendas_elsalvador = {
        "NE METROCENTRO LOURDES": "NE LOURDES",
        "NE METROCENTRO SAN MIGUEL": "NE SAN MIGUEL",
        "NE PLAZA MUNDO SOYAPANGO": "NE SOYAPANGO",
        "NE USULUTÁN": "NE USULUTAN",
        "NEW ERA EL PASEO": "NE EL PASEO",
        "NEW ERA METROCENTRO": "NE METROCENTRO",
        "NEW ERA METROCENTRO SANTA ANA": "NE SANTA ANA",
        "NEW ERA MULTIPLAZA": "NE MULTIPLAZA"
    }
    
    # Stock óptimo por códigos El Salvador
    optimos_data_elsalvador = {
        "10030709": {"NE MULTIPLAZA": 20, "NE EL PASEO": 10, "NE METROCENTRO": 20, "NE SANTA ANA": 20, "NE USULUTAN": 10, "NE LOURDES": 10, "NE SAN MIGUEL": 10, "NE SOYAPANGO": 10},
        "10030708": {"NE MULTIPLAZA": 10, "NE EL PASEO": 5, "NE METROCENTRO": 10, "NE SANTA ANA": 10, "NE USULUTAN": 5, "NE LOURDES": 5, "NE SAN MIGUEL": 5, "NE SOYAPANGO": 5},
        "10112874": {"NE MULTIPLAZA": 6, "NE EL PASEO": 1, "NE METROCENTRO": 6, "NE SANTA ANA": 6, "NE USULUTAN": 3, "NE LOURDES": 3, "NE SAN MIGUEL": 3, "NE SOYAPANGO": 1},
        "11591122": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11591128": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11591150": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11591175": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70331909": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70331911": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70331962": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "10975804": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "10975815": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "10975835": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70192970": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70353249": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70353266": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70360899": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70360903": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70428987": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70430338": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70457634": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11591024": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11591025": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11591026": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11591043": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11591046": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11591047": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11591077": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11591078": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11941921": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70556851": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70556867": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70556869": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "70558225": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "10047511": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "10047531": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "10047538": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11405605": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "11405614": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "12650335": {"NE MULTIPLAZA": 60, "NE EL PASEO": 24, "NE METROCENTRO": 60, "NE SANTA ANA": 60, "NE USULUTAN": 48, "NE LOURDES": 24, "NE SAN MIGUEL": 48, "NE SOYAPANGO": 48},
        "12650337": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "12650340": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "12650342": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "12650343": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36},
        "12650344": {"NE MULTIPLAZA": 48, "NE EL PASEO": 18, "NE METROCENTRO": 48, "NE SANTA ANA": 48, "NE USULUTAN": 36, "NE LOURDES": 18, "NE SAN MIGUEL": 36, "NE SOYAPANGO": 36}
    }
    
    # Convertir nombres de tiendas del mapeo
    optimos_dict = {}
    for tienda_real, tienda_codigos in mapeo_tiendas_elsalvador.items():
        for codigo, tiendas_optimos in optimos_data_elsalvador.items():
            if codigo not in optimos_dict:
                optimos_dict[codigo] = {}
            if tienda_codigos in tiendas_optimos:
                optimos_dict[codigo][tienda_real] = tiendas_optimos[tienda_codigos]
    
    return optimos_dict


def obtener_optimos_por_tallas_elsalvador() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por talla y tienda para El Salvador
    {tienda: {talla: cantidad_optima}}
    """
    # Mapeo de nombres de tiendas: Stock Real (dashboard) -> Stock Tallas
    mapeo_tiendas_tallas_elsalvador = {
        "NE METROCENTRO LOURDES": "NE LOURDES OUTLET",
        "NE METROCENTRO SAN MIGUEL": "NE SAN MIGUEL",
        "NE PLAZA MUNDO SOYAPANGO": "NE SOYAPANGO",
        "NE USULUTÁN": "NE USULUTAN",
        "NEW ERA EL PASEO": "NE EL PASEO",
        "NEW ERA METROCENTRO": "NE METROCENTRO",
        "NEW ERA METROCENTRO SANTA ANA": "NE SANTA ANA",
        "NEW ERA MULTIPLAZA": "NE MULTIPLAZA"
    }
    
    # Stock óptimo por tallas El Salvador - Datos correctos según tabla del usuario (todas suman 12)
    optimos_tallas_data_elsalvador = {
        "NE MULTIPLAZA": {"678": 0, "700": 1, "718": 1, "714": 2, "738": 4, "712": 2, "758": 2, "734": 0, "778": 0, "800": 0},      # suma: 12
        "NE EL PASEO": {"678": 0, "700": 1, "718": 1, "714": 2, "738": 3, "712": 3, "758": 2, "734": 0, "778": 0, "800": 0},        # suma: 12
        "NE METROCENTRO": {"678": 0, "700": 1, "718": 2, "714": 2, "738": 3, "712": 2, "758": 2, "734": 0, "778": 0, "800": 0},     # suma: 12
        "NE USULUTAN": {"678": 0, "700": 0, "718": 1, "714": 1, "738": 3, "712": 3, "758": 2, "734": 1, "778": 1, "800": 0},        # suma: 12
        "NE LOURDES OUTLET": {"678": 0, "700": 0, "718": 0, "714": 0, "738": 0, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},  # suma: 0 (sin distribución)
        "NE SANTA ANA": {"678": 0, "700": 1, "718": 1, "714": 2, "738": 3, "712": 3, "758": 2, "734": 0, "778": 0, "800": 0},       # suma: 12
        "NE SAN MIGUEL": {"678": 0, "700": 0, "718": 1, "714": 1, "738": 3, "712": 3, "758": 2, "734": 1, "778": 1, "800": 0},      # suma: 12
        "NE SOYAPANGO": {"678": 0, "700": 1, "718": 1, "714": 2, "738": 3, "712": 3, "758": 2, "734": 0, "778": 0, "800": 0}        # suma: 12
    }
    
    # Convertir nombres de tiendas de tallas a nombres del stock real
    optimos_dict = {}
    for tienda_real, tienda_tallas in mapeo_tiendas_tallas_elsalvador.items():
        if tienda_tallas in optimos_tallas_data_elsalvador:
            optimos_dict[tienda_real] = optimos_tallas_data_elsalvador[tienda_tallas]
    
    return optimos_dict


def obtener_optimos_por_tallas() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por talla y tienda
    {tienda: {talla: cantidad_optima}}
    """
    # Mapeo de nombres de tiendas: Stock Real (base) -> Stock Tallas 
    mapeo_tiendas_tallas = {
        "NE Miraflores": "NE MIRAFLORES",
        "NE Oakland": "NE OAKLAND",
        "NE Portales": "NE PORTALES", 
        "NE InterXela": "NE INTER XELA",
        "NE Concepcion": "NE CONCEPCIÓN",
        "NE Naranjo": "NE NARANJO",
        "NE Pradera Xela": "NE PRADERA XELA",
        "NE Peri Roosvelt": "NE PERI ROOSEVELT",
        "NE Cayala": "NE CAYALA",
        "NE Metronorte": "NE OUTLET METRONORTE",
        "NE Pradera Huehuetenango": "NE HUEHUETENANGO",
        "NE Interplaza Escuintla": "NE I ESCUINTLA",
        "NE Chimaltenango": "NE CHIMALTENANGO",
        "NE Metroplaza Jutiapa": "NE JUTIAPA",
        "NE Vistares": "NE VISTARES",
        "NE Pradera Escuintla": "NE PRADERA ESCUINTLA",
        "NE Pradera Chiquimula": "NE PRADERA CHIQUIMULA",
        "NE Paseo Antigua": "NE PASEO ANTIGUA",
        "NE Outlet Santa clara": "NE SANTA CLARA",
        "NE Plaza Magdalena": "NE PLAZA MAGDALENA",
        "NE Puerto Barrios": "NE PTO. BARRIOS"
    }
    
    # Stock óptimo por tallas (datos actualizados según tabla del usuario - tallas vacías = 0)
    optimos_tallas_data = {
        "NE MIRAFLORES": {"678": 1, "700": 2, "718": 4, "714": 3, "738": 2, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE OAKLAND": {"678": 0, "700": 2, "718": 4, "714": 4, "738": 2, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE PORTALES": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE INTER XELA": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE CONCEPCIÓN": {"678": 0, "700": 1, "718": 4, "714": 4, "738": 2, "712": 1, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE NARANJO": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE PRADERA XELA": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE PERI ROOSEVELT": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE CAYALA": {"678": 0, "700": 2, "718": 4, "714": 3, "738": 2, "712": 1, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE OUTLET METRONORTE": {"678": 0, "700": 0, "718": 0, "714": 0, "738": 0, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE HUEHUETENANGO": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE I ESCUINTLA": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE CHIMALTENANGO": {"678": 2, "700": 5, "718": 3, "714": 2, "738": 0, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE JUTIAPA": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE VISTARES": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE PRADERA ESCUINTLA": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE PRADERA CHIQUIMULA": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE PASEO ANTIGUA": {"678": 0, "700": 2, "718": 4, "714": 3, "738": 2, "712": 1, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE SANTA CLARA": {"678": 0, "700": 0, "718": 0, "714": 0, "738": 0, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE PLAZA MAGDALENA": {"678": 2, "700": 2, "718": 4, "714": 3, "738": 1, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE METROCENTRO VILLANUEVA": {"678": 0, "700": 0, "718": 0, "714": 0, "738": 0, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE PTO. BARRIOS": {"678": 0, "700": 3, "718": 4, "714": 3, "738": 2, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0}
    }
    
    # Convertir nombres de tiendas de tallas a nombres del stock real
    optimos_dict = {}
    for tienda_real, tienda_tallas in mapeo_tiendas_tallas.items():
        if tienda_tallas in optimos_tallas_data:
            optimos_dict[tienda_real] = optimos_tallas_data[tienda_tallas]
    
    return optimos_dict

def obtener_optimos_mvp_honduras() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por código y tienda para Honduras
    {codigo: {tienda: cantidad_optima}}
    """
    # Mapeo de nombres de tiendas: Stock Real -> Stock Códigos
    mapeo_tiendas_codigo = {
        "NE – Cascadas Mall Tegucigalpa": "CASCADAS",
        "NE – CITY MALL SP": "NE CITY MSLL SPS", 
        "NE – City Mall Tegucigalpa": "CITY MALL",
        "NE – Mega Mall SPS": "MEGA MALL",
        "NE – Multiplaza Tegucigalpa": "MULTIPLAZA",
        "NE –Multiplaza SPS": "NE MULTIPLAZA TEGU",
        "NEO – Megaplaza La Ceiba": "NEO CEIBA"
    }
    
    # Datos de stock óptimo por código para Honduras - Corregidos según tabla del usuario
    optimos_codigo_data = {
        '10030709': {'MID': 12, 'MULTIPLAZA': 20, 'MEGA MALL': 12, 'CITY MALL': 20, 'CASCADAS': 12, 'NE MULTIPLAZA TEGU': 12, 'NE CITY MSLL SPS': 20, 'NEO CEIBA': 0},
        '10030708': {'MID': 6, 'MULTIPLAZA': 10, 'MEGA MALL': 6, 'CITY MALL': 10, 'CASCADAS': 6, 'NE MULTIPLAZA TEGU': 6, 'NE CITY MSLL SPS': 10, 'NEO CEIBA': 0},
        '10112874': {'MID': 3, 'MULTIPLAZA': 6, 'MEGA MALL': 3, 'CITY MALL': 6, 'CASCADAS': 3, 'NE MULTIPLAZA TEGU': 3, 'NE CITY MSLL SPS': 6, 'NEO CEIBA': 0},
        '11591122': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11591128': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11591150': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11591175': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70331909': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70331911': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70331962': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '10975804': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '10975815': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '10975835': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70192970': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70353249': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70353266': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70360899': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70360903': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70428987': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70430338': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70457634': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11591024': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11591025': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11591026': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11591043': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11591046': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11591047': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11591077': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11591078': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11941921': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70556851': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70556867': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70556869': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '70558225': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '10047511': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '10047531': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '10047538': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11405605': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11405614': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '12650335': {'MID': 24, 'MULTIPLAZA': 60, 'MEGA MALL': 48, 'CITY MALL': 60, 'CASCADAS': 60, 'NE MULTIPLAZA TEGU': 24, 'NE CITY MSLL SPS': 48, 'NEO CEIBA': 0},
        '12650337': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '12650340': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '12650342': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '12650343': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '12650344': {'MID': 18, 'MULTIPLAZA': 48, 'MEGA MALL': 36, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 18, 'NE CITY MSLL SPS': 36, 'NEO CEIBA': 0},
        '11169822': {'MID': 18, 'MULTIPLAZA': 36, 'MEGA MALL': 48, 'CITY MALL': 48, 'CASCADAS': 48, 'NE MULTIPLAZA TEGU': 36, 'NE CITY MSLL SPS': 36}
    }
    
    # Convertir usando mapeo de tiendas
    optimos_dict = {}
    
    for codigo, tiendas_optimos in optimos_codigo_data.items():
        optimos_dict[codigo] = {}
        
        # Para cada tienda real (como aparece en stock), buscar su equivalente en datos óptimos
        for tienda_real, tienda_codigo in mapeo_tiendas_codigo.items():
            if tienda_codigo in tiendas_optimos:
                optimos_dict[codigo][tienda_real] = tiendas_optimos[tienda_codigo]
    
    return optimos_dict

def obtener_optimos_por_tallas_honduras() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por talla y tienda para Honduras
    {tienda: {talla: cantidad_optima}}
    """
    # Mapeo de nombres de tiendas: Stock Real -> Stock Tallas 
    mapeo_tiendas_tallas = {
        "NE – Cascadas Mall Tegucigalpa": "NE CASCADAS MALL TEGUCIGALPA",
        "NE – CITY MALL SP": "NE CITY MALL SAN PEDRO SULA", 
        "NE – City Mall Tegucigalpa": "NE CITY MALL TEGUCIGALPA",
        "NE – Mega Mall SPS": "NE MEGA MALL SAN PEDRO SULA",
        "NE – Multiplaza Tegucigalpa": "NE MULTIPLAZA TEGUCIGALPA",
        "NE –Multiplaza SPS": "NE MULTIPLAZA SAN PEDRO SULA",
        "NEO – Megaplaza La Ceiba": "NEO MEGAPLAZA LA CEIBA"
    }
    
    # Datos de stock óptimo por tallas para Honduras - Datos correctos según tabla del usuario (todas suman 12)
    optimos_tallas_data = {
        "NE CITY MALL TEGUCIGALPA": {'678': 0, '700': 1, '718': 2, '714': 2, '738': 3, '712': 3, '758': 1, '734': 0, '778': 0, '800': 0},       # suma: 12
        "NE MULTIPLAZA TEGUCIGALPA": {'678': 0, '700': 1, '718': 2, '714': 2, '738': 3, '712': 3, '758': 1, '734': 0, '778': 0, '800': 0},      # suma: 12
        "NE CASCADAS MALL TEGUCIGALPA": {'678': 0, '700': 1, '718': 2, '714': 2, '738': 3, '712': 3, '758': 1, '734': 0, '778': 0, '800': 0},  # suma: 12
        "NE MEGA MALL SAN PEDRO SULA": {'678': 0, '700': 1, '718': 1, '714': 2, '738': 3, '712': 3, '758': 1, '734': 1, '778': 0, '800': 0},    # suma: 12
        "NE MULTIPLAZA SAN PEDRO SULA": {'678': 0, '700': 1, '718': 1, '714': 2, '738': 3, '712': 3, '758': 1, '734': 1, '778': 0, '800': 0},   # suma: 12
        "NE CITY MALL SAN PEDRO SULA": {'678': 0, '700': 1, '718': 1, '714': 2, '738': 3, '712': 3, '758': 1, '734': 1, '778': 0, '800': 0},    # suma: 12
        "NEO MEGAPLAZA LA CEIBA": {'678': 0, '700': 0, '718': 0, '714': 0, '738': 0, '712': 0, '758': 0, '734': 0, '778': 0, '800': 0}           # suma: 0 (sin distribución)
    }
    
    # Convertir usando mapeo de tiendas
    optimos_dict = {}
    
    for tienda_real, tienda_tallas in mapeo_tiendas_tallas.items():
        if tienda_tallas in optimos_tallas_data:
            optimos_dict[tienda_real] = optimos_tallas_data[tienda_tallas]
    
    return optimos_dict

def obtener_optimos_mvp_costarica() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por código y tienda para Costa Rica
    {codigo: {tienda: cantidad_optima}}
    """
    # Mapeo de nombres de tiendas: Stock Real -> Stock Códigos
    mapeo_tiendas_codigo = {
        "NE City Mall": "NE CITY MALL ALAJUELA"
    }
    
    # Datos de stock óptimo por código para Costa Rica
    optimos_codigo_data = {
        '10030709': {'MID': 15, 'NE CITY MALL ALAJUELA': 15},
        '10030708': {'MID': 6, 'NE CITY MALL ALAJUELA': 6},
        '10112874': {'MID': 3, 'NE CITY MALL ALAJUELA': 3},
        '11591122': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11591128': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11591150': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11591175': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70331909': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70331911': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70331962': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '10975804': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '10975815': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '10975835': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70192970': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70353249': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70353266': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70360899': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70360903': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70428987': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70430338': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70457634': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11591024': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11591025': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11591026': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11591043': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11591046': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11591047': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11591077': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11591078': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11941921': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70556851': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70556867': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70556869': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '70558225': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '10047511': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '10047531': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '10047538': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11405605': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11405614': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '12650335': {'MID': 48, 'NE CITY MALL ALAJUELA': 48},
        '12650337': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '12650340': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '12650342': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '12650343': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '12650344': {'MID': 36, 'NE CITY MALL ALAJUELA': 36},
        '11169822': {'MID': 36, 'NE CITY MALL ALAJUELA': 36}
    }
    
    # Convertir usando mapeo de tiendas
    optimos_dict = {}
    
    for codigo, tiendas_optimos in optimos_codigo_data.items():
        optimos_dict[codigo] = {}
        
        # Para cada tienda real (como aparece en stock), buscar su equivalente en datos óptimos
        for tienda_real, tienda_codigo in mapeo_tiendas_codigo.items():
            if tienda_codigo in tiendas_optimos:
                optimos_dict[codigo][tienda_real] = tiendas_optimos[tienda_codigo]
    
    return optimos_dict

def obtener_optimos_por_tallas_costarica() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por talla y tienda para Costa Rica
    {tienda: {talla: cantidad_optima}}
    """
    # Mapeo de nombres de tiendas: Stock Real -> Stock Tallas 
    mapeo_tiendas_tallas = {
        "NE City Mall": "NE CITY MALL ALAJUELA"
    }
    
    # Datos de stock óptimo por tallas para Costa Rica (1 tienda) - Incluye todas las tallas específicas
    optimos_tallas_data = {
        "NE CITY MALL ALAJUELA": {'678': 0, '700': 1, '718': 2, '714': 3, '738': 3, '712': 2, '758': 1, '734': 0, '778': 0, '800': 0}
    }
    
    # Convertir usando mapeo de tiendas
    optimos_dict = {}
    
    for tienda_real, tienda_tallas in mapeo_tiendas_tallas.items():
        if tienda_tallas in optimos_tallas_data:
            optimos_dict[tienda_real] = optimos_tallas_data[tienda_tallas]
    
    return optimos_dict

def obtener_optimos_mvp_panama() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por código y bodega para Panamá
    {codigo: {bodega: cantidad_optima}}
    """
    # Mapeo de nombres de tiendas: Stock Real (archivo CSV) -> Stock Códigos
    mapeo_tiendas = {
        "NE Albrookmall": "ALBROOK",
        "NE Metromall": "METROMALL", 
        "NE Multiplaza Panamá": "MULTIPLAZA PANAMA",
        "NE Westland": "WESTLAND"
    }
    
    # Stock óptimo por códigos (datos de Panamá)
    optimos_data = {
        "10030709": {"MULTIPLAZA PANAMA": 20, "WESTLAND": 12, "METROMALL": 20, "ALBROOK": 18},
        "10030708": {"MULTIPLAZA PANAMA": 10, "WESTLAND": 6, "METROMALL": 10, "ALBROOK": 8},
        "10112874": {"MULTIPLAZA PANAMA": 6, "WESTLAND": 3, "METROMALL": 6, "ALBROOK": 3},
        "11591122": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11591128": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11591150": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11591175": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70331909": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70331911": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70331962": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "10975804": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "10975815": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "10975835": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70192970": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70353249": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70353266": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70360899": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70360903": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70428987": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70430338": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70457634": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11591024": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11591025": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11591026": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11591043": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11591046": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11591047": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11591077": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11591078": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11941921": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70556851": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70556867": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70556869": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "70558225": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "10047511": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "10047531": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "10047538": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11405605": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11405614": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "12650335": {"MULTIPLAZA PANAMA": 60, "WESTLAND": 24, "METROMALL": 48, "ALBROOK": 48},
        "12650337": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "12650340": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "12650342": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "12650343": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "12650344": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36},
        "11169822": {"MULTIPLAZA PANAMA": 48, "WESTLAND": 18, "METROMALL": 36, "ALBROOK": 36}
    }
    
    # Convertir nombres de tiendas de códigos a nombres del stock real
    optimos_dict = {}
    for codigo, tiendas_codigo in optimos_data.items():
        optimos_dict[codigo] = {}
        for tienda_real, tienda_codigo in mapeo_tiendas.items():
            if tienda_codigo in tiendas_codigo:
                optimos_dict[codigo][tienda_real] = tiendas_codigo[tienda_codigo]
    
    return optimos_dict


def obtener_optimos_por_tallas_panama() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por talla y tienda para Panamá
    {tienda: {talla: cantidad_optima}}
    """
    # Mapeo de nombres de tiendas: Stock Real (archivo CSV) -> Stock Tallas 
    mapeo_tiendas_tallas = {
        "NE Multiplaza Panamá": "NE MULTIPLAZA PTY",
        "NE Westland": "NE WESTLAND OUTLET",
        "NE Metromall": "NE METROMALL", 
        "NE Albrookmall": "NE ALBROOK MALL"
    }
    
    # Stock óptimo por tallas (datos de Panamá) - Incluye todas las tallas específicas
    optimos_tallas_data = {
        "NE MULTIPLAZA PTY": {"678": 0, "700": 1, "718": 1, "714": 3, "738": 3, "712": 1, "758": 2, "734": 1, "778": 0, "800": 0},
        "NE WESTLAND OUTLET": {"678": 0, "700": 0, "718": 0, "714": 0, "738": 0, "712": 0, "758": 0, "734": 0, "778": 0, "800": 0},
        "NE METROMALL": {"678": 0, "700": 1, "718": 1, "714": 2, "738": 4, "712": 3, "758": 1, "734": 0, "778": 0, "800": 0},
        "NE ALBROOK MALL": {"678": 0, "700": 1, "718": 1, "714": 2, "738": 4, "712": 3, "758": 1, "734": 0, "778": 0, "800": 0}
    }
    
    # Convertir nombres de tiendas de tallas a nombres del stock real
    optimos_dict = {}
    for tienda_real, tienda_tallas in mapeo_tiendas_tallas.items():
        if tienda_tallas in optimos_tallas_data:
            optimos_dict[tienda_real] = optimos_tallas_data[tienda_tallas]
    
    return optimos_dict

def validar_cuadre_sm_ml(codigo: str, bodega: str, stock_sm: int, stock_ml: int, stock_optimo_codigo: int) -> bool:
    """
    Valida que las tallas SM y ML cuadren exactamente con el stock óptimo por código
    SM + ML debe ser igual al stock óptimo por código
    """
    total_tallas = stock_sm + stock_ml
    cuadra = total_tallas == stock_optimo_codigo
    
    if not cuadra:
        print(f"⚠️  DESCUADRE - Código: {codigo}, Bodega: {bodega}")
        print(f"   SM: {stock_sm} + ML: {stock_ml} = {total_tallas} ≠ Óptimo: {stock_optimo_codigo}")
    
    return cuadra

def calcular_tallas_sm_ml(stock_optimo_codigo: int) -> tuple:
    """
    Calcula la distribución para tallas SM y ML según nueva lógica:
    - Base: SM=3, ML=9
    - Multiplicadores: 48→x4, 36→x3, 18→x1.5
    - Para 18: redondeo hacia abajo + 1 unidad extra a ML
    Retorna (stock_sm, stock_ml)
    """
    # Determinar multiplicador según stock óptimo del código
    if stock_optimo_codigo == 48:
        # Multiplicador x4
        stock_sm = 3 * 4  # = 12
        stock_ml = 9 * 4  # = 36
    elif stock_optimo_codigo == 36:
        # Multiplicador x3
        stock_sm = 3 * 3  # = 9
        stock_ml = 9 * 3  # = 27
    elif stock_optimo_codigo == 18:
        # Multiplicador x1.5 con redondeo especial
        stock_sm = int(3 * 1.5)  # = 4 (4.5 redondeado hacia abajo)
        stock_ml = int(9 * 1.5) + 1  # = 14 (13.5 redondeado hacia abajo + 1 extra)
    else:
        # Fallback a la lógica original para casos no contemplados
        if stock_optimo_codigo % 2 == 0:
            stock_sm = stock_optimo_codigo // 2
            stock_ml = stock_optimo_codigo // 2
        else:
            stock_sm = (stock_optimo_codigo // 2) + 1
            stock_ml = stock_optimo_codigo // 2
    
    return stock_sm, stock_ml

def obtener_optimos_mvp_puerto_rico() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por código y tienda para Puerto Rico
    {codigo: {tienda: cantidad_optima}}
    """
    # Stock óptimo por código y bodega para Puerto Rico
    optimos_data = {
        '10030709': {'NE BARCELONETA': 12, 'NE CAROLINA': 20},
        '10030708': {'NE BARCELONETA': 10, 'NE CAROLINA': 10},
        '10112874': {'NE BARCELONETA': 2, 'NE CAROLINA': 6},
        '11591122': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11591128': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11591150': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11591175': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70331909': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70331911': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70331962': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '10975804': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '10975815': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '10975835': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70192970': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70353249': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70353266': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70360899': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70360903': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70428987': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70430338': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70457634': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11591024': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11591025': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11591026': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11591043': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11591046': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11591047': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11591077': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11591078': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11941921': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70556851': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70556867': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70556869': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '70558225': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '10047511': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '10047531': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '10047538': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11405605': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '11405614': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '12650335': {'NE BARCELONETA': 24, 'NE CAROLINA': 60},
        '12650337': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '12650340': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '12650342': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '12650343': {'NE BARCELONETA': 18, 'NE CAROLINA': 48},
        '12650344': {'NE BARCELONETA': 18, 'NE CAROLINA': 48}
    }
    
    return optimos_data

def obtener_optimos_por_tallas_puerto_rico() -> Dict[str, Dict[str, int]]:
    """
    Retorna diccionario con cantidades óptimas por talla y tienda para Puerto Rico
    {tienda: {talla: cantidad_optima}}
    """
    # Mapeo de nombres de tiendas: Nombre final (NE BARCELONETA/NE CAROLINA) -> Stock Tallas 
    mapeo_tiendas_tallas = {
        "NE BARCELONETA": "NEO BARCELONETA PREMIUM OUTLETS",
        "NE CAROLINA": "NE PLAZA CAROLINA"
    }
    
    # Stock óptimo por tallas (datos de Puerto Rico) - Incluye todas las tallas específicas
    optimos_tallas_data = {
        "NE PLAZA CAROLINA": {"678": 0, "700": 0, "718": 1, "714": 1, "738": 2, "712": 3, "758": 2, "734": 1, "778": 1, "800": 1},
        "NEO BARCELONETA PREMIUM OUTLETS": {"678": 0, "700": 0, "718": 0, "714": 1, "738": 1, "712": 0, "758": 1, "734": 0, "778": 0, "800": 0}
    }
    
    # Convertir nombres de tiendas de tallas a nombres finales para mostrar
    optimos_dict = {}
    for tienda_final, tienda_tallas in mapeo_tiendas_tallas.items():
        if tienda_tallas in optimos_tallas_data:
            optimos_dict[tienda_final] = optimos_tallas_data[tienda_tallas]
    
    return optimos_dict

def procesar_stock_mvps_guatemala(df_stock: pd.DataFrame) -> pd.DataFrame:
    """
    Procesa los datos de stock para códigos MVP específicos en Guatemala con nueva lógica
    - Códigos con tallas específicas usan stock óptimo por tallas
    - Códigos con tallas SM/ML usan división 50%-50% del stock óptimo por código
    - Otros códigos usan stock óptimo por código general
    - Agrega filas faltantes para tallas requeridas con stock 0
    """
    if df_stock is None or df_stock.empty:
        return pd.DataFrame()
    
    # CÓDIGOS MVP ESPECÍFICOS - SOLO ESTOS 45 SE EXTRAEN DEL ARCHIVO DE STOCK
    codigos_mvp = [
        '10030708', '10030709', '10047511', '10047531', '10047538', '10112874', 
        '10975804', '10975815', '10975835', '11405605', '11405614', 
        '11591024', '11591025', '11591026', '11591043', '11591046', '11591047', 
        '11591077', '11591078', '11591122', '11591128', '11591150', '11591175', 
        '11941921', '12650335', '12650337', '12650340', '12650342', '12650343', 
        '12650344', '70192970', '70331909', '70331911', '70331962', '70353249', 
        '70353266', '70360899', '70360903', '70428987', '70430338', '70457634', 
        '70556851', '70556867', '70556869', '70558225'
    ]
    
    # Códigos que deben tener tallas específicas (678-800)
    codigos_con_tallas = ['11591122', '11591128', '11591150', '11591175', '70331909', '70331911', '70331962']
    
    # Códigos que deben tener tallas SM y ML (división 50%-50%)
    codigos_con_tallas_sm_ml = ['10975804', '10975815', '10975835', '70192970', '70353249', '70353266', 
                                '70360899', '70360903', '70428987', '70430338', '70457634']
    
    # Tallas específicas numéricas
    tallas_especificas = ['678', '700', '718', '714', '738', '712', '758', '734', '778', '800']
    
    # Tallas SM y ML
    tallas_sm_ml = ['SM', 'ML']
    
    # Filtrar por marca NEW ERA
    df_new_era = df_stock[df_stock['U_Marca'].str.upper() == 'NEW ERA'].copy()
    
    # Filtrar por códigos MVP específicos (SOLO ESTOS 45 CÓDIGOS)
    df_mvp = df_new_era[df_new_era['U_Estilo'].astype(str).isin(codigos_mvp)].copy()
    
    # Información de debug para verificar filtrado correcto
    print(f"DEBUG MVP GUATEMALA: Total registros NEW ERA: {len(df_new_era)}")
    print(f"DEBUG MVP GUATEMALA: Códigos MVP filtrados: {len(df_mvp)}")
    print(f"DEBUG MVP GUATEMALA: Columnas disponibles: {list(df_mvp.columns)}")
    
    # Verificar qué códigos MVP están presentes en el archivo
    codigos_encontrados = df_mvp['U_Estilo'].astype(str).unique().tolist()
    print(f"DEBUG MVP GUATEMALA: Códigos encontrados: {sorted(codigos_encontrados)}")
    
    if df_mvp.empty:
        print("DEBUG MVP GUATEMALA: No se encontraron códigos MVP en el archivo")
        return pd.DataFrame()
    
    # Verificar columnas necesarias - SOPORTE PARA AMBAS: 'Talla' y 'U_Talla'
    columna_talla = None
    if 'U_Talla' in df_mvp.columns:
        columna_talla = 'U_Talla'
        print("GUATEMALA: USANDO COLUMNA U_Talla")
    elif 'Talla' in df_mvp.columns:
        columna_talla = 'Talla'
        print("GUATEMALA: USANDO COLUMNA Talla")
    else:
        print("ERROR GUATEMALA: No se encontró columna de talla")
        return pd.DataFrame()
    
    columnas_necesarias = ['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Stock_Actual', 'Bodega']
    for col in columnas_necesarias:
        if col not in df_mvp.columns:
            print(f"ERROR GUATEMALA: Columna faltante: {col}")
            return pd.DataFrame()
    
    # Obtener bodegas de Guatemala
    bodegas_guatemala = [
        "NE Oakland", "NE Cayala", "NE Miraflores", "NE Portales", "NE InterXela",
        "NE Metronorte", "NE Concepcion", "NE Interplaza Escuintla", "NE Pradera Huehuetenango",
        "NE Naranjo", "NE Metrocentro Outlet", "NE Vistares", "NE Peri Roosvelt",
        "NE Outlet Santa clara", "NE Plaza Magdalena", "NE Pradera Chiquimula",
        "NE Pradera Escuintla", "NE Paseo Antigua", "NE Pradera Xela", "NE Chimaltenango",
        "NE Plaza Videre", "NE Metroplaza Jutiapa", "NE Puerto Barrios"
    ]
    
    # Filtrar solo bodegas de Guatemala
    df_mvp_guatemala = df_mvp[df_mvp['Bodega'].isin(bodegas_guatemala)].copy()
    
    if df_mvp_guatemala.empty:
        return pd.DataFrame()
    
    # Obtener datos de óptimos
    optimos_por_codigo = obtener_optimos_mvp()
    optimos_por_tallas = obtener_optimos_por_tallas()
    
    # NUEVA LÓGICA: Agregar filas faltantes para códigos con tallas específicas
    filas_adicionales = []
    
    # 1. Códigos con tallas numéricas (678-800)
    for codigo in codigos_con_tallas:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_guatemala[df_mvp_guatemala['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de Guatemala
            for bodega in bodegas_guatemala:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para TODAS las tallas específicas (678-800)
                # para que aparezcan en la tabla aunque tengan stock óptimo 0
                for talla_req in tallas_especificas:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    # 2. Códigos con tallas SM y ML (división 50%-50%)
    for codigo in codigos_con_tallas_sm_ml:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_guatemala[df_mvp_guatemala['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de Guatemala
            for bodega in bodegas_guatemala:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para SM y ML si no existen
                for talla_req in tallas_sm_ml:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    # Crear mapa de SAP solo para registros REALES (con stock > 0)
    df_sap_map = df_mvp_guatemala[df_mvp_guatemala['Stock_Actual'] > 0].groupby(['U_Estilo', columna_talla])['Codigo_SAP'].first().reset_index()
    sap_dict = {}
    for _, row in df_sap_map.iterrows():
        key = (str(row['U_Estilo']), str(row[columna_talla]))
        sap_dict[key] = row['Codigo_SAP']
    
    # Agregar filas adicionales al DataFrame
    if filas_adicionales:
        df_adicional = pd.DataFrame(filas_adicionales)
        df_mvp_guatemala = pd.concat([df_mvp_guatemala, df_adicional], ignore_index=True)
    
    # NUEVA FUNCIONALIDAD: Agregar códigos MVP faltantes con información "N/D"
    # Identificar códigos MVP que NO están en el archivo
    codigos_presentes = df_mvp_guatemala['U_Estilo'].astype(str).unique().tolist()
    codigos_faltantes = [codigo for codigo in codigos_mvp if codigo not in codigos_presentes]
    
    if codigos_faltantes:
        print(f"DEBUG MVP GUATEMALA: Agregando {len(codigos_faltantes)} códigos MVP faltantes: {sorted(codigos_faltantes)}")
        
        # Crear filas para códigos faltantes
        filas_faltantes = []
        for codigo_faltante in codigos_faltantes:
            for bodega in bodegas_guatemala:
                # Crear fila con información "N/D"
                fila_faltante = {
                    'U_Estilo': codigo_faltante,
                    'Codigo_SAP': '',  # Vacío porque no existe en archivo
                    'U_Segmento': 'N/D',
                    'U_Silueta': 'N/D', 
                    'U_Coleccion_NE': 'N/D',
                    'U_Descripcion': 'N/D',
                    columna_talla: 'N/D',
                    'Stock_Actual': 0,  # Sin stock porque no está en archivo
                    'Bodega': bodega
                }
                filas_faltantes.append(fila_faltante)
        
        # Agregar filas faltantes al DataFrame
        if filas_faltantes:
            df_faltantes = pd.DataFrame(filas_faltantes)
            df_mvp_guatemala = pd.concat([df_mvp_guatemala, df_faltantes], ignore_index=True)
    
    # Agrupar como tabla dinámica
    df_agrupado = df_mvp_guatemala.groupby([
        'U_Estilo', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Bodega'
    ])['Stock_Actual'].sum().reset_index()
    
    # Asignar SAP basado en mapa de códigos+tallas REALES
    df_agrupado['Codigo_SAP'] = df_agrupado.apply(
        lambda row: sap_dict.get((str(row['U_Estilo']), str(row[columna_talla])), ""), 
        axis=1
    )
    
    # Pivotar para tener bodegas como columnas
    tabla_pivoteada = df_agrupado.pivot_table(
        index=['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla],
        columns='Bodega',
        values='Stock_Actual',
        fill_value=0,
        aggfunc='sum'
    )
    
    # Asegurar que todas las bodegas estén presentes como columnas
    for bodega in bodegas_guatemala:
        if bodega not in tabla_pivoteada.columns:
            tabla_pivoteada[bodega] = 0
    
    # Reordenar columnas según el orden de bodegas_guatemala
    tabla_pivoteada = tabla_pivoteada.reindex(columns=bodegas_guatemala, fill_value=0)
    
    # Crear nueva tabla con columnas intercaladas (stock actual + stock óptimo nuevo)
    tabla_final = pd.DataFrame(index=tabla_pivoteada.index)
    
    # Agregar columnas intercaladas para cada bodega
    for bodega in bodegas_guatemala:
        # Columna de stock actual
        tabla_final[f"Real {bodega}"] = tabla_pivoteada[bodega]
        
        # Columna de stock óptimo nuevo
        col_optimo = f"Óptimo {bodega}"
        tabla_final[col_optimo] = 0
        
        # Llenar valores óptimos para cada código/talla
        for codigo_tuple in tabla_pivoteada.index:
            codigo = str(codigo_tuple[0])
            talla = str(codigo_tuple[6])  # Talla está en el índice 6 (después de agregar Codigo_SAP)
            
            # Determinar tipo de código y calcular stock óptimo
            if codigo in codigos_con_tallas and talla in tallas_especificas:
                # CASO 1: Códigos con tallas numéricas (678-800) - Calcular tallas basado en código ÷ 12
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo] and bodega in optimos_por_tallas:
                    # Paso 1: Obtener stock óptimo por código
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    
                    # Paso 2: Obtener distribución base de tallas para esta bodega
                    tallas_base = optimos_por_tallas[bodega]
                    
                    # Paso 3: Calcular factor multiplicador
                    suma_tallas_base = sum(tallas_base.values())
                    if suma_tallas_base > 0:
                        factor = stock_codigo / suma_tallas_base
                        
                        # Paso 4: Calcular nueva distribución
                        tallas_calculadas = {}
                        suma_redondeada = 0
                        
                        # Multiplicar y redondear hacia abajo
                        for talla_key, valor_base in tallas_base.items():
                            valor_calculado = int(valor_base * factor)  # Redondeo hacia abajo
                            tallas_calculadas[talla_key] = valor_calculado
                            suma_redondeada += valor_calculado
                        
                        # Paso 5: Ajustar para que cuadre exactamente
                        diferencia = stock_codigo - suma_redondeada
                        
                        if diferencia > 0:
                            # Encontrar talla con mayor valor (primera en caso de empate)
                            talla_maxima = max(tallas_calculadas.keys(), key=lambda k: tallas_calculadas[k])
                            tallas_calculadas[talla_maxima] += diferencia
                        
                        # Paso 6: Asignar valor para esta talla específica
                        if talla in tallas_calculadas:
                            tabla_final.loc[codigo_tuple, col_optimo] = tallas_calculadas[talla]
                        else:
                            tabla_final.loc[codigo_tuple, col_optimo] = 0
                    else:
                        tabla_final.loc[codigo_tuple, col_optimo] = 0
                else:
                    tabla_final.loc[codigo_tuple, col_optimo] = 0
                    
            elif codigo in codigos_con_tallas_sm_ml and talla in tallas_sm_ml:
                # CASO 2: Códigos con tallas SM y ML - División 50%-50%
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    stock_sm, stock_ml = calcular_tallas_sm_ml(stock_codigo)
                    
                    # Asignar según la talla
                    if talla == 'SM':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_sm
                    elif talla == 'ML':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_ml
                    
                    # Validar cuadre
                    validar_cuadre_sm_ml(codigo, bodega, stock_sm, stock_ml, stock_codigo)
                else:
                    tabla_final.loc[codigo_tuple, col_optimo] = 0
                    
            else:
                # CASO 3: Códigos sin tallas específicas - Usar stock óptimo por código directo
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    tabla_final.loc[codigo_tuple, col_optimo] = optimos_por_codigo[codigo][bodega]
    
    # Agregar fila de totales por columna
    fila_totales = tabla_final.sum(axis=0)
    fila_totales.name = ('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')
    tabla_final = pd.concat([tabla_final, fila_totales.to_frame().T])
    
    return tabla_final

def procesar_stock_mvps_elsalvador(df_stock: pd.DataFrame) -> pd.DataFrame:
    """
    Procesa los datos de stock para códigos MVP específicos en El Salvador con misma lógica que Guatemala
    - Códigos con tallas específicas usan stock óptimo por tallas
    - Otros códigos usan stock óptimo por código general
    - Agrega filas faltantes para tallas requeridas con stock 0
    """
    if df_stock is None or df_stock.empty:
        return pd.DataFrame()
    
    # CÓDIGOS MVP ESPECÍFICOS - MISMOS QUE GUATEMALA
    codigos_mvp = [
        '10030708', '10030709', '10047511', '10047531', '10047538', '10112874', 
        '10975804', '10975815', '10975835', '11405605', '11405614', 
        '11591024', '11591025', '11591026', '11591043', '11591046', '11591047', 
        '11591077', '11591078', '11591122', '11591128', '11591150', '11591175', 
        '11941921', '12650335', '12650337', '12650340', '12650342', '12650343', 
        '12650344', '70192970', '70331909', '70331911', '70331962', '70353249', 
        '70353266', '70360899', '70360903', '70428987', '70430338', '70457634', 
        '70556851', '70556867', '70556869', '70558225'
    ]
    
    # Códigos que deben tener tallas específicas (678-800)
    codigos_con_tallas = ['11591122', '11591128', '11591150', '11591175', '70331909', '70331911', '70331962']
    
    # Códigos que deben tener tallas SM y ML (división 50%-50%)
    codigos_con_tallas_sm_ml = ['10975804', '10975815', '10975835', '70192970', '70353249', '70353266', 
                                '70360899', '70360903', '70428987', '70430338', '70457634']
    
    # Tallas específicas numéricas
    tallas_especificas = ['678', '700', '718', '714', '738', '712', '758', '734', '778', '800']
    
    # Tallas SM y ML
    tallas_sm_ml = ['SM', 'ML']
    
    # Filtrar por marca NEW ERA
    df_new_era = df_stock[df_stock['U_Marca'].str.upper() == 'NEW ERA'].copy()
    
    # Filtrar por códigos MVP específicos
    df_mvp = df_new_era[df_new_era['U_Estilo'].astype(str).isin(codigos_mvp)].copy()
    
    # Información de debug
    print(f"DEBUG MVP EL SALVADOR: Total registros NEW ERA: {len(df_new_era)}")
    print(f"DEBUG MVP EL SALVADOR: Códigos MVP filtrados: {len(df_mvp)}")
    
    # Verificar qué códigos MVP están presentes
    codigos_encontrados = df_mvp['U_Estilo'].astype(str).unique().tolist()
    print(f"DEBUG MVP EL SALVADOR: Códigos encontrados: {sorted(codigos_encontrados)}")
    
    if df_mvp.empty:
        print("DEBUG MVP EL SALVADOR: No se encontraron códigos MVP en el archivo")
        return pd.DataFrame()
    
    # Verificar columnas necesarias - SOPORTE PARA AMBAS: 'Talla' y 'U_Talla'
    columna_talla = None
    if 'U_Talla' in df_mvp.columns:
        columna_talla = 'U_Talla'
    elif 'Talla' in df_mvp.columns:
        columna_talla = 'Talla'
    else:
        print("ERROR: No se encontró columna de talla ('U_Talla' o 'Talla')")
        return pd.DataFrame()
    
    columnas_necesarias = ['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Stock_Actual', 'Bodega']
    for col in columnas_necesarias:
        if col not in df_mvp.columns:
            print(f"ERROR: Columna faltante: {col}")
            return pd.DataFrame()
    
    # Obtener bodegas de El Salvador
    bodegas_elsalvador = [
        "NE METROCENTRO LOURDES", "NE METROCENTRO SAN MIGUEL", "NE PLAZA MUNDO SOYAPANGO",
        "NE USULUTÁN", "NEW ERA EL PASEO", "NEW ERA METROCENTRO", 
        "NEW ERA METROCENTRO SANTA ANA", "NEW ERA MULTIPLAZA"
    ]
    
    # Filtrar solo bodegas de El Salvador
    df_mvp_elsalvador = df_mvp[df_mvp['Bodega'].isin(bodegas_elsalvador)].copy()
    
    if df_mvp_elsalvador.empty:
        return pd.DataFrame()
    
    # Obtener datos de óptimos para El Salvador (IGUAL QUE GUATEMALA)
    optimos_por_codigo = obtener_optimos_mvp_elsalvador()
    optimos_por_tallas = obtener_optimos_por_tallas_elsalvador()
    
    # NUEVA LÓGICA: Agregar filas faltantes para códigos con tallas específicas
    filas_adicionales = []
    
    # 1. Códigos con tallas numéricas (678-800)
    for codigo in codigos_con_tallas:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_elsalvador[df_mvp_elsalvador['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de El Salvador
            for bodega in bodegas_elsalvador:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para TODAS las tallas específicas (678-800)
                # para que aparezcan en la tabla aunque tengan stock óptimo 0
                for talla_req in tallas_especificas:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    # 2. Códigos con tallas SM y ML (división 50%-50%)
    for codigo in codigos_con_tallas_sm_ml:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_elsalvador[df_mvp_elsalvador['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de El Salvador
            for bodega in bodegas_elsalvador:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para SM y ML si no existen
                for talla_req in tallas_sm_ml:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    # Crear mapa de SAP solo para registros REALES (con stock > 0) - IGUAL QUE GUATEMALA
    df_sap_map = df_mvp_elsalvador[df_mvp_elsalvador['Stock_Actual'] > 0].groupby(['U_Estilo', columna_talla])['Codigo_SAP'].first().reset_index()
    sap_dict = {}
    for _, row in df_sap_map.iterrows():
        key = (str(row['U_Estilo']), str(row[columna_talla]))
        sap_dict[key] = row['Codigo_SAP']
    
    # Agregar filas faltantes al DataFrame
    if filas_adicionales:
        df_adicional = pd.DataFrame(filas_adicionales)
        df_mvp_elsalvador = pd.concat([df_mvp_elsalvador, df_adicional], ignore_index=True)
    
    # NUEVA FUNCIONALIDAD: Agregar códigos MVP faltantes con información "N/D"
    # Identificar códigos MVP que NO están en el archivo
    codigos_presentes = df_mvp_elsalvador['U_Estilo'].astype(str).unique().tolist()
    codigos_faltantes = [codigo for codigo in codigos_mvp if codigo not in codigos_presentes]
    
    if codigos_faltantes:
        print(f"DEBUG MVP EL SALVADOR: Agregando {len(codigos_faltantes)} códigos MVP faltantes: {sorted(codigos_faltantes)}")
        
        # Crear filas para códigos faltantes
        filas_faltantes = []
        for codigo_faltante in codigos_faltantes:
            for bodega in bodegas_elsalvador:
                # Crear fila con información "N/D"
                fila_faltante = {
                    'U_Estilo': codigo_faltante,
                    'Codigo_SAP': '',  # Vacío porque no existe en archivo
                    'U_Segmento': 'N/D',
                    'U_Silueta': 'N/D', 
                    'U_Coleccion_NE': 'N/D',
                    'U_Descripcion': 'N/D',
                    columna_talla: 'N/D',
                    'Stock_Actual': 0,  # Sin stock porque no está en archivo
                    'Bodega': bodega
                }
                filas_faltantes.append(fila_faltante)
        
        # Agregar filas faltantes al DataFrame
        if filas_faltantes:
            df_faltantes = pd.DataFrame(filas_faltantes)
            df_mvp_elsalvador = pd.concat([df_mvp_elsalvador, df_faltantes], ignore_index=True)
    
    # Agrupar como tabla dinámica (IGUAL QUE GUATEMALA)
    df_agrupado = df_mvp_elsalvador.groupby([
        'U_Estilo', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Bodega'
    ])['Stock_Actual'].sum().reset_index()
    
    # Asignar SAP basado en mapa de códigos+tallas REALES
    df_agrupado['Codigo_SAP'] = df_agrupado.apply(
        lambda row: sap_dict.get((str(row['U_Estilo']), str(row[columna_talla])), ""), 
        axis=1
    )
    
    # Pivotar para tener bodegas como columnas (IGUAL QUE GUATEMALA)
    tabla_pivoteada = df_agrupado.pivot_table(
        index=['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla],
        columns='Bodega',
        values='Stock_Actual',
        fill_value=0,
        aggfunc='sum'  # ✅ AGREGADO - ESTO CORRIGE LAS DUPLICACIONES
    )
    
    # Asegurar que todas las bodegas estén presentes como columnas
    for bodega in bodegas_elsalvador:
        if bodega not in tabla_pivoteada.columns:
            tabla_pivoteada[bodega] = 0
    
    # Reordenar columnas según el orden de bodegas_elsalvador
    tabla_pivoteada = tabla_pivoteada.reindex(columns=bodegas_elsalvador, fill_value=0)
    
    # Crear nueva tabla con columnas intercaladas (stock actual + stock óptimo nuevo)
    tabla_final = pd.DataFrame(index=tabla_pivoteada.index)
    
    # Agregar columnas intercaladas para cada bodega
    for bodega in bodegas_elsalvador:
        # Columna de stock actual
        tabla_final[f"Real {bodega}"] = tabla_pivoteada[bodega]
        
        # Columna de stock óptimo nuevo
        col_optimo = f"Óptimo {bodega}"
        tabla_final[col_optimo] = 0
    
    # MISMA LÓGICA DE CÁLCULO QUE GUATEMALA
    for codigo_tuple in tabla_final.index:
        codigo = codigo_tuple[0]
        talla = str(codigo_tuple[6])  # Talla está en el índice 6 (corregido para coincidir con Guatemala)
        
        for bodega in bodegas_elsalvador:
            col_optimo = f"Óptimo {bodega}"
            
            # Determinar tipo de código y calcular stock óptimo
            if codigo in codigos_con_tallas and talla in tallas_especificas:
                # CASO 1: Códigos con tallas numéricas (678-800) - Calcular tallas basado en código ÷ 12
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo] and bodega in optimos_por_tallas:
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    tallas_base = optimos_por_tallas[bodega]
                    
                    suma_tallas_base = sum(tallas_base.values())
                    if suma_tallas_base > 0:
                        factor = stock_codigo / suma_tallas_base
                        
                        tallas_calculadas = {}
                        suma_redondeada = 0
                        
                        for talla_key, valor_base in tallas_base.items():
                            valor_calculado = int(valor_base * factor)
                            tallas_calculadas[talla_key] = valor_calculado
                            suma_redondeada += valor_calculado
                        
                        diferencia = stock_codigo - suma_redondeada
                        if diferencia > 0:
                            talla_maxima = max(tallas_calculadas.keys(), key=lambda k: tallas_calculadas[k])
                            tallas_calculadas[talla_maxima] += diferencia
                        
                        if talla in tallas_calculadas:
                            tabla_final.loc[codigo_tuple, col_optimo] = tallas_calculadas[talla]
                        else:
                            tabla_final.loc[codigo_tuple, col_optimo] = 0
                    else:
                        tabla_final.loc[codigo_tuple, col_optimo] = 0
                else:
                    tabla_final.loc[codigo_tuple, col_optimo] = 0
                    
            elif codigo in codigos_con_tallas_sm_ml and talla in tallas_sm_ml:
                # CASO 2: Códigos con tallas SM y ML - División 50%-50%
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    stock_sm, stock_ml = calcular_tallas_sm_ml(stock_codigo)
                    
                    # Asignar según la talla
                    if talla == 'SM':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_sm
                    elif talla == 'ML':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_ml
                    
                    # Validar cuadre
                    validar_cuadre_sm_ml(codigo, bodega, stock_sm, stock_ml, stock_codigo)
                else:
                    tabla_final.loc[codigo_tuple, col_optimo] = 0
                    
            else:
                # CASO 3: Códigos sin tallas específicas - Usar stock óptimo por código directo
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    tabla_final.loc[codigo_tuple, col_optimo] = optimos_por_codigo[codigo][bodega]
    
    # Agregar fila de totales
    fila_totales = tabla_final.sum(axis=0)
    fila_totales.name = ('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')
    tabla_final = pd.concat([tabla_final, fila_totales.to_frame().T])
    
    return tabla_final

def procesar_stock_mvps_honduras(df_stock: pd.DataFrame) -> pd.DataFrame:
    """
    Procesa los datos de stock para códigos MVP específicos en Honduras con misma lógica que Guatemala/El Salvador
    - Códigos con tallas específicas usan stock óptimo por tallas
    - Otros códigos usan stock óptimo por código general
    - Agrega filas faltantes para tallas requeridas con stock 0
    """
    if df_stock is None or df_stock.empty:
        return pd.DataFrame()
    
    # CÓDIGOS MVP ESPECÍFICOS - MISMOS QUE GUATEMALA Y EL SALVADOR
    codigos_mvp = [
        '10030708', '10030709', '10047511', '10047531', '10047538', '10112874', 
        '10975804', '10975815', '10975835', '11405605', '11405614', 
        '11591024', '11591025', '11591026', '11591043', '11591046', '11591047', 
        '11591077', '11591078', '11591122', '11591128', '11591150', '11591175', 
        '11941921', '12650335', '12650337', '12650340', '12650342', '12650343', 
        '12650344', '70192970', '70331909', '70331911', '70331962', '70353249', 
        '70353266', '70360899', '70360903', '70428987', '70430338', '70457634', 
        '70556851', '70556867', '70556869', '70558225'
    ]
    
    # Códigos que deben tener tallas específicas (678-800)
    codigos_con_tallas = ['11591122', '11591128', '11591150', '11591175', '70331909', '70331911', '70331962']
    
    # Códigos que deben tener tallas SM y ML (división 50%-50%)
    codigos_con_tallas_sm_ml = ['10975804', '10975815', '10975835', '70192970', '70353249', '70353266', 
                                '70360899', '70360903', '70428987', '70430338', '70457634']
    
    # Tallas específicas numéricas
    tallas_especificas = ['678', '700', '718', '714', '738', '712', '758', '734', '778', '800']
    
    # Tallas SM y ML
    tallas_sm_ml = ['SM', 'ML']
    
    # Filtrar por marca NEW ERA
    df_new_era = df_stock[df_stock['U_Marca'].str.upper() == 'NEW ERA'].copy()
    
    # Filtrar por códigos MVP específicos
    df_mvp = df_new_era[df_new_era['U_Estilo'].astype(str).isin(codigos_mvp)].copy()
    
    # Información de debug
    print(f"DEBUG MVP HONDURAS: Total registros NEW ERA: {len(df_new_era)}")
    print(f"DEBUG MVP HONDURAS: Códigos MVP filtrados: {len(df_mvp)}")
    
    # Verificar qué códigos MVP están presentes
    codigos_encontrados = df_mvp['U_Estilo'].astype(str).unique().tolist()
    print(f"DEBUG MVP HONDURAS: Códigos encontrados: {sorted(codigos_encontrados)}")
    
    if df_mvp.empty:
        print("DEBUG MVP HONDURAS: No se encontraron códigos MVP en el archivo")
        return pd.DataFrame()
    
    # Verificar columnas necesarias - SOPORTE PARA AMBAS: 'Talla' y 'U_Talla'
    columna_talla = None
    if 'U_Talla' in df_mvp.columns:
        columna_talla = 'U_Talla'
    elif 'Talla' in df_mvp.columns:
        columna_talla = 'Talla'
    else:
        print("ERROR: No se encontró columna de talla ('U_Talla' o 'Talla')")
        return pd.DataFrame()
    
    columnas_necesarias = ['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Stock_Actual', 'Bodega']
    for col in columnas_necesarias:
        if col not in df_mvp.columns:
            print(f"ERROR: Columna faltante: {col}")
            return pd.DataFrame()
    
    # Obtener bodegas de Honduras
    bodegas_honduras = [
        "NE – Cascadas Mall Tegucigalpa", "NE – CITY MALL SP", "NE – City Mall Tegucigalpa",
        "NE – Mega Mall SPS", "NE – Multiplaza Tegucigalpa", "NE –Multiplaza SPS",
        "NEO – Megaplaza La Ceiba"
    ]
    
    # Filtrar solo bodegas de Honduras
    df_mvp_honduras = df_mvp[df_mvp['Bodega'].isin(bodegas_honduras)].copy()
    
    if df_mvp_honduras.empty:
        return pd.DataFrame()
    
    # Crear mapa de SAP solo para registros REALES (con stock > 0)
    df_sap_map = df_mvp_honduras[df_mvp_honduras['Stock_Actual'] > 0].groupby(['U_Estilo', columna_talla])['Codigo_SAP'].first().reset_index()
    sap_dict = {}
    for _, row in df_sap_map.iterrows():
        key = (str(row['U_Estilo']), str(row[columna_talla]))
        sap_dict[key] = row['Codigo_SAP']
    
    # Obtener datos de óptimos para Honduras
    optimos_por_codigo = obtener_optimos_mvp_honduras()
    optimos_por_tallas = obtener_optimos_por_tallas_honduras()
    
    # NUEVA LÓGICA: Agregar filas faltantes para códigos con tallas específicas
    filas_adicionales = []
    
    # 1. Códigos con tallas numéricas (678-800)
    for codigo in codigos_con_tallas:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_honduras[df_mvp_honduras['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de Honduras
            for bodega in bodegas_honduras:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para TODAS las tallas específicas (678-800)
                # para que aparezcan en la tabla aunque tengan stock óptimo 0
                for talla_req in tallas_especificas:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    # 2. Códigos con tallas SM y ML (división 50%-50%)
    for codigo in codigos_con_tallas_sm_ml:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_honduras[df_mvp_honduras['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de Honduras
            for bodega in bodegas_honduras:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para SM y ML si no existen
                for talla_req in tallas_sm_ml:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    # Agregar filas faltantes al DataFrame
    if filas_adicionales:
        df_adicional = pd.DataFrame(filas_adicionales)
        df_mvp_honduras = pd.concat([df_mvp_honduras, df_adicional], ignore_index=True)
    
    # NUEVA FUNCIONALIDAD: Agregar códigos MVP faltantes con información "N/D"
    # Identificar códigos MVP que NO están en el archivo
    codigos_presentes = df_mvp_honduras['U_Estilo'].astype(str).unique().tolist()
    codigos_faltantes = [codigo for codigo in codigos_mvp if codigo not in codigos_presentes]
    
    if codigos_faltantes:
        print(f"DEBUG MVP HONDURAS: Agregando {len(codigos_faltantes)} códigos MVP faltantes: {sorted(codigos_faltantes)}")
        
        # Crear filas para códigos faltantes
        filas_faltantes = []
        for codigo_faltante in codigos_faltantes:
            for bodega in bodegas_honduras:
                # Crear fila con información "N/D"
                fila_faltante = {
                    'U_Estilo': codigo_faltante,
                    'Codigo_SAP': '',  # Vacío porque no existe en archivo
                    'U_Segmento': 'N/D',
                    'U_Silueta': 'N/D', 
                    'U_Coleccion_NE': 'N/D',
                    'U_Descripcion': 'N/D',
                    columna_talla: 'N/D',
                    'Stock_Actual': 0,  # Sin stock porque no está en archivo
                    'Bodega': bodega
                }
                filas_faltantes.append(fila_faltante)
        
        # Agregar filas faltantes al DataFrame
        if filas_faltantes:
            df_faltantes = pd.DataFrame(filas_faltantes)
            df_mvp_honduras = pd.concat([df_mvp_honduras, df_faltantes], ignore_index=True)
    
    # Agrupar como tabla dinámica (SIN Codigo_SAP en el groupby)
    df_agrupado = df_mvp_honduras.groupby([
        'U_Estilo', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Bodega'
    ])['Stock_Actual'].sum().reset_index()
    
    # Asignar SAP basado en mapa de códigos+tallas REALES
    df_agrupado['Codigo_SAP'] = df_agrupado.apply(
        lambda row: sap_dict.get((str(row['U_Estilo']), str(row[columna_talla])), ""), 
        axis=1
    )
    
    # Pivotar para tener bodegas como columnas
    tabla_pivoteada = df_agrupado.pivot_table(
        index=['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla],
        columns='Bodega',
        values='Stock_Actual',
        fill_value=0,
        aggfunc='sum'
    )
    
    # Asegurar que todas las bodegas estén presentes como columnas
    for bodega in bodegas_honduras:
        if bodega not in tabla_pivoteada.columns:
            tabla_pivoteada[bodega] = 0
    
    # Reordenar columnas según el orden de bodegas_honduras
    tabla_pivoteada = tabla_pivoteada.reindex(columns=bodegas_honduras, fill_value=0)
    
    # Crear nueva tabla con columnas intercaladas (stock actual + stock óptimo nuevo)
    tabla_final = pd.DataFrame(index=tabla_pivoteada.index)
    
    # Agregar columnas intercaladas para cada bodega
    for bodega in bodegas_honduras:
        # Columna de stock actual
        tabla_final[f"Real {bodega}"] = tabla_pivoteada[bodega]
        
        # Columna de stock óptimo nuevo
        col_optimo = f"Óptimo {bodega}"
        tabla_final[col_optimo] = 0
    
        # Llenar valores óptimos para cada código/talla
        for codigo_tuple in tabla_pivoteada.index:
            codigo = str(codigo_tuple[0])
            talla = str(codigo_tuple[6])  # Talla está en el índice 6 (después de agregar Codigo_SAP)
            
            # Determinar tipo de código y calcular stock óptimo
            if codigo in codigos_con_tallas and talla in tallas_especificas:
                # CASO 1: Códigos con tallas numéricas (678-800) - Calcular tallas basado en código ÷ 12
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo] and bodega in optimos_por_tallas:
                    # Paso 1: Obtener stock óptimo por código
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    
                    # Paso 2: Obtener distribución base de tallas para esta bodega
                    tallas_base = optimos_por_tallas[bodega]
                    
                    # Paso 3: Calcular factor multiplicador
                    suma_tallas_base = sum(tallas_base.values())
                    if suma_tallas_base > 0:
                        factor = stock_codigo / suma_tallas_base
                        
                        # Paso 4: Calcular nueva distribución
                        tallas_calculadas = {}
                        suma_redondeada = 0
                        
                        # Multiplicar y redondear hacia abajo
                        for talla_key, valor_base in tallas_base.items():
                            valor_calculado = int(valor_base * factor)  # Redondeo hacia abajo
                            tallas_calculadas[talla_key] = valor_calculado
                            suma_redondeada += valor_calculado
                        
                        # Paso 5: Ajustar para que cuadre exactamente
                        diferencia = stock_codigo - suma_redondeada
                        
                        if diferencia > 0:
                            # Distribuir diferencia positiva entre las tallas con mayor base
                            tallas_ordenadas = sorted(tallas_base.items(), key=lambda x: x[1], reverse=True)
                            for i, (talla_key, _) in enumerate(tallas_ordenadas[:diferencia]):
                                tallas_calculadas[talla_key] += 1
                        
                        # Paso 6: Asignar valor calculado si la talla coincide
                        if talla in tallas_calculadas:
                            tabla_final.loc[codigo_tuple, col_optimo] = tallas_calculadas[talla]
                else:
                    tabla_final.loc[codigo_tuple, col_optimo] = 0
                    
            elif codigo in codigos_con_tallas_sm_ml and talla in tallas_sm_ml:
                # CASO 2: Códigos con tallas SM y ML - División 50%-50%
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    stock_sm, stock_ml = calcular_tallas_sm_ml(stock_codigo)
                    
                    # Asignar según la talla
                    if talla == 'SM':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_sm
                    elif talla == 'ML':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_ml
                    
                    # Validar cuadre
                    validar_cuadre_sm_ml(codigo, bodega, stock_sm, stock_ml, stock_codigo)
                else:
                    tabla_final.loc[codigo_tuple, col_optimo] = 0
                    
            else:
                # CASO 3: Códigos sin tallas específicas - Usar stock óptimo por código directo
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    tabla_final.loc[codigo_tuple, col_optimo] = optimos_por_codigo[codigo][bodega]
    
    # Agregar fila de totales
    fila_totales = tabla_final.sum(axis=0)
    fila_totales.name = ('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')
    tabla_final = pd.concat([tabla_final, fila_totales.to_frame().T])
    
    return tabla_final

def procesar_stock_mvps_costarica(df_stock: pd.DataFrame) -> pd.DataFrame:
    """
    Procesa los datos de stock para códigos MVP específicos en Costa Rica con misma lógica que Guatemala/El Salvador/Honduras
    - Códigos con tallas específicas usan stock óptimo por tallas
    - Otros códigos usan stock óptimo por código general
    - Agrega filas faltantes para tallas requeridas con stock 0
    """
    if df_stock is None or df_stock.empty:
        return pd.DataFrame()
    
    # CÓDIGOS MVP ESPECÍFICOS - MISMOS QUE GUATEMALA, EL SALVADOR Y HONDURAS
    codigos_mvp = [
        '10030708', '10030709', '10047511', '10047531', '10047538', '10112874', 
        '10975804', '10975815', '10975835', '11405605', '11405614', 
        '11591024', '11591025', '11591026', '11591043', '11591046', '11591047', 
        '11591077', '11591078', '11591122', '11591128', '11591150', '11591175', 
        '11941921', '12650335', '12650337', '12650340', '12650342', '12650343', 
        '12650344', '70192970', '70331909', '70331911', '70331962', '70353249', 
        '70353266', '70360899', '70360903', '70428987', '70430338', '70457634', 
        '70556851', '70556867', '70556869', '70558225'
    ]
    
    # Códigos que deben tener tallas específicas (678-800)
    codigos_con_tallas = ['11591122', '11591128', '11591150', '11591175', '70331909', '70331911', '70331962']
    
    # Códigos que deben tener tallas SM y ML (división 50%-50%)
    codigos_con_tallas_sm_ml = ['10975804', '10975815', '10975835', '70192970', '70353249', '70353266', 
                                '70360899', '70360903', '70428987', '70430338', '70457634']
    
    # Tallas específicas numéricas
    tallas_especificas = ['678', '700', '718', '714', '738', '712', '758', '734', '778', '800']
    
    # Tallas SM y ML
    tallas_sm_ml = ['SM', 'ML']
    
    # Filtrar por marca NEW ERA
    df_new_era = df_stock[df_stock['U_Marca'].str.upper() == 'NEW ERA'].copy()
    
    # Filtrar por códigos MVP específicos
    df_mvp = df_new_era[df_new_era['U_Estilo'].astype(str).isin(codigos_mvp)].copy()
    
    # Información de debug
    print(f"DEBUG MVP COSTA RICA: Total registros NEW ERA: {len(df_new_era)}")
    print(f"DEBUG MVP COSTA RICA: Códigos MVP filtrados: {len(df_mvp)}")
    
    # Verificar qué códigos MVP están presentes
    codigos_encontrados = df_mvp['U_Estilo'].astype(str).unique().tolist()
    print(f"DEBUG MVP COSTA RICA: Códigos encontrados: {sorted(codigos_encontrados)}")
    
    if df_mvp.empty:
        print("DEBUG MVP COSTA RICA: No se encontraron códigos MVP en el archivo")
        return pd.DataFrame()
    
    # Verificar columnas necesarias - SOPORTE PARA AMBAS: 'Talla' y 'U_Talla'
    columna_talla = None
    if 'U_Talla' in df_mvp.columns:
        columna_talla = 'U_Talla'
    elif 'Talla' in df_mvp.columns:
        columna_talla = 'Talla'
    else:
        print("ERROR: No se encontró columna de talla ('U_Talla' o 'Talla')")
        return pd.DataFrame()
    
    columnas_necesarias = ['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Stock_Actual', 'Bodega']
    for col in columnas_necesarias:
        if col not in df_mvp.columns:
            print(f"ERROR: Columna faltante: {col}")
            return pd.DataFrame()
    
    # Obtener bodegas de Costa Rica
    bodegas_costarica = [
        "NE City Mall"
    ]
    
    # Filtrar solo bodegas de Costa Rica
    df_mvp_costarica = df_mvp[df_mvp['Bodega'].isin(bodegas_costarica)].copy()
    
    if df_mvp_costarica.empty:
        return pd.DataFrame()
    
    # Crear mapa de SAP solo para registros REALES (con stock > 0)
    df_sap_map = df_mvp_costarica[df_mvp_costarica['Stock_Actual'] > 0].groupby(['U_Estilo', columna_talla])['Codigo_SAP'].first().reset_index()
    sap_dict = {}
    for _, row in df_sap_map.iterrows():
        key = (str(row['U_Estilo']), str(row[columna_talla]))
        sap_dict[key] = row['Codigo_SAP']
    
    # Obtener datos de óptimos para Costa Rica
    optimos_por_codigo = obtener_optimos_mvp_costarica()
    optimos_por_tallas = obtener_optimos_por_tallas_costarica()
    
    # NUEVA LÓGICA: Agregar filas faltantes para códigos con tallas específicas
    filas_adicionales = []
    
    # 1. Códigos con tallas numéricas (678-800)
    for codigo in codigos_con_tallas:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_costarica[df_mvp_costarica['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de Costa Rica
            for bodega in bodegas_costarica:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para TODAS las tallas específicas (678-800)
                # para que aparezcan en la tabla aunque tengan stock óptimo 0
                for talla_req in tallas_especificas:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    # 2. Códigos con tallas SM y ML (división 50%-50%)
    for codigo in codigos_con_tallas_sm_ml:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_costarica[df_mvp_costarica['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de Costa Rica
            for bodega in bodegas_costarica:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para SM y ML si no existen
                for talla_req in tallas_sm_ml:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    
    # Agregar filas faltantes al DataFrame
    if filas_adicionales:
        df_adicional = pd.DataFrame(filas_adicionales)
        df_mvp_costarica = pd.concat([df_mvp_costarica, df_adicional], ignore_index=True)
    
    # NUEVA FUNCIONALIDAD: Agregar códigos MVP faltantes con información "N/D"
    # Identificar códigos MVP que NO están en el archivo
    codigos_presentes = df_mvp_costarica['U_Estilo'].astype(str).unique().tolist()
    codigos_faltantes = [codigo for codigo in codigos_mvp if codigo not in codigos_presentes]
    
    if codigos_faltantes:
        print(f"DEBUG MVP COSTA RICA: Agregando {len(codigos_faltantes)} códigos MVP faltantes: {sorted(codigos_faltantes)}")
        
        # Crear filas para códigos faltantes
        filas_faltantes = []
        for codigo_faltante in codigos_faltantes:
            for bodega in bodegas_costarica:
                # Crear fila con información "N/D"
                fila_faltante = {
                    'U_Estilo': codigo_faltante,
                    'Codigo_SAP': '',  # Vacío porque no existe en archivo
                    'U_Segmento': 'N/D',
                    'U_Silueta': 'N/D', 
                    'U_Coleccion_NE': 'N/D',
                    'U_Descripcion': 'N/D',
                    columna_talla: 'N/D',
                    'Stock_Actual': 0,  # Sin stock porque no está en archivo
                    'Bodega': bodega
                }
                filas_faltantes.append(fila_faltante)
        
        # Agregar filas faltantes al DataFrame
        if filas_faltantes:
            df_faltantes = pd.DataFrame(filas_faltantes)
            df_mvp_costarica = pd.concat([df_mvp_costarica, df_faltantes], ignore_index=True)
    
    # Agrupar como tabla dinámica (SIN Codigo_SAP en el groupby)
    df_agrupado = df_mvp_costarica.groupby([
        'U_Estilo', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Bodega'
    ])['Stock_Actual'].sum().reset_index()
    
    # Asignar SAP basado en mapa de códigos+tallas REALES
    df_agrupado['Codigo_SAP'] = df_agrupado.apply(
        lambda row: sap_dict.get((str(row['U_Estilo']), str(row[columna_talla])), ""), 
        axis=1
    )
    
    # Pivotar para tener bodegas como columnas
    tabla_pivoteada = df_agrupado.pivot_table(
        index=['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla],
        columns='Bodega',
        values='Stock_Actual',
        fill_value=0,
        aggfunc='sum'
    )
    
    # Asegurar que todas las bodegas estén presentes como columnas
    for bodega in bodegas_costarica:
        if bodega not in tabla_pivoteada.columns:
            tabla_pivoteada[bodega] = 0
    
    # Reordenar columnas según el orden de bodegas_costarica
    tabla_pivoteada = tabla_pivoteada.reindex(columns=bodegas_costarica, fill_value=0)
    
    # Crear nueva tabla con columnas intercaladas (stock actual + stock óptimo nuevo)
    tabla_final = pd.DataFrame(index=tabla_pivoteada.index)
    
    # Agregar columnas intercaladas para cada bodega
    for bodega in bodegas_costarica:
        # Columna de stock actual
        tabla_final[f"Real {bodega}"] = tabla_pivoteada[bodega]
        
        # Columna de stock óptimo nuevo
        col_optimo = f"Óptimo {bodega}"
        tabla_final[col_optimo] = 0
    
        # Llenar valores óptimos para cada código/talla
        for codigo_tuple in tabla_pivoteada.index:
            codigo = str(codigo_tuple[0])
            talla = str(codigo_tuple[6])  # Talla está en el índice 6 (después de agregar Codigo_SAP)
            
            # Determinar tipo de código y calcular stock óptimo
            if codigo in codigos_con_tallas and talla in tallas_especificas:
                # CASO 1: Códigos con tallas numéricas (678-800) - Calcular tallas basado en código ÷ 12
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo] and bodega in optimos_por_tallas:
                    # Paso 1: Obtener stock óptimo por código
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    
                    # Paso 2: Obtener distribución base de tallas para esta bodega
                    tallas_base = optimos_por_tallas[bodega]
                    
                    # Paso 3: Calcular factor multiplicador
                    suma_tallas_base = sum(tallas_base.values())
                    if suma_tallas_base > 0:
                        factor = stock_codigo / suma_tallas_base
                        
                        # Paso 4: Calcular nueva distribución
                        tallas_calculadas = {}
                        suma_redondeada = 0
                        
                        # Multiplicar y redondear hacia abajo
                        for talla_key, valor_base in tallas_base.items():
                            valor_calculado = int(valor_base * factor)  # Redondeo hacia abajo
                            tallas_calculadas[talla_key] = valor_calculado
                            suma_redondeada += valor_calculado
                        
                        # Paso 5: Ajustar para que cuadre exactamente
                        diferencia = stock_codigo - suma_redondeada
                        
                        if diferencia > 0:
                            # Distribuir diferencia positiva entre las tallas con mayor base
                            tallas_ordenadas = sorted(tallas_base.items(), key=lambda x: x[1], reverse=True)
                            for i, (talla_key, _) in enumerate(tallas_ordenadas[:diferencia]):
                                tallas_calculadas[talla_key] += 1
                        
                        # Paso 6: Asignar valor calculado si la talla coincide
                        if talla in tallas_calculadas:
                            tabla_final.loc[codigo_tuple, col_optimo] = tallas_calculadas[talla]
                else:
                    tabla_final.loc[codigo_tuple, col_optimo] = 0
                    
            elif codigo in codigos_con_tallas_sm_ml and talla in tallas_sm_ml:
                # CASO 2: Códigos con tallas SM y ML - División 50%-50%
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    stock_sm, stock_ml = calcular_tallas_sm_ml(stock_codigo)
                    
                    # Asignar según la talla
                    if talla == 'SM':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_sm
                    elif talla == 'ML':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_ml
                    
                    # Validar cuadre
                    validar_cuadre_sm_ml(codigo, bodega, stock_sm, stock_ml, stock_codigo)
                else:
                    tabla_final.loc[codigo_tuple, col_optimo] = 0
                    
            else:
                # CASO 3: Códigos sin tallas específicas - Usar stock óptimo por código directo
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    tabla_final.loc[codigo_tuple, col_optimo] = optimos_por_codigo[codigo][bodega]
    
    # Agregar fila de totales
    fila_totales = tabla_final.sum(axis=0)
    fila_totales.name = ('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')
    tabla_final = pd.concat([tabla_final, fila_totales.to_frame().T])
    
    return tabla_final

def procesar_stock_mvps_panama(df_stock: pd.DataFrame) -> pd.DataFrame:
    """
    Procesa los datos de stock para códigos MVP específicos en Panamá con nueva lógica
    - Códigos con tallas específicas usan stock óptimo por tallas
    - Otros códigos usan stock óptimo por código general
    - Agrega filas faltantes para tallas requeridas con stock 0
    """
    if df_stock is None or df_stock.empty:
        return pd.DataFrame()
    
    # CÓDIGOS MVP ESPECÍFICOS - SOLO ESTOS 46 SE EXTRAEN DEL ARCHIVO DE STOCK
    codigos_mvp = [
        '10030708', '10030709', '10047511', '10047531', '10047538', '10112874', 
        '10975804', '10975815', '10975835', '11405605', '11405614', 
        '11591024', '11591025', '11591026', '11591043', '11591046', '11591047', 
        '11591077', '11591078', '11591122', '11591128', '11591150', '11591175', 
        '11941921', '12650335', '12650337', '12650340', '12650342', '12650343', 
        '12650344', '70192970', '70331909', '70331911', '70331962', '70353249', 
        '70353266', '70360899', '70360903', '70428987', '70430338', '70457634', 
        '70556851', '70556867', '70556869', '70558225'
    ]
    
    # Códigos que deben tener tallas específicas (678-800)
    codigos_con_tallas = ['11591122', '11591128', '11591150', '11591175', '70331909', '70331911', '70331962']
    
    # Códigos que deben tener tallas SM y ML (división 50%-50%)
    codigos_con_tallas_sm_ml = ['10975804', '10975815', '10975835', '70192970', '70353249', '70353266', 
                                '70360899', '70360903', '70428987', '70430338', '70457634']
    
    # Tallas específicas numéricas
    tallas_especificas = ['678', '700', '718', '714', '738', '712', '758', '734', '778', '800']
    
    # Tallas SM y ML
    tallas_sm_ml = ['SM', 'ML']
    
    # Filtrar por marca NEW ERA
    df_new_era = df_stock[df_stock['U_Marca'].str.upper() == 'NEW ERA'].copy()
    
    # Filtrar por códigos MVP específicos (SOLO ESTOS 45 CÓDIGOS - SIN 11169822)
    df_mvp = df_new_era[df_new_era['U_Estilo'].astype(str).isin(codigos_mvp)].copy()
    
    # Información de debug para verificar filtrado correcto
    print(f"DEBUG MVP PANAMÁ: Total registros NEW ERA: {len(df_new_era)}")
    print(f"DEBUG MVP PANAMÁ: Códigos MVP filtrados: {len(df_mvp)}")
    print(f"DEBUG MVP PANAMÁ: Columnas disponibles: {list(df_mvp.columns)}")
    
    # Verificar qué códigos MVP están presentes en el archivo
    codigos_encontrados = df_mvp['U_Estilo'].astype(str).unique().tolist()
    print(f"DEBUG MVP PANAMÁ: Códigos encontrados: {sorted(codigos_encontrados)}")
    
    if df_mvp.empty:
        print("DEBUG MVP PANAMÁ: No se encontraron códigos MVP en el archivo")
        return pd.DataFrame()
    
    # Verificar columnas necesarias - SOPORTE PARA AMBAS: 'Talla' y 'U_Talla'
    columna_talla = None
    if 'U_Talla' in df_mvp.columns:
        columna_talla = 'U_Talla'
        print("PANAMÁ: USANDO COLUMNA U_Talla")
    elif 'Talla' in df_mvp.columns:
        columna_talla = 'Talla'
        print("PANAMÁ: USANDO COLUMNA Talla")
    else:
        print("ERROR PANAMÁ: No se encontró columna de talla")
        return pd.DataFrame()
    
    columnas_necesarias = ['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Stock_Actual', 'Bodega']
    for col in columnas_necesarias:
        if col not in df_mvp.columns:
            print(f"ERROR PANAMÁ: Columna faltante: {col}")
            return pd.DataFrame()
    
    # Obtener bodegas de Panamá
    bodegas_panama = [
        "NE Albrookmall", "NE Metromall", "NE Multiplaza Panamá", "NE Westland"
    ]
    
    # Filtrar solo bodegas de Panamá
    df_mvp_panama = df_mvp[df_mvp['Bodega'].isin(bodegas_panama)].copy()
    
    if df_mvp_panama.empty:
        return pd.DataFrame()
    
    # Crear mapa de SAP solo para registros REALES (con stock > 0)
    df_sap_map = df_mvp_panama[df_mvp_panama['Stock_Actual'] > 0].groupby(['U_Estilo', columna_talla])['Codigo_SAP'].first().reset_index()
    sap_dict = {}
    for _, row in df_sap_map.iterrows():
        key = (str(row['U_Estilo']), str(row[columna_talla]))
        sap_dict[key] = row['Codigo_SAP']
    
    # Obtener datos de óptimos
    optimos_por_codigo = obtener_optimos_mvp_panama()
    optimos_por_tallas = obtener_optimos_por_tallas_panama()
    
    # NUEVA LÓGICA: Agregar filas faltantes para códigos con tallas específicas
    filas_adicionales = []
    
    # 1. Códigos con tallas numéricas (678-800)
    for codigo in codigos_con_tallas:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_panama[df_mvp_panama['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de Panamá
            for bodega in bodegas_panama:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para TODAS las tallas específicas (678-800)
                # para que aparezcan en la tabla aunque tengan stock óptimo 0
                for talla_req in tallas_especificas:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    # 2. Códigos con tallas SM y ML (división 50%-50%)
    for codigo in codigos_con_tallas_sm_ml:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_panama[df_mvp_panama['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de Panamá
            for bodega in bodegas_panama:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para SM y ML si no existen
                for talla_req in tallas_sm_ml:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    # Agregar filas adicionales al DataFrame
    if filas_adicionales:
        df_adicional = pd.DataFrame(filas_adicionales)
        df_mvp_panama = pd.concat([df_mvp_panama, df_adicional], ignore_index=True)
    
    # NUEVA FUNCIONALIDAD: Agregar códigos MVP faltantes con información "N/D"
    # Identificar códigos MVP que NO están en el archivo
    codigos_presentes = df_mvp_panama['U_Estilo'].astype(str).unique().tolist()
    codigos_faltantes = [codigo for codigo in codigos_mvp if codigo not in codigos_presentes]
    
    if codigos_faltantes:
        print(f"DEBUG MVP PANAMA: Agregando {len(codigos_faltantes)} códigos MVP faltantes: {sorted(codigos_faltantes)}")
        
        # Crear filas para códigos faltantes
        filas_faltantes = []
        for codigo_faltante in codigos_faltantes:
            for bodega in bodegas_panama:
                # Crear fila con información "N/D"
                fila_faltante = {
                    'U_Estilo': codigo_faltante,
                    'Codigo_SAP': '',  # Vacío porque no existe en archivo
                    'U_Segmento': 'N/D',
                    'U_Silueta': 'N/D', 
                    'U_Coleccion_NE': 'N/D',
                    'U_Descripcion': 'N/D',
                    columna_talla: 'N/D',
                    'Stock_Actual': 0,  # Sin stock porque no está en archivo
                    'Bodega': bodega
                }
                filas_faltantes.append(fila_faltante)
        
        # Agregar filas faltantes al DataFrame
        if filas_faltantes:
            df_faltantes = pd.DataFrame(filas_faltantes)
            df_mvp_panama = pd.concat([df_mvp_panama, df_faltantes], ignore_index=True)
    
    # Agrupar como tabla dinámica (SIN Codigo_SAP en el groupby)
    df_agrupado = df_mvp_panama.groupby([
        'U_Estilo', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Bodega'
    ])['Stock_Actual'].sum().reset_index()
    
    # Asignar SAP basado en mapa de códigos+tallas REALES
    df_agrupado['Codigo_SAP'] = df_agrupado.apply(
        lambda row: sap_dict.get((str(row['U_Estilo']), str(row[columna_talla])), ""), 
        axis=1
    )
    
    # Pivotar para tener bodegas como columnas
    tabla_pivoteada = df_agrupado.pivot_table(
        index=['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla],
        columns='Bodega',
        values='Stock_Actual',
        fill_value=0,
        aggfunc='sum'
    )
    
    # Asegurar que todas las bodegas estén presentes como columnas
    for bodega in bodegas_panama:
        if bodega not in tabla_pivoteada.columns:
            tabla_pivoteada[bodega] = 0
    
    # Reordenar columnas según el orden de bodegas_panama
    tabla_pivoteada = tabla_pivoteada.reindex(columns=bodegas_panama, fill_value=0)
    
    # Crear nueva tabla con columnas intercaladas (stock actual + stock óptimo nuevo)
    tabla_final = pd.DataFrame(index=tabla_pivoteada.index)
    
    # Agregar columnas intercaladas para cada bodega
    for bodega in bodegas_panama:
        # Columna de stock actual
        tabla_final[f"Real {bodega}"] = tabla_pivoteada[bodega]
        
        # Columna de stock óptimo nuevo
        col_optimo = f"Óptimo {bodega}"
        tabla_final[col_optimo] = 0
        
        # Llenar valores óptimos para cada código/talla
        for codigo_tuple in tabla_pivoteada.index:
            codigo = str(codigo_tuple[0])
            talla = str(codigo_tuple[6])  # Talla está en el índice 6 (después de agregar Codigo_SAP)
            
            # Determinar tipo de código y calcular stock óptimo
            if codigo in codigos_con_tallas and talla in tallas_especificas:
                # CASO 1: Códigos con tallas numéricas (678-800) - Calcular tallas basado en código ÷ 12
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo] and bodega in optimos_por_tallas:
                    # Paso 1: Obtener stock óptimo por código
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    
                    # Paso 2: Obtener distribución base de tallas para esta bodega
                    tallas_base = optimos_por_tallas[bodega]
                    
                    # Paso 3: Calcular factor multiplicador
                    suma_tallas_base = sum(tallas_base.values())
                    if suma_tallas_base > 0:
                        factor = stock_codigo / suma_tallas_base
                        
                        # Paso 4: Calcular nueva distribución
                        tallas_calculadas = {}
                        suma_redondeada = 0
                        
                        # Multiplicar y redondear hacia abajo
                        for talla_key, valor_base in tallas_base.items():
                            valor_calculado = int(valor_base * factor)  # Redondeo hacia abajo
                            tallas_calculadas[talla_key] = valor_calculado
                            suma_redondeada += valor_calculado
                        
                        # Paso 5: Ajustar para que cuadre exactamente
                        diferencia = stock_codigo - suma_redondeada
                        
                        if diferencia > 0:
                            # Encontrar talla con mayor valor (primera en caso de empate)
                            talla_maxima = max(tallas_calculadas.keys(), key=lambda k: tallas_calculadas[k])
                            tallas_calculadas[talla_maxima] += diferencia
                        
                        # Paso 6: Asignar valor para esta talla específica
                        if talla in tallas_calculadas:
                            tabla_final.loc[codigo_tuple, col_optimo] = tallas_calculadas[talla]
                        else:
                            tabla_final.loc[codigo_tuple, col_optimo] = 0
                    else:
                        tabla_final.loc[codigo_tuple, col_optimo] = 0
                else:
                    tabla_final.loc[codigo_tuple, col_optimo] = 0
                    
            elif codigo in codigos_con_tallas_sm_ml and talla in tallas_sm_ml:
                # CASO 2: Códigos con tallas SM y ML - División 50%-50%
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    stock_sm, stock_ml = calcular_tallas_sm_ml(stock_codigo)
                    
                    # Asignar según la talla
                    if talla == 'SM':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_sm
                    elif talla == 'ML':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_ml
                    
                    # Validar cuadre
                    validar_cuadre_sm_ml(codigo, bodega, stock_sm, stock_ml, stock_codigo)
                else:
                    tabla_final.loc[codigo_tuple, col_optimo] = 0
                    
            else:
                # CASO 3: Códigos sin tallas específicas - Usar stock óptimo por código directo
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    tabla_final.loc[codigo_tuple, col_optimo] = optimos_por_codigo[codigo][bodega]
    
    # Agregar fila de totales por columna
    fila_totales = tabla_final.sum(axis=0)
    fila_totales.name = ('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')
    tabla_final = pd.concat([tabla_final, fila_totales.to_frame().T])
    
    return tabla_final

def procesar_stock_mvps_puerto_rico(df_stock: pd.DataFrame) -> pd.DataFrame:
    """
    Procesa los datos de stock para códigos MVP específicos en Puerto Rico con nueva lógica
    - Códigos con tallas específicas usan stock óptimo por tallas
    - Códigos con tallas SM/ML usan división 50%-50% del stock óptimo por código
    - Otros códigos usan stock óptimo por código general
    - Agrega filas faltantes para tallas requeridas con stock 0
    """
    if df_stock is None or df_stock.empty:
        print("❌ Error: Archivo de stock vacío o no válido")
        return pd.DataFrame()
    
    # CÓDIGOS MVP ESPECÍFICOS - SOLO ESTOS 45 SE EXTRAEN DEL ARCHIVO DE STOCK (IGUAL QUE TODOS LOS PAÍSES)
    codigos_mvp = [
        '10030708', '10030709', '10047511', '10047531', '10047538', '10112874', 
        '10975804', '10975815', '10975835', '11405605', '11405614', 
        '11591024', '11591025', '11591026', '11591043', '11591046', '11591047', 
        '11591077', '11591078', '11591122', '11591128', '11591150', '11591175', 
        '11941921', '12650335', '12650337', '12650340', '12650342', '12650343', 
        '12650344', '70192970', '70331909', '70331911', '70331962', '70353249', 
        '70353266', '70360899', '70360903', '70428987', '70430338', '70457634', 
        '70556851', '70556867', '70556869', '70558225'
    ]
    
    # Códigos que deben tener tallas específicas (678-800)
    codigos_con_tallas = ['11591122', '11591128', '11591150', '11591175', '70331909', '70331911', '70331962']
    
    # Códigos que deben tener tallas SM y ML (división 50%-50%)
    codigos_con_tallas_sm_ml = ['10975804', '10975815', '10975835', '70192970', '70353249', '70353266', 
                                '70360899', '70360903', '70428987', '70430338', '70457634']
    
    # Tallas específicas numéricas
    tallas_especificas = ['678', '700', '718', '714', '738', '712', '758', '734', '778', '800']
    
    # Tallas SM y ML
    tallas_sm_ml = ['SM', 'ML']
    
    # Filtrar por marca NEW ERA (IGUAL QUE GUATEMALA)
    df_new_era = df_stock[df_stock['U_Marca'].str.upper() == 'NEW ERA'].copy()
    
    # Filtrar por códigos MVP específicos (SOLO ESTOS 45 CÓDIGOS)
    df_mvp = df_new_era[df_new_era['U_Estilo'].astype(str).isin(codigos_mvp)].copy()
    
    # Información de debug para verificar filtrado correcto
    print(f"DEBUG MVP PUERTO RICO: Total registros NEW ERA: {len(df_new_era)}")
    print(f"DEBUG MVP PUERTO RICO: Códigos MVP filtrados: {len(df_mvp)}")
    print(f"DEBUG MVP PUERTO RICO: Columnas disponibles: {list(df_mvp.columns)}")
    
    # Verificar qué códigos MVP están presentes en el archivo
    codigos_encontrados = df_mvp['U_Estilo'].astype(str).unique().tolist()
    print(f"DEBUG MVP PUERTO RICO: Códigos encontrados: {sorted(codigos_encontrados)}")
    
    if df_mvp.empty:
        print("DEBUG MVP PUERTO RICO: No se encontraron códigos MVP en el archivo")
        return pd.DataFrame()
    
    # Verificar columnas necesarias - SOPORTE PARA AMBAS: 'Talla' y 'U_Talla'
    columna_talla = None
    if 'U_Talla' in df_mvp.columns:
        columna_talla = 'U_Talla'
        print("PUERTO RICO: USANDO COLUMNA U_Talla")
    elif 'Talla' in df_mvp.columns:
        columna_talla = 'Talla'
        print("PUERTO RICO: USANDO COLUMNA Talla")
    else:
        print("ERROR PUERTO RICO: No se encontró columna de talla")
        return pd.DataFrame()
    
    columnas_necesarias = ['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Stock_Actual', 'Bodega']
    for col in columnas_necesarias:
        if col not in df_mvp.columns:
            print(f"ERROR PUERTO RICO: Columna faltante: {col}")
            return pd.DataFrame()
    
    # Obtener bodegas de Puerto Rico (nombres del archivo CSV)
    bodegas_puerto_rico_archivo = [
        "NE Barceloneta Premium Outlet", "NE Plaza Carolina"
    ]
    
    # Mapeo completo de nombres de tiendas  
    mapeo_nombres_completo = {
        'NE Barceloneta Premium Outlet': 'NE BARCELONETA',  # Archivo CSV -> Nombre final
        'NE Plaza Carolina': 'NE CAROLINA'  # Archivo CSV -> Nombre final
    }
    
    # Filtrar solo bodegas de Puerto Rico
    df_mvp_puerto_rico = df_mvp[df_mvp['Bodega'].isin(bodegas_puerto_rico_archivo)].copy()
    
    if df_mvp_puerto_rico.empty:
        print("DEBUG MVP PUERTO RICO: No se encontraron productos en las bodegas de Puerto Rico")
        return pd.DataFrame()
    
    # Aplicar mapeo de nombres de tiendas
    df_mvp_puerto_rico['Bodega'] = df_mvp_puerto_rico['Bodega'].map(mapeo_nombres_completo)
    
    print(f"DEBUG MVP PUERTO RICO: Registros en bodegas de Puerto Rico: {len(df_mvp_puerto_rico)}")
    print(f"DEBUG MVP PUERTO RICO: Bodegas después del mapeo: {df_mvp_puerto_rico['Bodega'].unique()}")
    
    # Obtener datos de óptimos
    optimos_por_codigo = obtener_optimos_mvp_puerto_rico()
    optimos_por_tallas = obtener_optimos_por_tallas_puerto_rico()
    
    # NUEVA LÓGICA: Agregar filas faltantes para códigos con tallas específicas (IGUAL QUE GUATEMALA)
    filas_adicionales = []
    
    # 1. Códigos con tallas numéricas (678-800)
    for codigo in codigos_con_tallas:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_puerto_rico[df_mvp_puerto_rico['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de Puerto Rico
            bodegas_puerto_rico_finales = ['NE BARCELONETA', 'NE CAROLINA']
            for bodega in bodegas_puerto_rico_finales:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para TODAS las tallas específicas (678-800)
                # para que aparezcan en la tabla aunque tengan stock óptimo 0
                for talla_req in tallas_especificas:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    # 2. Códigos con tallas SM y ML (división 50%-50%)
    for codigo in codigos_con_tallas_sm_ml:
        # Obtener datos base del código (si existe)
        df_codigo = df_mvp_puerto_rico[df_mvp_puerto_rico['U_Estilo'].astype(str) == codigo]
        
        if not df_codigo.empty:
            # Obtener datos base del primer registro
            datos_base = df_codigo.iloc[0].copy()
            
            # Para cada bodega de Puerto Rico
            for bodega in bodegas_puerto_rico_finales:
                # Verificar qué tallas tiene este código en esta bodega (con limpieza robusta)
                tallas_existentes = df_codigo[df_codigo['Bodega'] == bodega][columna_talla].astype(str).str.strip().tolist()
                
                # Agregar filas para SM y ML si no existen
                for talla_req in tallas_sm_ml:
                    if talla_req not in tallas_existentes:
                        # Crear nueva fila con stock 0
                        nueva_fila = datos_base.copy()
                        nueva_fila['Bodega'] = bodega
                        nueva_fila[columna_talla] = talla_req
                        nueva_fila['Stock_Actual'] = 0
                        nueva_fila['Codigo_SAP'] = ""  # SAP vacío para filas artificiales (sin stock real)
                        filas_adicionales.append(nueva_fila)
    
    # Crear mapa de SAP solo para registros REALES (con stock > 0)
    df_sap_map = df_mvp_puerto_rico[df_mvp_puerto_rico['Stock_Actual'] > 0].groupby(['U_Estilo', columna_talla])['Codigo_SAP'].first().reset_index()
    sap_dict = {}
    for _, row in df_sap_map.iterrows():
        key = (str(row['U_Estilo']), str(row[columna_talla]))
        sap_dict[key] = row['Codigo_SAP']
    
    # Aplicar códigos SAP reales a las filas artificiales que correspondan
    for fila in filas_adicionales:
        key = (str(fila['U_Estilo']), str(fila[columna_talla]))
        if key in sap_dict:
            fila['Codigo_SAP'] = sap_dict[key]
    
    # Concatenar filas adicionales
    if filas_adicionales:
        df_adicional = pd.DataFrame(filas_adicionales)
        df_mvp_puerto_rico = pd.concat([df_mvp_puerto_rico, df_adicional], ignore_index=True)
        print(f"DEBUG MVP PUERTO RICO: Agregadas {len(filas_adicionales)} filas para tallas faltantes")
    
    # Verificar códigos faltantes y crear filas con stock 0 (CÓDIGOS COMPLETAMENTE FALTANTES)
    codigos_presentes = set(df_mvp_puerto_rico['U_Estilo'].astype(str).unique())
    codigos_faltantes = set(codigos_mvp) - codigos_presentes
    
    # CÓDIGOS COMPLETAMENTE FALTANTES: Crear filas para códigos que no están en el archivo
    if codigos_faltantes:
        print(f"DEBUG MVP PUERTO RICO: Códigos MVP faltantes en archivo: {len(codigos_faltantes)}")
        
        # Crear filas para códigos faltantes con información "N/D" (IGUAL QUE GUATEMALA)
        filas_faltantes = []
        for codigo in codigos_faltantes:
            for bodega in bodegas_puerto_rico_finales:
                # Determinar tallas según el tipo de código
                if codigo in codigos_con_tallas:
                    # Códigos con tallas específicas (678-800)
                    tallas = tallas_especificas
                elif codigo in codigos_con_tallas_sm_ml:
                    # Códigos con tallas SM/ML
                    tallas = tallas_sm_ml
                else:
                    # Otros códigos (usar una talla genérica)
                    tallas = ['N/D']
                
                for talla in tallas:
                    fila = {
                        'U_Estilo': int(codigo) if codigo.isdigit() else codigo,  # Mantener tipo correcto
                        'Codigo_SAP': "",  # Vacío para códigos artificiales
                        'U_Marca': 'NEW ERA',
                        'U_Segmento': 'N/D',
                        'U_Silueta': 'N/D', 
                        'U_Coleccion_NE': 'N/D',
                        'U_Descripcion': 'N/D',
                        columna_talla: str(talla),
                        'Bodega': bodega,
                        'Stock_Actual': 0
                    }
                    filas_faltantes.append(fila)
        
        # Agregar filas faltantes al DataFrame
        if filas_faltantes:
            df_faltantes = pd.DataFrame(filas_faltantes)
            df_mvp_puerto_rico = pd.concat([df_mvp_puerto_rico, df_faltantes], ignore_index=True)
            print(f"DEBUG MVP PUERTO RICO: Agregadas {len(filas_faltantes)} filas para códigos completamente faltantes")
    
    # Asegurar que la columna de talla sea string y limpiar espacios (IGUAL QUE GUATEMALA)
    df_mvp_puerto_rico[columna_talla] = df_mvp_puerto_rico[columna_talla].astype(str).str.strip()
    
    # Agrupar por las dimensiones principales
    tabla_agrupada = df_mvp_puerto_rico.groupby([
        'U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 
        'U_Coleccion_NE', 'U_Descripcion', columna_talla, 'Bodega'
    ])['Stock_Actual'].sum().reset_index()
    
    # Crear tabla pivoteada con bodegas como columnas
    tabla_pivoteada = tabla_agrupada.pivot_table(
        index=['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', columna_talla],
        columns='Bodega',
        values='Stock_Actual',
        fill_value=0,
        aggfunc='sum'
    )
    
    # Renombrar columna de talla
    if columna_talla != 'Talla':
        tabla_pivoteada.index.names = ['U_Estilo', 'Codigo_SAP', 'U_Segmento', 'U_Silueta', 'U_Coleccion_NE', 'U_Descripcion', 'Talla']
    
    # Asegurar que todas las bodegas estén presentes
    for bodega in bodegas_puerto_rico_finales:
        if bodega not in tabla_pivoteada.columns:
            tabla_pivoteada[bodega] = 0
    
    # Crear columnas de stock real
    tabla_final = tabla_pivoteada.copy()
    for bodega in bodegas_puerto_rico_finales:
        nueva_columna = f"Real {bodega}"
        tabla_final[nueva_columna] = tabla_pivoteada[bodega]
    
    # Eliminar columnas originales de bodega
    tabla_final = tabla_final.drop(columns=bodegas_puerto_rico_finales)
    
    # (Los datos de óptimos y las listas ya están definidas arriba)
    
    # Crear columnas de stock óptimo
    for bodega in bodegas_puerto_rico_finales:
        col_optimo = f"Óptimo {bodega}"
        tabla_final[col_optimo] = 0
        
        for codigo_tuple in tabla_final.index:
            codigo = str(codigo_tuple[0])  # Convertir a string para que coincida con las claves del diccionario
            talla = str(codigo_tuple[6])
            
            if codigo in codigos_con_tallas and talla in tallas_especificas:
                # CASO 1: Códigos con tallas específicas - Calcular tallas basado en código ÷ 12
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo] and bodega in optimos_por_tallas:
                    # Paso 1: Obtener stock óptimo por código
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    
                    # Paso 2: Obtener distribución base de tallas para esta bodega
                    tallas_base = optimos_por_tallas[bodega]
                    
                    # Paso 3: Calcular factor multiplicador
                    suma_tallas_base = sum(tallas_base.values())
                    if suma_tallas_base > 0:
                        factor = stock_codigo / suma_tallas_base
                        
                        # Paso 4: Calcular nueva distribución
                        tallas_calculadas = {}
                        suma_redondeada = 0
                        
                        # Multiplicar y redondear hacia abajo
                        for talla_key, valor_base in tallas_base.items():
                            valor_calculado = int(valor_base * factor)  # Redondeo hacia abajo
                            tallas_calculadas[talla_key] = valor_calculado
                            suma_redondeada += valor_calculado
                        
                        # Paso 5: Ajustar para que cuadre exactamente
                        diferencia = stock_codigo - suma_redondeada
                        
                        if diferencia > 0:
                            # Encontrar talla con mayor valor (primera en caso de empate)
                            talla_mayor = max(tallas_base.items(), key=lambda x: (x[1], x[0]))[0]
                            tallas_calculadas[talla_mayor] += diferencia
                        elif diferencia < 0:
                            # Reducir de la talla con menor valor base > 0
                            for talla_key in sorted(tallas_base.items(), key=lambda x: (x[1], x[0])):
                                if tallas_calculadas[talla_key[0]] > 0 and diferencia < 0:
                                    reduccion = min(tallas_calculadas[talla_key[0]], abs(diferencia))
                                    tallas_calculadas[talla_key[0]] -= reduccion
                                    diferencia += reduccion
                        
                        # Asignar valor calculado específico para esta talla
                        if talla in tallas_calculadas:
                            tabla_final.loc[codigo_tuple, col_optimo] = tallas_calculadas[talla]
            
            elif codigo in codigos_con_tallas_sm_ml and talla in tallas_sm_ml:
                # CASO 2: Códigos con tallas SM y ML - División 50%-50%
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    stock_codigo = optimos_por_codigo[codigo][bodega]
                    stock_sm, stock_ml = calcular_tallas_sm_ml(stock_codigo)
                    # Asignar según la talla
                    if talla == 'SM':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_sm
                    elif talla == 'ML':
                        tabla_final.loc[codigo_tuple, col_optimo] = stock_ml
                    # Validar cuadre
                    validar_cuadre_sm_ml(codigo, bodega, stock_sm, stock_ml, stock_codigo)
            
            else:
                # CASO 3: Otros códigos - usar stock óptimo por código general
                if codigo in optimos_por_codigo and bodega in optimos_por_codigo[codigo]:
                    tabla_final.loc[codigo_tuple, col_optimo] = optimos_por_codigo[codigo][bodega]
    
    # Crear mapeo SAP solo para registros con stock real > 0
    columnas_real = [col for col in tabla_final.columns if col.startswith('Real ')]
    filas_con_stock = tabla_final[
        tabla_final[columnas_real].apply(
            lambda row: row.sum() > 0, axis=1
        )
    ]
    
    mapeo_sap = {}
    for codigo_tuple in filas_con_stock.index:
        codigo = codigo_tuple[0]
        codigo_sap_original = codigo_tuple[1]
        if codigo_sap_original and str(codigo_sap_original).strip():
            mapeo_sap[codigo] = str(codigo_sap_original)
    
    # Aplicar mapeo SAP a todos los registros del mismo código
    for codigo_tuple in tabla_final.index:
        codigo = codigo_tuple[0]
        if codigo in mapeo_sap:
            # Actualizar el Codigo_SAP en el índice
            nuevo_tuple = (codigo_tuple[0], mapeo_sap[codigo], *codigo_tuple[2:])
            tabla_final.rename(index={codigo_tuple: nuevo_tuple}, inplace=True)
    
    # Reorganizar columnas intercalando Real y Óptimo por bodega
    columnas_ordenadas = []
    
    # Agregar columnas de índice al final como columnas normales para mejor visualización
    for bodega in bodegas_puerto_rico_finales:
        columnas_ordenadas.append(f"Real {bodega}")
        columnas_ordenadas.append(f"Óptimo {bodega}")
    
    tabla_final = tabla_final[columnas_ordenadas]
    
    # Agregar fila de totales por columna
    fila_totales = tabla_final.sum(axis=0)
    fila_totales.name = ('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')
    tabla_final = pd.concat([tabla_final, fila_totales.to_frame().T])
    
    print(f"✅ Procesamiento completado - Puerto Rico: {len(tabla_final)} filas en tabla final")
    
    return tabla_final

def procesar_archivo_optimos_gt(df_optimos: pd.DataFrame) -> Dict[str, Dict[str, float]]:
    """
    Procesa el archivo CSV de cantidades óptimas para Guatemala
    Retorna: {codigo: {bodega: cantidad_optima}}
    """
    if df_optimos is None or df_optimos.empty:
        print("DataFrame de óptimos está vacío o es None")
        return {}
    
    print(f"Archivo de óptimos cargado con {len(df_optimos)} filas y {len(df_optimos.columns)} columnas")
    print(f"Columnas disponibles: {list(df_optimos.columns)}")
    
    # Buscar columna de código
    codigo_col = None
    for col in df_optimos.columns:
        if 'codigo' in col.lower() or 'estilo' in col.lower() or col.lower() in ['u_estilo', 'código']:
            codigo_col = col
            break
    
    if codigo_col is None:
        print("No se encontró columna de código en archivo de óptimos")
        print("Buscando en la primera columna como código por defecto...")
        if len(df_optimos.columns) > 0:
            codigo_col = df_optimos.columns[0]
            print(f"Usando primera columna como código: {codigo_col}")
        else:
            print("No hay columnas disponibles")
            return {}
    
    # Obtener bodegas de Guatemala
    bodegas_guatemala = [
        "NE Oakland", "NE Cayala", "NE Miraflores", "NE Portales", "NE InterXela",
        "NE Metronorte", "NE Concepcion", "NE Interplaza Escuintla", "NE Pradera Huehuetenango",
        "NE Naranjo", "NE Metrocentro Outlet", "NE Vistares", "NE Peri Roosvelt",
        "NE Outlet Santa clara", "NE Plaza Magdalena", "NE Pradera Chiquimula",
        "NE Pradera Escuintla", "NE Paseo Antigua", "NE Pradera Xela", "NE Chimaltenango",
        "NE Plaza Videre", "NE Metroplaza Jutiapa", "NE Puerto Barrios"
    ]
    
    # Mapear columnas de archivo con bodegas (nombres similares)
    def encontrar_bodega_similar(col_name: str) -> str:
        """Encuentra la bodega más similar basada en nombres"""
        col_clean = col_name.lower().replace('ne ', '').replace('_', ' ').strip()
        
        # Primero buscar coincidencia exacta
        for bodega in bodegas_guatemala:
            if col_name == bodega:
                return bodega
        
        # Luego buscar coincidencias parciales
        for bodega in bodegas_guatemala:
            bodega_clean = bodega.lower().replace('ne ', '').replace('_', ' ').strip()
            # Buscar coincidencias parciales
            if col_clean in bodega_clean or bodega_clean in col_clean:
                return bodega
            # Buscar palabras clave
            col_words = col_clean.split()
            bodega_words = bodega_clean.split()
            if len(set(col_words) & set(bodega_words)) >= 1:
                return bodega
        
        # Si no encuentra nada, intentar mapeo más flexible
        print(f"No se pudo mapear columna '{col_name}' con ninguna bodega")
        return None
    
    # Debug: Mostrar mapeo de columnas con bodegas
    print("Mapeo de columnas de archivo con bodegas:")
    bodegas_mapeadas = {}
    for col in df_optimos.columns:
        if col != codigo_col:
            bodega_mapeada = encontrar_bodega_similar(col)
            bodegas_mapeadas[col] = bodega_mapeada
            print(f"  '{col}' -> '{bodega_mapeada}'")
    
    # Crear mapeo de códigos a óptimos por bodega
    optimos_dict = {}
    
    for _, row in df_optimos.iterrows():
        codigo = str(row[codigo_col]).strip()
        if not codigo or codigo == 'nan':
            continue
            
        optimos_dict[codigo] = {}
        
        # Mapear cada columna con bodegas
        for col in df_optimos.columns:
            if col != codigo_col:
                bodega_mapeada = bodegas_mapeadas[col]
                if bodega_mapeada:
                    try:
                        cantidad = float(row[col]) if pd.notnull(row[col]) else 0
                        optimos_dict[codigo][bodega_mapeada] = cantidad
                    except Exception as e:
                        print(f"Error procesando {codigo}-{bodega_mapeada}: {e}")
                        optimos_dict[codigo][bodega_mapeada] = 0
    
    print(f"Procesados {len(optimos_dict)} códigos con cantidades óptimas")
    
    if len(optimos_dict) > 0:
        # Mostrar ejemplo del primer código procesado
        primer_codigo = list(optimos_dict.keys())[0]
        print(f"Ejemplo - Código {primer_codigo}: {optimos_dict[primer_codigo]}")
    
    return optimos_dict

def calcular_color_semaforo_mvp(real: float, optimo: float) -> str:
    """
    Calcula el color del semáforo basado en cumplimiento del óptimo
    Verde: Stock real >= Stock óptimo
    Amarillo: Stock real < Stock óptimo pero >= 80% del óptimo
    Rojo: Stock real < 80% del óptimo
    """
    if optimo == 0:
        # Aplicar misma lógica cuando stock óptimo = 0
        if real == 0:
            return "#d4edda"  # Verde - cumple perfectamente (ambos son 0)
        else:
            # Cualquier stock real > 0 cuando óptimo = 0 es desviación significativa
            return "#f8d7da"  # Rojo - no debería tener stock
    
    # Nueva lógica basada en cumplimiento
    if real >= optimo:
        return "#d4edda"  # Verde - cumple o supera el óptimo
    elif real >= optimo * 0.8:  # 80% o más del óptimo
        return "#fff3cd"  # Amarillo - falta máximo 20%
    else:
        return "#f8d7da"  # Rojo - falta más del 20%

def contar_celdas_semaforo_mvp(tabla_mvp: pd.DataFrame, columnas_real: List[str], columnas_optimo: List[str]) -> dict:
    """
    Cuenta las celdas por color de semáforo en la tabla MVP
    Retorna: {'verde': count, 'amarillo': count, 'rojo': count}
    """
    print(f"DEBUG: Iniciando conteo de celdas semáforo. Filas: {len(tabla_mvp)}, Columnas Real: {len(columnas_real)}")
    contadores = {'verde': 0, 'amarillo': 0, 'rojo': 0}
    
    # Iterar por todas las filas excluyendo la fila TOTAL
    for idx, row in tabla_mvp.iterrows():
        # Saltar la fila TOTAL
        if isinstance(idx, tuple) and idx[0] == 'TOTAL':
            continue
            
        # Iterar por cada par de columnas Real/Óptimo
        for col_real, col_optimo in zip(columnas_real, columnas_optimo):
            if col_real in tabla_mvp.columns and col_optimo in tabla_mvp.columns:
                valor_real = row[col_real] if pd.notna(row[col_real]) else 0
                valor_optimo = row[col_optimo] if pd.notna(row[col_optimo]) else 0
                
                # Calcular color según la lógica del semáforo
                color_hex = calcular_color_semaforo_mvp(float(valor_real), float(valor_optimo))
                
                # Contar según el color
                if color_hex == "#d4edda":  # Verde
                    contadores['verde'] += 1
                elif color_hex == "#fff3cd":  # Amarillo
                    contadores['amarillo'] += 1
                elif color_hex == "#f8d7da":  # Rojo
                    contadores['rojo'] += 1
    
    return contadores

def mostrar_stock_mvps_guatemala(df_stock: pd.DataFrame, key_suffix: str = ""):
    """Muestra la tabla de stock de códigos MVP para Guatemala con nueva funcionalidad"""
    if df_stock is None or df_stock.empty:
        st.warning("No se pudo cargar el archivo de stock para mostrar MVP")
        return
    
    # Crear sección
    professional_design.create_section_header(
        "Stock de MVPS - Guatemala", 
        "Stock actual vs Stock óptimo nuevo de códigos MVP específicos",
        "🏆"
    )
    
    # Procesar datos
    tabla_mvp = procesar_stock_mvps_guatemala(df_stock)
    
    if tabla_mvp.empty:
        st.warning("No se encontraron datos de códigos MVP en el stock de Guatemala")
        return
    
    # Obtener columnas Real y Óptimo
    columnas_real = [col for col in tabla_mvp.columns if col.startswith('Real ')]
    columnas_optimo = [col for col in tabla_mvp.columns if col.startswith('Óptimo ')]
    
    
    # Mostrar métricas resumen
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        # Contar códigos únicos excluyendo la fila TOTAL
        codigos_unicos = tabla_mvp.index.get_level_values(0).unique()
        total_productos = len([codigo for codigo in codigos_unicos if codigo != 'TOTAL'])
        st.metric("Total Productos MVP", f"{total_productos:,}")
    
    with col2:
        total_stock_real = tabla_mvp[columnas_real].loc[('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')].sum()
        st.metric("Total Stock Real", f"{int(total_stock_real):,}")
    
    with col3:
        total_stock_optimo = tabla_mvp[columnas_optimo].loc[('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')].sum()
        st.metric("Total Stock Óptimo", f"{int(total_stock_optimo):,}")
    
    with col4:
        if total_stock_optimo > 0:
            cumplimiento = (total_stock_real / total_stock_optimo) * 100
            st.metric("Cumplimiento de unidades totales", f"{cumplimiento:.1f}%")
    
    # Nueva fila de métricas - Contar celdas de semáforo
    contadores_semaforo = contar_celdas_semaforo_mvp(tabla_mvp, columnas_real, columnas_optimo)
    print(f"DEBUG: Contadores semáforo - Verde: {contadores_semaforo['verde']}, Amarillo: {contadores_semaforo['amarillo']}, Rojo: {contadores_semaforo['rojo']}")
    
    col5, col6, col7, col8 = st.columns(4)
    
    with col5:
        # Columna vacía para mantener simetría
        st.metric("", "")
    
    with col6:
        celdas_verde_amarillo = contadores_semaforo['verde'] + contadores_semaforo['amarillo']
        st.metric("Cantidad de celdas en verde y amarillo", f"{celdas_verde_amarillo:,}")
    
    with col7:
        celdas_rojo = contadores_semaforo['rojo']
        st.metric("Cantidad de celdas en rojo", f"{celdas_rojo:,}")
    
    with col8:
        # Calcular % de Cumplimiento con nueva fórmula
        total_celdas = celdas_verde_amarillo + celdas_rojo
        if total_celdas > 0:
            cumplimiento_semaforo = (celdas_verde_amarillo / total_celdas) * 100
            st.metric("% de Cumplimiento", f"{cumplimiento_semaforo:.1f}%")
        else:
            st.metric("% de Cumplimiento", "0.0%")
    
    # Crear tabla HTML con formato profesional
    def crear_tabla_html_mvp(df):
        """Crea tabla HTML con formato profesional para MVP"""
        # Resetear índice para mostrar las columnas de información
        df_display = df.reset_index()
        
        # Crear HTML de la tabla
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 9px; width: 100%; margin-top: 20px;">'
        
        # Obtener bodegas únicas
        bodegas = []
        for col in columnas_real:
            bodega = col.replace('Real ', '')
            bodegas.append(bodega)
        
        # Fila 1: Encabezado principal con bodegas
        html += '<tr style="background-color: #000000; color: white; font-weight: bold; height: 40px;">'
        
        # Columnas de información
        info_headers = ['Código', 'Codigo_SAP', 'Segmento', 'Silueta', 'Colección', 'Descripción', 'Talla']
        widths = [80, 80, 70, 70, 90, 150, 60]
        
        for header, width in zip(info_headers, widths):
            html += f'<td style="border: 1px solid #fff; padding: 8px; width: {width}px; vertical-align: middle;">{header}</td>'
        
        # Columnas de bodegas (2 columnas por bodega: Real y Óptimo)
        for bodega in bodegas:
            html += f'<td colspan="2" style="border: 1px solid #fff; padding: 8px; vertical-align: middle; min-width: 120px;">{bodega}</td>'
        
        html += '</tr>'
        
        # Fila 2: Sub-encabezados (Real / Óptimo)
        html += '<tr style="background-color: #333333; color: white; font-weight: bold; height: 30px;">'
        
        # Espacios vacíos para columnas de información
        for i in range(7):
            html += '<td style="border: 1px solid #fff; padding: 4px;"></td>'
        
        # Sub-encabezados para cada bodega
        for _ in bodegas:
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #28a745; font-size: 8px;">Real</td>'
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #007bff; font-size: 8px;">Óptimo</td>'
        
        html += '</tr>'
        
        # Filas de datos
        for idx, row in df_display.iterrows():
            # Verificar si es la fila TOTAL
            es_total = (str(row.iloc[0]) == 'TOTAL')
            
            if es_total:
                html += '<tr style="background-color: #f8f9fa; font-weight: bold; border-top: 3px solid #000;">'
            else:
                html += '<tr style="background-color: white;">'
            
            # Columnas de información
            for i in range(7):
                valor = str(row.iloc[i]) if i < len(row) else ""
                style = "border: 1px solid #ddd; padding: 4px; text-align: left;" if i in [4, 5] else "border: 1px solid #ddd; padding: 4px;"
                html += f'<td style="{style}">{valor}</td>'
            
            # Columnas de bodegas (Real y Óptimo)
            for bodega in bodegas:
                col_real = f'Real {bodega}'
                col_optimo = f'Óptimo {bodega}'
                
                # Valores reales y óptimos
                valor_real = row[col_real] if col_real in row else 0
                valor_optimo = row[col_optimo] if col_optimo in row else 0
                
                try:
                    valor_real_num = float(str(valor_real).replace(',', '')) if valor_real != 'TOTAL' else 0
                    valor_optimo_num = float(str(valor_optimo).replace(',', '')) if valor_optimo != 'TOTAL' else 0
                except:
                    valor_real_num = 0
                    valor_optimo_num = 0
                
                # Color del semáforo solo para la columna Real
                if not es_total:
                    color_fondo = calcular_color_semaforo_mvp(valor_real_num, valor_optimo_num)
                else:
                    color_fondo = "white"
                
                # Formatear valores
                valor_real_display = f"{int(valor_real_num):,}" if valor_real_num > 0 else "0"
                valor_optimo_display = f"{int(valor_optimo_num):,}" if valor_optimo_num > 0 else "0"
                
                # Celda Real (con semáforo)
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: {color_fondo}; font-weight: bold;">{valor_real_display}</td>'
                
                # Celda Óptimo (sin semáforo)
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: #f8f9fa; color: #007bff;">{valor_optimo_display}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        
        return html
    
    # Leyenda del semáforo (antes de la tabla)
    st.markdown("""
    <div style="margin-bottom: 20px; padding: 15px; background-color: #f8f9fa; border-radius: 8px;">
        <h4 style="margin-bottom: 10px; color: #333;">🚦 Leyenda del Semáforo</h4>
        <div style="display: flex; gap: 20px; align-items: center;">
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #d4edda; border: 1px solid #c3e6cb; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Verde: Stock real >= Stock óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Amarillo: Stock real entre 80%-99% del óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #f8d7da; border: 1px solid #f1b0b7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Rojo: Stock real < 80% del óptimo</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Mostrar tabla
    tabla_html = crear_tabla_html_mvp(tabla_mvp)
    st.markdown(tabla_html, unsafe_allow_html=True)
    
    # Botón de exportación a Excel con colores
    st.markdown("---")
    if st.button("📊 Exportar Tabla MVP a Excel", type="primary", key=f"export_mvp_excel_{key_suffix}"):
        excel_data = exportar_mvp_excel_con_colores(tabla_mvp, columnas_real, columnas_optimo, "Guatemala")
        
        if excel_data:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            st.download_button(
                label="📥 Descargar",
                data=excel_data,
                file_name=f"MVP_Guatemala_Semaforo_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_mvp_excel_{key_suffix}",
                type="primary",
                use_container_width=True
            )

def mostrar_stock_mvps_honduras(df_stock: pd.DataFrame, key_suffix: str = ""):
    """Muestra la tabla de stock de códigos MVP para Honduras con mismo formato que Guatemala y El Salvador"""
    if df_stock is None or df_stock.empty:
        st.warning("No se pudo cargar el archivo de stock para mostrar MVP")
        return
    
    # Crear sección
    professional_design.create_section_header(
        "Stock de MVPS - Honduras", 
        "Stock actual vs Stock óptimo nuevo de códigos MVP específicos",
        "🏆"
    )
    
    # Procesar datos
    tabla_mvp = procesar_stock_mvps_honduras(df_stock)
    
    if tabla_mvp.empty:
        st.warning("No se encontraron datos de códigos MVP en el stock de Honduras")
        return
    
    # Obtener columnas Real y Óptimo
    columnas_real = [col for col in tabla_mvp.columns if col.startswith('Real ')]
    columnas_optimo = [col for col in tabla_mvp.columns if col.startswith('Óptimo ')]
    
    
    # Mostrar métricas resumen
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        # Contar códigos únicos excluyendo la fila TOTAL
        codigos_unicos = tabla_mvp.index.get_level_values(0).unique()
        total_productos = len([codigo for codigo in codigos_unicos if codigo != 'TOTAL'])
        st.metric("Total Productos MVP", f"{total_productos:,}")
    
    with col2:
        total_stock_real = tabla_mvp[columnas_real].loc[('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')].sum()
        st.metric("Total Stock Real", f"{int(total_stock_real):,}")
    
    with col3:
        total_stock_optimo = tabla_mvp[columnas_optimo].loc[('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')].sum()
        st.metric("Total Stock Óptimo", f"{int(total_stock_optimo):,}")
    
    with col4:
        if total_stock_optimo > 0:
            cumplimiento = (total_stock_real / total_stock_optimo) * 100
            st.metric("Cumplimiento de unidades totales", f"{cumplimiento:.1f}%")
    
    # Nueva fila de métricas - Contar celdas de semáforo
    contadores_semaforo = contar_celdas_semaforo_mvp(tabla_mvp, columnas_real, columnas_optimo)
    print(f"DEBUG: Contadores semáforo - Verde: {contadores_semaforo['verde']}, Amarillo: {contadores_semaforo['amarillo']}, Rojo: {contadores_semaforo['rojo']}")
    
    col5, col6, col7, col8 = st.columns(4)
    
    with col5:
        # Columna vacía para mantener simetría
        st.metric("", "")
    
    with col6:
        celdas_verde_amarillo = contadores_semaforo['verde'] + contadores_semaforo['amarillo']
        st.metric("Cantidad de celdas en verde y amarillo", f"{celdas_verde_amarillo:,}")
    
    with col7:
        celdas_rojo = contadores_semaforo['rojo']
        st.metric("Cantidad de celdas en rojo", f"{celdas_rojo:,}")
    
    with col8:
        # Calcular % de Cumplimiento con nueva fórmula
        total_celdas = celdas_verde_amarillo + celdas_rojo
        if total_celdas > 0:
            cumplimiento_semaforo = (celdas_verde_amarillo / total_celdas) * 100
            st.metric("% de Cumplimiento", f"{cumplimiento_semaforo:.1f}%")
        else:
            st.metric("% de Cumplimiento", "0.0%")
    
    # Crear tabla HTML con formato profesional
    def crear_tabla_html_mvp_honduras(df):
        """Crea tabla HTML con formato profesional para MVP Honduras"""
        # Resetear índice para mostrar las columnas de información
        df_display = df.reset_index()
        
        # Crear HTML de la tabla
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 9px; width: 100%; margin-top: 20px;">'
        
        # Obtener bodegas únicas
        bodegas = []
        for col in columnas_real:
            bodega = col.replace('Real ', '')
            bodegas.append(bodega)
        
        # Fila 1: Encabezado principal con bodegas
        html += '<tr style="background-color: #000000; color: white; font-weight: bold; height: 40px;">'
        
        # Columnas de información
        info_headers = ['Código', 'Codigo_SAP', 'Segmento', 'Silueta', 'Colección', 'Descripción', 'Talla']
        widths = [80, 80, 70, 70, 90, 150, 60]
        
        for header, width in zip(info_headers, widths):
            html += f'<td style="border: 1px solid #fff; padding: 8px; width: {width}px; vertical-align: middle;">{header}</td>'
        
        # Columnas de bodegas (2 columnas por bodega: Real y Óptimo)
        for bodega in bodegas:
            html += f'<td colspan="2" style="border: 1px solid #fff; padding: 8px; vertical-align: middle; min-width: 120px;">{bodega}</td>'
        
        html += '</tr>'
        
        # Fila 2: Sub-encabezados (Real / Óptimo)
        html += '<tr style="background-color: #333333; color: white; font-weight: bold; height: 30px;">'
        
        # Espacios vacíos para columnas de información
        for i in range(7):
            html += '<td style="border: 1px solid #fff; padding: 4px;"></td>'
        
        # Sub-encabezados para cada bodega
        for _ in bodegas:
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #28a745; font-size: 8px;">Real</td>'
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #007bff; font-size: 8px;">Óptimo</td>'
        
        html += '</tr>'
        
        # Filas de datos
        for idx, row in df_display.iterrows():
            # Verificar si es la fila TOTAL
            es_total = (str(row.iloc[0]) == 'TOTAL')
            
            if es_total:
                html += '<tr style="background-color: #f8f9fa; font-weight: bold; border-top: 3px solid #000;">'
            else:
                html += '<tr style="background-color: white;">'
            
            # Columnas de información
            for i in range(7):
                valor = str(row.iloc[i]) if i < len(row) else ""
                style = "border: 1px solid #ddd; padding: 4px; text-align: left;" if i in [4, 5] else "border: 1px solid #ddd; padding: 4px;"
                html += f'<td style="{style}">{valor}</td>'
            
            # Columnas de bodegas (Real y Óptimo)
            for bodega in bodegas:
                col_real = f'Real {bodega}'
                col_optimo = f'Óptimo {bodega}'
                
                # Valores reales y óptimos
                valor_real = row[col_real] if col_real in row else 0
                valor_optimo = row[col_optimo] if col_optimo in row else 0
                
                try:
                    valor_real_num = float(str(valor_real).replace(',', '')) if valor_real != 'TOTAL' else 0
                    valor_optimo_num = float(str(valor_optimo).replace(',', '')) if valor_optimo != 'TOTAL' else 0
                except:
                    valor_real_num = 0
                    valor_optimo_num = 0
                
                # Color del semáforo solo para la columna Real
                if not es_total:
                    color_fondo = calcular_color_semaforo_mvp(valor_real_num, valor_optimo_num)
                else:
                    color_fondo = "white"
                
                # Formatear valores
                valor_real_display = f"{int(valor_real_num):,}" if valor_real_num > 0 else "0"
                valor_optimo_display = f"{int(valor_optimo_num):,}" if valor_optimo_num > 0 else "0"
                
                # Celda Real (con semáforo)
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: {color_fondo}; font-weight: bold;">{valor_real_display}</td>'
                
                # Celda Óptimo (sin semáforo)
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: #f8f9fa; color: #007bff;">{valor_optimo_display}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        
        return html
    
    # Leyenda del semáforo (antes de la tabla)
    st.markdown("""
    <div style="margin-bottom: 20px; padding: 15px; background-color: #f8f9fa; border-radius: 8px;">
        <h4 style="margin-bottom: 10px; color: #333;">🚦 Leyenda del Semáforo</h4>
        <div style="display: flex; gap: 20px; align-items: center;">
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #d4edda; border: 1px solid #c3e6cb; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Verde: Stock real >= Stock óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Amarillo: Stock real entre 80%-99% del óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #f8d7da; border: 1px solid #f1b0b7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Rojo: Stock real < 80% del óptimo</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Mostrar tabla
    tabla_html = crear_tabla_html_mvp_honduras(tabla_mvp)
    st.markdown(tabla_html, unsafe_allow_html=True)
    
    # Botón de exportación a Excel con colores
    st.markdown("---")
    if st.button("📊 Exportar Tabla MVP a Excel", type="primary", key=f"export_mvp_excel_{key_suffix}"):
        excel_data = exportar_mvp_excel_con_colores(tabla_mvp, columnas_real, columnas_optimo, "Honduras")
        
        if excel_data:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            st.download_button(
                label="📥 Descargar",
                data=excel_data,
                file_name=f"MVP_Honduras_Semaforo_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_mvp_excel_{key_suffix}",
                type="primary",
                use_container_width=True
            )

def mostrar_stock_mvps_costarica(df_stock: pd.DataFrame, key_suffix: str = ""):
    """Muestra la tabla de stock de códigos MVP para Costa Rica con mismo formato que Guatemala, El Salvador y Honduras"""
    if df_stock is None or df_stock.empty:
        st.warning("No se pudo cargar el archivo de stock para mostrar MVP")
        return
    
    # Crear sección
    professional_design.create_section_header(
        "Stock de MVPS - Costa Rica", 
        "Stock actual vs Stock óptimo nuevo de códigos MVP específicos",
        "🏆"
    )
    
    # Procesar datos
    tabla_mvp = procesar_stock_mvps_costarica(df_stock)
    
    if tabla_mvp.empty:
        st.warning("No se encontraron datos de códigos MVP en el stock de Costa Rica")
        return
    
    # Obtener columnas Real y Óptimo
    columnas_real = [col for col in tabla_mvp.columns if col.startswith('Real ')]
    columnas_optimo = [col for col in tabla_mvp.columns if col.startswith('Óptimo ')]
    
    
    # Mostrar métricas resumen
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        # Contar códigos únicos excluyendo la fila TOTAL
        codigos_unicos = tabla_mvp.index.get_level_values(0).unique()
        total_productos = len([codigo for codigo in codigos_unicos if codigo != 'TOTAL'])
        st.metric("Total Productos MVP", f"{total_productos:,}")
    
    with col2:
        total_stock_real = tabla_mvp[columnas_real].loc[('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')].sum()
        st.metric("Total Stock Real", f"{int(total_stock_real):,}")
    
    with col3:
        total_stock_optimo = tabla_mvp[columnas_optimo].loc[('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')].sum()
        st.metric("Total Stock Óptimo", f"{int(total_stock_optimo):,}")
    
    with col4:
        if total_stock_optimo > 0:
            cumplimiento = (total_stock_real / total_stock_optimo) * 100
            st.metric("Cumplimiento de unidades totales", f"{cumplimiento:.1f}%")
    
    # Nueva fila de métricas - Contar celdas de semáforo
    contadores_semaforo = contar_celdas_semaforo_mvp(tabla_mvp, columnas_real, columnas_optimo)
    print(f"DEBUG: Contadores semáforo - Verde: {contadores_semaforo['verde']}, Amarillo: {contadores_semaforo['amarillo']}, Rojo: {contadores_semaforo['rojo']}")
    
    col5, col6, col7, col8 = st.columns(4)
    
    with col5:
        # Columna vacía para mantener simetría
        st.metric("", "")
    
    with col6:
        celdas_verde_amarillo = contadores_semaforo['verde'] + contadores_semaforo['amarillo']
        st.metric("Cantidad de celdas en verde y amarillo", f"{celdas_verde_amarillo:,}")
    
    with col7:
        celdas_rojo = contadores_semaforo['rojo']
        st.metric("Cantidad de celdas en rojo", f"{celdas_rojo:,}")
    
    with col8:
        # Calcular % de Cumplimiento con nueva fórmula
        total_celdas = celdas_verde_amarillo + celdas_rojo
        if total_celdas > 0:
            cumplimiento_semaforo = (celdas_verde_amarillo / total_celdas) * 100
            st.metric("% de Cumplimiento", f"{cumplimiento_semaforo:.1f}%")
        else:
            st.metric("% de Cumplimiento", "0.0%")
    
    # Crear tabla HTML con formato profesional
    def crear_tabla_html_mvp_costarica(df):
        """Crea tabla HTML con formato profesional para MVP Costa Rica"""
        # Resetear índice para mostrar las columnas de información
        df_display = df.reset_index()
        
        # Crear HTML de la tabla
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 9px; width: 100%; margin-top: 20px;">'
        
        # Obtener bodegas únicas
        bodegas = []
        for col in columnas_real:
            bodega = col.replace('Real ', '')
            bodegas.append(bodega)
        
        # Fila 1: Encabezado principal con bodegas
        html += '<tr style="background-color: #000000; color: white; font-weight: bold; height: 40px;">'
        
        # Columnas de información
        info_headers = ['Código', 'Codigo_SAP', 'Segmento', 'Silueta', 'Colección', 'Descripción', 'Talla']
        widths = [80, 80, 70, 70, 90, 150, 60]
        
        for header, width in zip(info_headers, widths):
            html += f'<td style="border: 1px solid #fff; padding: 8px; width: {width}px; vertical-align: middle;">{header}</td>'
        
        # Columnas de bodegas (2 columnas por bodega: Real y Óptimo)
        for bodega in bodegas:
            html += f'<td colspan="2" style="border: 1px solid #fff; padding: 8px; vertical-align: middle; min-width: 120px;">{bodega}</td>'
        
        html += '</tr>'
        
        # Fila 2: Sub-encabezados (Real / Óptimo)
        html += '<tr style="background-color: #333333; color: white; font-weight: bold; height: 30px;">'
        
        # Espacios vacíos para columnas de información
        for i in range(7):
            html += '<td style="border: 1px solid #fff; padding: 4px;"></td>'
        
        # Sub-encabezados para cada bodega
        for _ in bodegas:
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #28a745; font-size: 8px;">Real</td>'
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #007bff; font-size: 8px;">Óptimo</td>'
        
        html += '</tr>'
        
        # Filas de datos
        for idx, row in df_display.iterrows():
            # Verificar si es la fila TOTAL
            es_total = (str(row.iloc[0]) == 'TOTAL')
            
            if es_total:
                html += '<tr style="background-color: #f8f9fa; font-weight: bold; border-top: 3px solid #000;">'
            else:
                html += '<tr style="background-color: white;">'
            
            # Columnas de información
            for i in range(7):
                valor = str(row.iloc[i]) if i < len(row) else ""
                style = "border: 1px solid #ddd; padding: 4px; text-align: left;" if i in [4, 5] else "border: 1px solid #ddd; padding: 4px;"
                html += f'<td style="{style}">{valor}</td>'
            
            # Columnas de bodegas (Real y Óptimo)
            for bodega in bodegas:
                col_real = f'Real {bodega}'
                col_optimo = f'Óptimo {bodega}'
                
                # Valores reales y óptimos
                valor_real = row[col_real] if col_real in row else 0
                valor_optimo = row[col_optimo] if col_optimo in row else 0
                
                try:
                    valor_real_num = float(str(valor_real).replace(',', '')) if valor_real != 'TOTAL' else 0
                    valor_optimo_num = float(str(valor_optimo).replace(',', '')) if valor_optimo != 'TOTAL' else 0
                except:
                    valor_real_num = 0
                    valor_optimo_num = 0
                
                # Color del semáforo solo para la columna Real
                if not es_total:
                    color_fondo = calcular_color_semaforo_mvp(valor_real_num, valor_optimo_num)
                else:
                    color_fondo = "white"
                
                # Formatear valores
                valor_real_display = f"{int(valor_real_num):,}" if valor_real_num > 0 else "0"
                valor_optimo_display = f"{int(valor_optimo_num):,}" if valor_optimo_num > 0 else "0"
                
                # Celda Real (con semáforo)
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: {color_fondo}; font-weight: bold;">{valor_real_display}</td>'
                
                # Celda Óptimo (sin semáforo)
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: #f8f9fa; color: #007bff;">{valor_optimo_display}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        
        return html
    
    # Leyenda del semáforo (antes de la tabla)
    st.markdown("""
    <div style="margin-bottom: 20px; padding: 15px; background-color: #f8f9fa; border-radius: 8px;">
        <h4 style="margin-bottom: 10px; color: #333;">🚦 Leyenda del Semáforo</h4>
        <div style="display: flex; gap: 20px; align-items: center;">
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #d4edda; border: 1px solid #c3e6cb; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Verde: Stock real >= Stock óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Amarillo: Stock real entre 80%-99% del óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #f8d7da; border: 1px solid #f1b0b7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Rojo: Stock real < 80% del óptimo</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Mostrar tabla
    tabla_html = crear_tabla_html_mvp_costarica(tabla_mvp)
    st.markdown(tabla_html, unsafe_allow_html=True)
    
    # Botón de exportación a Excel con colores
    st.markdown("---")
    if st.button("📊 Exportar Tabla MVP a Excel", type="primary", key=f"export_mvp_excel_{key_suffix}"):
        excel_data = exportar_mvp_excel_con_colores(tabla_mvp, columnas_real, columnas_optimo, "CostaRica")
        
        if excel_data:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            st.download_button(
                label="📥 Descargar",
                data=excel_data,
                file_name=f"MVP_CostaRica_Semaforo_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_mvp_excel_{key_suffix}",
                type="primary",
                use_container_width=True
            )

def exportar_mvp_excel_con_colores(tabla_mvp: pd.DataFrame, columnas_real: List[str], columnas_optimo: List[str], pais: str = "Guatemala") -> bytes:
    """
    Exporta la tabla MVP a Excel con formato profesional y colores de semáforo
    """
    try:
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Crear archivo Excel en memoria
        output = BytesIO()
        
        # Resetear índice para tener las columnas de información como columnas normales
        df_export = tabla_mvp.reset_index()
        
        # Renombrar columnas de información
        df_export.columns = ['Código', 'Codigo_SAP', 'Segmento', 'Silueta', 'Colección', 'Descripción', 'Talla'] + list(df_export.columns[7:])
        
        # Crear workbook y worksheet
        sheet_name = f'MVP_{pais}'
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Obtener worksheet para formatear
            worksheet = writer.sheets[sheet_name]
            
            # Ocultar líneas de cuadrícula en celdas no utilizadas
            worksheet.sheet_view.showGridLines = False
            
            # Configurar estilos
            # Fuentes
            font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            font_subheader = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            font_normal = Font(name='Arial', size=10)
            font_total = Font(name='Arial', size=10, bold=True, color='FFFFFF')
            
            # Alineación
            align_center = Alignment(horizontal='center', vertical='center')
            align_left = Alignment(horizontal='left', vertical='center')
            
            # Colores de fondo
            fill_header = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            fill_subheader_real = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')  # Verde
            fill_subheader_optimo = PatternFill(start_color='007BFF', end_color='007BFF', fill_type='solid')  # Azul
            fill_total = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
            
            # Colores semáforo
            fill_semaforo_verde = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
            fill_semaforo_amarillo = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
            fill_semaforo_rojo = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            fill_optimo = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            
            # Bordes
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Obtener bodegas y crear mapeo de columnas
            bodegas = []
            col_mapping = {}
            col_idx = 8  # Las primeras 7 son información (Código, Codigo_SAP, Segmento, Silueta, Colección, Descripción, Talla)
            
            for col_real in columnas_real:
                bodega = col_real.replace('Real ', '')
                bodegas.append(bodega)
                col_mapping[col_real] = col_idx
                col_mapping[f'Óptimo {bodega}'] = col_idx + 1
                col_idx += 2
            
            # 1. FORMATEAR ENCABEZADOS PRINCIPALES
            # Fila 1: Información + Bodegas
            for col_num in range(1, 8):  # Columnas de información (incluye Codigo_SAP)
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border
            
            # Agregar encabezados de bodegas (fusionar celdas para Real + Óptimo)
            for i, bodega in enumerate(bodegas):
                start_col = 8 + (i * 2)
                end_col = start_col + 1
                
                # Fusionar celdas para la bodega
                worksheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                
                # Formatear celda fusionada
                cell = worksheet.cell(row=1, column=start_col)
                cell.value = bodega
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border
                
                # Aplicar bordes a la celda fusionada
                for col in range(start_col, end_col + 1):
                    worksheet.cell(row=1, column=col).border = border
            
            # 2. AGREGAR FILA DE SUB-ENCABEZADOS (Real / Óptimo)
            worksheet.insert_rows(2)
            
            # Sub-encabezados para información (vacíos)
            for col_num in range(1, 8):
                cell = worksheet.cell(row=2, column=col_num)
                cell.font = font_subheader
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = border
            
            # Sub-encabezados Real/Óptimo
            for i, bodega in enumerate(bodegas):
                start_col = 8 + (i * 2)
                
                # Columna Real
                cell_real = worksheet.cell(row=2, column=start_col)
                cell_real.value = "Real"
                cell_real.font = font_subheader
                cell_real.fill = fill_subheader_real
                cell_real.alignment = align_center
                cell_real.border = border
                
                # Columna Óptimo
                cell_optimo = worksheet.cell(row=2, column=start_col + 1)
                cell_optimo.value = "Óptimo"
                cell_optimo.font = font_subheader
                cell_optimo.fill = fill_subheader_optimo
                cell_optimo.alignment = align_center
                cell_optimo.border = border
            
            # 3. FORMATEAR DATOS Y APLICAR SEMÁFORO
            total_rows = worksheet.max_row
            
            for row_num in range(3, total_rows + 1):  # Empezar desde fila 3 (datos)
                # Verificar si es fila TOTAL
                codigo_cell = worksheet.cell(row=row_num, column=1)
                es_fila_total = str(codigo_cell.value) == 'TOTAL'
                
                # Formatear columnas de información
                for col_num in range(1, 8):  # Corregido: incluir las 7 columnas de información
                    cell = worksheet.cell(row=row_num, column=col_num)
                    
                    if es_fila_total:
                        cell.font = font_total
                        cell.fill = fill_total
                    else:
                        cell.font = font_normal
                    
                    # Alineación según tipo de columna
                    if col_num in [5, 6]:  # Colección y Descripción (columnas E y F)
                        cell.alignment = align_left
                    else:
                        cell.alignment = align_center
                    
                    cell.border = border
                
                # Formatear columnas de bodegas con semáforo
                for i, bodega in enumerate(bodegas):
                    col_real = 8 + (i * 2)  # Corregido: empezar desde columna 8
                    col_optimo = col_real + 1
                    
                    cell_real = worksheet.cell(row=row_num, column=col_real)
                    cell_optimo = worksheet.cell(row=row_num, column=col_optimo)
                    
                    if es_fila_total:
                        # Fila TOTAL: fondo negro
                        cell_real.font = font_total
                        cell_real.fill = fill_total
                        cell_optimo.font = font_total
                        cell_optimo.fill = fill_total
                    else:
                        # Datos normales: aplicar semáforo solo a columna Real
                        try:
                            valor_real = float(str(cell_real.value).replace(',', '')) if cell_real.value else 0
                            valor_optimo = float(str(cell_optimo.value).replace(',', '')) if cell_optimo.value else 0
                            
                            # Aplicar semáforo a columna Real
                            color_semaforo = calcular_color_semaforo_mvp(valor_real, valor_optimo)
                            if color_semaforo == "#d4edda":  # Verde
                                cell_real.fill = fill_semaforo_verde
                            elif color_semaforo == "#fff3cd":  # Amarillo
                                cell_real.fill = fill_semaforo_amarillo
                            elif color_semaforo == "#f8d7da":  # Rojo
                                cell_real.fill = fill_semaforo_rojo
                            
                            # Columna Óptimo: fondo gris claro
                            cell_optimo.fill = fill_optimo
                            
                        except:
                            # En caso de error, usar colores por defecto
                            pass
                        
                        cell_real.font = font_normal
                        cell_optimo.font = font_normal
                    
                    cell_real.alignment = align_center
                    cell_optimo.alignment = align_center
                    cell_real.border = border
                    cell_optimo.border = border
            
            # 4. AJUSTAR ANCHOS DE COLUMNAS
            # Columnas de información
            column_widths = {
                'A': 12,  # Código
                'B': 12,  # Codigo_SAP
                'C': 12,  # Segmento
                'D': 12,  # Silueta
                'E': 20,  # Colección
                'F': 25,  # Descripción
                'G': 8,   # Talla
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # Columnas de bodegas (más estrechas) - ahora empiezan desde columna 8
            for i in range(len(bodegas) * 2):
                col_letter = get_column_letter(8 + i)
                worksheet.column_dimensions[col_letter].width = 10
            
            # 5. AGREGAR INFORMACIÓN DE LEYENDA
            leyenda_row = total_rows + 3
            
            # Título de leyenda
            worksheet.cell(row=leyenda_row, column=1, value="LEYENDA DEL SEMÁFORO (Solo columna Real):")
            worksheet.cell(row=leyenda_row, column=1).font = Font(name='Arial', size=11, bold=True)
            
            # Elementos de leyenda
            leyenda_items = [
                ("Verde: Stock real >= Stock óptimo", fill_semaforo_verde),
                ("Amarillo: Stock real entre 80%-99% del óptimo", fill_semaforo_amarillo), 
                ("Rojo: Stock real < 80% del óptimo", fill_semaforo_rojo)
            ]
            
            for i, (texto, fill) in enumerate(leyenda_items):
                row = leyenda_row + i + 1
                cell = worksheet.cell(row=row, column=1, value=texto)
                cell.font = Font(name='Arial', size=10)
                cell.fill = fill
                cell.border = border
        
        # Retornar datos del archivo
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error al generar Excel: {str(e)}")
        return None

def mostrar_stock_mvps_elsalvador(df_stock: pd.DataFrame, key_suffix: str = ""):
    """Muestra la tabla de stock de códigos MVP para El Salvador con mismo formato que Guatemala"""
    if df_stock is None or df_stock.empty:
        st.warning("No se pudo cargar el archivo de stock para mostrar MVP")
        return
    
    # Crear sección
    professional_design.create_section_header(
        "Stock de MVPS - El Salvador", 
        "Stock actual vs Stock óptimo nuevo de códigos MVP específicos",
        "🏆"
    )
    
    # Procesar datos
    tabla_mvp = procesar_stock_mvps_elsalvador(df_stock)
    
    if tabla_mvp.empty:
        st.warning("No se encontraron datos de códigos MVP en el stock de El Salvador")
        return
    
    # Obtener columnas Real y Óptimo
    columnas_real = [col for col in tabla_mvp.columns if col.startswith('Real ')]
    columnas_optimo = [col for col in tabla_mvp.columns if col.startswith('Óptimo ')]
    
    
    # Mostrar métricas resumen
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        # Contar códigos únicos excluyendo la fila TOTAL
        codigos_unicos = tabla_mvp.index.get_level_values(0).unique()
        total_productos = len([codigo for codigo in codigos_unicos if codigo != 'TOTAL'])
        st.metric("Total Productos MVP", f"{total_productos:,}")
    
    with col2:
        total_stock_real = tabla_mvp[columnas_real].loc[('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')].sum()
        st.metric("Total Stock Real", f"{int(total_stock_real):,}")
    
    with col3:
        total_stock_optimo = tabla_mvp[columnas_optimo].loc[('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')].sum()
        st.metric("Total Stock Óptimo", f"{int(total_stock_optimo):,}")
    
    with col4:
        if total_stock_optimo > 0:
            cumplimiento = (total_stock_real / total_stock_optimo) * 100
            st.metric("Cumplimiento de unidades totales", f"{cumplimiento:.1f}%")
    
    # Nueva fila de métricas - Contar celdas de semáforo
    contadores_semaforo = contar_celdas_semaforo_mvp(tabla_mvp, columnas_real, columnas_optimo)
    print(f"DEBUG: Contadores semáforo - Verde: {contadores_semaforo['verde']}, Amarillo: {contadores_semaforo['amarillo']}, Rojo: {contadores_semaforo['rojo']}")
    
    col5, col6, col7, col8 = st.columns(4)
    
    with col5:
        # Columna vacía para mantener simetría
        st.metric("", "")
    
    with col6:
        celdas_verde_amarillo = contadores_semaforo['verde'] + contadores_semaforo['amarillo']
        st.metric("Cantidad de celdas en verde y amarillo", f"{celdas_verde_amarillo:,}")
    
    with col7:
        celdas_rojo = contadores_semaforo['rojo']
        st.metric("Cantidad de celdas en rojo", f"{celdas_rojo:,}")
    
    with col8:
        # Calcular % de Cumplimiento con nueva fórmula
        total_celdas = celdas_verde_amarillo + celdas_rojo
        if total_celdas > 0:
            cumplimiento_semaforo = (celdas_verde_amarillo / total_celdas) * 100
            st.metric("% de Cumplimiento", f"{cumplimiento_semaforo:.1f}%")
        else:
            st.metric("% de Cumplimiento", "0.0%")
    
    # Crear tabla HTML con formato profesional
    def crear_tabla_html_mvp_elsalvador(df):
        """Crea tabla HTML con formato profesional para MVP El Salvador"""
        # Resetear índice para mostrar las columnas de información
        df_display = df.reset_index()
        
        # Crear HTML de la tabla
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 9px; width: 100%; margin-top: 20px;">'
        
        # Obtener bodegas únicas
        bodegas = []
        for col in columnas_real:
            bodega = col.replace('Real ', '')
            bodegas.append(bodega)
        
        # Fila 1: Encabezado principal con bodegas
        html += '<tr style="background-color: #000000; color: white; font-weight: bold; height: 40px;">'
        
        # Columnas de información
        info_headers = ['Código', 'Codigo_SAP', 'Segmento', 'Silueta', 'Colección', 'Descripción', 'Talla']
        widths = [80, 80, 70, 70, 90, 150, 60]
        
        for header, width in zip(info_headers, widths):
            html += f'<td style="border: 1px solid #fff; padding: 8px; width: {width}px; vertical-align: middle;">{header}</td>'
        
        # Columnas de bodegas (2 columnas por bodega: Real y Óptimo)
        for bodega in bodegas:
            html += f'<td colspan="2" style="border: 1px solid #fff; padding: 8px; vertical-align: middle; min-width: 120px;">{bodega}</td>'
        
        html += '</tr>'
        
        # Fila 2: Sub-encabezados (Real / Óptimo)
        html += '<tr style="background-color: #333333; color: white; font-weight: bold; height: 30px;">'
        
        # Espacios vacíos para columnas de información
        for i in range(7):
            html += '<td style="border: 1px solid #fff; padding: 4px;"></td>'
        
        # Sub-encabezados para cada bodega
        for _ in bodegas:
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #28a745; font-size: 8px;">Real</td>'
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #007bff; font-size: 8px;">Óptimo</td>'
        
        html += '</tr>'
        
        # Filas de datos
        for idx, row in df_display.iterrows():
            # Verificar si es la fila TOTAL
            es_total = (str(row.iloc[0]) == 'TOTAL')
            
            if es_total:
                html += '<tr style="background-color: #f8f9fa; font-weight: bold; border-top: 3px solid #000;">'
            else:
                html += '<tr style="background-color: white;">'
            
            # Columnas de información
            for i in range(7):
                valor = str(row.iloc[i]) if i < len(row) else ""
                style = "border: 1px solid #ddd; padding: 4px; text-align: left;" if i in [4, 5] else "border: 1px solid #ddd; padding: 4px;"
                html += f'<td style="{style}">{valor}</td>'
            
            # Columnas de bodegas (Real y Óptimo)
            for bodega in bodegas:
                col_real = f'Real {bodega}'
                col_optimo = f'Óptimo {bodega}'
                
                # Valores reales y óptimos
                valor_real = row[col_real] if col_real in row else 0
                valor_optimo = row[col_optimo] if col_optimo in row else 0
                
                try:
                    valor_real_num = float(str(valor_real).replace(',', '')) if valor_real != 'TOTAL' else 0
                    valor_optimo_num = float(str(valor_optimo).replace(',', '')) if valor_optimo != 'TOTAL' else 0
                except:
                    valor_real_num = 0
                    valor_optimo_num = 0
                
                # Color del semáforo solo para la columna Real
                if not es_total:
                    color_fondo = calcular_color_semaforo_mvp(valor_real_num, valor_optimo_num)
                else:
                    color_fondo = "white"
                
                # Formatear valores
                valor_real_display = f"{int(valor_real_num):,}" if valor_real_num > 0 else "0"
                valor_optimo_display = f"{int(valor_optimo_num):,}" if valor_optimo_num > 0 else "0"
                
                # Celda Real (con semáforo)
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: {color_fondo}; font-weight: bold;">{valor_real_display}</td>'
                
                # Celda Óptimo (sin semáforo)
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: #f8f9fa; color: #007bff;">{valor_optimo_display}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        
        return html
    
    # Leyenda del semáforo (antes de la tabla)
    st.markdown("""
    <div style="margin-bottom: 20px; padding: 15px; background-color: #f8f9fa; border-radius: 8px;">
        <h4 style="margin-bottom: 10px; color: #333;">🚦 Leyenda del Semáforo</h4>
        <div style="display: flex; gap: 20px; align-items: center;">
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #d4edda; border: 1px solid #c3e6cb; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Verde: Stock real >= Stock óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Amarillo: Stock real entre 80%-99% del óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #f8d7da; border: 1px solid #f1b0b7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Rojo: Stock real < 80% del óptimo</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Mostrar tabla
    tabla_html = crear_tabla_html_mvp_elsalvador(tabla_mvp)
    st.markdown(tabla_html, unsafe_allow_html=True)
    
    # Botón de exportación a Excel con colores
    st.markdown("---")
    if st.button("📊 Exportar Tabla MVP a Excel", type="primary", key=f"export_mvp_excel_{key_suffix}"):
        excel_data = exportar_mvp_excel_con_colores(tabla_mvp, columnas_real, columnas_optimo, "ElSalvador")
        
        if excel_data:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            st.download_button(
                label="📥 Descargar",
                data=excel_data,
                file_name=f"MVP_ElSalvador_Semaforo_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_mvp_excel_{key_suffix}",
                type="primary",
                use_container_width=True
            )

def mostrar_stock_mvps_panama(df_stock: pd.DataFrame, key_suffix: str = ""):
    """Muestra la tabla de stock de códigos MVP para Panamá"""
    if df_stock is None or df_stock.empty:
        st.warning("No se pudo cargar el archivo de stock para mostrar MVP")
        return
    
    # Crear sección
    professional_design.create_section_header(
        "Stock de MVPS - Panamá", 
        "Stock actual vs Stock óptimo nuevo de códigos MVP específicos",
        "🏆"
    )
    
    # Procesar datos
    tabla_mvp = procesar_stock_mvps_panama(df_stock)
    
    if tabla_mvp.empty:
        st.warning("No se encontraron datos de códigos MVP en el stock de Panamá")
        return
    
    # Obtener columnas Real y Óptimo
    columnas_real = [col for col in tabla_mvp.columns if col.startswith('Real ')]
    columnas_optimo = [col for col in tabla_mvp.columns if col.startswith('Óptimo ')]
    
    
    # Mostrar métricas resumen
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        # Contar códigos únicos excluyendo la fila TOTAL
        codigos_unicos = tabla_mvp.index.get_level_values(0).unique()
        total_productos = len([codigo for codigo in codigos_unicos if codigo != 'TOTAL'])
        st.metric("Total Productos MVP", f"{total_productos:,}")
    
    with col2:
        total_stock_real = tabla_mvp[columnas_real].loc[tabla_mvp.index[-1]].sum()
        st.metric("Total Stock Real", f"{int(total_stock_real):,}")
    
    with col3:
        total_stock_optimo = tabla_mvp[columnas_optimo].loc[tabla_mvp.index[-1]].sum()
        st.metric("Total Stock Óptimo", f"{int(total_stock_optimo):,}")
    
    with col4:
        if total_stock_optimo > 0:
            cumplimiento = (total_stock_real / total_stock_optimo) * 100
            st.metric("Cumplimiento de unidades totales", f"{cumplimiento:.1f}%")
    
    # Nueva fila de métricas - Contar celdas de semáforo
    contadores_semaforo = contar_celdas_semaforo_mvp(tabla_mvp, columnas_real, columnas_optimo)
    print(f"DEBUG: Contadores semáforo - Verde: {contadores_semaforo['verde']}, Amarillo: {contadores_semaforo['amarillo']}, Rojo: {contadores_semaforo['rojo']}")
    
    col5, col6, col7, col8 = st.columns(4)
    
    with col5:
        # Columna vacía para mantener simetría
        st.metric("", "")
    
    with col6:
        celdas_verde_amarillo = contadores_semaforo['verde'] + contadores_semaforo['amarillo']
        st.metric("Cantidad de celdas en verde y amarillo", f"{celdas_verde_amarillo:,}")
    
    with col7:
        celdas_rojo = contadores_semaforo['rojo']
        st.metric("Cantidad de celdas en rojo", f"{celdas_rojo:,}")
    
    with col8:
        # Calcular % de Cumplimiento con nueva fórmula
        total_celdas = celdas_verde_amarillo + celdas_rojo
        if total_celdas > 0:
            cumplimiento_semaforo = (celdas_verde_amarillo / total_celdas) * 100
            st.metric("% de Cumplimiento", f"{cumplimiento_semaforo:.1f}%")
        else:
            st.metric("% de Cumplimiento", "0.0%")
    
    # Crear tabla HTML simplificada
    def crear_tabla_html_mvp_panama(df):
        """Crea tabla HTML con formato profesional para MVP Panamá - MISMO FORMATO QUE GUATEMALA"""
        # Resetear índice para mostrar las columnas de información
        df_display = df.reset_index()
        
        # Crear HTML de la tabla
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 9px; width: 100%; margin-top: 20px;">'
        
        # Obtener bodegas únicas
        bodegas = []
        for col in df.columns:
            if col.startswith('Real '):
                bodega = col.replace('Real ', '')
                bodegas.append(bodega)
        
        # Fila 1: Encabezado principal con bodegas
        html += '<tr style="background-color: #000000; color: white; font-weight: bold; height: 40px;">'
        
        # Columnas de información
        info_headers = ['Código', 'Codigo_SAP', 'Segmento', 'Silueta', 'Colección', 'Descripción', 'Talla']
        widths = [80, 80, 70, 70, 90, 150, 60]
        
        for header, width in zip(info_headers, widths):
            html += f'<td style="border: 1px solid #fff; padding: 8px; width: {width}px; vertical-align: middle;">{header}</td>'
        
        # Columnas de bodegas (2 columnas por bodega: Real y Óptimo)
        for bodega in bodegas:
            html += f'<td colspan="2" style="border: 1px solid #fff; padding: 8px; vertical-align: middle; min-width: 120px;">{bodega}</td>'
        
        html += '</tr>'
        
        # Fila 2: Sub-encabezados (Real / Óptimo)
        html += '<tr style="background-color: #333333; color: white; font-weight: bold; height: 30px;">'
        
        # Espacios vacíos para columnas de información
        for i in range(7):
            html += '<td style="border: 1px solid #fff; padding: 4px;"></td>'
        
        # Sub-encabezados para cada bodega
        for _ in bodegas:
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #28a745; font-size: 8px;">Real</td>'
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #007bff; font-size: 8px;">Óptimo</td>'
        
        html += '</tr>'
        
        # Filas de datos
        for idx, row in df_display.iterrows():
            # Verificar si es la fila TOTAL
            es_total = (str(row.iloc[0]) == 'TOTAL')
            
            if es_total:
                html += '<tr style="background-color: #f8f9fa; font-weight: bold; border-top: 3px solid #000;">'
            else:
                html += '<tr style="background-color: white;">'
            
            # Columnas de información
            for i in range(7):
                valor = str(row.iloc[i]) if i < len(row) else ""
                style = "border: 1px solid #ddd; padding: 4px; text-align: left;" if i in [4, 5] else "border: 1px solid #ddd; padding: 4px;"
                html += f'<td style="{style}">{valor}</td>'
            
            # Columnas de bodegas (Real y Óptimo)
            for bodega in bodegas:
                col_real = f'Real {bodega}'
                col_optimo = f'Óptimo {bodega}'
                
                # Valores reales y óptimos
                valor_real = row[col_real] if col_real in row else 0
                valor_optimo = row[col_optimo] if col_optimo in row else 0
                
                try:
                    valor_real_num = float(str(valor_real).replace(',', '')) if valor_real != 'TOTAL' else 0
                    valor_optimo_num = float(str(valor_optimo).replace(',', '')) if valor_optimo != 'TOTAL' else 0
                except:
                    valor_real_num = 0
                    valor_optimo_num = 0
                
                # Color del semáforo solo para la columna Real
                if not es_total:
                    color_fondo = calcular_color_semaforo_mvp(valor_real_num, valor_optimo_num)
                else:
                    color_fondo = "#f8f9fa"
                
                # Celda Real con color de semáforo
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: {color_fondo}; font-weight: bold;">{valor_real}</td>'
                
                # Celda Óptimo con color azul claro y texto azul
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: #f8f9fa; color: #007bff; font-weight: bold;">{valor_optimo}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        return html
    
    # Leyenda del semáforo (antes de la tabla)
    st.markdown("""
    <div style="margin-bottom: 20px; padding: 15px; background-color: #f8f9fa; border-radius: 8px;">
        <h4 style="margin-bottom: 10px; color: #333;">🚦 Leyenda del Semáforo</h4>
        <div style="display: flex; gap: 20px; align-items: center;">
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #d4edda; border: 1px solid #c3e6cb; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Verde: Stock real >= Stock óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Amarillo: Stock real entre 80%-99% del óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #f8d7da; border: 1px solid #f1b0b7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Rojo: Stock real < 80% del óptimo</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Mostrar tabla
    st.subheader("📊 Tabla de Stock MVP - Panamá")
    tabla_html = crear_tabla_html_mvp_panama(tabla_mvp)
    st.markdown(tabla_html, unsafe_allow_html=True)
    
    # Botón de exportación a Excel con colores (mismo formato que Guatemala)
    st.markdown("---")
    if st.button("📊 Exportar Tabla MVP a Excel", type="primary", key=f"export_mvp_excel_{key_suffix}"):
        excel_data = exportar_mvp_excel_con_colores(tabla_mvp, columnas_real, columnas_optimo, "Panama")
        
        if excel_data:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            st.download_button(
                label="📥 Descargar",
                data=excel_data,
                file_name=f"MVP_Panama_Semaforo_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_mvp_excel_{key_suffix}",
                type="primary",
                use_container_width=True
            )

def mostrar_stock_mvps_puerto_rico(df_stock: pd.DataFrame, key_suffix: str = ""):
    """Muestra la tabla de stock de códigos MVP para Puerto Rico con mismo formato que Guatemala"""
    if df_stock is None or df_stock.empty:
        st.warning("No se pudo cargar el archivo de stock para mostrar MVP")
        return
    
    # Crear sección
    professional_design.create_section_header(
        "Stock de MVPS - Puerto Rico", 
        "Stock actual vs Stock óptimo nuevo de códigos MVP específicos",
        "🏆"
    )
    
    # Procesar datos
    tabla_mvp = procesar_stock_mvps_puerto_rico(df_stock)
    
    if tabla_mvp.empty:
        st.warning("No se encontraron datos de códigos MVP en el stock de Puerto Rico")
        return
    
    # Obtener columnas Real y Óptimo
    columnas_real = [col for col in tabla_mvp.columns if col.startswith('Real ')]
    columnas_optimo = [col for col in tabla_mvp.columns if col.startswith('Óptimo ')]
    
    
    # Mostrar métricas resumen
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        # Contar códigos únicos excluyendo la fila TOTAL
        codigos_unicos = tabla_mvp.index.get_level_values(0).unique()
        total_productos = len([codigo for codigo in codigos_unicos if codigo != 'TOTAL'])
        st.metric("Total Productos MVP", f"{total_productos:,}")
    
    with col2:
        total_stock_real = tabla_mvp[columnas_real].loc[('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')].sum()
        st.metric("Total Stock Real", f"{int(total_stock_real):,}")
    
    with col3:
        total_stock_optimo = tabla_mvp[columnas_optimo].loc[('TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL', 'TOTAL')].sum()
        st.metric("Total Stock Óptimo", f"{int(total_stock_optimo):,}")
    
    with col4:
        if total_stock_optimo > 0:
            cumplimiento = (total_stock_real / total_stock_optimo) * 100
            st.metric("Cumplimiento de unidades totales", f"{cumplimiento:.1f}%")
    
    # Nueva fila de métricas - Contar celdas de semáforo
    contadores_semaforo = contar_celdas_semaforo_mvp(tabla_mvp, columnas_real, columnas_optimo)
    print(f"DEBUG: Contadores semáforo - Verde: {contadores_semaforo['verde']}, Amarillo: {contadores_semaforo['amarillo']}, Rojo: {contadores_semaforo['rojo']}")
    
    col5, col6, col7, col8 = st.columns(4)
    
    with col5:
        # Columna vacía para mantener simetría
        st.metric("", "")
    
    with col6:
        celdas_verde_amarillo = contadores_semaforo['verde'] + contadores_semaforo['amarillo']
        st.metric("Cantidad de celdas en verde y amarillo", f"{celdas_verde_amarillo:,}")
    
    with col7:
        celdas_rojo = contadores_semaforo['rojo']
        st.metric("Cantidad de celdas en rojo", f"{celdas_rojo:,}")
    
    with col8:
        # Calcular % de Cumplimiento con nueva fórmula
        total_celdas = celdas_verde_amarillo + celdas_rojo
        if total_celdas > 0:
            cumplimiento_semaforo = (celdas_verde_amarillo / total_celdas) * 100
            st.metric("% de Cumplimiento", f"{cumplimiento_semaforo:.1f}%")
        else:
            st.metric("% de Cumplimiento", "0.0%")
    
    # Crear tabla HTML con formato profesional
    def crear_tabla_html_mvp(df):
        """Crea tabla HTML con formato profesional para MVP"""
        # Resetear índice para mostrar las columnas de información
        df_display = df.reset_index()
        
        # Crear HTML de la tabla
        html = '<table style="border-collapse: collapse; text-align: center; font-size: 9px; width: 100%; margin-top: 20px;">'
        
        # Obtener bodegas únicas
        bodegas = []
        for col in columnas_real:
            bodega = col.replace('Real ', '')
            bodegas.append(bodega)
        
        # Fila 1: Encabezado principal con bodegas
        html += '<tr style="background-color: #000000; color: white; font-weight: bold; height: 40px;">'
        
        # Columnas de información
        info_headers = ['Código', 'Codigo_SAP', 'Segmento', 'Silueta', 'Colección', 'Descripción', 'Talla']
        widths = [80, 80, 70, 70, 90, 150, 60]
        
        for header, width in zip(info_headers, widths):
            html += f'<td style="border: 1px solid #fff; padding: 8px; width: {width}px; vertical-align: middle;">{header}</td>'
        
        # Columnas de bodegas (2 columnas por bodega: Real y Óptimo)
        for bodega in bodegas:
            html += f'<td colspan="2" style="border: 1px solid #fff; padding: 8px; vertical-align: middle; min-width: 120px;">{bodega}</td>'
        
        html += '</tr>'
        
        # Fila 2: Sub-encabezados (Real / Óptimo)
        html += '<tr style="background-color: #333333; color: white; font-weight: bold; height: 30px;">'
        
        # Espacios vacíos para columnas de información
        for i in range(7):
            html += '<td style="border: 1px solid #fff; padding: 4px;"></td>'
        
        # Sub-encabezados para cada bodega
        for _ in bodegas:
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #28a745; font-size: 8px;">Real</td>'
            html += '<td style="border: 1px solid #fff; padding: 4px; background-color: #007bff; font-size: 8px;">Óptimo</td>'
        
        html += '</tr>'
        
        # Filas de datos
        for idx, row in df_display.iterrows():
            # Verificar si es la fila TOTAL
            es_total = (str(row.iloc[0]) == 'TOTAL')
            
            if es_total:
                html += '<tr style="background-color: #f8f9fa; font-weight: bold; border-top: 3px solid #000;">'
            else:
                html += '<tr style="background-color: white;">'
            
            # Columnas de información
            for i in range(7):
                valor = str(row.iloc[i]) if i < len(row) else ""
                style = "border: 1px solid #ddd; padding: 4px; text-align: left;" if i in [4, 5] else "border: 1px solid #ddd; padding: 4px;"
                html += f'<td style="{style}">{valor}</td>'
            
            # Columnas de bodegas (Real y Óptimo)
            for bodega in bodegas:
                col_real = f'Real {bodega}'
                col_optimo = f'Óptimo {bodega}'
                
                # Valores reales y óptimos
                valor_real = row[col_real] if col_real in row else 0
                valor_optimo = row[col_optimo] if col_optimo in row else 0
                
                try:
                    valor_real_num = float(str(valor_real).replace(',', '')) if valor_real != 'TOTAL' else 0
                    valor_optimo_num = float(str(valor_optimo).replace(',', '')) if valor_optimo != 'TOTAL' else 0
                except:
                    valor_real_num = 0
                    valor_optimo_num = 0
                
                # Color del semáforo solo para la columna Real
                if not es_total:
                    color_fondo = calcular_color_semaforo_mvp(valor_real_num, valor_optimo_num)
                else:
                    color_fondo = "white"
                
                # Formatear valores
                valor_real_display = f"{int(valor_real_num):,}" if valor_real_num > 0 else "0"
                valor_optimo_display = f"{int(valor_optimo_num):,}" if valor_optimo_num > 0 else "0"
                
                # Celda Real (con semáforo)
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: {color_fondo}; font-weight: bold;">{valor_real_display}</td>'
                
                # Celda Óptimo (sin semáforo)
                html += f'<td style="border: 1px solid #ddd; padding: 4px; background-color: #f8f9fa; color: #007bff;">{valor_optimo_display}</td>'
            
            html += '</tr>'
        
        html += '</table>'
        
        return html
    
    # Leyenda del semáforo (antes de la tabla)
    st.markdown("""
    <div style="margin-bottom: 20px; padding: 15px; background-color: #f8f9fa; border-radius: 8px;">
        <h4 style="margin-bottom: 10px; color: #333;">🚦 Leyenda del Semáforo</h4>
        <div style="display: flex; gap: 20px; align-items: center;">
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #d4edda; border: 1px solid #c3e6cb; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Verde: Stock real >= Stock óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #fff3cd; border: 1px solid #ffeaa7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Amarillo: Stock real entre 80%-99% del óptimo</span>
            </div>
            <div style="display: flex; align-items: center; gap: 8px;">
                <div style="width: 20px; height: 20px; background-color: #f8d7da; border: 1px solid #f1b0b7; border-radius: 4px;"></div>
                <span style="font-size: 14px;">Rojo: Stock real < 80% del óptimo</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Mostrar tabla
    tabla_html = crear_tabla_html_mvp(tabla_mvp)
    st.markdown(tabla_html, unsafe_allow_html=True)
    
    # Botón de exportación a Excel con colores
    st.markdown("---")
    if st.button("📊 Exportar Tabla MVP a Excel", type="primary", key=f"export_mvp_excel_{key_suffix}"):
        excel_data = exportar_mvp_excel_con_colores(tabla_mvp, columnas_real, columnas_optimo, "Puerto Rico")
        
        if excel_data:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            st.download_button(
                label="📥 Descargar",
                data=excel_data,
                file_name=f"MVP_Puerto_Rico_Semaforo_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_mvp_excel_{key_suffix}",
                type="primary",
                use_container_width=True
            )


def main():
    """Función principal"""
    logger.info("Iniciando aplicación New Era Analytics Dashboard")
    
    # Inyectar CSS personalizado
    professional_design.inject_custom_css()
    
    # Crear header principal con hora en tiempo real
    professional_design.create_main_header()
    
    # Descripción con diseño mejorado
    professional_design.create_leagues_section()
    
    # Crear pestañas para cada país con iconos mejorados + pestaña temporal MVPs
    tab_guatemala, tab_el_salvador, tab_honduras, tab_costa_rica, tab_panama, tab_mvps_temporal = st.tabs([
        "Guatemala", 
        "El Salvador", 
        "Honduras", 
        "Costa Rica",
        "Panama",
        "MVPs (Temporal)"
    ])
    
    # PESTAÑA GUATEMALA
    with tab_guatemala:
        professional_design.create_section_header(
            "Análisis de Stock - Guatemala", 
            "Gestión de inventario para 24 tiendas en territorio guatemalteco",
            "GT"
        )
        
        # Crear dos columnas para los espacios de carga
        col_guatemala, col_ventas = st.columns(2)
        
        with col_guatemala:
            archivo_guatemala = data_loader.cargar_archivo("📁 Subir archivo GUATEMALA.csv", "GUATEMALA")
            
        with col_ventas:
            archivo_ventas_guatemala = data_loader.cargar_archivo_ventas("📁 Subir archivo VENTAS_GUATEMALA.csv", "Guatemala_ventas", "GUATEMALA")
        
        if archivo_guatemala is not None:
            # CASO 1: Archivo de stock cargado (lógica original)
            # Guardar nombre del archivo en session state para la exportación
            if hasattr(archivo_guatemala, 'name'):
                st.session_state.archivo_guatemala_name = archivo_guatemala.name
            
            # Crear hash del DataFrame para cache
            df_hash = archivo_guatemala.to_dict('records')
            
            # Procesar datos Guatemala (con cache)
            selected_league = st.session_state.get('selected_league', None)
            # Convertir "Todas" a None para mostrar todas las ligas
            if selected_league == "Todas":
                selected_league = None
            df_ventas_hash = archivo_ventas_guatemala.to_dict('records') if archivo_ventas_guatemala is not None else None
            
            # Limpiar cache si hay cambios
            if 'cache_cleared' not in st.session_state:
                st.cache_data.clear()
                st.session_state.cache_cleared = True
                
            tabla_guatemala = data_processor.procesar_datos_consolidados(df_hash, "Guatemala", selected_league, df_ventas_hash)
            
            # Mostrar resultados Guatemala
            mostrar_tabla_consolidada(tabla_guatemala, "Guatemala")
            
            # Nueva sección: Stock de MVPs para Guatemala
            st.markdown("---")
            mostrar_stock_mvps_guatemala(archivo_guatemala, "_main")
            
        elif archivo_ventas_guatemala is not None:
            # CASO 2: Solo archivo de ventas cargado (NUEVA FUNCIONALIDAD)
            st.info("📊 **Modo Solo-Ventas activado:** Mostrando análisis basado únicamente en datos de cantidad vendida")
            
            # Crear hash del DataFrame de ventas para cache
            df_ventas_hash = archivo_ventas_guatemala.to_dict('records')
            
            # Procesar datos solo-ventas Guatemala
            selected_league = st.session_state.get('selected_league', None)
            if selected_league == "Todas":
                selected_league = None
                
            # Limpiar cache si hay cambios
            if 'cache_cleared_ventas' not in st.session_state:
                st.cache_data.clear()
                st.session_state.cache_cleared_ventas = True
            
            tabla_solo_ventas = data_processor.procesar_solo_ventas_guatemala(df_ventas_hash, selected_league)
            
            if tabla_solo_ventas is not None:
                # Mostrar tabla consolidada adaptada (sin capacidades ni % cumplimiento)
                mostrar_tabla_solo_ventas_guatemala(tabla_solo_ventas)
            else:
                st.error("❌ No se pudieron procesar los datos de ventas. Verifica el formato del archivo.")
            
        # Mostrar mensajes de bienvenida en columnas cuando no hay archivos
        if archivo_guatemala is None or archivo_ventas_guatemala is None:
            col_msg_guatemala, col_msg_ventas = st.columns(2)
            
            with col_msg_guatemala:
                if archivo_guatemala is None:
                    st.markdown("""
                    <div class="country-card country-card-gt">
                        <div class="country-flag">🇬🇹</div>
                        <h3 class="country-title" style="color: #000000; font-size: 1.75rem; font-weight: 700; margin-bottom: 1rem;">Guatemala - Sistema de <span style="color: #1e3a8a;">Stock</span></h3>
                        <p class="country-description" style="color: #64748b; font-size: 1rem; font-weight: 500; line-height: 1.6; margin-bottom: 0; background: rgba(191, 219, 254, 0.1); padding: 1rem; border-radius: 12px; border: 1px solid rgba(191, 219, 254, 0.3);">
                            Selecciona tu archivo GUATEMALA.csv para comenzar el análisis completo del inventario<br>
                            <strong style="color: #1e3a8a;">24 tiendas</strong> en operación
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                    
            with col_msg_ventas:
                if archivo_ventas_guatemala is None:
                    st.markdown("""
                    <div class="country-card country-card-gt">
                        <div class="country-flag">🇬🇹</div>
                        <h3 class="country-title" style="color: #000000; font-size: 1.75rem; font-weight: 700; margin-bottom: 1rem;">Guatemala - Sistema de <span style="color: #22c55e;">Ventas</span></h3>
                        <p class="country-description" style="color: #64748b; font-size: 1rem; font-weight: 500; line-height: 1.6; margin-bottom: 0; background: rgba(134, 239, 172, 0.1); padding: 1rem; border-radius: 12px; border: 1px solid rgba(134, 239, 172, 0.3);">
                            Selecciona tu archivo VENTAS_GUATEMALA.csv para comenzar el análisis completo de ventas<br>
                            <strong style="color: #22c55e;">24 tiendas</strong> en operación
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.success("✅ Archivo VENTAS_GUATEMALA.csv cargado correctamente")

    # CONTINUACIÓN DE MAIN() - Las pestañas restantes
    
    # Las pestañas restantes continúan aquí:
    
    # PESTAÑA PANAMA
    with tab_panama:
        professional_design.create_section_header(
            "Análisis de Stock - Panamá", 
            "Gestión de inventario para 6 tiendas estratégicas en Panamá",
            "PA"
        )
        
        # Crear dos columnas para los espacios de carga (igual que Guatemala)
        col_panama, col_ventas_pa = st.columns(2)
        
        with col_panama:
            archivo_panama = data_loader.cargar_archivo("📁 Subir archivo PANAMA.csv", "PANAMA")
            
        with col_ventas_pa:
            archivo_ventas_panama = data_loader.cargar_archivo_ventas("📁 Subir archivo VENTAS_PANAMA.csv", "Panama_ventas", "PANAMA")
        
        if archivo_panama is not None:
            # Guardar nombre del archivo en session state para la exportación
            if hasattr(archivo_panama, 'name'):
                st.session_state.archivo_panama_name = archivo_panama.name
            
            # Crear hash del DataFrame para cache
            df_hash = archivo_panama.to_dict('records')
            
            # Procesar datos PANAMA (con cache)
            selected_league = st.session_state.get('selected_league', None)
            # Convertir "Todas" a None para mostrar todas las ligas
            if selected_league == "Todas":
                selected_league = None
            df_ventas_hash = archivo_ventas_panama.to_dict('records') if archivo_ventas_panama is not None else None
            
            # Limpiar cache si hay cambios
            if 'cache_cleared' not in st.session_state:
                st.cache_data.clear()
                st.session_state.cache_cleared = True
            
            tabla_panama = data_processor.procesar_datos_consolidados(df_hash, "PANAMA", selected_league, df_ventas_hash)
            
            # Mostrar resultados PANAMA
            mostrar_tabla_consolidada(tabla_panama, "PANAMA")
            
        elif archivo_ventas_panama is not None:
            # CASO 2: Solo archivo de ventas cargado (NUEVA FUNCIONALIDAD)
            st.info("📊 **Modo Solo-Ventas activado:** Mostrando análisis basado únicamente en datos de cantidad vendida")
            
            # Crear hash del DataFrame de ventas para cache
            df_ventas_hash = archivo_ventas_panama.to_dict('records')
            
            # Procesar datos solo-ventas Panama
            selected_league = st.session_state.get('selected_league', None)
            if selected_league == "Todas":
                selected_league = None
                
            # Limpiar cache si hay cambios
            if 'cache_cleared_ventas_pa' not in st.session_state:
                st.cache_data.clear()
                st.session_state.cache_cleared_ventas_pa = True
            
            tabla_solo_ventas = data_processor.procesar_solo_ventas_panama(df_ventas_hash, selected_league)
            
            if tabla_solo_ventas is not None:
                # Mostrar tabla consolidada adaptada (sin capacidades ni % cumplimiento)
                mostrar_tabla_solo_ventas_panama(tabla_solo_ventas)
            else:
                st.error("❌ No se pudieron procesar los datos de ventas. Verifica el formato del archivo.")
        
        # Mostrar mensajes de bienvenida en columnas cuando no hay archivos
        if archivo_panama is None or archivo_ventas_panama is None:
            col_msg_panama, col_msg_ventas_pa = st.columns(2)
            
            with col_msg_panama:
                if archivo_panama is None:
                    st.markdown("""
                    <div class="country-card country-card-pa">
                        <div class="country-flag">🇵🇦</div>
                        <h3 class="country-title" style="color: #000000; font-size: 1.75rem; font-weight: 700; margin-bottom: 1rem;">Panamá - Sistema de <span style="color: #dc2626;">Stock</span></h3>
                        <p class="country-description" style="color: #64748b; font-size: 1rem; font-weight: 500; line-height: 1.6; margin-bottom: 0; background: rgba(254, 202, 202, 0.1); padding: 1rem; border-radius: 12px; border: 1px solid rgba(254, 202, 202, 0.3);">
                            Selecciona tu archivo PANAMA.csv para comenzar el análisis completo del inventario<br>
                            <strong style="color: #dc2626;">6 tiendas</strong> en operación
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
            
            with col_msg_ventas_pa:
                if archivo_ventas_panama is None:
                    st.markdown("""
                    <div class="country-card country-card-pa">
                        <div class="country-flag">🇵🇦</div>
                        <h3 class="country-title" style="color: #000000; font-size: 1.75rem; font-weight: 700; margin-bottom: 1rem;">Panamá - Sistema de <span style="color: #22c55e;">Ventas</span></h3>
                        <p class="country-description" style="color: #64748b; font-size: 1rem; font-weight: 500; line-height: 1.6; margin-bottom: 0; background: rgba(134, 239, 172, 0.1); padding: 1rem; border-radius: 12px; border: 1px solid rgba(134, 239, 172, 0.3);">
                            Selecciona tu archivo VENTAS_PANAMA.csv para análisis de ventas en USD<br>
                            <strong style="color: #22c55e;">Análisis de ingresos</strong> por tienda
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.success("✅ Archivo VENTAS_PANAMA.csv cargado correctamente")

    # PESTAÑA HONDURAS
    with tab_honduras:
        professional_design.create_section_header(
            "Análisis de Stock - Honduras", 
            "Gestión de inventario para 5 tiendas en Honduras",
            "HN"
        )
        
        # Crear dos columnas para los espacios de carga (igual que Guatemala, El Salvador y Costa Rica)
        col_honduras, col_ventas_hn = st.columns(2)
        
        with col_honduras:
            archivo_honduras = data_loader.cargar_archivo("📁 Subir archivo HONDURAS.csv", "HONDURAS")
            
        with col_ventas_hn:
            archivo_ventas_honduras = data_loader.cargar_archivo_ventas("📁 Subir archivo VENTAS_HONDURAS.csv", "Honduras_ventas", "HONDURAS")
        
        if archivo_honduras is not None:
            # Guardar nombre del archivo en session state para la exportación
            if hasattr(archivo_honduras, 'name'):
                st.session_state.archivo_honduras_name = archivo_honduras.name
            
            # Crear hash del DataFrame para cache
            df_hash = archivo_honduras.to_dict('records')
            
            # Procesar datos Honduras (con cache)
            selected_league = st.session_state.get('selected_league', None)
            # Convertir "Todas" a None para mostrar todas las ligas
            if selected_league == "Todas":
                selected_league = None
            df_ventas_hash = archivo_ventas_honduras.to_dict('records') if archivo_ventas_honduras is not None else None
            
            # Limpiar cache si hay cambios
            if 'cache_cleared_hn' not in st.session_state:
                st.cache_data.clear()
                st.session_state.cache_cleared_hn = True
                
            tabla_honduras = data_processor.procesar_datos_consolidados(df_hash, "Honduras", selected_league, df_ventas_hash)
            
            # Mostrar resultados Honduras
            mostrar_tabla_consolidada(tabla_honduras, "Honduras")
            
        elif archivo_ventas_honduras is not None:
            # CASO 2: Solo archivo de ventas cargado para Honduras (NUEVA FUNCIONALIDAD)
            st.info("📊 **Modo Solo-Ventas activado:** Mostrando análisis basado únicamente en datos de cantidad vendida")
            
            # Crear hash del DataFrame de ventas para cache
            df_ventas_hash = archivo_ventas_honduras.to_dict('records')
            
            # Procesar datos solo-ventas Honduras
            selected_league = st.session_state.get('selected_league', None)
            if selected_league == "Todas":
                selected_league = None
                
            # Limpiar cache si hay cambios
            if 'cache_cleared_ventas_hn' not in st.session_state:
                st.cache_data.clear()
                st.session_state.cache_cleared_ventas_hn = True
            
            tabla_solo_ventas = data_processor.procesar_solo_ventas_honduras(df_ventas_hash, selected_league)
            
            if tabla_solo_ventas is not None:
                # Mostrar tabla consolidada adaptada para Honduras (sin capacidades ni % cumplimiento)
                mostrar_tabla_solo_ventas_honduras(tabla_solo_ventas)
            else:
                st.error("❌ No se pudieron procesar los datos de ventas. Verifica el formato del archivo.")
            
        # Mostrar mensajes de bienvenida en columnas cuando no hay archivos
        if archivo_honduras is None or archivo_ventas_honduras is None:
            col_msg_honduras, col_msg_ventas_hn = st.columns(2)
            
            with col_msg_honduras:
                if archivo_honduras is None:
                    st.markdown("""
                    <div class="country-card country-card-hn">
                        <div class="country-flag">🇭🇳</div>
                        <h3 class="country-title" style="color: #000000; font-size: 1.75rem; font-weight: 700; margin-bottom: 1rem;">Honduras - Sistema de <span style="color: #1e3a8a;">Stock</span></h3>
                        <p class="country-description" style="color: #64748b; font-size: 1rem; font-weight: 500; line-height: 1.6; margin-bottom: 0; background: rgba(191, 219, 254, 0.1); padding: 1rem; border-radius: 12px; border: 1px solid rgba(191, 219, 254, 0.3);">
                            Selecciona tu archivo HONDURAS.csv para comenzar el análisis completo del inventario<br>
                            <strong style="color: #1e3a8a;">5 tiendas</strong> en operación
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
            
            with col_msg_ventas_hn:
                if archivo_ventas_honduras is None:
                    st.markdown("""
                    <div class="country-card country-card-hn">
                        <div class="country-flag">🇭🇳</div>
                        <h3 class="country-title" style="color: #000000; font-size: 1.75rem; font-weight: 700; margin-bottom: 1rem;">Honduras - Sistema de <span style="color: #22c55e;">Ventas</span></h3>
                        <p class="country-description" style="color: #64748b; font-size: 1rem; font-weight: 500; line-height: 1.6; margin-bottom: 0; background: rgba(134, 239, 172, 0.1); padding: 1rem; border-radius: 12px; border: 1px solid rgba(134, 239, 172, 0.3);">
                            Selecciona tu archivo VENTAS_HONDURAS.csv para análisis de ventas en USD<br>
                            <strong style="color: #22c55e;">Análisis de ingresos</strong> por tienda
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
    
    # PESTAÑA EL SALVADOR
    with tab_el_salvador:
        professional_design.create_section_header(
            "Análisis de Stock - El Salvador", 
            "Gestión de inventario para 9 tiendas en territorio salvadoreño",
            "SV"
        )
        
        # Crear dos columnas para los espacios de carga (igual que Guatemala)
        col_el_salvador, col_ventas_sv = st.columns(2)
        
        with col_el_salvador:
            archivo_el_salvador = data_loader.cargar_archivo("📁 Subir archivo EL_SALVADOR.csv", "EL_SALVADOR")
            
        with col_ventas_sv:
            archivo_ventas_el_salvador = data_loader.cargar_archivo_ventas("📁 Subir archivo VENTAS_EL_SALVADOR.csv", "El_Salvador_ventas", "EL_SALVADOR")
        
        if archivo_el_salvador is not None:
            # Guardar nombre del archivo en session state para la exportación
            if hasattr(archivo_el_salvador, 'name'):
                st.session_state.archivo_el_salvador_name = archivo_el_salvador.name
            
            # Crear hash del DataFrame para cache
            df_hash = archivo_el_salvador.to_dict('records')
            
            # Procesar datos El Salvador (con cache)
            selected_league = st.session_state.get('selected_league', None)
            # Convertir "Todas" a None para mostrar todas las ligas
            if selected_league == "Todas":
                selected_league = None
            df_ventas_hash = archivo_ventas_el_salvador.to_dict('records') if archivo_ventas_el_salvador is not None else None
            
            # Limpiar cache si hay cambios
            if 'cache_cleared_sv' not in st.session_state:
                st.cache_data.clear()
                st.session_state.cache_cleared_sv = True
                
            tabla_el_salvador = data_processor.procesar_datos_consolidados(df_hash, "El Salvador", selected_league, df_ventas_hash)
            
            # Mostrar resultados El Salvador
            mostrar_tabla_consolidada(tabla_el_salvador, "El Salvador")
            
        elif archivo_ventas_el_salvador is not None:
            # CASO 2: Solo archivo de ventas cargado para El Salvador (NUEVA FUNCIONALIDAD)
            st.info("📊 **Modo Solo-Ventas activado:** Mostrando análisis basado únicamente en datos de cantidad vendida")
            
            # Crear hash del DataFrame de ventas para cache
            df_ventas_hash = archivo_ventas_el_salvador.to_dict('records')
            
            # Procesar datos solo-ventas El Salvador
            selected_league = st.session_state.get('selected_league', None)
            if selected_league == "Todas":
                selected_league = None
                
            # Limpiar cache si hay cambios
            if 'cache_cleared_ventas_sv' not in st.session_state:
                st.cache_data.clear()
                st.session_state.cache_cleared_ventas_sv = True
            
            tabla_solo_ventas = data_processor.procesar_solo_ventas_el_salvador(df_ventas_hash, selected_league)
            
            if tabla_solo_ventas is not None:
                # Mostrar tabla consolidada adaptada para El Salvador (sin capacidades ni % cumplimiento)
                mostrar_tabla_solo_ventas_el_salvador(tabla_solo_ventas)
            else:
                st.error("❌ No se pudieron procesar los datos de ventas. Verifica el formato del archivo.")
            
        # Mostrar mensajes de bienvenida en columnas cuando no hay archivos
        if archivo_el_salvador is None or archivo_ventas_el_salvador is None:
            col_msg_el_salvador, col_msg_ventas_sv = st.columns(2)
            
            with col_msg_el_salvador:
                if archivo_el_salvador is None:
                    st.markdown("""
                    <div class="country-card country-card-sv">
                        <div class="country-flag">🇸🇻</div>
                        <h3 class="country-title" style="color: #000000; font-size: 1.75rem; font-weight: 700; margin-bottom: 1rem;">El Salvador - Sistema de <span style="color: #1e3a8a;">Stock</span></h3>
                        <p class="country-description" style="color: #64748b; font-size: 1rem; font-weight: 500; line-height: 1.6; margin-bottom: 0; background: rgba(191, 219, 254, 0.1); padding: 1rem; border-radius: 12px; border: 1px solid rgba(191, 219, 254, 0.3);">
                            Selecciona tu archivo EL_SALVADOR.csv para comenzar el análisis completo del inventario<br>
                            <strong style="color: #1e3a8a;">9 tiendas</strong> en operación
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
            
            with col_msg_ventas_sv:
                if archivo_ventas_el_salvador is None:
                    st.markdown("""
                    <div class="country-card country-card-sv">
                        <div class="country-flag">🇸🇻</div>
                        <h3 class="country-title" style="color: #000000; font-size: 1.75rem; font-weight: 700; margin-bottom: 1rem;">El Salvador - Sistema de <span style="color: #22c55e;">Ventas</span></h3>
                        <p class="country-description" style="color: #64748b; font-size: 1rem; font-weight: 500; line-height: 1.6; margin-bottom: 0; background: rgba(134, 239, 172, 0.1); padding: 1rem; border-radius: 12px; border: 1px solid rgba(134, 239, 172, 0.3);">
                            Selecciona tu archivo VENTAS_EL_SALVADOR.csv para análisis de ventas en USD<br>
                            <strong style="color: #22c55e;">Datos de ventas</strong> opcionales
                        </p>
                    </div>
                    """, unsafe_allow_html=True)

    # PESTAÑA COSTA RICA
    with tab_costa_rica:
        professional_design.create_section_header(
            "Análisis de Stock - Costa Rica", 
            "Gestión de inventario para 2 tiendas en Costa Rica",
            "CR"
        )
        
        # Crear dos columnas para los espacios de carga (igual que Guatemala y El Salvador)
        col_costa_rica, col_ventas_cr = st.columns(2)
        
        with col_costa_rica:
            archivo_costa_rica = data_loader.cargar_archivo("📁 Subir archivo COSTA_RICA.csv", "COSTA_RICA")
            
        with col_ventas_cr:
            archivo_ventas_costa_rica = data_loader.cargar_archivo_ventas("📁 Subir archivo VENTAS_COSTA_RICA.csv", "Costa_Rica_ventas", "COSTA_RICA")
        
        if archivo_costa_rica is not None:
            # Guardar nombre del archivo en session state para la exportación
            if hasattr(archivo_costa_rica, 'name'):
                st.session_state.archivo_costa_rica_name = archivo_costa_rica.name
            
            # Crear hash del DataFrame para cache
            df_hash = archivo_costa_rica.to_dict('records')
            
            # Procesar datos Costa Rica (con cache)
            selected_league = st.session_state.get('selected_league', None)
            # Convertir "Todas" a None para mostrar todas las ligas
            if selected_league == "Todas":
                selected_league = None
            df_ventas_hash = archivo_ventas_costa_rica.to_dict('records') if archivo_ventas_costa_rica is not None else None
            
            # Limpiar cache si hay cambios
            if 'cache_cleared_cr' not in st.session_state:
                st.cache_data.clear()
                st.session_state.cache_cleared_cr = True
                
            tabla_costa_rica = data_processor.procesar_datos_consolidados(df_hash, "Costa Rica", selected_league, df_ventas_hash)
            
            # Mostrar resultados Costa Rica
            mostrar_tabla_consolidada(tabla_costa_rica, "Costa Rica")
            
        elif archivo_ventas_costa_rica is not None:
            # CASO 2: Solo archivo de ventas cargado (NUEVA FUNCIONALIDAD)
            st.info("📊 **Modo Solo-Ventas activado:** Mostrando análisis basado únicamente en datos de cantidad vendida")
            
            # Crear hash del DataFrame de ventas para cache
            df_ventas_hash = archivo_ventas_costa_rica.to_dict('records')
            
            # Procesar datos solo-ventas Costa Rica
            selected_league = st.session_state.get('selected_league', None)
            if selected_league == "Todas":
                selected_league = None
                
            # Limpiar cache si hay cambios
            if 'cache_cleared_ventas_cr' not in st.session_state:
                st.cache_data.clear()
                st.session_state.cache_cleared_ventas_cr = True
            
            tabla_solo_ventas = data_processor.procesar_solo_ventas_costa_rica(df_ventas_hash, selected_league)
            
            if tabla_solo_ventas is not None:
                # Mostrar tabla consolidada adaptada (sin capacidades ni % cumplimiento)
                mostrar_tabla_solo_ventas_costa_rica(tabla_solo_ventas)
            else:
                st.error("❌ No se pudieron procesar los datos de ventas. Verifica el formato del archivo.")
            
        # Mostrar mensajes de bienvenida en columnas cuando no hay archivos
        if archivo_costa_rica is None or archivo_ventas_costa_rica is None:
            col_msg_costa_rica, col_msg_ventas_cr = st.columns(2)
            
            with col_msg_costa_rica:
                if archivo_costa_rica is None:
                    st.markdown("""
                    <div class="country-card country-card-cr">
                        <div class="country-flag">🇨🇷</div>
                        <h3 class="country-title" style="color: #000000; font-size: 1.75rem; font-weight: 700; margin-bottom: 1rem;">Costa Rica - Sistema de <span style="color: #1e3a8a;">Stock</span></h3>
                        <p class="country-description" style="color: #64748b; font-size: 1rem; font-weight: 500; line-height: 1.6; margin-bottom: 0; background: rgba(191, 219, 254, 0.1); padding: 1rem; border-radius: 12px; border: 1px solid rgba(191, 219, 254, 0.3);">
                            Selecciona tu archivo COSTA_RICA.csv para comenzar el análisis completo del inventario<br>
                            <strong style="color: #1e3a8a;">2 tiendas</strong> en operación
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
            
            with col_msg_ventas_cr:
                if archivo_ventas_costa_rica is None:
                    st.markdown("""
                    <div class="country-card country-card-cr">
                        <div class="country-flag">🇨🇷</div>
                        <h3 class="country-title" style="color: #000000; font-size: 1.75rem; font-weight: 700; margin-bottom: 1rem;">Costa Rica - Sistema de <span style="color: #22c55e;">Ventas</span></h3>
                        <p class="country-description" style="color: #64748b; font-size: 1rem; font-weight: 500; line-height: 1.6; margin-bottom: 0; background: rgba(134, 239, 172, 0.1); padding: 1rem; border-radius: 12px; border: 1px solid rgba(134, 239, 172, 0.3);">
                            Selecciona tu archivo VENTAS_COSTA_RICA.csv para análisis de ventas en USD<br>
                            <strong style="color: #22c55e;">Datos de ventas</strong> opcionales
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.success("✅ Archivo VENTAS_COSTA_RICA.csv cargado correctamente")
    
    # PESTAÑA TEMPORAL - ESPACIO ADICIONAL COMPLETO
    with tab_mvps_temporal:
        st.markdown("""
        <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                    padding: 25px; border-radius: 15px; margin: 20px 0; text-align: center;
                    box-shadow: 0 10px 30px rgba(0,0,0,0.3);">
            <h2 style="color: white; margin: 0; font-size: 2.2rem; font-weight: 700;">
                🌎 ESPACIO ADICIONAL COMPLETO 🌎
            </h2>
            <p style="color: white; margin: 10px 0 0 0; font-size: 1.1rem; opacity: 0.9;">
                📊 Dashboard completo adicional - Todos los países + MVPs
            </p>
            <p style="color: #ffeb3b; margin: 5px 0 0 0; font-size: 0.9rem; font-weight: 500;">
                ⚠️ NOTA: Misma funcionalidad que pestañas principales, en espacio dedicado
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Crear sub-pestañas para cada país en el espacio temporal
        st.markdown("### 📁 Selecciona el País para Análisis:")
        
        sub_tab_gt, sub_tab_sv, sub_tab_hn, sub_tab_cr, sub_tab_pa, sub_tab_pr = st.tabs([
            "🇬🇹 Guatemala",
            "🇸🇻 El Salvador", 
            "🇭🇳 Honduras",
            "🇨🇷 Costa Rica",
            "🇵🇦 Panamá",
            "🇵🇷 Puerto Rico"
        ])
        
        # SUB-PESTAÑA GUATEMALA TEMPORAL
        with sub_tab_gt:
            professional_design.create_section_header(
                "Stock MVPs - Guatemala", 
                "Análisis exclusivo de códigos MVP (Most Valuable Products)",
                "🏆"
            )
            
            # Solo carga de archivo de stock (sin ventas)
            archivo_guatemala_temp = data_loader.cargar_archivo("📁 Subir archivo GUATEMALA.csv", "GUATEMALA_TEMP")
            
            if archivo_guatemala_temp is not None:
                # Guardar nombre del archivo en session state
                if hasattr(archivo_guatemala_temp, 'name'):
                    st.session_state.archivo_guatemala_temp_name = archivo_guatemala_temp.name
                
                # Solo mostrar tabla de MVPs (sin tabla consolidada)
                mostrar_stock_mvps_guatemala(archivo_guatemala_temp, "_temp_guatemala")
        
        # SUB-PESTAÑA EL SALVADOR TEMPORAL
        with sub_tab_sv:
            professional_design.create_section_header(
                "Stock MVPs - El Salvador", 
                "Análisis exclusivo de códigos MVP (Most Valuable Products)",
                "🏆"
            )
            
            # Solo carga de archivo de stock (sin ventas)
            archivo_el_salvador_temp = data_loader.cargar_archivo("📁 Subir archivo EL_SALVADOR.csv", "EL_SALVADOR_TEMP")
            
            if archivo_el_salvador_temp is not None:
                if hasattr(archivo_el_salvador_temp, 'name'):
                    st.session_state.archivo_el_salvador_temp_name = archivo_el_salvador_temp.name
                
                # Solo mostrar tabla de MVPs (igual que Guatemala)
                mostrar_stock_mvps_elsalvador(archivo_el_salvador_temp, "_temp_elsalvador")
        
        # SUB-PESTAÑA HONDURAS TEMPORAL
        with sub_tab_hn:
            professional_design.create_section_header(
                "Análisis Temporal - Honduras", 
                "Análisis de códigos MVP específicos para Honduras",
                "HN"
            )
            
            # Solo carga de archivo de stock (sin ventas)
            archivo_honduras_temp = data_loader.cargar_archivo("📁 Subir archivo HONDURAS.csv", "HONDURAS_TEMP")
            
            if archivo_honduras_temp is not None:
                if hasattr(archivo_honduras_temp, 'name'):
                    st.session_state.archivo_honduras_temp_name = archivo_honduras_temp.name
                
                # Solo mostrar tabla de MVPs (igual que Guatemala y El Salvador)
                mostrar_stock_mvps_honduras(archivo_honduras_temp, "_temp_honduras")
        
        # SUB-PESTAÑA COSTA RICA TEMPORAL
        with sub_tab_cr:
            professional_design.create_section_header(
                "Análisis Temporal - Costa Rica", 
                "Análisis de códigos MVP específicos para Costa Rica",
                "CR"
            )
            
            # Solo carga de archivo de stock (sin ventas)
            archivo_costa_rica_temp = data_loader.cargar_archivo("📁 Subir archivo COSTA_RICA.csv", "COSTA_RICA_TEMP")
            
            if archivo_costa_rica_temp is not None:
                if hasattr(archivo_costa_rica_temp, 'name'):
                    st.session_state.archivo_costa_rica_temp_name = archivo_costa_rica_temp.name
                
                # Solo mostrar tabla de MVPs (igual que Guatemala, El Salvador y Honduras)
                mostrar_stock_mvps_costarica(archivo_costa_rica_temp, "_temp_costarica")
        
        # SUB-PESTAÑA PANAMÁ TEMPORAL
        with sub_tab_pa:
            professional_design.create_section_header(
                "Análisis Temporal - Panamá", 
                "Espacio adicional para análisis completo de Panamá",
                "PA"
            )
            
            # Solo carga de archivo de stock (sin ventas)
            archivo_panama_temp = data_loader.cargar_archivo("📁 Subir archivo PANAMA.csv", "PANAMA_TEMP")
            
            if archivo_panama_temp is not None:
                if hasattr(archivo_panama_temp, 'name'):
                    st.session_state.archivo_panama_temp_name = archivo_panama_temp.name
                
                # Solo mostrar tabla de MVPs (igual que Guatemala, El Salvador, Honduras y Costa Rica)
                mostrar_stock_mvps_panama(archivo_panama_temp, "_temp_panama")
        
        # SUB-PESTAÑA PUERTO RICO TEMPORAL
        with sub_tab_pr:
            professional_design.create_section_header(
                "Análisis Temporal - Puerto Rico", 
                "Espacio adicional para análisis completo de Puerto Rico",
                "PR"
            )
            
            # Solo carga de archivo de stock (sin ventas)
            archivo_puerto_rico_temp = data_loader.cargar_archivo("📁 Subir archivo PUERTO_RICO.csv", "PUERTO_RICO_TEMP")
            
            if archivo_puerto_rico_temp is not None:
                if hasattr(archivo_puerto_rico_temp, 'name'):
                    st.session_state.archivo_puerto_rico_temp_name = archivo_puerto_rico_temp.name
                
                # Solo mostrar tabla de MVPs (igual que otros países)
                mostrar_stock_mvps_puerto_rico(archivo_puerto_rico_temp, "_temp_puerto_rico")
        

if __name__ == "__main__":
    main()